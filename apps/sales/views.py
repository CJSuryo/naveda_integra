"""Sales views — restructured with SalesEntitasBisnis groups (like purchase module)."""
import json
from decimal import Decimal, InvalidOperation

from naveda_integra.json_utils import safe_json

from django.contrib import messages as dj_messages
from django.contrib.auth.decorators import login_required
from django_ratelimit.decorators import ratelimit

from naveda_integra.ratelimit_utils import rate_from
from django.db import transaction
from django.db.models import Q, Sum
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from apps.entitas_bisnis.models import EntitasBisnis
from apps.inventory.models import InventoryRecord, Warehouse
from apps.master_data.models import Akun
from apps.master_data.utils import get_akun_sorted
from apps.purchase.models import ItemMasterPurchase, SubTransactionType
from apps.purchase.views import _get_eb_tree, _resolve_eb_lv1_ids, _get_warehouses_data, _get_item_uoms_data
from apps.uom.conversion import (
    convert_input_to_base, resolvable_uoms_for_item, to_stock_uom,
)
from apps.uom.models import UnitOfMeasure

from .models import SalesHeader, SalesEntitasBisnis, SalesItem, SalesTaxLine, SalesItemFIFOAllocation, SalesEventLog
from .services import (
    get_available_stock,
    get_fifo_unit_cost,
    process_sales_fifo,
    create_sales_automated_journals,
    reverse_sales_automated_journals,
    reverse_sales_fifo,
    _cancel_sales_pajak,
)


def _get_eb_dropdown_options(user) -> list[dict[str, str]]:
    """Build hierarchical EntitasBisnis dropdown options (lv1/lv2/lv3),
    scoped to the entities the user may access."""
    from apps.entitas_bisnis.models import EntitasBisnisLv2, EntitasBisnisLv3
    from apps.purchase.views import accessible_eb_lv1_ids

    allowed = accessible_eb_lv1_ids(user)
    lv1_list = list(EntitasBisnis.objects.filter(status_aktif=True).order_by('nama'))
    lv2_list = list(
        EntitasBisnisLv2.objects.filter(status_aktif=True)
        .select_related('entitas_bisnis').order_by('nama')
    )
    lv3_list = list(
        EntitasBisnisLv3.objects.filter(status_aktif=True)
        .select_related('parent_lv2__entitas_bisnis').order_by('nama')
    )
    if allowed is not None:
        lv1_list = [eb for eb in lv1_list if eb.pk in allowed]
        lv2_list = [lv2 for lv2 in lv2_list if lv2.entitas_bisnis_id in allowed]
        lv3_list = [lv3 for lv3 in lv3_list if lv3.parent_lv2.entitas_bisnis_id in allowed]

    lv2_by_lv1: dict[int, list] = {}
    for lv2 in lv2_list:
        lv2_by_lv1.setdefault(lv2.entitas_bisnis_id, []).append(lv2)
    lv3_by_lv2: dict[int, list] = {}
    for lv3 in lv3_list:
        lv3_by_lv2.setdefault(lv3.parent_lv2_id, []).append(lv3)

    options: list[dict[str, str]] = []
    for eb in lv1_list:
        options.append({'value': f'lv1:{eb.pk}', 'label': eb.nama})
        for lv2 in lv2_by_lv1.get(eb.pk, []):
            options.append({'value': f'lv2:{lv2.pk}', 'label': f'\u00a0\u00a0\u00a0\u00a0↳ {lv2.nama}'})
            for lv3 in lv3_by_lv2.get(lv2.pk, []):
                options.append({'value': f'lv3:{lv3.pk}', 'label': f'\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0↳ {lv3.nama}'})
    return options


def _resolve_eb_selection(selection: str | int | None, user) -> dict | None:
    """Thin wrapper over the scoped resolver in purchase.views, kept so existing
    sales imports/usages keep working. Rejects entities the user cannot access."""
    from apps.purchase.views import _resolve_eb_selection as _impl
    return _impl(selection, user)


# ── Sales List ───────────────────────────────────────────────────────────────

@login_required
def sales_list(request: HttpRequest) -> HttpResponse:
    """List all sales transactions with filtering."""
    qs = (
        SalesHeader.objects
        .prefetch_related(
            'entitas_groups__entitas_bisnis',
            'entitas_groups__payment_account',
            'entitas_groups__items__item',
            'entitas_groups__items__sub_transaction_type',
        )
        .order_by('-tanggal', '-created_at')
    )

    tanggal_dari = request.GET.get('tanggal_dari', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    item_filter = request.GET.get('item', '')
    stt_filter = request.GET.get('sub_transaction_type', '')
    eb_filter_list = [v for v in request.GET.getlist('entitas_bisnis') if v]

    if tanggal_dari:
        qs = qs.filter(tanggal__gte=tanggal_dari)
    if tanggal_sampai:
        qs = qs.filter(tanggal__lte=tanggal_sampai)
    if item_filter:
        qs = qs.filter(entitas_groups__items__item_id=item_filter).distinct()
    if stt_filter:
        qs = qs.filter(entitas_groups__items__sub_transaction_type_id=stt_filter).distinct()
    if eb_filter_list:
        qs = qs.filter(entitas_groups__entitas_bisnis_id__in=_resolve_eb_lv1_ids(eb_filter_list, request.user)).distinct()

    # Build flat rows
    rows = []
    for sh in qs:
        for eg in sh.entitas_groups.all():
            for si in eg.items.all():
                if stt_filter and str(si.sub_transaction_type_id) != str(stt_filter):
                    continue
                if item_filter and str(si.item_id) != str(item_filter):
                    continue
                # Compute HPP & profit
                is_bulk = si.item.tipe_item in ('RMB', 'FGB', 'ITMB') if si.item else False
                hpp = si.hpp_terpakai if is_bulk and si.hpp_terpakai else (si.cogs_amount or Decimal('0'))
                profit = (si.total_sales or Decimal('0')) - hpp
                rows.append({
                    'sales_header': sh,
                    'entitas_group': eg,
                    'item': si,
                    'hpp': hpp,
                    'profit': profit,
                })

    # Compute rowspan
    trx_rowspan: dict[str, int] = {}
    trx_seen: dict[str, bool] = {}
    for row in rows:
        tid = row['sales_header'].transaction_id
        trx_rowspan[tid] = trx_rowspan.get(tid, 0) + 1
    for row in rows:
        tid = row['sales_header'].transaction_id
        if tid not in trx_seen:
            trx_seen[tid] = True
            row['show_trx_cell'] = True
            row['trx_rowspan'] = trx_rowspan[tid]
        else:
            row['show_trx_cell'] = False
            row['trx_rowspan'] = 0

    inventory_items = ItemMasterPurchase.objects.filter(
        tipe_item__in=['RM', 'FG', 'ITM', 'RMB', 'FGB', 'ITMB']
    ).order_by('nama')

    return render(request, 'sales/sales_list.html', {
        'rows': rows,
        'tanggal_dari': tanggal_dari,
        'tanggal_sampai': tanggal_sampai,
        'items': inventory_items,
        'sub_transaction_types': SubTransactionType.objects.filter(module='sales').order_by('nama'),
        'eb_tree': _get_eb_tree(request.user),
        'item_filter': item_filter,
        'stt_filter': stt_filter,
        'eb_filter_list': eb_filter_list,
    })


@login_required
@ratelimit(key='user', rate=rate_from('export'), method='GET', block=True)
def sales_export(request: HttpRequest) -> HttpResponse:
    """Export sales list as XLSX with same filters as the list page."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    qs = (
        SalesHeader.objects
        .prefetch_related(
            'entitas_groups__entitas_bisnis',
            'entitas_groups__items__item',
            'entitas_groups__items__sub_transaction_type',
        )
        .order_by('-tanggal', '-created_at')
    )

    tanggal_dari = request.GET.get('tanggal_dari', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    item_filter = request.GET.get('item', '')
    stt_filter = request.GET.get('sub_transaction_type', '')
    eb_filter = request.GET.get('entitas_bisnis', '')

    if tanggal_dari:
        qs = qs.filter(tanggal__gte=tanggal_dari)
    if tanggal_sampai:
        qs = qs.filter(tanggal__lte=tanggal_sampai)
    if item_filter:
        qs = qs.filter(entitas_groups__items__item_id=item_filter).distinct()
    if stt_filter:
        qs = qs.filter(entitas_groups__items__sub_transaction_type_id=stt_filter).distinct()
    if eb_filter:
        qs = qs.filter(entitas_groups__entitas_bisnis_id=eb_filter).distinct()

    rows = []
    for sh in qs:
        for eg in sh.entitas_groups.all():
            for si in eg.items.all():
                if stt_filter and str(si.sub_transaction_type_id) != str(stt_filter):
                    continue
                if item_filter and str(si.item_id) != str(item_filter):
                    continue
                is_bulk = si.item.tipe_item in ('RMB', 'FGB', 'ITMB') if si.item else False
                hpp = si.hpp_terpakai if is_bulk and si.hpp_terpakai else (si.cogs_amount or Decimal('0'))
                profit = (si.total_sales or Decimal('0')) - hpp
                rows.append({
                    'transaction_id': sh.transaction_id or '',
                    'tanggal': str(sh.tanggal),
                    'item': str(si.item) if si.item else '-',
                    'stt': si.sub_transaction_type.nama if si.sub_transaction_type else '-',
                    'qty': float(si.quantity or 0),
                    'selling_price': float(si.selling_price or 0),
                    'total_sales': float(si.total_sales or 0),
                    'hpp': float(hpp),
                    'profit': float(profit),
                    'entitas_bisnis': eg.entitas_bisnis.nama if eg.entitas_bisnis else '-',
                })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales'

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    right_align = Alignment(horizontal='right')

    col_headers = ['Transaction ID', 'Tanggal', 'Item', 'Sub-Transaction Type', 'Qty',
                   'Harga Jual', 'Total', 'HPP', 'Profit', 'Entitas Bisnis']
    for col_idx, title in enumerate(col_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    for row_num, row in enumerate(rows, 2):
        vals = [row['transaction_id'], row['tanggal'], row['item'], row['stt'], row['qty'],
                row['selling_price'], row['total_sales'], row['hpp'], row['profit'], row['entitas_bisnis']]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx in (5, 6, 7, 8, 9):
                cell.alignment = right_align
                cell.number_format = '#,##0'

    # Auto column width
    if rows:
        for col_idx in range(1, len(col_headers) + 1):
            max_len = len(col_headers[col_idx - 1])
            for ws_row in ws.iter_rows(min_row=2, max_row=len(rows) + 1, min_col=col_idx, max_col=col_idx):
                for cell in ws_row:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales.xlsx"'
    wb.save(response)
    return response


@login_required
@ratelimit(key='user', rate=rate_from('export'), method='GET', block=True)
def sales_export_pdf(request: HttpRequest) -> HttpResponse:
    """Render a print-friendly sales list for PDF export via browser print dialog."""
    qs = (
        SalesHeader.objects
        .prefetch_related(
            'entitas_groups__entitas_bisnis',
            'entitas_groups__items__item',
            'entitas_groups__items__sub_transaction_type',
        )
        .order_by('-tanggal', '-created_at')
    )

    tanggal_dari = request.GET.get('tanggal_dari', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    item_filter = request.GET.get('item', '')
    stt_filter = request.GET.get('sub_transaction_type', '')
    eb_filter = request.GET.get('entitas_bisnis', '')

    if tanggal_dari:
        qs = qs.filter(tanggal__gte=tanggal_dari)
    if tanggal_sampai:
        qs = qs.filter(tanggal__lte=tanggal_sampai)
    if item_filter:
        qs = qs.filter(entitas_groups__items__item_id=item_filter).distinct()
    if stt_filter:
        qs = qs.filter(entitas_groups__items__sub_transaction_type_id=stt_filter).distinct()
    if eb_filter:
        qs = qs.filter(entitas_groups__entitas_bisnis_id=eb_filter).distinct()

    rows = []
    total_sales_sum = Decimal('0')
    total_hpp_sum = Decimal('0')
    total_profit_sum = Decimal('0')

    for sh in qs:
        for eg in sh.entitas_groups.all():
            for si in eg.items.all():
                if stt_filter and str(si.sub_transaction_type_id) != str(stt_filter):
                    continue
                if item_filter and str(si.item_id) != str(item_filter):
                    continue
                is_bulk = si.item.tipe_item in ('RMB', 'FGB', 'ITMB') if si.item else False
                hpp = si.hpp_terpakai if is_bulk and si.hpp_terpakai else (si.cogs_amount or Decimal('0'))
                profit = (si.total_sales or Decimal('0')) - hpp
                total_sales_sum += si.total_sales or Decimal('0')
                total_hpp_sum += hpp
                total_profit_sum += profit
                rows.append({
                    'sales_header': sh,
                    'entitas_group': eg,
                    'item': si,
                    'hpp': hpp,
                    'profit': profit,
                })

    return render(request, 'sales/sales_export_pdf.html', {
        'rows': rows,
        'tanggal_dari': tanggal_dari,
        'tanggal_sampai': tanggal_sampai,
        'eb_filter': eb_filter,
        'total_sales_sum': total_sales_sum,
        'total_hpp_sum': total_hpp_sum,
        'total_profit_sum': total_profit_sum,
        'generated_at': timezone.now(),
    })


# ── Sales Create ─────────────────────────────────────────────────────────────

@login_required
def sales_create(request: HttpRequest) -> HttpResponse:
    """Create a new sales transaction."""
    if request.method == 'POST':
        return _handle_sales_save(request)

    inventory_items = ItemMasterPurchase.objects.filter(
        tipe_item__in=['RM', 'FG', 'ITM', 'RMB', 'FGB', 'ITMB']
    ).select_related('coa_account').order_by('nama')

    return render(request, 'sales/sales_form.html', {
        'title': 'Tambah Penjualan',
        'today': timezone.now().date(),
        'items_master': inventory_items,
        'sub_transaction_types': SubTransactionType.objects.filter(module='sales').order_by('nama'),
        'akun_list': get_akun_sorted(),
        'eb_options': _get_eb_dropdown_options(request.user),
        'warehouses_json': safe_json(_get_warehouses_data()),
        'item_uoms_json': safe_json(_get_item_uoms_data('sales')),
    })


# ── Sales Update ─────────────────────────────────────────────────────────────

@login_required
def sales_update(request: HttpRequest, pk: int) -> HttpResponse:
    """Edit an existing sales transaction."""
    sales = get_object_or_404(SalesHeader, pk=pk)
    if sales.is_locked:
        dj_messages.error(request, 'Transaksi ini sudah di-lock dan tidak bisa diedit.')
        return redirect('sales:detail', pk=pk)

    if request.method == 'POST':
        return _handle_sales_save(request, existing=sales)

    # Build existing EB groups data for JS
    eb_groups_data = []
    for eg in sales.entitas_groups.select_related(
        'entitas_bisnis', 'entitas_bisnis_lv2', 'entitas_bisnis_lv3', 'payment_account',
    ).all():
        # Determine the eb_selection value
        if eg.entitas_bisnis_lv3_id:
            eb_val = f'lv3:{eg.entitas_bisnis_lv3_id}'
        elif eg.entitas_bisnis_lv2_id:
            eb_val = f'lv2:{eg.entitas_bisnis_lv2_id}'
        else:
            eb_val = f'lv1:{eg.entitas_bisnis_id}'

        items_data = []
        for si in eg.items.select_related(
            'item', 'sub_transaction_type', 'offset_coa_account',
            'revenue_account', 'payment_account', 'tax_account', 'tax_payment_account',
        ).prefetch_related('tax_lines').all():
            items_data.append({
                'item_id': si.item_id,
                'item_name': f'{si.item.item_id} - {si.item.nama}',
                'tipe_item': si.item.tipe_item,
                'sub_transaction_type_id': si.sub_transaction_type_id,
                'quantity': str(si.input_qty) if si.input_qty is not None else str(si.quantity),
                'selling_price': (
                    str((si.selling_price * si.quantity) / si.input_qty)
                    if si.input_qty else str(si.selling_price)
                ),
                'hpp_terpakai': str(si.hpp_terpakai) if si.hpp_terpakai else str(si.cogs_amount or ''),
                'offset_coa_account_id': si.offset_coa_account_id,
                'revenue_account_id': si.revenue_account_id,
                'payment_account_id': si.payment_account_id or eg.payment_account_id or '',
                'tax': str(si.tax) if si.tax else '',
                'tax_type': si.tax_type or '',
                'tax_account_id': si.tax_account_id or '',
                'tax_payment': si.tax_payment or '',
                'tax_payment_account_id': si.tax_payment_account_id or '',
                'warehouse_id': si.warehouse_id or '',
                'input_uom_id': si.input_uom_id or '',
                # New: tax lines array
                'tax_lines': [
                    {
                        'tax_type': tl.tax_type,
                        'tax': str(tl.tax) if tl.tax else '',
                        'is_manual': tl.is_manual,
                        'tax_account_id': tl.tax_account_id,
                        'tax_payment_account_id': tl.tax_payment_account_id,
                    }
                    for tl in si.tax_lines.all()
                ],
            })

        eb_groups_data.append({
            'eb_selection': eb_val,
            'payment_account_id': eg.payment_account_id,
            'items': items_data,
        })

    inventory_items = ItemMasterPurchase.objects.filter(
        tipe_item__in=['RM', 'FG', 'ITM', 'RMB', 'FGB', 'ITMB']
    ).select_related('coa_account').order_by('nama')

    return render(request, 'sales/sales_form.html', {
        'title': 'Edit Penjualan',
        'today': sales.tanggal,
        'sales': sales,
        'items_master': inventory_items,
        'sub_transaction_types': SubTransactionType.objects.filter(module='sales').order_by('nama'),
        'akun_list': get_akun_sorted(),
        'eb_options': _get_eb_dropdown_options(request.user),
        'eb_groups_json': safe_json(eb_groups_data),
        'warehouses_json': safe_json(_get_warehouses_data()),
        'item_uoms_json': safe_json(_get_item_uoms_data('sales')),
    })


# ── Sales Detail ─────────────────────────────────────────────────────────────

@login_required
def sales_detail(request: HttpRequest, pk: int) -> HttpResponse:
    """View sales transaction detail."""
    sales = get_object_or_404(SalesHeader, pk=pk)
    eb_groups = sales.entitas_groups.select_related(
        'entitas_bisnis', 'entitas_bisnis_lv2', 'entitas_bisnis_lv3', 'payment_account',
    ).prefetch_related(
        'items__item', 'items__sub_transaction_type',
        'items__offset_coa_account', 'items__revenue_account',
        'items__payment_account',
        'items__tax_account', 'items__tax_payment_account',
        'items__tax_lines__tax_account', 'items__tax_lines__tax_payment_account',
    ).all()

    # Build inventory mutations from per-batch FIFO allocations
    inventory_mutations = []
    allocations = (
        SalesItemFIFOAllocation.objects
        .filter(sales_item__sales_eb__sales_header=sales)
        .select_related(
            'inventory_record__item',
            'inventory_record__entitas_bisnis',
            'sales_item__sales_eb__entitas_bisnis',
            'sales_item__item',
        )
        .order_by('inventory_record__tanggal', 'inventory_record__inventory_number')
    )
    for alloc in allocations:
        inv_rec = alloc.inventory_record
        is_bulk_alloc = alloc.sales_item.item.tipe_item in ('RMB', 'FGB', 'ITMB')
        inventory_mutations.append({
            'inventory_number': inv_rec.inventory_number,
            'inventory_pk': inv_rec.pk,
            'item': str(inv_rec.item),
            'entitas_bisnis': alloc.sales_item.sales_eb.entitas_bisnis.nama,
            'quantity_sold': alloc.quantity_consumed,
            'cogs': alloc.cogs_amount,
            'is_bulk': is_bulk_alloc,
        })

    event_logs = sales.event_logs.select_related('actor').order_by('timestamp')

    from apps.jurnal.models import JurnalHeader
    uraian_match = f'Penjualan {sales.transaction_id} —'
    journals = list(
        JurnalHeader.objects
        .filter(uraian_transaksi__startswith=uraian_match, is_penyesuaian=False)
        .prefetch_related('details__akun')
        .order_by('tanggal', 'id')
    )
    for jh in journals:
        jh.source_label = 'sales'

    from apps.pajak.models import PajakTransaksi
    si_ids = [si.pk for eg in eb_groups for si in eg.items.all()]
    pajak_jurnal_ids = []
    if si_ids:
        pajak_jurnal_ids = list(
            PajakTransaksi.objects
            .filter(source_type='sales_item', source_id__in=si_ids, jurnal_header__isnull=False)
            .values_list('jurnal_header_id', flat=True)
        )
    if pajak_jurnal_ids:
        pajak_journals = list(
            JurnalHeader.objects
            .filter(pk__in=pajak_jurnal_ids)
            .prefetch_related('details__akun')
        )
        for jh in pajak_journals:
            jh.source_label = 'pajak'
        journals = sorted(journals + pajak_journals, key=lambda j: (j.tanggal, j.id))

    return render(request, 'sales/sales_detail.html', {
        'sales': sales,
        'eb_groups': eb_groups,
        'inventory_mutations': inventory_mutations,
        'event_logs': event_logs,
        'journals': journals,
    })


# ── Sales Invoice ─────────────────────────────────────────────────────────────

@login_required
def sales_invoice(request: HttpRequest, pk: int) -> HttpResponse:
    """Render a printable invoice/receipt for a sales transaction."""
    from apps.pajak.models import PajakTransaksi
    from apps.pajak.services import compute_pajak
    from .services import SIFAT_PAJAK_MAP, TAX_TYPE_MAP
    sales = get_object_or_404(SalesHeader, pk=pk)
    eb_groups = sales.entitas_groups.select_related(
        'entitas_bisnis', 'entitas_bisnis_lv2', 'entitas_bisnis_lv3', 'payment_account',
    ).prefetch_related(
        'items__item', 'items__sub_transaction_type',
        'items__offset_coa_account', 'items__revenue_account',
        'items__payment_account',
        'items__tax_account', 'items__tax_payment_account',
        'items__tax_lines',
    ).all()

    company = EntitasBisnis.objects.filter(is_company_profile=True).first()

    # Nominal pajak untuk baris otomatis (is_manual=False) tidak disimpan di
    # SalesTaxLine.tax — hanya di PajakTransaksi yang dibuat saat konfirmasi.
    si_ids = [si.pk for eg in eb_groups for si in eg.items.all()]
    pajak_by_key = {}
    for pt in PajakTransaksi.objects.filter(
        source_type='sales_item', source_id__in=si_ids,
    ).exclude(status='dibatalkan'):
        pajak_by_key[(pt.source_id, pt.jenis_pajak)] = pt.jumlah_pajak

    def _tax_amount(si, tl):
        """Nominal pajak efektif: override manual → PajakTransaksi → hitung dari tarif."""
        if tl.is_manual and tl.tax:
            return tl.tax
        jenis_pajak = TAX_TYPE_MAP.get(tl.tax_type)
        if not jenis_pajak:
            return tl.tax or Decimal('0')
        amount = pajak_by_key.get((si.pk, jenis_pajak))
        if amount is not None:
            return amount
        if tl.tax:
            return tl.tax
        # Belum dikonfirmasi (belum ada PajakTransaksi) → hitung dari tarif berlaku.
        try:
            return compute_pajak(
                jenis_pajak, si.total_sales or Decimal('0'), sales.tanggal,
            )['jumlah_pajak']
        except Exception:
            # Tarif belum tersedia untuk periode ini — jangan gagalkan invoice.
            return Decimal('0')

    # Pre-compute per-group totals for the template
    eb_groups_with_totals = []
    grand_total = Decimal('0')
    grand_tax = Decimal('0')
    for eg in eb_groups:
        subtotal = Decimal('0')
        pungut_total = Decimal('0')   # PPN keluaran — ditambahkan ke tagihan pelanggan
        potong_total = Decimal('0')   # PPh dipotong pelanggan — pengurang kas yang dibayar
        items_data = []
        for si in eg.items.all():
            subtotal += si.total_sales or Decimal('0')
            si_pungut = Decimal('0')
            si_potong = Decimal('0')
            # Use tax_lines (new) or fallback to deprecated si.tax (old records)
            tax_lines_list = list(si.tax_lines.all())
            if tax_lines_list:
                for tl in tax_lines_list:
                    amt = _tax_amount(si, tl) or Decimal('0')
                    tl.amount = amt
                    tl.is_dipotong = SIFAT_PAJAK_MAP.get(tl.tax_type) == 'prepaid'
                    if tl.is_dipotong:
                        si_potong += amt
                    else:
                        si_pungut += amt
            elif si.tax:
                if SIFAT_PAJAK_MAP.get(si.tax_type) == 'prepaid':
                    si_potong += si.tax
                else:
                    si_pungut += si.tax
            pungut_total += si_pungut
            potong_total += si_potong
            if si.payment_account_id is None:
                payment_label = '-'
            elif si.payment_account.is_kas_setara:
                payment_label = 'Kas'
            else:
                payment_label = 'Kredit'
            items_data.append({
                'si': si,
                'tax_lines': tax_lines_list,
                'si_pungut': si_pungut,
                'si_potong': si_potong,
                'payment_label': payment_label,
            })
        tax_total = pungut_total - potong_total
        group_total = subtotal + tax_total
        grand_total += group_total
        grand_tax += tax_total
        group_payment_labels = {d['payment_label'] for d in items_data if d['payment_label'] != '-'}
        if not group_payment_labels:
            group_payment_label = '-'
        elif group_payment_labels == {'Kas'}:
            group_payment_label = 'Kas'
        elif group_payment_labels == {'Kredit'}:
            group_payment_label = 'Kredit'
        else:
            group_payment_label = 'Kas dan Kredit'
        eb_groups_with_totals.append({
            'group': eg,
            'items_data': items_data,
            'subtotal': subtotal,
            'pungut_total': pungut_total,
            'potong_total': potong_total,
            'tax_total': tax_total,
            'group_total': group_total,
            'payment_label': group_payment_label,
        })

    if sales.payment_type == 'cash':
        lunas_status = 'Lunas'
    else:
        piutang = sales.piutang_headers.order_by('-tanggal', '-id').first()
        lunas_status = 'Lunas' if piutang and piutang.status == 'paid' else 'Belum Lunas'

    return render(request, 'sales/sales_invoice.html', {
        'sales': sales,
        'eb_groups_data': eb_groups_with_totals,
        'company': company,
        'grand_total': grand_total,
        'grand_tax': grand_tax,
        'lunas_status': lunas_status,
    })


# ── Sales Delete ─────────────────────────────────────────────────────────────

@login_required
def sales_delete(request: HttpRequest, pk: int) -> HttpResponse:
    """Delete a sales transaction, reversing journals and FIFO."""
    from apps.jurnal.models import JurnalHeader
    from apps.jurnal.utils import log_jurnal_terhapus

    sales = get_object_or_404(SalesHeader, pk=pk)
    if sales.is_locked:
        dj_messages.error(request, 'Transaksi ini sudah di-lock dan tidak bisa dihapus.')
        return redirect('sales:detail', pk=pk)

    # Find journals that will be deleted
    uraian_match = f'Penjualan {sales.transaction_id} \u2014'
    related_journals = list(
        JurnalHeader.objects.filter(
            uraian_transaksi__startswith=uraian_match,
            is_penyesuaian=False,
        ).order_by('nomor_transaksi')
    )

    if request.method == 'POST':
        tid = sales.transaction_id
        with transaction.atomic():
            for journal in related_journals:
                log_jurnal_terhapus(journal, 'sales', request)
            _cancel_sales_pajak(sales)
            reverse_sales_automated_journals(sales)
            reverse_sales_fifo(sales)
            SalesEventLog.objects.create(
                sales_header=sales,
                event_type='VOIDED',
                description=f'Transaksi {tid} dihapus.',
                actor=request.user,
            )
            sales.delete()
        dj_messages.success(request, f'Sales {tid} berhasil dihapus.')
        return redirect('sales:list')

    return render(request, 'sales/sales_delete.html', {
        'sales': sales,
        'related_journals': related_journals,
    })


# ── API Endpoints ────────────────────────────────────────────────────────────

@login_required
def api_stock_check(request: HttpRequest) -> JsonResponse:
    """Check available stock for an item."""
    item_id = request.GET.get('item_id', '')
    if not item_id:
        return JsonResponse({'available_stock': '0', 'is_bulk': False, 'unit_cost': '0'})
    try:
        item = ItemMasterPurchase.objects.only('tipe_item').get(pk=int(item_id))
        is_bulk = item.tipe_item in ('RMB', 'FGB', 'ITMB')
        if is_bulk:
            # For bulk items, return total inventory value
            from django.db.models import Sum
            total_value = (
                InventoryRecord.objects
                .filter(item_id=int(item_id), total_value__gt=0)
                .aggregate(total=Sum('total_value'))['total'] or 0
            )
            return JsonResponse({
                'available_stock': str(total_value),
                'is_bulk': True,
                'unit_cost': '0',
            })
        else:
            stock = get_available_stock(int(item_id))
            unit_cost = get_fifo_unit_cost(int(item_id))
            return JsonResponse({
                'available_stock': str(stock),
                'is_bulk': False,
                'unit_cost': str(unit_cost),
            })
    except (ValueError, TypeError, ItemMasterPurchase.DoesNotExist):
        return JsonResponse({'available_stock': '0', 'is_bulk': False, 'unit_cost': '0'})


@login_required
def api_stt_offset(request: HttpRequest) -> JsonResponse:
    """Return default offset account for a sales sub-transaction type."""
    stt_id = request.GET.get('stt_id', '')
    if not stt_id:
        return JsonResponse({'offset_account_id': '', 'offset_account_text': ''})
    try:
        stt = SubTransactionType.objects.select_related('default_offset_account').get(
            pk=stt_id, module='sales',
        )
        return JsonResponse({
            'offset_account_id': stt.default_offset_account_id,
            'offset_account_text': str(stt.default_offset_account),
        })
    except SubTransactionType.DoesNotExist:
        return JsonResponse({'offset_account_id': '', 'offset_account_text': ''})


@login_required
def api_stt_defaults(request: HttpRequest) -> JsonResponse:
    """Return all default accounts for a sales SubTransactionType (for POS auto-fill)."""
    stt_id = request.GET.get('stt_id', '')
    if not stt_id:
        return JsonResponse({'ok': False})
    try:
        stt = SubTransactionType.objects.select_related(
            'default_offset_account',
            'default_revenue_account',
            'default_payment_account',
            'default_inventory_account',
            'default_tax_account',
        ).get(pk=stt_id, module='sales')
        return JsonResponse({
            'ok': True,
            'offset_account_id': stt.default_offset_account_id or '',
            'revenue_account_id': stt.default_revenue_account_id or '',
            'payment_account_id': stt.default_payment_account_id or '',
            'inventory_account_id': stt.default_inventory_account_id or '',
            'tax_account_id': stt.default_tax_account_id or '',
            'tax_type': stt.default_tax_type or '',
        })
    except (SubTransactionType.DoesNotExist, ValueError, TypeError):
        return JsonResponse({'ok': False})


# ── Internal Helpers ─────────────────────────────────────────────────────────

def _handle_sales_save(request: HttpRequest, existing: SalesHeader | None = None) -> HttpResponse:
    """Process the sales form submission (create or update)."""
    tanggal_str = request.POST.get('tanggal', '')
    deskripsi = request.POST.get('deskripsi', '')
    eb_groups_json = request.POST.get('eb_groups_data', '[]')

    try:
        eb_groups_list = json.loads(eb_groups_json)
    except (ValueError, TypeError):
        eb_groups_list = []

    # Validation
    errors: dict[str, str] = {}

    tanggal = None
    if tanggal_str:
        try:
            from datetime import date as dt_date
            tanggal = dt_date.fromisoformat(tanggal_str)
        except ValueError:
            errors['tanggal'] = 'Tanggal tidak valid.'
    else:
        errors['tanggal'] = 'Tanggal wajib diisi.'

    if not eb_groups_list:
        errors['eb_groups'] = 'Minimal satu Entitas Bisnis harus ditambahkan.'

    # Pre-compute total demand per (item, eb) for cross-row stock validation.
    # Keyed by (item_id, eb_lv1_id, eb_lv2_id, eb_lv3_id) to match the EB
    # isolation enforced by process_sales_fifo / apps.inventory.ledger.
    item_demands: dict[tuple, Decimal] = {}
    all_items_flat: list[dict] = []

    for g_idx, group in enumerate(eb_groups_list):
        eb_sel = group.get('eb_selection', '')
        if not eb_sel:
            errors[f'group_{g_idx}_eb'] = f'Grup {g_idx + 1}: Entitas Bisnis wajib dipilih.'
            continue

        eb_resolved = _resolve_eb_selection(eb_sel, request.user)
        if not eb_resolved:
            errors[f'group_{g_idx}_eb'] = f'Grup {g_idx + 1}: Entitas Bisnis tidak valid.'
            continue

        items = group.get('items', [])
        if not items:
            errors[f'group_{g_idx}_items'] = f'Grup {g_idx + 1}: Minimal satu item harus ditambahkan.'
            continue

        for i, item_data in enumerate(items):
            item_id = item_data.get('item_id')
            if not item_id:
                errors[f'group_{g_idx}_item_{i}_id'] = f'Grup {g_idx + 1}, Item {i + 1}: Item wajib dipilih.'
                continue

            is_bulk = item_data.get('is_bulk') == '1'

            if not item_data.get('payment_account_id'):
                errors[f'group_{g_idx}_item_{i}_payment'] = f'Grup {g_idx + 1}, Item {i + 1}: Payment Account wajib diisi.'

            try:
                if is_bulk:
                    hpp = Decimal(str(item_data.get('hpp_terpakai', 0)))
                    if hpp <= 0:
                        errors[f'group_{g_idx}_item_{i}_hpp'] = f'Grup {g_idx + 1}, Item {i + 1}: HPP Terpakai harus > 0.'
                    qty = Decimal('0')
                    # Bulk: selling_price is optional (for revenue entry)
                    try:
                        price = Decimal(str(item_data.get('selling_price', 0)))
                    except (InvalidOperation, ValueError):
                        price = Decimal('0')
                else:
                    qty = Decimal(str(item_data.get('quantity', 0)))
                    if qty <= 0:
                        errors[f'group_{g_idx}_item_{i}_qty'] = f'Grup {g_idx + 1}, Item {i + 1}: Quantity harus > 0.'
            except (InvalidOperation, ValueError):
                errors[f'group_{g_idx}_item_{i}_qty'] = f'Grup {g_idx + 1}, Item {i + 1}: Quantity tidak valid.'
                continue

            try:
                if not is_bulk:
                    price = Decimal(str(item_data.get('selling_price', 0)))
                    if price <= 0:
                        errors[f'group_{g_idx}_item_{i}_price'] = f'Grup {g_idx + 1}, Item {i + 1}: Harga jual harus > 0.'
            except (InvalidOperation, ValueError):
                errors[f'group_{g_idx}_item_{i}_price'] = f'Grup {g_idx + 1}, Item {i + 1}: Harga jual tidak valid.'

            # Stock validation (skip for bulk items — value-based tracking)
            if not is_bulk:
                iid = int(item_id)
                demand_key = (
                    iid, eb_resolved['lv1_id'],
                    eb_resolved.get('lv2_id'), eb_resolved.get('lv3_id'),
                )
                # Convert to the item's stock UOM before accumulating demand —
                # the raw input-unit qty understates demand whenever the row
                # uses a packaging UOM (conversion factor >= 1), which would
                # silently weaken this pre-validation check.
                demand_qty = qty
                demand_input_uom_id = item_data.get('input_uom_id') or None
                if demand_input_uom_id:
                    demand_input_uom = UnitOfMeasure.objects.filter(pk=demand_input_uom_id).first()
                    if demand_input_uom is not None:
                        demand_item_obj = ItemMasterPurchase.objects.filter(pk=iid).first()
                        if demand_item_obj is not None:
                            if demand_input_uom.pk not in {
                                    u.pk for u in resolvable_uoms_for_item(demand_item_obj)}:
                                errors[f'group_{g_idx}_item_{i}_uom'] = (
                                    f'Grup {g_idx + 1}, Item {i + 1}: satuan '
                                    f'{demand_input_uom.kode} tidak punya konversi ke '
                                    f'satuan stok item. Tetapkan konversi di master item dulu.'
                                )
                            else:
                                demand_qty = to_stock_uom(qty, demand_input_uom, demand_item_obj)
                item_demands[demand_key] = item_demands.get(demand_key, Decimal('0')) + demand_qty

            # Bulk inventory value validation
            if is_bulk:
                try:
                    hpp_val = Decimal(str(item_data.get('hpp_terpakai', 0)))
                    iid_bulk = int(item_id)
                    from django.db.models import Sum as DjSum
                    avail_val = InventoryRecord.objects.filter(
                        item_id=iid_bulk,
                        entitas_bisnis_id=eb_resolved['lv1_id'],
                    ).aggregate(total=DjSum('total_value'))['total'] or Decimal('0')
                    if existing:
                        for eg_old in existing.entitas_groups.all():
                            for si_old in eg_old.items.filter(item_id=iid_bulk):
                                if si_old.hpp_terpakai:
                                    avail_val += si_old.hpp_terpakai
                    if hpp_val > avail_val:
                        errors[f'group_{g_idx}_item_{i}_bulk_stock'] = (
                            f'Grup {g_idx + 1}, Item {i + 1}: '
                            f'Nilai yang dijual (Rp {hpp_val:,.0f}) melebihi '
                            f'Nilai Persediaan tersedia (Rp {avail_val:,.0f}).'
                        )
                except (InvalidOperation, ValueError, TypeError):
                    pass

            if not item_data.get('offset_coa_account_id'):
                errors[f'group_{g_idx}_item_{i}_offset'] = f'Grup {g_idx + 1}, Item {i + 1}: Offset CoA (HPP) wajib diisi.'
            if not item_data.get('revenue_account_id'):
                errors[f'group_{g_idx}_item_{i}_revenue'] = f'Grup {g_idx + 1}, Item {i + 1}: Revenue Account wajib diisi.'

            # Validate item has coa_account set (required for balanced COGS journal entry)
            try:
                item_obj = ItemMasterPurchase.objects.only('tipe_item', 'coa_account_id', 'nama').get(pk=int(item_id))
                if item_obj.tipe_item in ('RM', 'FG', 'ITM', 'RMB', 'FGB', 'ITMB') and not item_obj.coa_account_id:
                    errors[f'group_{g_idx}_item_{i}_coa'] = (
                        f'Grup {g_idx + 1}, Item {i + 1}: Item "{item_obj.nama}" belum memiliki '
                        'CoA Account (Persediaan). Harap set CoA Account di Item Master terlebih dahulu.'
                    )
            except (ItemMasterPurchase.DoesNotExist, ValueError, TypeError):
                pass

            all_items_flat.append(item_data)

    # Check stock availability — EB-isolated, matching the consumption scope
    # used by process_sales_fifo (apps.inventory.ledger.get_available_stock).
    # A stale global (non-EB-isolated) check here would pass validation for a
    # sale whose own EB branch lacks stock but a sibling EB has some, only to
    # blow up later as an unhandled InsufficientStockError inside the save
    # transaction.
    if item_demands:
        from apps.entitas_bisnis.models import EntitasBisnisLv2, EntitasBisnisLv3
        from apps.inventory.ledger import get_available_stock as ledger_available_stock

        eb_lv1_cache: dict[int, EntitasBisnis | None] = {}
        eb_lv2_cache: dict[int, EntitasBisnisLv2 | None] = {}
        eb_lv3_cache: dict[int, EntitasBisnisLv3 | None] = {}
        item_cache: dict[int, ItemMasterPurchase | None] = {}

        for (iid, lv1_id, lv2_id, lv3_id), total_demand in item_demands.items():
            if iid not in item_cache:
                item_cache[iid] = ItemMasterPurchase.objects.filter(pk=iid).first()
            item_obj = item_cache[iid]
            if item_obj is None:
                continue

            if lv1_id not in eb_lv1_cache:
                eb_lv1_cache[lv1_id] = EntitasBisnis.objects.filter(pk=lv1_id).first()
            eb_lv1_obj = eb_lv1_cache[lv1_id]
            if eb_lv1_obj is None:
                continue

            eb_lv2_obj = None
            if lv2_id:
                if lv2_id not in eb_lv2_cache:
                    eb_lv2_cache[lv2_id] = EntitasBisnisLv2.objects.filter(pk=lv2_id).first()
                eb_lv2_obj = eb_lv2_cache[lv2_id]

            eb_lv3_obj = None
            if lv3_id:
                if lv3_id not in eb_lv3_cache:
                    eb_lv3_cache[lv3_id] = EntitasBisnisLv3.objects.filter(pk=lv3_id).first()
                eb_lv3_obj = eb_lv3_cache[lv3_id]

            available = ledger_available_stock(item_obj, eb_lv1_obj, eb_lv2_obj, eb_lv3_obj)
            if existing:
                for eg in existing.entitas_groups.filter(
                    entitas_bisnis_id=lv1_id,
                    entitas_bisnis_lv2_id=lv2_id,
                    entitas_bisnis_lv3_id=lv3_id,
                ):
                    for old_si in eg.items.filter(item_id=iid):
                        available += old_si.quantity
            if total_demand > available:
                errors[f'item_stock_{iid}_{lv1_id}_{lv2_id or 0}_{lv3_id or 0}'] = (
                    f'Stok tidak mencukupi untuk "{item_obj.nama}" di entitas bisnis '
                    f'"{eb_lv3_obj.nama if eb_lv3_obj else (eb_lv2_obj.nama if eb_lv2_obj else eb_lv1_obj.nama)}". '
                    f'Total permintaan: {total_demand} unit, Stok tersedia: {available} unit.'
                )

    if errors:
        dj_messages.error(request, 'Terdapat kesalahan pada form. Silakan periksa kembali.')
        inventory_items = ItemMasterPurchase.objects.filter(
            tipe_item__in=['RM', 'FG', 'ITM', 'RMB', 'FGB', 'ITMB']
        ).select_related('coa_account').order_by('nama')
        return render(request, 'sales/sales_form.html', {
            'title': 'Edit Penjualan' if existing else 'Tambah Penjualan',
            'today': tanggal or timezone.now().date(),
            'sales': existing,
            'items_master': inventory_items,
            'sub_transaction_types': SubTransactionType.objects.filter(module='sales').order_by('nama'),
            'akun_list': get_akun_sorted(),
            'eb_options': _get_eb_dropdown_options(request.user),
            'errors': errors,
            'eb_groups_json': safe_json(eb_groups_list),
            'warehouses_json': safe_json(_get_warehouses_data()),
            'item_uoms_json': safe_json(_get_item_uoms_data('sales')),
        })

    # Save
    from apps.inventory.ledger import InsufficientStockError
    try:
        with transaction.atomic():
            if existing:
                _cancel_sales_pajak(existing)
                reverse_sales_automated_journals(existing)
                reverse_sales_fifo(existing)
                existing.entitas_groups.all().delete()
                existing.tanggal = tanggal
                existing.deskripsi = deskripsi
                existing.save()
                sales = existing
            else:
                sales = SalesHeader.objects.create(
                    tanggal=tanggal,
                    deskripsi=deskripsi,
                    created_by=request.user,
                )

            for group in eb_groups_list:
                eb_resolved = _resolve_eb_selection(group['eb_selection'], request.user)
                eb_group = SalesEntitasBisnis.objects.create(
                    sales_header=sales,
                    entitas_bisnis_id=eb_resolved['lv1_id'],
                    entitas_bisnis_lv2_id=eb_resolved.get('lv2_id'),
                    entitas_bisnis_lv3_id=eb_resolved.get('lv3_id'),
                    payment_account_id=None,  # now per-item on SalesItem
                )

                for item_data in group.get('items', []):
                    tax_val = item_data.get('tax')
                    tax_amount = Decimal(str(tax_val)) if tax_val else None
                    is_bulk = item_data.get('is_bulk') == '1'

                    wh_id = item_data.get('warehouse_id') or None
                    if wh_id and not Warehouse.objects.filter(
                            pk=wh_id, entitas_bisnis_id=eb_resolved['lv1_id']).exists():
                        raise ValueError('Gudang tidak valid untuk bisnis ini.')

                    input_uom_id = item_data.get('input_uom_id') or None
                    if input_uom_id and not UnitOfMeasure.objects.filter(pk=input_uom_id).exists():
                        raise ValueError('Satuan input tidak valid.')
                    input_uom = UnitOfMeasure.objects.filter(pk=input_uom_id).first() if input_uom_id else None
                    if is_bulk:
                        qty_base = Decimal('0')
                        input_qty_raw = None
                        selling_base = Decimal(str(item_data.get('selling_price') or '0'))
                        # Bulk items are value-based (not unit-based) and must
                        # never persist a UOM, regardless of what the client
                        # submitted — a real UOM with input_qty=None would be
                        # a misleading audit record.
                        input_uom = None
                    else:
                        item_obj = ItemMasterPurchase.objects.get(pk=item_data['item_id'])
                        input_qty_raw = Decimal(str(item_data['quantity']))
                        qty_base, selling_base = convert_input_to_base(
                            item_obj, input_uom, input_qty_raw,
                            Decimal(str(item_data.get('selling_price') or '0')))

                    si = SalesItem.objects.create(
                        sales_eb=eb_group,
                        item_id=item_data['item_id'],
                        sub_transaction_type_id=item_data['sub_transaction_type_id'],
                        quantity=qty_base,
                        selling_price=selling_base,
                        hpp_terpakai=Decimal(str(item_data['hpp_terpakai'])) if is_bulk else None,
                        offset_coa_account_id=item_data['offset_coa_account_id'],
                        revenue_account_id=item_data['revenue_account_id'],
                        payment_account_id=item_data.get('payment_account_id') or None,
                        tax=tax_amount,
                        tax_type=item_data.get('tax_type', ''),
                        tax_account_id=item_data.get('tax_account_id') or None,
                        tax_payment=item_data.get('tax_payment', ''),
                        tax_payment_account_id=item_data.get('tax_payment_account_id') or None,
                        warehouse_id=wh_id,
                        input_uom=input_uom,
                        input_qty=input_qty_raw,
                    )

                    # Buat SalesTaxLine dari tax_lines array (format baru)
                    for tl_data in item_data.get('tax_lines', []):
                        tl_tax_type = tl_data.get('tax_type', '')
                        if not tl_tax_type:
                            continue
                        tl_tax_raw = tl_data.get('tax')
                        tl_tax_val = Decimal(str(tl_tax_raw)) if tl_tax_raw else None
                        tl_is_manual = bool(tl_data.get('is_manual', False)) or bool(tl_tax_val)
                        tl_tax_account_id = tl_data.get('tax_account_id') or None
                        tl_tax_payment_account_id = tl_data.get('tax_payment_account_id') or None
                        if not tl_tax_account_id or not tl_tax_payment_account_id:
                            continue
                        SalesTaxLine.objects.create(
                            sales_item=si,
                            tax_type=tl_tax_type,
                            tax=tl_tax_val,
                            is_manual=tl_is_manual,
                            tax_account_id=tl_tax_account_id,
                            tax_payment_account_id=tl_tax_payment_account_id,
                        )

            # Process FIFO outflow
            reports = process_sales_fifo(sales)
            for rep in reports:
                if rep.used_fallback:
                    sumber = ', '.join(
                        f"{row['qty']} dari {row['eb_name']} ({row['level']})"
                        for row in rep.by_level if row['level'] != rep.requested_level)
                    dj_messages.warning(
                        request,
                        f'Stok di level {rep.requested_level} tidak mencukupi; '
                        f'sebagian diambil dari induk: {sumber}.')

            # Generate automated journals
            create_sales_automated_journals(sales, user=request.user)

            event_type = 'EDITED' if existing else 'CREATED'
            SalesEventLog.objects.create(
                sales_header=sales,
                event_type=event_type,
                description=f'Transaksi {sales.transaction_id} {"diperbarui" if existing else "dibuat"} via form.',
                actor=request.user,
            )
            SalesEventLog.objects.create(
                sales_header=sales,
                event_type='FIFO_PROCESSED',
                actor=None,
            )
            SalesEventLog.objects.create(
                sales_header=sales,
                event_type='JOURNAL_CREATED',
                actor=None,
            )
    except InsufficientStockError as exc:
        dj_messages.error(request, f'Stok tidak mencukupi untuk memproses transaksi ini: {exc}')
        inventory_items = ItemMasterPurchase.objects.filter(
            tipe_item__in=['RM', 'FG', 'ITM', 'RMB', 'FGB', 'ITMB']
        ).select_related('coa_account').order_by('nama')
        return render(request, 'sales/sales_form.html', {
            'title': 'Edit Penjualan' if existing else 'Tambah Penjualan',
            'today': tanggal or timezone.now().date(),
            'sales': existing,
            'items_master': inventory_items,
            'sub_transaction_types': SubTransactionType.objects.filter(module='sales').order_by('nama'),
            'akun_list': get_akun_sorted(),
            'eb_options': _get_eb_dropdown_options(request.user),
            'errors': {'stock': str(exc)},
            'eb_groups_json': safe_json(eb_groups_list),
            'warehouses_json': safe_json(_get_warehouses_data()),
            'item_uoms_json': safe_json(_get_item_uoms_data('sales')),
        })

    action = 'diperbarui' if existing else 'dibuat'
    dj_messages.success(request, f'Sales {sales.transaction_id} berhasil {action}.')
    return redirect('sales:detail', pk=sales.pk)


# ── POS Cashier ──────────────────────────────────────────────────────────────

@login_required
def pos_cashier(request: HttpRequest) -> HttpResponse:
    """POS cashier screen — mobile-friendly sales entry at /sales/pos/."""
    from apps.entitas_bisnis.models import EntitasBisnisLv3
    stores = EntitasBisnisLv3.objects.filter(status_aktif=True).select_related(
        'parent_lv2__entitas_bisnis'
    ).order_by('nama')
    return render(request, 'sales/pos_cashier.html', {
        'stores': stores,
        'sub_transaction_types': SubTransactionType.objects.filter(module='sales').order_by('nama'),
        'today': timezone.now().date(),
    })


@login_required
def api_pos_items(request: HttpRequest) -> JsonResponse:
    """Return inventory items for a given Lv3 store (for POS item grid)."""
    lv3_pk = request.GET.get('lv3_pk', '')
    if not lv3_pk:
        return JsonResponse({'items': []})
    try:
        lv3_pk = int(lv3_pk)
    except (ValueError, TypeError):
        return JsonResponse({'items': []})

    from apps.entitas_bisnis.models import EntitasBisnisLv3
    from pos_catalog.models import ProductModifierGroup

    # Get inventory records for this store (lv3 specific, then fall back to lv1)
    inv_qs = (
        InventoryRecord.objects
        .filter(entitas_bisnis_lv3_id=lv3_pk, quantity__gt=0)
        .exclude(item__tipe_item__in=['RMB', 'FGB', 'ITMB'])
        .select_related('item')
        .order_by('item__nama')
    )
    if not inv_qs.exists():
        # Fallback: get lv3 → lv2 → lv1, then filter inventory by lv1
        try:
            lv3 = EntitasBisnisLv3.objects.select_related('parent_lv2__entitas_bisnis').get(pk=lv3_pk)
            lv1_id = lv3.parent_lv2.entitas_bisnis_id
            inv_qs = (
                InventoryRecord.objects
                .filter(entitas_bisnis_id=lv1_id, quantity__gt=0)
                .exclude(item__tipe_item__in=['RMB', 'FGB', 'ITMB'])
                .select_related('item')
                .order_by('item__nama')
            )
        except EntitasBisnisLv3.DoesNotExist:
            return JsonResponse({'items': []})

    # Get modifier groups per item
    item_ids = list(inv_qs.values_list('item_id', flat=True).distinct())
    modifier_map: dict[int, list] = {}
    for pmg in (
        ProductModifierGroup.objects
        .filter(item_id__in=item_ids)
        .select_related('modifier_group')
        .prefetch_related('modifier_group__options')
    ):
        modifier_map.setdefault(pmg.item_id, []).append({
            'pk': pmg.modifier_group_id,
            'nama': pmg.modifier_group.name,
            'is_required': pmg.modifier_group.is_required,
            'min_selections': pmg.modifier_group.min_selections,
            'max_selections': pmg.modifier_group.max_selections,
            'options': [
                {
                    'pk': opt.pk,
                    'name': opt.name,
                    'additional_price': str(opt.additional_price),
                    'is_default': opt.is_default,
                }
                for opt in pmg.modifier_group.options.all() if opt.is_available
            ],
        })

    # Deduplicate by item (one entry per item)
    seen_items: set[int] = set()
    items_data = []
    for inv in inv_qs:
        if inv.item_id in seen_items:
            continue
        seen_items.add(inv.item_id)
        items_data.append({
            'item_pk': inv.item_id,
            'name': inv.item.nama,
            'kode_item': inv.item.item_id,
            'selling_price': str(inv.selling_price) if inv.selling_price is not None else '0',
            'unit': inv.item.tipe_item,
            'is_bulk': False,
            'modifier_groups': modifier_map.get(inv.item_id, []),
        })

    return JsonResponse({'items': items_data})
