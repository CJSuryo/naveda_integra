"""Jurnal views."""
import json
from decimal import Decimal

from django.contrib import messages as dj_messages
from django.contrib.auth.decorators import login_required
from django.db.models import Max, Q, Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from apps.entitas_bisnis.models import EntitasBisnis as EBModel
from apps.master_data.models import Akun, AsetLv2, KewajibanLv2, EkuitasLv2
from apps.aset_tetap.models import AsetTetapRecord
from apps.aset_lainnya.models import AsetLainnyaRecord

from .forms import (
    ItemForm,
    JurnalHeaderForm, JurnalDetailForm,
    JurnalAutomasiForm, JurnalAutomasiAkunForm,
    AutomasiEntryForm,
)
from .models import (
    Item, TransactionPrefix,
    JurnalHeader, JurnalDetail,
    JurnalAutomasi, JurnalAutomasiAkun,
)


def _parse_int_list(values: list[str]) -> list[int]:
    """Parse a list of string values into integers, ignoring non-digit values."""
    return [int(x) for x in values if x.isdigit()]


# ── Rekap Jurnal (replaces old index / header_list) ─────────────────────────


def _rekap_jurnal_qs(request: HttpRequest):
    """Build filtered JurnalHeader queryset from request GET params. Returns (qs, filters_dict)."""
    from apps.entitas_bisnis.models import EntitasBisnis, EntitasBisnisLv2, EntitasBisnisLv3

    qs = (
        JurnalHeader.objects
        .select_related('tipe_transaksi', 'entitas_bisnis', 'item', 'transaction_prefix', 'no_bukti')
        .prefetch_related('details__akun')
        .order_by('-tanggal', 'nomor_transaksi')
    )

    tanggal_dari = request.GET.get('tanggal_dari', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    prefix_filter = request.GET.get('prefix', '')
    akun_filter = request.GET.get('akun', '')
    uraian_filter = request.GET.get('uraian', '')
    eb_lv1_ids = request.GET.getlist('eb_lv1')
    eb_lv2_ids = request.GET.getlist('eb_lv2')
    eb_lv3_ids = request.GET.getlist('eb_lv3')

    if tanggal_dari:
        qs = qs.filter(tanggal__gte=tanggal_dari)
    if tanggal_sampai:
        qs = qs.filter(tanggal__lte=tanggal_sampai)
    if prefix_filter:
        qs = qs.filter(nomor_transaksi__startswith=prefix_filter)
    if akun_filter:
        qs = qs.filter(details__akun_id=akun_filter).distinct()
    if uraian_filter:
        qs = qs.filter(uraian_transaksi__icontains=uraian_filter)

    eb_filter_ids = set()
    if eb_lv1_ids:
        eb_filter_ids.update(_parse_int_list(eb_lv1_ids))
    if eb_lv2_ids:
        lv2_parent_ids = EntitasBisnisLv2.objects.filter(
            pk__in=_parse_int_list(eb_lv2_ids)
        ).values_list('entitas_bisnis_id', flat=True)
        eb_filter_ids.update(lv2_parent_ids)
    if eb_lv3_ids:
        lv3_parent_ids = EntitasBisnisLv3.objects.filter(
            pk__in=_parse_int_list(eb_lv3_ids)
        ).select_related('parent_lv2').values_list('parent_lv2__entitas_bisnis_id', flat=True)
        eb_filter_ids.update(lv3_parent_ids)

    if eb_filter_ids:
        qs = qs.filter(entitas_bisnis_id__in=eb_filter_ids)

    return qs, {
        'tanggal_dari': tanggal_dari,
        'tanggal_sampai': tanggal_sampai,
        'prefix_filter': prefix_filter,
        'akun_filter': akun_filter,
        'uraian_filter': uraian_filter,
        'eb_lv1_ids': eb_lv1_ids,
        'eb_lv2_ids': eb_lv2_ids,
        'eb_lv3_ids': eb_lv3_ids,
    }


@login_required
def rekap_jurnal(request: HttpRequest) -> HttpResponse:
    """Read-only journal recap with date range and entitas bisnis filtering."""
    from apps.entitas_bisnis.models import EntitasBisnis, EntitasBisnisLv2, EntitasBisnisLv3

    qs, filters = _rekap_jurnal_qs(request)
    headers = list(qs)

    # Prepare entitas bisnis hierarchy for filter modal
    eb_lv1_all = EntitasBisnis.objects.filter(status_aktif=True).order_by('nama')
    eb_lv2_all = EntitasBisnisLv2.objects.filter(status_aktif=True).select_related('entitas_bisnis').order_by('nama')
    eb_lv3_all = EntitasBisnisLv3.objects.filter(status_aktif=True).select_related('parent_lv2__entitas_bisnis').order_by('nama')

    # Distinct transaction prefixes for filter dropdown
    prefix_choices = (
        JurnalHeader.objects
        .values_list('nomor_transaksi', flat=True)
        .order_by('nomor_transaksi')
    )
    prefix_set = sorted({trx.rsplit('-', 1)[0] for trx in prefix_choices if trx and '-' in trx})

    # All akun for filter — only needed to resolve current_akun_label for TomSelect
    from apps.master_data.utils import get_akun_sorted
    akun_list = get_akun_sorted()
    current_akun_label = ''
    if filters['akun_filter']:
        for _a in akun_list:
            if str(_a.pk) == str(filters['akun_filter']):
                current_akun_label = f"{_a.kode_akun} — {_a.nama}"
                break

    return render(request, 'jurnal/rekap_jurnal.html', {
        'headers': headers,
        'tanggal_dari': filters['tanggal_dari'],
        'tanggal_sampai': filters['tanggal_sampai'],
        'prefix_filter': filters['prefix_filter'],
        'prefix_choices': prefix_set,
        'akun_filter': filters['akun_filter'],
        'uraian_filter': filters['uraian_filter'],
        'current_akun_label': current_akun_label,
        'eb_lv1_all': eb_lv1_all,
        'eb_lv2_all': eb_lv2_all,
        'eb_lv3_all': eb_lv3_all,
        'selected_lv1': filters['eb_lv1_ids'],
        'selected_lv2': filters['eb_lv2_ids'],
        'selected_lv3': filters['eb_lv3_ids'],
    })


@login_required
def rekap_jurnal_export(request: HttpRequest) -> HttpResponse:
    """Export filtered Rekap Jurnal as XLSX."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    qs, filters = _rekap_jurnal_qs(request)
    headers = list(qs)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Rekap Jurnal'

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    right_align = Alignment(horizontal='right')

    # Column headers
    col_headers = ['Tanggal', 'No. Transaksi', 'Uraian', 'Entitas Bisnis', 'Akun', 'Debit', 'Kredit']
    for col_idx, title in enumerate(col_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row_num = 2
    for header in headers:
        details = list(header.details.all())
        if not details:
            for col_idx, val in enumerate([
                header.tanggal,
                header.nomor_transaksi,
                header.uraian_transaksi,
                str(header.entitas_bisnis) if header.entitas_bisnis else '-',
                '-', 0, 0,
            ], 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = thin_border
                if col_idx in (6, 7):
                    cell.alignment = right_align
                    cell.number_format = '#,##0'
            row_num += 1
        else:
            for i, detail in enumerate(details):
                for col_idx, val in enumerate([
                    header.tanggal if i == 0 else '',
                    header.nomor_transaksi if i == 0 else '',
                    header.uraian_transaksi if i == 0 else '',
                    (str(header.entitas_bisnis) if header.entitas_bisnis else '-') if i == 0 else '',
                    f'{detail.akun.kode_akun} - {detail.akun.nama}',
                    detail.debit or 0,
                    detail.kredit or 0,
                ], 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=val)
                    cell.border = thin_border
                    if col_idx in (6, 7):
                        cell.alignment = right_align
                        cell.number_format = '#,##0'
                row_num += 1

    # Auto-width
    for col_idx in range(1, 8):
        max_len = len(col_headers[col_idx - 1])
        for row in ws.iter_rows(min_row=2, max_row=row_num - 1, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 50)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="rekap_jurnal.xlsx"'
    wb.save(response)
    return response


# ── JurnalHeader detail (read-only) ─────────────────────────────────────────

@login_required
def header_detail(request: HttpRequest, pk: int) -> HttpResponse:
    obj = get_object_or_404(
        JurnalHeader.objects.select_related('tipe_transaksi', 'entitas_bisnis', 'item', 'transaction_prefix'),
        pk=pk,
    )
    details = obj.details.select_related('akun').order_by('pk')
    return render(request, 'jurnal/header_detail.html', {'object': obj, 'details': details})


# ── Item CRUD ─────────────────────────────────────────────────────────────────

@login_required
def item_list(request: HttpRequest) -> HttpResponse:
    return render(request, 'jurnal/item_list.html', {'object_list': Item.objects.all().order_by('kode')})


@login_required
def item_create(request: HttpRequest) -> HttpResponse:
    form = ItemForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('jurnal:item_list')
    return render(request, 'jurnal/item_form.html', {'form': form, 'title': 'Tambah Item'})


@login_required
def item_update(request: HttpRequest, pk: int) -> HttpResponse:
    obj = get_object_or_404(Item, pk=pk)
    form = ItemForm(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('jurnal:item_list')
    return render(request, 'jurnal/item_form.html', {'form': form, 'title': 'Edit Item', 'object': obj})


@login_required
def item_delete(request: HttpRequest, pk: int) -> HttpResponse:
    obj = get_object_or_404(Item, pk=pk)
    if request.method == 'POST':
        obj.delete()
        return redirect('jurnal:item_list')
    return redirect('jurnal:item_list')


# ── Automasi / Jurnal Manual CRUD ────────────────────────────────────────────

@login_required
def automasi_list(request: HttpRequest) -> HttpResponse:
    return render(request, 'jurnal/automasi_list.html', {
        'object_list': JurnalAutomasi.objects.all().order_by('nama'),
    })


@login_required
def automasi_create(request: HttpRequest) -> HttpResponse:
    form = JurnalAutomasiForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('jurnal:automasi_list')
    return render(request, 'jurnal/automasi_form.html', {'form': form, 'title': 'Tambah Jurnal Manual'})


@login_required
def automasi_update(request: HttpRequest, pk: int) -> HttpResponse:
    obj = get_object_or_404(JurnalAutomasi, pk=pk)
    form = JurnalAutomasiForm(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('jurnal:automasi_list')
    return render(request, 'jurnal/automasi_form.html', {'form': form, 'title': 'Edit Jurnal Manual', 'object': obj})


@login_required
def automasi_delete(request: HttpRequest, pk: int) -> HttpResponse:
    obj = get_object_or_404(JurnalAutomasi, pk=pk)
    if request.method == 'POST':
        obj.delete()
        return redirect('jurnal:automasi_list')
    return redirect('jurnal:automasi_list')


@login_required
def automasi_detail(request: HttpRequest, pk: int) -> HttpResponse:
    obj = get_object_or_404(JurnalAutomasi, pk=pk)
    mappings = obj.akun_mappings.select_related('akun').order_by('pk')
    add_form = JurnalAutomasiAkunForm()
    return render(request, 'jurnal/automasi_detail.html', {
        'object': obj, 'mappings': mappings, 'add_form': add_form,
    })


@login_required
def automasi_add_akun(request: HttpRequest, pk: int) -> HttpResponse:
    obj = get_object_or_404(JurnalAutomasi, pk=pk)
    if request.method == 'POST':
        form = JurnalAutomasiAkunForm(request.POST)
        if form.is_valid():
            mapping = form.save(commit=False)
            mapping.automasi = obj
            mapping.save()
    return redirect('jurnal:automasi_detail', pk=pk)


@login_required
def automasi_remove_akun(request: HttpRequest, pk: int, mapping_pk: int) -> HttpResponse:
    obj = get_object_or_404(JurnalAutomasi, pk=pk)
    mapping = get_object_or_404(JurnalAutomasiAkun, pk=mapping_pk, automasi=obj)
    if request.method == 'POST':
        mapping.delete()
    return redirect('jurnal:automasi_detail', pk=pk)


def _next_nomor_transaksi(prefix_kode: str) -> str:
    """Generate the next sequential transaction number for a prefix.

    Uses database-level locking to avoid race conditions.
    """
    from django.db import transaction

    with transaction.atomic():
        last = (
            JurnalHeader.objects
            .select_for_update()
            .filter(nomor_transaksi__startswith=prefix_kode + '-')
            .order_by('-nomor_transaksi')
            .values_list('nomor_transaksi', flat=True)
            .first()
        )
        if last:
            try:
                seq = int(last.rsplit('-', 1)[1]) + 1
            except (ValueError, IndexError):
                seq = 1
        else:
            seq = 1
        return f'{prefix_kode}-{seq:03d}'


@login_required
def automasi_entry(request: HttpRequest, pk: int) -> HttpResponse:
    """Show the form to create an automated journal entry based on the mapping."""
    automasi = get_object_or_404(JurnalAutomasi, pk=pk)
    mappings = automasi.akun_mappings.select_related('akun').order_by('pk')

    if request.method == 'POST':
        form = AutomasiEntryForm(request.POST)
        if form.is_valid():
            tanggal = form.cleaned_data['tanggal']
            uraian = form.cleaned_data['uraian_transaksi']
            item = form.cleaned_data['item']
            prefix = form.cleaned_data['transaction_prefix']
            nomor = _next_nomor_transaksi(prefix.kode)

            header = JurnalHeader.objects.create(
                tanggal=tanggal,
                nomor_transaksi=nomor,
                uraian_transaksi=uraian,
                item=item,
                transaction_prefix=prefix,
                is_penyesuaian=True,
            )

            # Create detail lines: read amounts from POST for each mapped akun
            for mapping in mappings:
                debit_key = f'debit_{mapping.pk}'
                kredit_key = f'kredit_{mapping.pk}'
                debit = Decimal(request.POST.get(debit_key, '0') or '0')
                kredit = Decimal(request.POST.get(kredit_key, '0') or '0')
                if debit or kredit:
                    JurnalDetail.objects.create(
                        jurnal_header=header,
                        akun=mapping.akun,
                        debit=debit,
                        kredit=kredit,
                    )
            return redirect('jurnal:header_detail', pk=header.pk)
    else:
        from django.utils import timezone
        form = AutomasiEntryForm(initial={'tanggal': timezone.now().date()})

    return render(request, 'jurnal/automasi_entry.html', {
        'automasi': automasi, 'mappings': mappings, 'form': form,
    })


# ── Neraca Saldo (Trial Balance) ────────────────────────────────────────────

@login_required
def neraca_saldo(request: HttpRequest) -> HttpResponse:
    """Trial balance with saldo awal, mutasi modul, sebelum & setelah penyesuaian."""
    tanggal_dari = request.GET.get('tanggal_dari', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')

    from apps.master_data.utils import get_akun_sorted
    akun_list = get_akun_sorted()

    # Helper: aggregate debit/kredit for a queryset of JurnalDetail
    def _aggregate(qs):
        result = (
            qs.values('akun_id')
            .annotate(
                total_debit=Coalesce(Sum('debit'), Value(Decimal('0')), output_field=DecimalField()),
                total_kredit=Coalesce(Sum('kredit'), Value(Decimal('0')), output_field=DecimalField()),
            )
        )
        return {row['akun_id']: row for row in result}

    # Period filter
    period_filter = {}
    if tanggal_dari:
        period_filter['jurnal_header__tanggal__gte'] = tanggal_dari
    if tanggal_sampai:
        period_filter['jurnal_header__tanggal__lte'] = tanggal_sampai

    # Saldo Awal: entries created via the Saldo Awal menu (is_saldo_awal=True)
    # PLUS all entries (any type) that fall BEFORE tanggal_dari when a date range is active.
    if tanggal_dari:
        sa_qs = JurnalDetail.objects.filter(
            jurnal_header__tanggal__lt=tanggal_dari
        ) | JurnalDetail.objects.filter(
            jurnal_header__is_saldo_awal=True, **period_filter
        )
    else:
        sa_qs = JurnalDetail.objects.filter(jurnal_header__is_saldo_awal=True)
    saldo_awal_map = _aggregate(sa_qs)

    # Mutasi Modul: automated entries within period, excluding saldo awal entries
    modul_map = _aggregate(
        JurnalDetail.objects.filter(
            jurnal_header__is_penyesuaian=False,
            jurnal_header__is_saldo_awal=False,
            **period_filter
        )
    )

    # Penyesuaian: manual adjustment entries within period, excluding saldo awal
    penyesuaian_map = _aggregate(
        JurnalDetail.objects.filter(
            jurnal_header__is_penyesuaian=True,
            jurnal_header__is_saldo_awal=False,
            **period_filter
        )
    )

    # Akun prefixes whose entries (regardless of is_penyesuaian flag) belong
    # in Mutasi Modul because they represent periodic operating activity, not
    # manual adjustments (e.g. Beban Penyusutan 5.1.19.xx).
    MODUL_OVERRIDE_PREFIXES = ('5.1.19',)

    def _get(m, akun_id, field):
        return m.get(akun_id, {}).get(field, Decimal('0'))

    rows = []
    for akun in akun_list:
        sa_d = _get(saldo_awal_map, akun.pk, 'total_debit')
        sa_k = _get(saldo_awal_map, akun.pk, 'total_kredit')
        modul_d = _get(modul_map, akun.pk, 'total_debit')
        modul_k = _get(modul_map, akun.pk, 'total_kredit')
        peny_d = _get(penyesuaian_map, akun.pk, 'total_debit')
        peny_k = _get(penyesuaian_map, akun.pk, 'total_kredit')

        # For overridden prefixes, move penyesuaian amounts into mutasi modul
        if any(akun.kode_akun.startswith(p) for p in MODUL_OVERRIDE_PREFIXES):
            modul_d += peny_d
            modul_k += peny_k
            peny_d = Decimal('0')
            peny_k = Decimal('0')

        # Sebelum penyesuaian = saldo awal + mutasi modul
        sblm_d = sa_d + modul_d
        sblm_k = sa_k + modul_k

        # Setelah penyesuaian = sebelum + penyesuaian manual
        stlh_d = sblm_d + peny_d
        stlh_k = sblm_k + peny_k

        rows.append({
            'akun': akun,
            'sa_d': sa_d, 'sa_k': sa_k,
            'modul_d': modul_d, 'modul_k': modul_k,
            'sblm_d': sblm_d, 'sblm_k': sblm_k,
            'peny_d': peny_d, 'peny_k': peny_k,
            'stlh_d': stlh_d, 'stlh_k': stlh_k,
        })

    # Compute totals
    totals = {
        'sa_d': Decimal('0'), 'sa_k': Decimal('0'),
        'modul_d': Decimal('0'), 'modul_k': Decimal('0'),
        'sblm_d': Decimal('0'), 'sblm_k': Decimal('0'),
        'peny_d': Decimal('0'), 'peny_k': Decimal('0'),
        'stlh_d': Decimal('0'), 'stlh_k': Decimal('0'),
    }
    for row in rows:
        for key in totals:
            totals[key] += row[key]

    return render(request, 'jurnal/neraca_saldo.html', {
        'rows': rows,
        'totals': totals,
        'tanggal_dari': tanggal_dari,
        'tanggal_sampai': tanggal_sampai,
    })


# ── Akun autocomplete API ───────────────────────────────────────────────────

@login_required
def akun_autocomplete(request: HttpRequest) -> JsonResponse:
    """Return Akun options matching a search term, for autocomplete widgets.

    Optional query params:
    - term: search text
    - eb_id: filter to akuns available for this entitas bisnis (via EntitasBisnisAkun)
    - all: if '1', return all matching akuns (no limit)
    """
    from apps.master_data.models import EntitasBisnisAkun

    term = request.GET.get('term', '')
    eb_id = request.GET.get('eb_id', '')
    return_all = request.GET.get('all', '')
    prefix = request.GET.get('prefix', '')

    from apps.master_data.utils import natural_sort_key

    qs = Akun.objects.filter(
        Q(nama__icontains=term) | Q(kode_akun__icontains=term)
    )
    if prefix:
        qs = qs.filter(kode_akun__startswith=prefix)

    # If eb_id is provided, filter to only available akuns for that EB
    if eb_id:
        resolved_eb_id = _resolve_entitas_bisnis_id(eb_id)
        available_akun_ids = list(EntitasBisnisAkun.objects.filter(
            entitas_bisnis_id=resolved_eb_id,
        ).values_list('akun_id', flat=True)) if resolved_eb_id else []
        if available_akun_ids:
            qs = qs.filter(pk__in=available_akun_ids)

    if not return_all:
        akun_objs = sorted(qs[:200], key=lambda a: natural_sort_key(a.kode_akun))
    else:
        akun_objs = sorted(qs, key=lambda a: natural_sort_key(a.kode_akun))

    results = [
        {
            'id': a.pk,
            'text': f'{a.kode_akun} - {a.nama}',
            'kode': a.kode_akun,
            'nama': a.nama,
        }
        for a in akun_objs
    ]
    return JsonResponse(results, safe=False)


@login_required
def rekap_jurnal_get(request: HttpRequest, pk: int) -> JsonResponse:
    """Return JurnalHeader data + details for the edit modal in rekap jurnal."""
    header = get_object_or_404(
        JurnalHeader.objects.select_related('entitas_bisnis', 'transaction_prefix'),
        pk=pk,
    )
    details = header.details.select_related('akun').order_by('pk')
    return JsonResponse({
        'id': header.pk,
        'tanggal': str(header.tanggal),
        'nomor_transaksi': header.nomor_transaksi,
        'uraian_transaksi': header.uraian_transaksi,
        'entitas_bisnis_id': header.entitas_bisnis_id or '',
        'is_penyesuaian': header.is_penyesuaian,
        'details': [
            {
                'id': d.pk,
                'akun_id': d.akun_id,
                'akun_text': f'{d.akun.kode_akun} - {d.akun.nama}',
                'debit': str(d.debit),
                'kredit': str(d.kredit),
            }
            for d in details
        ],
    })


@login_required
def rekap_jurnal_update(request: HttpRequest, pk: int) -> JsonResponse:
    """Update a JurnalHeader + details via AJAX from the rekap jurnal edit modal."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    header = get_object_or_404(JurnalHeader, pk=pk)

    try:
        data = json.loads(request.body)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    tanggal = data.get('tanggal')
    uraian = (data.get('uraian_transaksi') or '').strip()
    eb_id = data.get('entitas_bisnis_id') or None

    if not tanggal:
        return JsonResponse({'error': 'Tanggal wajib diisi.'}, status=400)
    if not uraian:
        return JsonResponse({'error': 'Uraian transaksi wajib diisi.'}, status=400)

    rows = data.get('details', [])
    rows = [r for r in rows if r.get('akun_id')]
    if not rows:
        return JsonResponse({'error': 'Minimal 1 baris akun wajib diisi.'}, status=400)

    total_debit = sum(Decimal(str(r.get('debit') or 0)) for r in rows)
    total_kredit = sum(Decimal(str(r.get('kredit') or 0)) for r in rows)
    if total_debit != total_kredit:
        return JsonResponse({'error': f'Total Debit ({total_debit}) harus sama dengan Total Kredit ({total_kredit}).'}, status=400)

    header.tanggal = tanggal
    header.uraian_transaksi = uraian
    header.entitas_bisnis_id = eb_id
    header.save()

    # Replace all details
    header.details.all().delete()
    JurnalDetail.objects.bulk_create([
        JurnalDetail(
            jurnal_header=header,
            akun_id=row['akun_id'],
            debit=Decimal(str(row.get('debit') or 0)),
            kredit=Decimal(str(row.get('kredit') or 0)),
        )
        for row in rows
    ])

    return JsonResponse({'ok': True})


@login_required
def rekap_jurnal_delete(request: HttpRequest, pk: int) -> JsonResponse:
    """Delete a JurnalHeader via AJAX from the rekap jurnal page."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    header = get_object_or_404(JurnalHeader, pk=pk)
    try:
        header.delete()
    except Exception as e:
        from django.db.models import ProtectedError
        if isinstance(e, ProtectedError):
            return JsonResponse(
                {'error': 'Jurnal tidak bisa dihapus karena ada data lain yang terkait (modal disetor, dll). Hapus data terkait terlebih dahulu.'},
                status=400,
            )
        raise
    return JsonResponse({'ok': True})


def _get_entitas_bisnis_dropdown_options() -> list[dict[str, str]]:
    """Return EB options grouped hierarchically: Lv1, then its Lv2, then Lv3 under each Lv2."""
    from apps.entitas_bisnis.models import EntitasBisnisLv2, EntitasBisnisLv3

    # Pre-fetch all levels
    lv1_list = list(EBModel.objects.filter(status_aktif=True).order_by('nama'))
    lv2_list = list(
        EntitasBisnisLv2.objects.filter(status_aktif=True)
        .select_related('entitas_bisnis')
        .order_by('nama')
    )
    lv3_list = list(
        EntitasBisnisLv3.objects.filter(status_aktif=True)
        .select_related('parent_lv2__entitas_bisnis')
        .order_by('nama')
    )

    # Group lv2 by lv1, lv3 by lv2
    lv2_by_lv1: dict[int, list[EntitasBisnisLv2]] = {}
    for lv2 in lv2_list:
        lv2_by_lv1.setdefault(lv2.entitas_bisnis_id, []).append(lv2)
    lv3_by_lv2: dict[int, list[EntitasBisnisLv3]] = {}
    for lv3 in lv3_list:
        lv3_by_lv2.setdefault(lv3.parent_lv2_id, []).append(lv3)

    options: list[dict[str, str]] = []
    for eb in lv1_list:
        options.append({'value': f'lv1:{eb.pk}', 'label': eb.nama})
        for lv2 in lv2_by_lv1.get(eb.pk, []):
            options.append({
                'value': f'lv2:{lv2.pk}',
                'label': f'\u00a0\u00a0\u00a0\u00a0↳ {lv2.nama}',
            })
            for lv3 in lv3_by_lv2.get(lv2.pk, []):
                options.append({
                    'value': f'lv3:{lv3.pk}',
                    'label': f'\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0\u00a0↳ {lv3.nama}',
                })

    return options


def _resolve_entitas_bisnis_id(selection: str | None) -> int | None:
    from apps.entitas_bisnis.models import EntitasBisnisLv2, EntitasBisnisLv3

    if not selection:
        return None

    try:
        level, raw_pk = selection.split(':', 1)
        pk = int(raw_pk)
    except (ValueError, TypeError):
        return None

    if level == 'lv1':
        return pk

    if level == 'lv2':
        lv2 = EntitasBisnisLv2.objects.filter(pk=pk).select_related('entitas_bisnis').first()
        return lv2.entitas_bisnis_id if lv2 else None

    if level == 'lv3':
        lv3 = EntitasBisnisLv3.objects.filter(pk=pk).select_related('parent_lv2__entitas_bisnis').first()
        return lv3.parent_lv2.entitas_bisnis_id if lv3 else None

    return None


@login_required
def manual_jurnal_create(request: HttpRequest) -> HttpResponse:
    """Spreadsheet-like direct manual journal entry form."""
    from django.db import transaction as db_transaction
    from apps.purchase.models import FIFOBatch, ItemMasterPurchase
    from apps.inventory.models import InventoryRecord

    AKUN_DETAIL_PREFIXES: dict[str, str] = {
        '1.1.7': 'persediaan',
        '1.1.8': 'persediaan',
        '1.1.9': 'persediaan',
        '1.1.10': 'persediaan',
        '1.2': 'aset_tetap',
        '1.2.3': 'aset_tetap',
        '1.3': 'aset_lainnya',
        '3.1.1': 'modal_disetor',
    }

    if request.method == 'POST':
        tanggal = request.POST.get('tanggal')
        uraian = request.POST.get('uraian_transaksi', '').strip()
        eb_selection = request.POST.get('entitas_bisnis') or ''
        eb_id = _resolve_entitas_bisnis_id(eb_selection)
        rows_json = request.POST.get('rows_data', '[]')

        errors = {}
        if not tanggal:
            errors['tanggal'] = 'Tanggal wajib diisi.'
        if not uraian:
            errors['uraian_transaksi'] = 'Deskripsi wajib diisi.'
        if eb_selection and eb_id is None:
            errors['entitas_bisnis'] = 'Pilihan entitas bisnis tidak valid.'

        try:
            rows = json.loads(rows_json)
        except (ValueError, TypeError):
            rows = []

        rows = [r for r in rows if r.get('akun_id')]
        if not rows:
            errors['rows'] = 'Minimal 1 baris akun wajib diisi.'

        total_debit = sum(Decimal(str(r.get('debit') or 0)) for r in rows)
        total_kredit = sum(Decimal(str(r.get('kredit') or 0)) for r in rows)

        if not errors and total_debit != total_kredit:
            errors['balance'] = f'Total Debit ({total_debit}) harus sama dengan Total Kredit ({total_kredit}).'

        if errors:
            eb_list = _get_entitas_bisnis_dropdown_options()
            return render(request, 'jurnal/manual_jurnal.html', {
                'today': tanggal or timezone.now().date(),
                'eb_list': eb_list,
                'errors': errors,
                'posted': True,
                'selected_entitas_bisnis': eb_selection,
                'akun_detail_prefixes_json': json.dumps(AKUN_DETAIL_PREFIXES),
            })

        with db_transaction.atomic():
            nomor = _next_nomor_transaksi('TRX-MAN')
            header = JurnalHeader.objects.create(
                tanggal=tanggal,
                nomor_transaksi=nomor,
                uraian_transaksi=uraian,
                entitas_bisnis_id=eb_id,
                is_penyesuaian=True,
            )
            JurnalDetail.objects.bulk_create([
                JurnalDetail(
                    jurnal_header=header,
                    akun_id=row['akun_id'],
                    debit=Decimal(str(row.get('debit') or 0)),
                    kredit=Decimal(str(row.get('kredit') or 0)),
                )
                for row in rows
            ])

            # ── Persediaan detail: InventoryRecord + FIFOBatch ─────────────
            persediaan_rows = [
                r for r in rows
                if r.get('detail_type') == 'persediaan' and r.get('detail_rows')
            ]
            if persediaan_rows:
                all_item_ids = {
                    int(d['item_id'])
                    for r in persediaan_rows
                    for d in r.get('detail_rows', [])
                    if str(d.get('item_id', '')).isdigit()
                }
                items_map = {
                    item.pk: item
                    for item in ItemMasterPurchase.objects.filter(pk__in=all_item_ids)
                }
                for row in persediaan_rows:
                    for d in row.get('detail_rows', []):
                        try:
                            item_pk = int(str(d.get('item_id', '')))
                            qty = Decimal(str(d.get('qty') or 0))
                            unit_price = Decimal(str(d.get('unit_price') or 0))
                        except (ValueError, TypeError):
                            continue
                        if qty <= 0 or unit_price < 0:
                            continue
                        item = items_map.get(item_pk)
                        if not item:
                            continue
                        InventoryRecord.objects.create(
                            item=item,
                            purchase_item=None,
                            entitas_bisnis_id=eb_id,
                            quantity=qty,
                            unit_price=unit_price,
                            tanggal=tanggal,
                        )
                        FIFOBatch.objects.create(
                            purchase_item=None,
                            item=item,
                            tanggal=tanggal,
                            quantity_in=qty,
                            unit_price=unit_price,
                            remaining_qty=qty,
                        )

            # ── Aset Tetap detail: AsetTetapRecord ─────────────────────────
            aset_tetap_rows = [
                r for r in rows
                if r.get('detail_type') == 'aset_tetap' and r.get('detail_rows')
            ]
            if aset_tetap_rows:
                all_item_ids = {
                    int(d['item_id'])
                    for r in aset_tetap_rows
                    for d in r.get('detail_rows', [])
                    if str(d.get('item_id', '')).isdigit()
                }
                items_map = {
                    item.pk: item
                    for item in ItemMasterPurchase.objects.filter(pk__in=all_item_ids)
                }
                for row in aset_tetap_rows:
                    for d in row.get('detail_rows', []):
                        try:
                            item_pk = int(str(d.get('item_id', '')))
                            qty = Decimal(str(d.get('qty') or 0))
                            unit_price = Decimal(str(d.get('unit_price') or 0))
                        except (ValueError, TypeError):
                            continue
                        if qty <= 0 or unit_price < 0:
                            continue
                        item = items_map.get(item_pk)
                        if not item:
                            continue
                        AsetTetapRecord.objects.create(
                            item=item,
                            purchase_item=None,
                            entitas_bisnis_id=eb_id,
                            quantity=qty,
                            harga_perolehan=unit_price,
                            tanggal_perolehan=tanggal,
                            masa_manfaat=item.masa_manfaat,
                            metode_penyusutan=item.metode_penyusutan,
                        )

            # ── Aset Lainnya detail: AsetLainnyaRecord ─────────────────────
            aset_lainnya_rows = [
                r for r in rows
                if r.get('detail_type') == 'aset_lainnya' and r.get('detail_rows')
            ]
            if aset_lainnya_rows:
                all_item_ids = {
                    int(d['item_id'])
                    for r in aset_lainnya_rows
                    for d in r.get('detail_rows', [])
                    if str(d.get('item_id', '')).isdigit()
                }
                items_map = {
                    item.pk: item
                    for item in ItemMasterPurchase.objects.filter(pk__in=all_item_ids)
                }
                for row in aset_lainnya_rows:
                    for d in row.get('detail_rows', []):
                        try:
                            item_pk = int(str(d.get('item_id', '')))
                            qty = Decimal(str(d.get('qty') or 0))
                            unit_price = Decimal(str(d.get('unit_price') or 0))
                        except (ValueError, TypeError):
                            continue
                        if qty <= 0 or unit_price < 0:
                            continue
                        item = items_map.get(item_pk)
                        if not item:
                            continue
                        AsetLainnyaRecord.objects.create(
                            item=item,
                            purchase_item=None,
                            entitas_bisnis_id=eb_id,
                            quantity=qty,
                            harga_perolehan=unit_price,
                            tanggal_perolehan=tanggal,
                            masa_manfaat=item.masa_manfaat,
                            metode_amortisasi=item.metode_amortisasi,
                        )

            # ── Modal Disetor detail: ModalDisetor + Pemilik ───────────────
            from apps.ekuitas.models import ModalDisetor as ModalDisetorModel, Pemilik as PemilikModel
            from apps.ekuitas.services import get_or_create_pemilik
            modal_disetor_rows = [
                r for r in rows
                if r.get('detail_type') == 'modal_disetor' and r.get('detail_rows')
            ]
            for row in modal_disetor_rows:
                for d in row.get('detail_rows', []):
                    nama_pemilik = str(d.get('nama_pemilik', '')).strip()
                    if not nama_pemilik:
                        continue
                    try:
                        jumlah = Decimal(str(d.get('jumlah', 0)))
                    except Exception:
                        continue
                    if jumlah <= 0:
                        continue
                    pemilik = get_or_create_pemilik(nama_pemilik)
                    ModalDisetorModel.objects.create(
                        entitas_bisnis_id=eb_id,
                        pemilik=pemilik,
                        jumlah_modal=jumlah,
                        tanggal_setor=tanggal,
                        keterangan=str(d.get('keterangan', '')).strip(),
                        jurnal_header=header,
                    )

        dj_messages.success(request, f'Jurnal manual {nomor} berhasil dibuat.')
        return redirect('jurnal:header_detail', pk=header.pk)

    eb_list = _get_entitas_bisnis_dropdown_options()
    return render(request, 'jurnal/manual_jurnal.html', {
        'today': timezone.now().date(),
        'eb_list': eb_list,
        'errors': {},
        'akun_detail_prefixes_json': json.dumps(AKUN_DETAIL_PREFIXES),
    })


# ── Laporan Laba Rugi (Income Statement) ────────────────────────────────────

@login_required
def laporan_laba_rugi(request: HttpRequest) -> HttpResponse:
    """Read-only Income Statement (Laporan Operasional / Laba Rugi).

    Sections:
    - Pendapatan Operasional: akun prefix 4.1.x (kredit)
    - HPP / COGS: akun prefix 5.1.1 s/d 5.1.5 (debit)
    - Laba Kotor = Pendapatan - HPP
    - Beban Operasional: akun prefix 5.1.6+ (debit)
    - Laba Operasional = Laba Kotor - Beban Operasional
    - Pendapatan Non-Operasional: akun prefix 4.2.x (kredit)
    - Beban Non-Operasional: akun prefix 5.2.x (debit)
    - Laba Bersih = Laba Operasional + (Pendapatan Non-Op - Beban Non-Op)
    """
    from apps.master_data.models import PendapatanLv1, PendapatanLv2, BebanLv1, BebanLv2

    tanggal_dari = request.GET.get('tanggal_dari', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    eb_filter = request.GET.get('entitas_bisnis', '')
    tampilan = request.GET.get('tampilan', 'ringkas')  # 'ringkas' or 'detail'

    # Build period filter
    period_filter: dict = {}
    if tanggal_dari:
        period_filter['jurnal_header__tanggal__gte'] = tanggal_dari
    if tanggal_sampai:
        period_filter['jurnal_header__tanggal__lte'] = tanggal_sampai
    if eb_filter:
        period_filter['jurnal_header__entitas_bisnis_id'] = eb_filter

    # Aggregate by akun
    agg = (
        JurnalDetail.objects
        .filter(**period_filter)
        .values('akun_id', 'akun__kode_akun', 'akun__nama', 'akun__kategori_id')
        .annotate(
            total_debit=Coalesce(Sum('debit'), Value(Decimal('0')), output_field=DecimalField()),
            total_kredit=Coalesce(Sum('kredit'), Value(Decimal('0')), output_field=DecimalField()),
        )
    )

    # Build lookup by akun
    akun_data: dict[str, dict] = {}
    for row in agg:
        kode = row['akun__kode_akun'] or ''
        akun_data[kode] = {
            'kode': kode,
            'nama': row['akun__nama'] or '',
            'kategori': row['akun__kategori_id'] or '',
            'debit': row['total_debit'],
            'kredit': row['total_kredit'],
        }

    def _is_prefix(kode: str, prefix: str) -> bool:
        return kode.startswith(prefix + '.') or kode == prefix

    def _collect(prefix: str) -> list[dict]:
        """Collect all akun rows matching prefix, sorted by kode."""
        return sorted(
            [v for k, v in akun_data.items() if _is_prefix(k, prefix)],
            key=lambda x: x['kode'],
        )

    def _sum_net(items: list[dict], net_type: str = 'kredit') -> Decimal:
        """Sum net value. For pendapatan: kredit - debit. For beban: debit - kredit.
        Also stamps each item with its 'net' value for template display."""
        total = Decimal('0')
        for i in items:
            if net_type == 'kredit':
                i['net'] = i['kredit'] - i['debit']
            else:
                i['net'] = i['debit'] - i['kredit']
            total += i['net']
        return total

    # 1. Pendapatan Operasional (4.1.x)
    pendapatan_op_items = _collect('4.1')
    total_pendapatan_op = _sum_net(pendapatan_op_items, 'kredit')

    # 2. HPP / COGS (5.1.1 through 5.1.5)
    hpp_items = []
    for sub in ['5.1.1', '5.1.2', '5.1.3', '5.1.4', '5.1.5']:
        hpp_items.extend(_collect(sub))
    total_hpp = _sum_net(hpp_items, 'debit')

    # 3. Laba Kotor
    laba_kotor = total_pendapatan_op - total_hpp

    # 4. Beban Operasional (5.1.6 through 5.1.30)
    hpp_kodes = {item['kode'] for item in hpp_items}
    beban_op_items = []
    beban_op_kodes: set[str] = set()
    for i in range(6, 31):
        for item in _collect(f'5.1.{i}'):
            if item['kode'] not in beban_op_kodes:
                beban_op_kodes.add(item['kode'])
                beban_op_items.append(item)
    # Also include any 5.1.x that aren't HPP (kode > 5.1.5)
    for kode, data in akun_data.items():
        if kode.startswith('5.1.') and kode not in hpp_kodes and kode not in beban_op_kodes:
            parts = kode.split('.')
            if len(parts) >= 3:
                try:
                    sub_num = int(parts[2])
                    if sub_num >= 6:
                        beban_op_kodes.add(kode)
                        beban_op_items.append(data)
                except ValueError:
                    pass
    beban_op_items.sort(key=lambda x: x['kode'])
    total_beban_op = _sum_net(beban_op_items, 'debit')

    # 5. Laba Operasional
    laba_operasional = laba_kotor - total_beban_op

    # 6. Pendapatan Non-Operasional (4.2.x)
    pendapatan_non_op_items = _collect('4.2')
    total_pendapatan_non_op = _sum_net(pendapatan_non_op_items, 'kredit')

    # 7. Beban Non-Operasional (5.2.x)
    beban_non_op_items = _collect('5.2')
    total_beban_non_op = _sum_net(beban_non_op_items, 'debit')

    # 8. Net Non-Operasional
    net_non_op = total_pendapatan_non_op - total_beban_non_op

    # 9. Laba Bersih
    laba_bersih = laba_operasional + net_non_op

    eb_list = EBModel.objects.filter(status_aktif=True).order_by('nama')

    return render(request, 'jurnal/laporan_laba_rugi.html', {
        'tanggal_dari': tanggal_dari,
        'tanggal_sampai': tanggal_sampai,
        'eb_filter': eb_filter,
        'eb_list': eb_list,
        'tampilan': tampilan,
        # Data
        'pendapatan_op_items': pendapatan_op_items,
        'total_pendapatan_op': total_pendapatan_op,
        'hpp_items': hpp_items,
        'total_hpp': total_hpp,
        'laba_kotor': laba_kotor,
        'beban_op_items': beban_op_items,
        'total_beban_op': total_beban_op,
        'laba_operasional': laba_operasional,
        'pendapatan_non_op_items': pendapatan_non_op_items,
        'total_pendapatan_non_op': total_pendapatan_non_op,
        'beban_non_op_items': beban_non_op_items,
        'total_beban_non_op': total_beban_non_op,
        'net_non_op': net_non_op,
        'laba_bersih': laba_bersih,
    })


# ── Saldo Awal (Opening Balance) ────────────────────────────────────────────

def _reverse_saldo_awal_records(header) -> None:
    """Reverse all sub-records created by a saldo awal journal (excluding JurnalDetails).

    Matches records by purchase_item=None + entitas_bisnis + tanggal, which is
    the identifying footprint of saldo-awal-created sub-records.
    """
    from apps.inventory.models import InventoryRecord
    from apps.aset_tetap.models import AsetTetapRecord
    from apps.aset_lainnya.models import AsetLainnyaRecord
    from apps.ekuitas.models import ModalDisetor
    from apps.purchase.models import FIFOBatch

    eb_id = header.entitas_bisnis_id
    tanggal = header.tanggal

    # ModalDisetor — exact FK match
    ModalDisetor.objects.filter(jurnal_header=header).delete()

    # InventoryRecord + FIFOBatch — matched by purchase_item=None + entity + date
    inv_qs = InventoryRecord.objects.filter(
        purchase_item=None,
        entitas_bisnis_id=eb_id,
        tanggal=tanggal,
    )
    item_ids = list(inv_qs.values_list('item_id', flat=True))
    FIFOBatch.objects.filter(
        purchase_item=None,
        item_id__in=item_ids,
        tanggal=tanggal,
    ).delete()
    inv_qs.delete()

    # AsetTetapRecord
    AsetTetapRecord.objects.filter(
        purchase_item=None,
        entitas_bisnis_id=eb_id,
        tanggal_perolehan=tanggal,
    ).delete()

    # AsetLainnyaRecord
    AsetLainnyaRecord.objects.filter(
        purchase_item=None,
        entitas_bisnis_id=eb_id,
        tanggal_perolehan=tanggal,
    ).delete()


@login_required
def saldo_awal_list(request: HttpRequest) -> HttpResponse:
    """List all saldo awal journal headers."""
    qs = (
        JurnalHeader.objects
        .filter(is_saldo_awal=True)
        .select_related('entitas_bisnis')
        .order_by('-tanggal', '-nomor_transaksi')
    )

    # RBAC: business users see only their linked entities
    if request.user.is_business_user:
        from apps.accounts.models import UserEntitasBisnis
        allowed_eb_ids = UserEntitasBisnis.objects.filter(
            user=request.user
        ).values_list('entitas_bisnis_id', flat=True)
        qs = qs.filter(entitas_bisnis_id__in=allowed_eb_ids)

    eb_filter = request.GET.get('entitas_bisnis', '')
    tanggal_dari = request.GET.get('tanggal_dari', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    if eb_filter:
        qs = qs.filter(entitas_bisnis_id=eb_filter)
    if tanggal_dari:
        qs = qs.filter(tanggal__gte=tanggal_dari)
    if tanggal_sampai:
        qs = qs.filter(tanggal__lte=tanggal_sampai)

    eb_list = EBModel.objects.order_by('nama')
    return render(request, 'jurnal/saldo_awal_list.html', {
        'headers': qs,
        'eb_filter': eb_filter,
        'tanggal_dari': tanggal_dari,
        'tanggal_sampai': tanggal_sampai,
        'eb_list': eb_list,
    })


@login_required
def saldo_awal_delete(request: HttpRequest, pk: int) -> HttpResponse:
    """Confirm and delete a saldo awal journal, reversing all related records."""
    from apps.jurnal.utils import log_jurnal_terhapus

    header = get_object_or_404(JurnalHeader, pk=pk, is_saldo_awal=True)

    if request.method == 'POST':
        nomor = header.nomor_transaksi
        from django.db import transaction as db_transaction
        with db_transaction.atomic():
            log_jurnal_terhapus(header, 'saldo_awal', request)
            _reverse_saldo_awal_records(header)
            header.details.all().delete()
            header.delete()
        dj_messages.success(request, f'Saldo Awal {nomor} berhasil dihapus.')
        return redirect('jurnal:saldo_awal_list')

    # GET: load detail lines for the confirmation page
    details = header.details.select_related('akun').order_by('pk')
    return render(request, 'jurnal/saldo_awal_delete.html', {
        'header': header,
        'details': details,
    })


@login_required
def saldo_awal(request: HttpRequest) -> HttpResponse:
    """Spreadsheet-style opening balance form.

    Rows can carry per-akun detail records.  Currently supported:
    - ``detail_type='persediaan'``: creates InventoryRecord + FIFOBatch per item.

    Extend by adding a new prefix to AKUN_DETAIL_PREFIXES and a matching
    creation block in the POST handler below.
    """
    from django.db import transaction as db_transaction
    from apps.purchase.models import FIFOBatch, ItemMasterPurchase
    from apps.inventory.models import InventoryRecord

    # Registry: kode_akun prefix → detail type.  Extendable without touching the core flow.
    AKUN_DETAIL_PREFIXES: dict[str, str] = {
        '1.1.7': 'persediaan',
        '1.1.8': 'persediaan',
        '1.1.9': 'persediaan',
        '1.1.10': 'persediaan',
        '1.2': 'aset_tetap',
        '1.2.3': 'aset_tetap',
        '1.3': 'aset_lainnya',
        '3.1.1': 'modal_disetor',
    }

    if request.method == 'POST':
        tanggal = request.POST.get('tanggal')
        eb_selection = request.POST.get('entitas_bisnis') or ''
        eb_id = _resolve_entitas_bisnis_id(eb_selection)
        rows_json = request.POST.get('rows_data', '[]')

        errors: dict = {}
        if not tanggal:
            errors['tanggal'] = 'Tanggal wajib diisi.'
        if not eb_selection:
            errors['entitas_bisnis'] = 'Entitas Bisnis wajib dipilih.'
        elif eb_id is None:
            errors['entitas_bisnis'] = 'Pilihan entitas bisnis tidak valid.'

        try:
            rows = json.loads(rows_json)
        except (ValueError, TypeError):
            rows = []

        rows = [r for r in rows if r.get('akun_id')]
        if not rows:
            errors['rows'] = 'Minimal 1 baris akun wajib diisi.'

        total_debit = sum(Decimal(str(r.get('debit') or 0)) for r in rows)
        total_kredit = sum(Decimal(str(r.get('kredit') or 0)) for r in rows)

        if not errors and total_debit != total_kredit:
            errors['balance'] = (
                f'Total Debit ({total_debit}) harus sama dengan Total Kredit ({total_kredit}).'
            )

        if errors:
            eb_list = _get_entitas_bisnis_dropdown_options()
            return render(request, 'jurnal/saldo_awal.html', {
                'today': tanggal or timezone.now().date(),
                'eb_list': eb_list,
                'errors': errors,
                'posted': True,
                'selected_entitas_bisnis': eb_selection,
                'akun_detail_prefixes_json': json.dumps(AKUN_DETAIL_PREFIXES),
            })

        with db_transaction.atomic():
            nomor = _next_nomor_transaksi('SA')
            header = JurnalHeader.objects.create(
                tanggal=tanggal,
                nomor_transaksi=nomor,
                uraian_transaksi=f'Saldo Awal — {tanggal}',
                entitas_bisnis_id=eb_id,
                is_penyesuaian=True,
                is_saldo_awal=True,
            )
            JurnalDetail.objects.bulk_create([
                JurnalDetail(
                    jurnal_header=header,
                    akun_id=row['akun_id'],
                    debit=Decimal(str(row.get('debit') or 0)),
                    kredit=Decimal(str(row.get('kredit') or 0)),
                )
                for row in rows
            ])

            # ── Persediaan detail: InventoryRecord + FIFOBatch ─────────────
            persediaan_rows = [
                r for r in rows
                if r.get('detail_type') == 'persediaan' and r.get('detail_rows')
            ]
            if persediaan_rows:
                all_item_ids = {
                    int(d['item_id'])
                    for r in persediaan_rows
                    for d in r.get('detail_rows', [])
                    if str(d.get('item_id', '')).isdigit()
                }
                items_map = {
                    item.pk: item
                    for item in ItemMasterPurchase.objects.filter(pk__in=all_item_ids)
                }
                for row in persediaan_rows:
                    for d in row.get('detail_rows', []):
                        try:
                            item_pk = int(str(d.get('item_id', '')))
                            qty = Decimal(str(d.get('qty') or 0))
                            unit_price = Decimal(str(d.get('unit_price') or 0))
                        except (ValueError, TypeError):
                            continue
                        if qty <= 0 or unit_price < 0:
                            continue
                        item = items_map.get(item_pk)
                        if not item:
                            continue
                        InventoryRecord.objects.create(
                            item=item,
                            purchase_item=None,
                            entitas_bisnis_id=eb_id,
                            quantity=qty,
                            unit_price=unit_price,
                            tanggal=tanggal,
                        )
                        FIFOBatch.objects.create(
                            purchase_item=None,
                            item=item,
                            tanggal=tanggal,
                            quantity_in=qty,
                            unit_price=unit_price,
                            remaining_qty=qty,
                        )

            # ── Aset Tetap detail: AsetTetapRecord ─────────────
            aset_tetap_rows = [
                r for r in rows
                if r.get('detail_type') == 'aset_tetap' and r.get('detail_rows')
            ]
            if aset_tetap_rows:
                all_item_ids = {
                    int(d['item_id'])
                    for r in aset_tetap_rows
                    for d in r.get('detail_rows', [])
                    if str(d.get('item_id', '')).isdigit()
                }
                items_map = {
                    item.pk: item
                    for item in ItemMasterPurchase.objects.filter(pk__in=all_item_ids)
                }
                for row in aset_tetap_rows:
                    for d in row.get('detail_rows', []):
                        try:
                            item_pk = int(str(d.get('item_id', '')))
                            qty = Decimal(str(d.get('qty') or 0))
                            unit_price = Decimal(str(d.get('unit_price') or 0))
                        except (ValueError, TypeError):
                            continue
                        if qty <= 0 or unit_price < 0:
                            continue
                        item = items_map.get(item_pk)
                        if not item:
                            continue
                        AsetTetapRecord.objects.create(
                            item=item,
                            purchase_item=None,
                            entitas_bisnis_id=eb_id,
                            quantity=qty,
                            harga_perolehan=unit_price,
                            tanggal_perolehan=tanggal,
                            masa_manfaat=item.masa_manfaat,
                            metode_penyusutan=item.metode_penyusutan,
                        )

            # ── Aset Lainnya detail: AsetLainnyaRecord ─────────────
            aset_lainnya_rows = [
                r for r in rows
                if r.get('detail_type') == 'aset_lainnya' and r.get('detail_rows')
            ]
            if aset_lainnya_rows:
                all_item_ids = {
                    int(d['item_id'])
                    for r in aset_lainnya_rows
                    for d in r.get('detail_rows', [])
                    if str(d.get('item_id', '')).isdigit()
                }
                items_map = {
                    item.pk: item
                    for item in ItemMasterPurchase.objects.filter(pk__in=all_item_ids)
                }
                for row in aset_lainnya_rows:
                    for d in row.get('detail_rows', []):
                        try:
                            item_pk = int(str(d.get('item_id', '')))
                            qty = Decimal(str(d.get('qty') or 0))
                            unit_price = Decimal(str(d.get('unit_price') or 0))
                        except (ValueError, TypeError):
                            continue
                        if qty <= 0 or unit_price < 0:
                            continue
                        item = items_map.get(item_pk)
                        if not item:
                            continue
                        AsetLainnyaRecord.objects.create(
                            item=item,
                            purchase_item=None,
                            entitas_bisnis_id=eb_id,
                            quantity=qty,
                            harga_perolehan=unit_price,
                            tanggal_perolehan=tanggal,
                            masa_manfaat=item.masa_manfaat,
                            metode_amortisasi=item.metode_amortisasi,
                        )

            # ── Modal Disetor detail: ModalDisetor + Pemilik ─────────────
            from apps.ekuitas.models import ModalDisetor as ModalDisetorModel, Pemilik as PemilikModel
            from apps.ekuitas.services import get_or_create_pemilik
            modal_disetor_rows = [
                r for r in rows
                if r.get('detail_type') == 'modal_disetor' and r.get('detail_rows')
            ]
            for row in modal_disetor_rows:
                for d in row.get('detail_rows', []):
                    nama_pemilik = str(d.get('nama_pemilik', '')).strip()
                    if not nama_pemilik:
                        continue
                    try:
                        jumlah = Decimal(str(d.get('jumlah', 0)))
                    except Exception:
                        continue
                    if jumlah <= 0:
                        continue
                    pemilik = get_or_create_pemilik(nama_pemilik)
                    ModalDisetorModel.objects.create(
                        entitas_bisnis_id=eb_id,
                        pemilik=pemilik,
                        jumlah_modal=jumlah,
                        tanggal_setor=tanggal,
                        keterangan=str(d.get('keterangan', '')).strip(),
                        jurnal_header=header,
                    )

        dj_messages.success(request, f'Saldo Awal {nomor} berhasil disimpan.')
        return redirect('jurnal:saldo_awal_list')

    eb_list = _get_entitas_bisnis_dropdown_options()
    return render(request, 'jurnal/saldo_awal.html', {
        'today': timezone.now().date(),
        'eb_list': eb_list,
        'errors': {},
        'akun_detail_prefixes_json': json.dumps(AKUN_DETAIL_PREFIXES),
        'initial_rows_json': '[]',
    })


@login_required
def saldo_awal_edit(request: HttpRequest, pk: int) -> HttpResponse:
    """Edit an existing saldo awal journal.

    On GET: load existing rows into the spreadsheet form.
    On POST: reverse old sub-records, re-create everything from submitted data.

    Note: detail sub-records (persediaan, aset_tetap, dll) cannot be automatically
    reconstructed in the form, so the user must re-enter them if applicable.
    """
    from django.db import transaction as db_transaction
    from apps.purchase.models import FIFOBatch, ItemMasterPurchase
    from apps.inventory.models import InventoryRecord

    header = get_object_or_404(JurnalHeader, pk=pk, is_saldo_awal=True)

    AKUN_DETAIL_PREFIXES: dict[str, str] = {
        '1.1.7': 'persediaan',
        '1.1.8': 'persediaan',
        '1.1.9': 'persediaan',
        '1.1.10': 'persediaan',
        '1.2': 'aset_tetap',
        '1.2.3': 'aset_tetap',
        '1.3': 'aset_lainnya',
        '3.1.1': 'modal_disetor',
    }

    if request.method == 'POST':
        tanggal = request.POST.get('tanggal')
        eb_selection = request.POST.get('entitas_bisnis') or ''
        eb_id = _resolve_entitas_bisnis_id(eb_selection)
        rows_json = request.POST.get('rows_data', '[]')

        errors: dict = {}
        if not tanggal:
            errors['tanggal'] = 'Tanggal wajib diisi.'
        if not eb_selection:
            errors['entitas_bisnis'] = 'Entitas Bisnis wajib dipilih.'
        elif eb_id is None:
            errors['entitas_bisnis'] = 'Pilihan entitas bisnis tidak valid.'

        try:
            rows = json.loads(rows_json)
        except (ValueError, TypeError):
            rows = []

        rows = [r for r in rows if r.get('akun_id')]
        if not rows:
            errors['rows'] = 'Minimal 1 baris akun wajib diisi.'

        total_debit = sum(Decimal(str(r.get('debit') or 0)) for r in rows)
        total_kredit = sum(Decimal(str(r.get('kredit') or 0)) for r in rows)

        if not errors and total_debit != total_kredit:
            errors['balance'] = (
                f'Total Debit ({total_debit}) harus sama dengan Total Kredit ({total_kredit}).'
            )

        if errors:
            eb_list = _get_entitas_bisnis_dropdown_options()
            initial_rows = _build_initial_rows_json(header)
            return render(request, 'jurnal/saldo_awal.html', {
                'today': tanggal or header.tanggal,
                'eb_list': eb_list,
                'errors': errors,
                'posted': True,
                'selected_entitas_bisnis': eb_selection,
                'akun_detail_prefixes_json': json.dumps(AKUN_DETAIL_PREFIXES),
                'edit_header': header,
                'initial_rows_json': initial_rows,
            })

        with db_transaction.atomic():
            # Reverse old sub-records and delete the journal
            _reverse_saldo_awal_records(header)
            header.details.all().delete()
            header.tanggal = tanggal
            header.entitas_bisnis_id = eb_id
            header.uraian_transaksi = f'Saldo Awal — {tanggal}'
            header.save()

            JurnalDetail.objects.bulk_create([
                JurnalDetail(
                    jurnal_header=header,
                    akun_id=row['akun_id'],
                    debit=Decimal(str(row.get('debit') or 0)),
                    kredit=Decimal(str(row.get('kredit') or 0)),
                )
                for row in rows
            ])

            # ── Persediaan ─────────────────────────────────────────────────
            persediaan_rows = [
                r for r in rows
                if r.get('detail_type') == 'persediaan' and r.get('detail_rows')
            ]
            if persediaan_rows:
                all_item_ids = {
                    int(d['item_id'])
                    for r in persediaan_rows
                    for d in r.get('detail_rows', [])
                    if str(d.get('item_id', '')).isdigit()
                }
                items_map = {
                    item.pk: item
                    for item in ItemMasterPurchase.objects.filter(pk__in=all_item_ids)
                }
                for row in persediaan_rows:
                    for d in row.get('detail_rows', []):
                        try:
                            item_pk = int(str(d.get('item_id', '')))
                            qty = Decimal(str(d.get('qty') or 0))
                            unit_price = Decimal(str(d.get('unit_price') or 0))
                        except (ValueError, TypeError):
                            continue
                        if qty <= 0 or unit_price < 0:
                            continue
                        item = items_map.get(item_pk)
                        if not item:
                            continue
                        InventoryRecord.objects.create(
                            item=item,
                            purchase_item=None,
                            entitas_bisnis_id=eb_id,
                            quantity=qty,
                            unit_price=unit_price,
                            tanggal=tanggal,
                        )
                        FIFOBatch.objects.create(
                            purchase_item=None,
                            item=item,
                            tanggal=tanggal,
                            quantity_in=qty,
                            unit_price=unit_price,
                            remaining_qty=qty,
                        )

            # ── Aset Tetap ─────────────────────────────────────────────────
            aset_tetap_rows = [
                r for r in rows
                if r.get('detail_type') == 'aset_tetap' and r.get('detail_rows')
            ]
            if aset_tetap_rows:
                all_item_ids = {
                    int(d['item_id'])
                    for r in aset_tetap_rows
                    for d in r.get('detail_rows', [])
                    if str(d.get('item_id', '')).isdigit()
                }
                items_map = {
                    item.pk: item
                    for item in ItemMasterPurchase.objects.filter(pk__in=all_item_ids)
                }
                for row in aset_tetap_rows:
                    for d in row.get('detail_rows', []):
                        try:
                            item_pk = int(str(d.get('item_id', '')))
                            qty = Decimal(str(d.get('qty') or 0))
                            unit_price = Decimal(str(d.get('unit_price') or 0))
                        except (ValueError, TypeError):
                            continue
                        if qty <= 0 or unit_price < 0:
                            continue
                        item = items_map.get(item_pk)
                        if not item:
                            continue
                        from apps.aset_tetap.models import AsetTetapRecord
                        AsetTetapRecord.objects.create(
                            item=item,
                            purchase_item=None,
                            entitas_bisnis_id=eb_id,
                            quantity=qty,
                            harga_perolehan=unit_price,
                            tanggal_perolehan=tanggal,
                            masa_manfaat=item.masa_manfaat,
                            metode_penyusutan=item.metode_penyusutan,
                        )

            # ── Aset Lainnya ───────────────────────────────────────────────
            aset_lainnya_rows = [
                r for r in rows
                if r.get('detail_type') == 'aset_lainnya' and r.get('detail_rows')
            ]
            if aset_lainnya_rows:
                all_item_ids = {
                    int(d['item_id'])
                    for r in aset_lainnya_rows
                    for d in r.get('detail_rows', [])
                    if str(d.get('item_id', '')).isdigit()
                }
                items_map = {
                    item.pk: item
                    for item in ItemMasterPurchase.objects.filter(pk__in=all_item_ids)
                }
                for row in aset_lainnya_rows:
                    for d in row.get('detail_rows', []):
                        try:
                            item_pk = int(str(d.get('item_id', '')))
                            qty = Decimal(str(d.get('qty') or 0))
                            unit_price = Decimal(str(d.get('unit_price') or 0))
                        except (ValueError, TypeError):
                            continue
                        if qty <= 0 or unit_price < 0:
                            continue
                        item = items_map.get(item_pk)
                        if not item:
                            continue
                        from apps.aset_lainnya.models import AsetLainnyaRecord
                        AsetLainnyaRecord.objects.create(
                            item=item,
                            purchase_item=None,
                            entitas_bisnis_id=eb_id,
                            quantity=qty,
                            harga_perolehan=unit_price,
                            tanggal_perolehan=tanggal,
                            masa_manfaat=item.masa_manfaat,
                            metode_amortisasi=item.metode_amortisasi,
                        )

            # ── Modal Disetor ──────────────────────────────────────────────
            from apps.ekuitas.models import ModalDisetor as ModalDisetorModel
            from apps.ekuitas.services import get_or_create_pemilik
            modal_disetor_rows = [
                r for r in rows
                if r.get('detail_type') == 'modal_disetor' and r.get('detail_rows')
            ]
            for row in modal_disetor_rows:
                for d in row.get('detail_rows', []):
                    nama_pemilik = str(d.get('nama_pemilik', '')).strip()
                    if not nama_pemilik:
                        continue
                    try:
                        jumlah = Decimal(str(d.get('jumlah', 0)))
                    except Exception:
                        continue
                    if jumlah <= 0:
                        continue
                    pemilik = get_or_create_pemilik(nama_pemilik)
                    ModalDisetorModel.objects.create(
                        entitas_bisnis_id=eb_id,
                        pemilik=pemilik,
                        jumlah_modal=jumlah,
                        tanggal_setor=tanggal,
                        keterangan=str(d.get('keterangan', '')).strip(),
                        jurnal_header=header,
                    )

        dj_messages.success(request, f'Saldo Awal {header.nomor_transaksi} berhasil diperbarui.')
        return redirect('jurnal:saldo_awal_list')

    # GET: pre-populate form with existing rows
    eb_list = _get_entitas_bisnis_dropdown_options()
    initial_rows = _build_initial_rows_json(header)

    # Build selected_entitas_bisnis value string to match the dropdown format
    selected_eb = f'lv1:{header.entitas_bisnis_id}' if header.entitas_bisnis_id else ''

    return render(request, 'jurnal/saldo_awal.html', {
        'today': header.tanggal,
        'eb_list': eb_list,
        'errors': {},
        'selected_entitas_bisnis': selected_eb,
        'akun_detail_prefixes_json': json.dumps(AKUN_DETAIL_PREFIXES),
        'edit_header': header,
        'initial_rows_json': initial_rows,
    })


def _build_initial_rows_json(header) -> str:
    """Serialise existing JurnalDetail rows to JSON for pre-populating the saldo awal form.

    For saldo awal journals, also fetches sub-records (persediaan, aset_tetap,
    aset_lainnya, modal_disetor) and attaches them as ``detail_rows`` so the edit
    form can pre-populate the detail modals.
    """
    from collections import defaultdict
    from apps.inventory.models import InventoryRecord
    from apps.aset_tetap.models import AsetTetapRecord as ATRecord
    from apps.aset_lainnya.models import AsetLainnyaRecord as ALRecord
    from apps.ekuitas.models import ModalDisetor

    AKUN_DETAIL_PREFIXES: dict[str, str] = {
        '1.1.7': 'persediaan',
        '1.1.8': 'persediaan',
        '1.1.9': 'persediaan',
        '1.1.10': 'persediaan',
        '1.2': 'aset_tetap',
        '1.2.3': 'aset_tetap',
        '1.3': 'aset_lainnya',
        '3.1.1': 'modal_disetor',
    }

    def _detail_type(kode: str) -> str | None:
        for prefix, dtype in AKUN_DETAIL_PREFIXES.items():
            if kode == prefix or kode.startswith(prefix + '.'):
                return dtype
        return None

    eb_id = header.entitas_bisnis_id
    tanggal = header.tanggal

    # Build lookup maps: coa_account_id → list of sub-records
    inv_by_coa: dict = defaultdict(list)
    for inv in InventoryRecord.objects.filter(
        purchase_item=None, entitas_bisnis_id=eb_id, tanggal=tanggal
    ).select_related('item'):
        inv_by_coa[inv.item.coa_account_id].append(inv)

    at_by_coa: dict = defaultdict(list)
    for at in ATRecord.objects.filter(
        purchase_item=None, entitas_bisnis_id=eb_id, tanggal_perolehan=tanggal
    ).select_related('item'):
        at_by_coa[at.item.coa_account_id].append(at)

    al_by_coa: dict = defaultdict(list)
    for al in ALRecord.objects.filter(
        purchase_item=None, entitas_bisnis_id=eb_id, tanggal_perolehan=tanggal
    ).select_related('item'):
        al_by_coa[al.item.coa_account_id].append(al)

    modal_rows = [
        {
            'nama_pemilik': md.pemilik.nama,
            'jumlah': float(md.jumlah_modal),
            'keterangan': md.keterangan or '',
        }
        for md in ModalDisetor.objects.filter(jurnal_header=header).select_related('pemilik')
    ]

    # Track whether modal_disetor rows have been attached (attach to first matching row only)
    modal_attached = False

    rows = []
    for detail in header.details.select_related('akun').order_by('pk'):
        kode = detail.akun.kode_akun or ''
        dtype = _detail_type(kode)
        row: dict = {
            'akun_id': str(detail.akun_id),
            'kode_akun': kode,
            'nama_akun': detail.akun.nama,
            'debit': str(detail.debit),
            'kredit': str(detail.kredit),
        }

        if dtype == 'persediaan' and inv_by_coa.get(detail.akun_id):
            row['detail_type'] = 'persediaan'
            row['detail_rows'] = [
                {
                    'item_id': str(inv.item.pk),
                    'item_text': f'{inv.item.item_id} — {inv.item.nama}',
                    'qty': float(inv.quantity),
                    'unit_price': float(inv.unit_price),
                    'total': float(inv.total_value),
                }
                for inv in inv_by_coa[detail.akun_id]
            ]

        elif dtype == 'aset_tetap' and at_by_coa.get(detail.akun_id):
            row['detail_type'] = 'aset_tetap'
            row['detail_rows'] = [
                {
                    'item_id': str(at.item.pk),
                    'item_text': f'{at.item.item_id} — {at.item.nama}',
                    'qty': float(at.quantity),
                    'unit_price': float(at.harga_perolehan),
                    'total': float(at.total_value),
                }
                for at in at_by_coa[detail.akun_id]
            ]

        elif dtype == 'aset_lainnya' and al_by_coa.get(detail.akun_id):
            row['detail_type'] = 'aset_lainnya'
            row['detail_rows'] = [
                {
                    'item_id': str(al.item.pk),
                    'item_text': f'{al.item.item_id} — {al.item.nama}',
                    'qty': float(al.quantity),
                    'unit_price': float(al.harga_perolehan),
                    'total': float(al.total_value),
                }
                for al in al_by_coa[detail.akun_id]
            ]

        elif dtype == 'modal_disetor' and modal_rows and not modal_attached:
            row['detail_type'] = 'modal_disetor'
            row['detail_rows'] = modal_rows
            modal_attached = True

        rows.append(row)
    return json.dumps(rows)


# ── Neraca (Balance Sheet) ──────────────────────────────────────────────────

@login_required
def neraca(request: HttpRequest) -> HttpResponse:
    """Balance Sheet: Aset, Kewajiban, Ekuitas with balance check."""
    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    eb_filter = request.GET.get('entitas_bisnis', '')

    period_filter: dict = {}
    if tanggal_sampai:
        period_filter['jurnal_header__tanggal__lte'] = tanggal_sampai
    if eb_filter:
        period_filter['jurnal_header__entitas_bisnis_id'] = eb_filter

    # Aggregate all journal details up to tanggal_sampai
    agg = (
        JurnalDetail.objects
        .filter(**period_filter)
        .values('akun_id', 'akun__kode_akun', 'akun__nama')
        .annotate(
            total_debit=Coalesce(Sum('debit'), Value(Decimal('0')), output_field=DecimalField()),
            total_kredit=Coalesce(Sum('kredit'), Value(Decimal('0')), output_field=DecimalField()),
        )
    )

    akun_balances: dict[str, dict] = {}
    for row in agg:
        kode = row['akun__kode_akun'] or ''
        akun_balances[kode] = {
            'kode': kode,
            'nama': row['akun__nama'] or '',
            'debit': row['total_debit'],
            'kredit': row['total_kredit'],
        }

    def _collect_net(prefix: str, normal: str = 'debit') -> tuple[list[dict], Decimal]:
        """Collect akun items matching prefix and compute net balance.
        normal='debit': balance = debit - kredit (assets)
        normal='kredit': balance = kredit - debit (liabilities/equity)
        """
        items = []
        total = Decimal('0')
        for kode, data in sorted(akun_balances.items(), key=lambda x: x[0]):
            if kode.startswith(prefix + '.') or kode == prefix:
                if normal == 'debit':
                    net = data['debit'] - data['kredit']
                else:
                    net = data['kredit'] - data['debit']
                items.append({'kode': kode, 'nama': data['nama'], 'saldo': net})
                total += net
        return items, total

    # ASET
    aset_lancar_items, total_aset_lancar = _collect_net('1.1', 'debit')
    aset_tetap_items, total_aset_tetap = _collect_net('1.2', 'debit')
    aset_lain_items, total_aset_lain = _collect_net('1.3', 'debit')
    total_aset = total_aset_lancar + total_aset_tetap + total_aset_lain

    # KEWAJIBAN
    kwj_pendek_items, total_kwj_pendek = _collect_net('2.1', 'kredit')
    kwj_panjang_items, total_kwj_panjang = _collect_net('2.2', 'kredit')
    total_kewajiban = total_kwj_pendek + total_kwj_panjang

    # EKUITAS (3.1.x)
    ekuitas_items, total_ekuitas_akun = _collect_net('3.1', 'kredit')

    # Laba Tahun Berjalan = Pendapatan (4.x) - Beban (5.x)
    pendapatan_total = Decimal('0')
    beban_total = Decimal('0')
    for kode, data in akun_balances.items():
        if kode.startswith('4.'):
            pendapatan_total += data['kredit'] - data['debit']
        elif kode.startswith('5.'):
            beban_total += data['debit'] - data['kredit']

    laba_berjalan = pendapatan_total - beban_total
    total_ekuitas = total_ekuitas_akun + laba_berjalan

    total_kewajiban_ekuitas = total_kewajiban + total_ekuitas
    selisih = total_aset - total_kewajiban_ekuitas
    is_balanced = abs(selisih) < Decimal('0.01')

    eb_list = EBModel.objects.filter(status_aktif=True).order_by('nama')

    return render(request, 'jurnal/neraca.html', {
        'tanggal_sampai': tanggal_sampai,
        'eb_filter': eb_filter,
        'eb_list': eb_list,
        # Aset
        'aset_lancar_items': aset_lancar_items, 'total_aset_lancar': total_aset_lancar,
        'aset_tetap_items': aset_tetap_items, 'total_aset_tetap': total_aset_tetap,
        'aset_lain_items': aset_lain_items, 'total_aset_lain': total_aset_lain,
        'total_aset': total_aset,
        # Kewajiban
        'kwj_pendek_items': kwj_pendek_items, 'total_kwj_pendek': total_kwj_pendek,
        'kwj_panjang_items': kwj_panjang_items, 'total_kwj_panjang': total_kwj_panjang,
        'total_kewajiban': total_kewajiban,
        # Ekuitas
        'ekuitas_items': ekuitas_items, 'total_ekuitas_akun': total_ekuitas_akun,
        'laba_berjalan': laba_berjalan, 'total_ekuitas': total_ekuitas,
        # Balance
        'total_kewajiban_ekuitas': total_kewajiban_ekuitas,
        'selisih': selisih,
        'is_balanced': is_balanced,
    })


# ── Laporan Perubahan Ekuitas (LPE) ─────────────────────────────────────────

@login_required
def laporan_perubahan_ekuitas(request: HttpRequest) -> HttpResponse:
    """Statement of Changes in Equity."""
    tanggal_dari = request.GET.get('tanggal_dari', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    eb_filter = request.GET.get('entitas_bisnis', '')

    # Build period filter for the reporting period
    period_filter: dict = {}
    if tanggal_dari:
        period_filter['jurnal_header__tanggal__gte'] = tanggal_dari
    if tanggal_sampai:
        period_filter['jurnal_header__tanggal__lte'] = tanggal_sampai
    if eb_filter:
        period_filter['jurnal_header__entitas_bisnis_id'] = eb_filter

    # Build filter for before-period (saldo awal)
    before_filter: dict = {}
    if eb_filter:
        before_filter['jurnal_header__entitas_bisnis_id'] = eb_filter

    # 1. Saldo Awal Ekuitas = net balance of 3.x accounts BEFORE tanggal_dari
    saldo_awal_ekuitas = Decimal('0')
    if tanggal_dari:
        agg_before = (
            JurnalDetail.objects
            .filter(
                akun__kode_akun__startswith='3.',
                jurnal_header__tanggal__lt=tanggal_dari,
                **before_filter,
            )
            .aggregate(
                total_debit=Coalesce(Sum('debit'), Value(Decimal('0')), output_field=DecimalField()),
                total_kredit=Coalesce(Sum('kredit'), Value(Decimal('0')), output_field=DecimalField()),
            )
        )
        saldo_awal_ekuitas = agg_before['total_kredit'] - agg_before['total_debit']

    # 2. Laba Bersih Periode = Pendapatan (4.x) - Beban (5.x) within period
    pendapatan_agg = (
        JurnalDetail.objects
        .filter(akun__kode_akun__startswith='4.', **period_filter)
        .aggregate(
            total_debit=Coalesce(Sum('debit'), Value(Decimal('0')), output_field=DecimalField()),
            total_kredit=Coalesce(Sum('kredit'), Value(Decimal('0')), output_field=DecimalField()),
        )
    )
    beban_agg = (
        JurnalDetail.objects
        .filter(akun__kode_akun__startswith='5.', **period_filter)
        .aggregate(
            total_debit=Coalesce(Sum('debit'), Value(Decimal('0')), output_field=DecimalField()),
            total_kredit=Coalesce(Sum('kredit'), Value(Decimal('0')), output_field=DecimalField()),
        )
    )
    total_pendapatan = pendapatan_agg['total_kredit'] - pendapatan_agg['total_debit']
    total_beban = beban_agg['total_debit'] - beban_agg['total_kredit']
    laba_bersih = total_pendapatan - total_beban

    # 3. Prive / Drawing = debit balance of 3.1.4.x accounts within period
    prive_agg = (
        JurnalDetail.objects
        .filter(akun__kode_akun__startswith='3.1.4', **period_filter)
        .aggregate(
            total_debit=Coalesce(Sum('debit'), Value(Decimal('0')), output_field=DecimalField()),
            total_kredit=Coalesce(Sum('kredit'), Value(Decimal('0')), output_field=DecimalField()),
        )
    )
    prive = prive_agg['total_debit'] - prive_agg['total_kredit']

    # 4. Saldo Akhir
    saldo_akhir_ekuitas = saldo_awal_ekuitas + laba_bersih - prive

    eb_list = EBModel.objects.filter(status_aktif=True).order_by('nama')

    return render(request, 'jurnal/laporan_perubahan_ekuitas.html', {
        'tanggal_dari': tanggal_dari,
        'tanggal_sampai': tanggal_sampai,
        'eb_filter': eb_filter,
        'eb_list': eb_list,
        'saldo_awal_ekuitas': saldo_awal_ekuitas,
        'laba_bersih': laba_bersih,
        'prive': prive,
        'saldo_akhir_ekuitas': saldo_akhir_ekuitas,
    })


# ── Analisis Keuangan (Financial Ratio Analysis) ─────────────────────────────

@login_required
def analisis_keuangan(request: HttpRequest) -> HttpResponse:
    """Financial ratio analysis: Likuiditas, Profitabilitas, Efisiensi, Solvabilitas."""
    import json as _json

    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    tanggal_dari = request.GET.get('tanggal_dari', '')
    eb_filter = request.GET.get('entitas_bisnis', '')

    # ── Balance-sheet period filter (cumulative up to tanggal_sampai) ──
    bs_filter: dict = {}
    if tanggal_sampai:
        bs_filter['jurnal_header__tanggal__lte'] = tanggal_sampai
    if eb_filter:
        bs_filter['jurnal_header__entitas_bisnis_id'] = eb_filter

    # ── P&L period filter (tanggal_dari – tanggal_sampai) ──
    pl_filter: dict = {}
    if tanggal_dari:
        pl_filter['jurnal_header__tanggal__gte'] = tanggal_dari
    if tanggal_sampai:
        pl_filter['jurnal_header__tanggal__lte'] = tanggal_sampai
    if eb_filter:
        pl_filter['jurnal_header__entitas_bisnis_id'] = eb_filter

    # ── Aggregate balance sheet ──
    bs_agg = (
        JurnalDetail.objects
        .filter(**bs_filter)
        .values('akun__kode_akun')
        .annotate(
            total_debit=Coalesce(Sum('debit'), Value(Decimal('0')), output_field=DecimalField()),
            total_kredit=Coalesce(Sum('kredit'), Value(Decimal('0')), output_field=DecimalField()),
        )
    )
    bs_balances: dict[str, tuple[Decimal, Decimal]] = {}
    for row in bs_agg:
        kode = row['akun__kode_akun'] or ''
        bs_balances[kode] = (row['total_debit'], row['total_kredit'])

    def _bs_net(prefix: str, normal: str = 'debit') -> Decimal:
        total = Decimal('0')
        for kode, (d, k) in bs_balances.items():
            if kode.startswith(prefix + '.') or kode == prefix:
                total += (d - k) if normal == 'debit' else (k - d)
        return total

    total_aset_lancar = _bs_net('1.1', 'debit')
    # Persediaan typically lives in 1.1.3.x; use best-effort prefix
    total_persediaan = _bs_net('1.1.3', 'debit')
    total_aset_tetap = _bs_net('1.2', 'debit')
    total_aset_lain = _bs_net('1.3', 'debit')
    total_aset = total_aset_lancar + total_aset_tetap + total_aset_lain

    total_kwj_pendek = _bs_net('2.1', 'kredit')
    total_kwj_panjang = _bs_net('2.2', 'kredit')
    total_kewajiban = total_kwj_pendek + total_kwj_panjang

    total_ekuitas_akun = _bs_net('3.1', 'kredit')
    # Laba berjalan is included in ekuitas for ratio purposes
    for kode, (d, k) in bs_balances.items():
        pass  # pendapatan/beban computed separately below
    pendapatan_bs = Decimal('0')
    beban_bs = Decimal('0')
    for kode, (d, k) in bs_balances.items():
        if kode.startswith('4.'):
            pendapatan_bs += k - d
        elif kode.startswith('5.'):
            beban_bs += d - k
    laba_berjalan_bs = pendapatan_bs - beban_bs
    total_ekuitas = total_ekuitas_akun + laba_berjalan_bs

    # ── Aggregate P&L ──
    pl_agg = (
        JurnalDetail.objects
        .filter(**pl_filter)
        .values('akun__kode_akun')
        .annotate(
            total_debit=Coalesce(Sum('debit'), Value(Decimal('0')), output_field=DecimalField()),
            total_kredit=Coalesce(Sum('kredit'), Value(Decimal('0')), output_field=DecimalField()),
        )
    )
    pl_balances: dict[str, tuple[Decimal, Decimal]] = {}
    for row in pl_agg:
        kode = row['akun__kode_akun'] or ''
        pl_balances[kode] = (row['total_debit'], row['total_kredit'])

    def _pl_sum(prefix: str, net_type: str = 'kredit') -> Decimal:
        total = Decimal('0')
        for kode, (d, k) in pl_balances.items():
            if kode.startswith(prefix + '.') or kode == prefix:
                total += (k - d) if net_type == 'kredit' else (d - k)
        return total

    total_pendapatan_op = _pl_sum('4.1', 'kredit')
    hpp_total = sum(
        _pl_sum(f'5.1.{i}', 'debit') for i in range(1, 6)
    )
    laba_kotor = total_pendapatan_op - hpp_total
    total_pendapatan_non_op = _pl_sum('4.2', 'kredit')
    total_pendapatan = total_pendapatan_op + total_pendapatan_non_op
    total_beban = sum(
        (d - k) for kode, (d, k) in pl_balances.items() if kode.startswith('5.')
    )
    laba_bersih = total_pendapatan - total_beban

    # ── Compute ratios (safe division) ──
    def _ratio(num: Decimal, den: Decimal, pct: bool = False, decimals: int = 2) -> float | None:
        if den == 0:
            return None
        val = float(num / den)
        return round(val * 100 if pct else val, decimals)

    ratios: dict[str, float | None] = {
        # Likuiditas
        'current_ratio': _ratio(total_aset_lancar, total_kwj_pendek),
        'quick_ratio': _ratio(total_aset_lancar - total_persediaan, total_kwj_pendek),
        # Profitabilitas
        'gross_profit_margin': _ratio(laba_kotor, total_pendapatan, pct=True),
        'net_profit_margin': _ratio(laba_bersih, total_pendapatan, pct=True),
        'roe': _ratio(laba_bersih, total_ekuitas, pct=True),
        'roa': _ratio(laba_bersih, total_aset, pct=True),
        # Efisiensi
        'asset_turnover': _ratio(total_pendapatan, total_aset),
        # Solvabilitas
        'der': _ratio(total_kewajiban, total_ekuitas),
        'dar': _ratio(total_kewajiban, total_aset, pct=True),
    }

    # Radar chart data (normalize each axis 0–100 for display)
    def _radar_score(key: str, ideal_min: float, ideal_max: float, invert: bool = False) -> float:
        val = ratios.get(key)
        if val is None:
            return 0.0
        if invert:
            score = max(0.0, min(100.0, (ideal_max - val) / max(ideal_max - ideal_min, 0.01) * 100))
        else:
            score = max(0.0, min(100.0, (val - ideal_min) / max(ideal_max - ideal_min, 0.01) * 100))
        return round(score, 1)

    radar_data = [
        _radar_score('current_ratio', 0, 3),           # Likuiditas
        _radar_score('der', 0, 3, invert=True),         # Solvabilitas (lower DER = better)
        _radar_score('net_profit_margin', -20, 30),     # Profitabilitas
        _radar_score('asset_turnover', 0, 2),           # Efisiensi
        _radar_score('roe', -10, 30),                   # ROE
        _radar_score('roa', -5, 20),                    # ROA
    ]

    eb_list = EBModel.objects.filter(status_aktif=True).order_by('nama')

    return render(request, 'jurnal/analisis_keuangan.html', {
        'tanggal_dari': tanggal_dari,
        'tanggal_sampai': tanggal_sampai,
        'eb_filter': eb_filter,
        'eb_list': eb_list,
        # Balance sheet intermediate values
        'total_aset_lancar': total_aset_lancar,
        'total_aset': total_aset,
        'total_kwj_pendek': total_kwj_pendek,
        'total_kewajiban': total_kewajiban,
        'total_ekuitas': total_ekuitas,
        # P&L intermediate values
        'total_pendapatan': total_pendapatan,
        'laba_kotor': laba_kotor,
        'laba_bersih': laba_bersih,
        # Ratios
        'ratios': ratios,
        'ratios_json': _json.dumps(ratios),
        'radar_json': _json.dumps(radar_data),
    })


# ── History Jurnal Terhapus ──────────────────────────────────────────────────

@login_required
def jurnal_terhapus_list(request: HttpRequest) -> HttpResponse:
    """List all deleted journal records with filters."""
    from .models import JurnalTerhapus

    qs = JurnalTerhapus.objects.select_related('deleted_by').order_by('-deleted_at')

    # Filters
    module_filter = request.GET.get('module', '')
    tanggal_dari = request.GET.get('tanggal_dari', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    search = request.GET.get('q', '')

    if module_filter:
        qs = qs.filter(module=module_filter)
    if tanggal_dari:
        qs = qs.filter(tanggal__gte=tanggal_dari)
    if tanggal_sampai:
        qs = qs.filter(tanggal__lte=tanggal_sampai)
    if search:
        qs = qs.filter(
            Q(nomor_transaksi__icontains=search) |
            Q(uraian_transaksi__icontains=search) |
            Q(entitas_bisnis_nama__icontains=search)
        )

    module_choices = JurnalTerhapus.MODULE_CHOICES

    return render(request, 'jurnal/jurnal_terhapus.html', {
        'records': qs,
        'module_filter': module_filter,
        'tanggal_dari': tanggal_dari,
        'tanggal_sampai': tanggal_sampai,
        'search': search,
        'module_choices': module_choices,
    })


# ── Buku Besar (General Ledger) ─────────────────────────────────────────────

@login_required
def buku_besar(request: HttpRequest) -> HttpResponse:
    """General Ledger — show all journal entries for a specific account with running balance."""
    from apps.master_data.utils import get_akun_sorted

    akun_id = request.GET.get('akun', '')
    tanggal_dari = request.GET.get('tanggal_dari', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    eb_filter = request.GET.get('entitas_bisnis', '')

    akun_list = get_akun_sorted()
    eb_list = EBModel.objects.filter(status_aktif=True).order_by('nama')

    rows = []
    selected_akun = None
    saldo_awal = Decimal('0')

    if akun_id:
        from apps.master_data.models import Akun
        try:
            selected_akun = Akun.objects.get(pk=akun_id)
        except Akun.DoesNotExist:
            selected_akun = None

    if selected_akun:
        # Determine normal balance direction: debit-normal for aset(1)/beban(5), kredit-normal for others
        kat = selected_akun.kategori_akun or ''
        is_debit_normal = kat in ('aset', 'beban')

        # Build filters
        filters: dict = {'akun': selected_akun}
        if eb_filter:
            filters['jurnal_header__entitas_bisnis_id'] = eb_filter

        # Saldo awal: sum of all entries before tanggal_dari
        if tanggal_dari:
            sa_qs = JurnalDetail.objects.filter(
                jurnal_header__tanggal__lt=tanggal_dari,
                **filters,
            ).aggregate(
                total_debit=Coalesce(Sum('debit'), Value(Decimal('0')), output_field=DecimalField()),
                total_kredit=Coalesce(Sum('kredit'), Value(Decimal('0')), output_field=DecimalField()),
            )
            if is_debit_normal:
                saldo_awal = sa_qs['total_debit'] - sa_qs['total_kredit']
            else:
                saldo_awal = sa_qs['total_kredit'] - sa_qs['total_debit']

        # Period entries
        period_filters = dict(filters)
        if tanggal_dari:
            period_filters['jurnal_header__tanggal__gte'] = tanggal_dari
        if tanggal_sampai:
            period_filters['jurnal_header__tanggal__lte'] = tanggal_sampai

        entries = (
            JurnalDetail.objects
            .filter(**period_filters)
            .select_related('jurnal_header', 'jurnal_header__entitas_bisnis')
            .order_by('jurnal_header__tanggal', 'jurnal_header__pk')
        )

        running = saldo_awal
        for entry in entries:
            if is_debit_normal:
                running += entry.debit - entry.kredit
            else:
                running += entry.kredit - entry.debit
            rows.append({
                'tanggal': entry.jurnal_header.tanggal,
                'nomor_transaksi': entry.jurnal_header.nomor_transaksi,
                'uraian': entry.jurnal_header.uraian_transaksi,
                'entitas_bisnis': entry.jurnal_header.entitas_bisnis,
                'debit': entry.debit,
                'kredit': entry.kredit,
                'saldo': running,
                'header_pk': entry.jurnal_header.pk,
            })

    return render(request, 'jurnal/buku_besar.html', {
        'akun_list': akun_list,
        'eb_list': eb_list,
        'selected_akun': selected_akun,
        'akun_id': akun_id,
        'tanggal_dari': tanggal_dari,
        'tanggal_sampai': tanggal_sampai,
        'eb_filter': eb_filter,
        'saldo_awal': saldo_awal,
        'rows': rows,
        'total_debit': sum((r['debit'] for r in rows), Decimal('0')),
        'total_kredit': sum((r['kredit'] for r in rows), Decimal('0')),
        'saldo_akhir': rows[-1]['saldo'] if rows else saldo_awal,
    })
