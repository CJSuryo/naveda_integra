from datetime import date, timedelta
from decimal import Decimal

from django.test import TestCase

from apps.entitas_bisnis.models import EntitasBisnis, TipeEntitas
from apps.master_data.models import Akun

from .models import PiutangHeader, PiutangDetail, PiutangAuditLog
from .services import create_manual_piutang, create_piutang_payment
from apps.piutang.services import compute_penyisihan_for_piutang
from apps.piutang.models import PenyisihanRateConfig
from apps.piutang.services import compute_present_value, compute_amortization_schedule_pv


def make_fixtures():
    tipe = TipeEntitas.objects.create(nama='Pelanggan')
    eb = EntitasBisnis.objects.create(nama='PT Klien', tipe_entitas=tipe, relasi='pelanggan')
    coa_piutang = Akun.objects.create(kategori_id='aset', nama='Piutang Dagang', kode_akun='1.2.1')
    coa_kas = Akun.objects.create(kategori_id='aset', nama='Kas', kode_akun='1.1.1')
    coa_pendapatan = Akun.objects.create(kategori_id='pendapatan', nama='Pendapatan Jasa', kode_akun='4.1.1')
    return {
        'tipe': tipe, 'eb': eb,
        'coa_piutang': coa_piutang, 'coa_kas': coa_kas, 'coa_pendapatan': coa_pendapatan,
    }


class CreateManualPiutangTests(TestCase):
    def setUp(self):
        self.f = make_fixtures()

    def test_creates_header_with_correct_fields(self):
        piutang = create_manual_piutang(
            tanggal=date(2026, 6, 7),
            entitas_bisnis=self.f['eb'],
            debitur='PT Klien',
            deskripsi='Test piutang',
            coa_piutang_account=self.f['coa_piutang'],
            jatuh_tempo=date(2026, 7, 7),
            details=[{'deskripsi': 'Jasa konsultasi', 'jumlah': Decimal('1000000')}],
        )
        self.assertIsNotNone(piutang.pk)
        self.assertTrue(piutang.nomor_piutang.startswith('TRX-PIU-'))
        self.assertEqual(piutang.jumlah_pokok, Decimal('1000000'))
        self.assertEqual(piutang.status, 'draft')
        self.assertEqual(piutang.details.count(), 1)

    def test_creates_audit_log(self):
        create_manual_piutang(
            tanggal=date(2026, 6, 7),
            entitas_bisnis=self.f['eb'],
            debitur='',
            deskripsi='',
            coa_piutang_account=self.f['coa_piutang'],
            jatuh_tempo=None,
            details=[{'deskripsi': 'Item', 'jumlah': Decimal('500000')}],
        )
        self.assertEqual(PiutangAuditLog.objects.filter(action='CREATED').count(), 1)

    def test_raises_if_no_details(self):
        with self.assertRaises(ValueError):
            create_manual_piutang(
                tanggal=date(2026, 6, 7), entitas_bisnis=self.f['eb'],
                debitur='', deskripsi='', coa_piutang_account=self.f['coa_piutang'],
                jatuh_tempo=None, details=[],
            )

    def test_auto_number_increments(self):
        p1 = create_manual_piutang(
            tanggal=date(2026, 6, 7), entitas_bisnis=None, debitur='X',
            deskripsi='', coa_piutang_account=self.f['coa_piutang'], jatuh_tempo=None,
            details=[{'deskripsi': 'A', 'jumlah': Decimal('100')}],
        )
        p2 = create_manual_piutang(
            tanggal=date(2026, 6, 7), entitas_bisnis=None, debitur='Y',
            deskripsi='', coa_piutang_account=self.f['coa_piutang'], jatuh_tempo=None,
            details=[{'deskripsi': 'B', 'jumlah': Decimal('200')}],
        )
        n1 = int(p1.nomor_piutang.rsplit('-', 1)[1])
        n2 = int(p2.nomor_piutang.rsplit('-', 1)[1])
        self.assertEqual(n2, n1 + 1)


class CreatePiutangPaymentTests(TestCase):
    def setUp(self):
        self.f = make_fixtures()
        self.piutang = create_manual_piutang(
            tanggal=date(2026, 6, 7), entitas_bisnis=self.f['eb'],
            debitur='PT Klien', deskripsi='', coa_piutang_account=self.f['coa_piutang'],
            jatuh_tempo=date(2026, 7, 7),
            details=[{'deskripsi': 'Jasa', 'jumlah': Decimal('1000000')}],
        )
        self.piutang.status = 'open'
        self.piutang.save()

    def test_creates_penerimaan_record(self):
        create_piutang_payment(
            self.piutang,
            {'tanggal_terima': date(2026, 6, 10), 'jumlah_diterima': Decimal('400000'),
             'payment_account': self.f['coa_kas'], 'metode_penerimaan': 'transfer',
             'nomor_referensi': 'TRF-001', 'catatan': ''},
        )
        self.assertEqual(self.piutang.penerimaan.count(), 1)

    def test_updates_jumlah_terbayar(self):
        create_piutang_payment(
            self.piutang,
            {'tanggal_terima': date(2026, 6, 10), 'jumlah_diterima': Decimal('400000'),
             'payment_account': self.f['coa_kas'], 'metode_penerimaan': 'transfer',
             'nomor_referensi': '', 'catatan': ''},
        )
        self.piutang.refresh_from_db()
        self.assertEqual(self.piutang.jumlah_terbayar, Decimal('400000'))
        self.assertEqual(self.piutang.status, 'partial')

    def test_status_becomes_paid_when_fully_settled(self):
        create_piutang_payment(
            self.piutang,
            {'tanggal_terima': date(2026, 6, 10), 'jumlah_diterima': Decimal('1000000'),
             'payment_account': self.f['coa_kas'], 'metode_penerimaan': 'tunai',
             'nomor_referensi': '', 'catatan': ''},
        )
        self.piutang.refresh_from_db()
        self.assertEqual(self.piutang.status, 'paid')

    def test_raises_if_exceeds_sisa(self):
        with self.assertRaises(ValueError):
            create_piutang_payment(
                self.piutang,
                {'tanggal_terima': date(2026, 6, 10), 'jumlah_diterima': Decimal('2000000'),
                 'payment_account': self.f['coa_kas'], 'metode_penerimaan': 'transfer',
                 'nomor_referensi': '', 'catatan': ''},
            )

    def test_generates_journal(self):
        from apps.jurnal.models import JurnalHeader
        create_piutang_payment(
            self.piutang,
            {'tanggal_terima': date(2026, 6, 10), 'jumlah_diterima': Decimal('500000'),
             'payment_account': self.f['coa_kas'], 'metode_penerimaan': 'transfer',
             'nomor_referensi': '', 'catatan': ''},
        )
        penerimaan = self.piutang.penerimaan.first()
        self.assertIsNotNone(penerimaan.jurnal_header)
        details = penerimaan.jurnal_header.details.all()
        debits = [d for d in details if d.debit > 0]
        credits = [d for d in details if d.kredit > 0]
        self.assertEqual(len(debits), 1)
        self.assertEqual(len(credits), 1)
        self.assertEqual(debits[0].akun, self.f['coa_kas'])
        self.assertEqual(credits[0].akun, self.f['coa_piutang'])


from .services import compute_bagian_lancar, write_off_piutang, reverse_piutang_payment


class ComputeBagianLancarTests(TestCase):
    def setUp(self):
        self.f = make_fixtures()

    def test_full_amount_when_due_within_12_months(self):
        p = create_manual_piutang(
            tanggal=date(2026, 1, 1), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'],
            jatuh_tempo=date(2026, 6, 1),
            details=[{'deskripsi': 'X', 'jumlah': Decimal('500000')}],
        )
        bagian = compute_bagian_lancar(p)
        self.assertEqual(bagian, Decimal('500000'))

    def test_zero_when_due_beyond_12_months(self):
        p = create_manual_piutang(
            tanggal=date(2026, 1, 1), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'],
            jatuh_tempo=date(2028, 1, 1),
            details=[{'deskripsi': 'X', 'jumlah': Decimal('500000')}],
        )
        bagian = compute_bagian_lancar(p)
        self.assertEqual(bagian, Decimal('0'))


# ── Task 4: Aging bucket tests ────────────────────────────────────────────────

from apps.piutang.services import _classify_bucket, get_piutang_aging


class ClassifyBucketTest(TestCase):
    def test_future(self):
        future = date(date.today().year + 1, 1, 1)
        assert _classify_bucket(future, date.today()) == 'current'

    def test_today(self):
        assert _classify_bucket(date.today(), date.today()) == 'current'

    def test_1_day_overdue(self):
        assert _classify_bucket(date.today() - timedelta(days=1), date.today()) == '1_30'

    def test_31_days(self):
        assert _classify_bucket(date.today() - timedelta(days=31), date.today()) == '31_60'

    def test_over_365(self):
        assert _classify_bucket(date.today() - timedelta(days=400), date.today()) == 'over_365'

    def test_none_returns_current(self):
        assert _classify_bucket(None, date.today()) == 'current'


class GetPiutangAgingTest(TestCase):
    def test_returns_7_buckets(self):
        result = get_piutang_aging()
        expected_keys = {'current', '1_30', '31_60', '61_90', '91_180', '181_365', 'over_365'}
        assert set(result.keys()) == expected_keys

    def test_each_bucket_is_list(self):
        result = get_piutang_aging()
        for v in result.values():
            assert isinstance(v, list)


# ── Task 3: Schedule helpers ───────────────────────────────────────────────────

from apps.piutang.services import _add_months, compute_angsuran_schedule


class AddMonthsTest(TestCase):
    def test_simple(self):
        assert _add_months(date(2026, 1, 31), 1) == date(2026, 2, 28)

    def test_year_rollover(self):
        assert _add_months(date(2025, 11, 30), 3) == date(2026, 2, 28)


class ComputeAngsuranScheduleTest(TestCase):
    def _make_piutang(self, jumlah_pokok, jenis_bunga='tanpa_bunga', suku_bunga=0,
                      periode='bulanan', tanggal=None, jatuh_tempo=None):
        from apps.master_data.models import Akun
        akun, _ = Akun.objects.get_or_create(
            kode_akun='1100', defaults={'nama': 'Piutang Usaha', 'kategori_id': 'aset'}
        )
        return PiutangHeader(
            nomor_piutang='TEST-001',
            tanggal=tanggal or date(2026, 1, 1),
            jatuh_tempo=jatuh_tempo or date(2026, 12, 31),
            jumlah_pokok=Decimal(str(jumlah_pokok)),
            jumlah_terbayar=Decimal('0'),
            jenis_jangka_waktu='long_term',
            jenis_bunga=jenis_bunga,
            suku_bunga=Decimal(str(suku_bunga)),
            periode_angsuran=periode,
            coa_piutang_account=akun,
            status='open',
        )

    def test_tanpa_bunga_12_bulan(self):
        p = self._make_piutang(12_000_000, tanggal=date(2026, 1, 1), jatuh_tempo=date(2026, 12, 31))
        rows = compute_angsuran_schedule(p)
        assert len(rows) == 11  # 11 monthly periods from Jan to Dec
        assert all(r['bunga'] == 0 for r in rows)
        total_pokok = sum(r['pokok'] for r in rows)
        assert total_pokok == Decimal('12000000')

    def test_no_schedule_without_jatuh_tempo(self):
        p = self._make_piutang(1_000_000, jatuh_tempo=None)
        p.jatuh_tempo = None
        assert compute_angsuran_schedule(p) == []

    def test_flat_interest(self):
        p = self._make_piutang(
            12_000_000, jenis_bunga='flat', suku_bunga=12,
            tanggal=date(2026, 1, 1), jatuh_tempo=date(2026, 12, 31),
        )
        rows = compute_angsuran_schedule(p)
        assert all(r['bunga'] > 0 for r in rows)

    def test_status_akan_datang(self):
        future = date.today().replace(year=date.today().year + 1)
        p = self._make_piutang(
            6_000_000, tanggal=date.today(), jatuh_tempo=future,
        )
        rows = compute_angsuran_schedule(p)
        assert all(r['status'] == 'akan_datang' for r in rows)

    def test_anuitas_last_row_zero_sisa_pokok(self):
        p = self._make_piutang(
            12_000_000, jenis_bunga='anuitas', suku_bunga=12,
            tanggal=date(2026, 1, 1), jatuh_tempo=date(2026, 12, 31),
        )
        rows = compute_angsuran_schedule(p)
        assert len(rows) == 11
        assert rows[-1]['sisa_pokok'] == Decimal('0')
        assert all(r['bunga'] > 0 for r in rows)


class WriteOffPiutangTests(TestCase):
    def setUp(self):
        self.f = make_fixtures()
        self.coa_beban = Akun.objects.create(
            kategori_id='beban', nama='Beban Piutang Tak Tertagih', kode_akun='6.1.1',
        )
        self.piutang = create_manual_piutang(
            tanggal=date(2026, 1, 1), entitas_bisnis=self.f['eb'], debitur='X', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'], jatuh_tempo=date(2026, 3, 1),
            details=[{'deskripsi': 'X', 'jumlah': Decimal('500000')}],
        )
        self.piutang.status = 'overdue'
        self.piutang.save()

    def test_creates_write_off_record(self):
        write_off_piutang(
            self.piutang,
            {'tanggal': date(2026, 6, 7), 'metode': 'langsung',
             'bad_debt_account': self.coa_beban, 'alasan': 'Tidak tertagih'},
        )
        self.assertTrue(hasattr(self.piutang, 'write_off'))

    def test_status_becomes_written_off(self):
        write_off_piutang(
            self.piutang,
            {'tanggal': date(2026, 6, 7), 'metode': 'langsung',
             'bad_debt_account': self.coa_beban, 'alasan': ''},
        )
        self.piutang.refresh_from_db()
        self.assertEqual(self.piutang.status, 'written_off')

    def test_langsung_journal_dr_bad_debt_cr_piutang(self):
        write_off_piutang(
            self.piutang,
            {'tanggal': date(2026, 6, 7), 'metode': 'langsung',
             'bad_debt_account': self.coa_beban, 'alasan': ''},
        )
        wo = self.piutang.write_off
        details = wo.jurnal.details.all()
        dr = next(d for d in details if d.debit > 0)
        cr = next(d for d in details if d.kredit > 0)
        self.assertEqual(dr.akun, self.coa_beban)
        self.assertEqual(cr.akun, self.f['coa_piutang'])


class ReversePiutangPaymentTests(TestCase):
    def setUp(self):
        self.f = make_fixtures()
        self.piutang = create_manual_piutang(
            tanggal=date(2026, 6, 1), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'], jatuh_tempo=None,
            details=[{'deskripsi': 'X', 'jumlah': Decimal('1000000')}],
        )
        self.piutang.status = 'open'
        self.piutang.save()
        self.penerimaan = create_piutang_payment(
            self.piutang,
            {'tanggal_terima': date(2026, 6, 7), 'jumlah_diterima': Decimal('600000'),
             'payment_account': self.f['coa_kas'], 'metode_penerimaan': 'transfer',
             'nomor_referensi': '', 'catatan': ''},
        )

    def test_reversal_creates_counter_journal(self):
        from apps.jurnal.models import JurnalHeader
        initial_count = JurnalHeader.objects.count()
        reverse_piutang_payment(self.penerimaan)
        self.assertEqual(JurnalHeader.objects.count(), initial_count + 1)

    def test_jumlah_terbayar_reverts(self):
        reverse_piutang_payment(self.penerimaan)
        self.piutang.refresh_from_db()
        self.assertEqual(self.piutang.jumlah_terbayar, Decimal('0'))
        self.assertEqual(self.piutang.status, 'open')


# ── Task 5: compute_penyisihan_for_piutang ────────────────────────────────────

class ComputePenyisihanForPiutangTest(TestCase):
    def setUp(self):
        defaults = [
            ('current', 'Belum Jatuh Tempo', '0.00', 1),
            ('1_30', 'Lewat 1–30 Hari', '5.00', 2),
            ('31_60', 'Lewat 31–60 Hari', '15.00', 3),
            ('61_90', 'Lewat 61–90 Hari', '25.00', 4),
            ('91_180', 'Lewat 91–180 Hari', '50.00', 5),
            ('181_365', 'Lewat 181–365 Hari', '75.00', 6),
            ('over_365', 'Lewat > 365 Hari', '100.00', 7),
        ]
        for key, label, rate, urutan in defaults:
            PenyisihanRateConfig.objects.get_or_create(
                bucket_key=key, defaults={'label': label, 'rate_percent': rate, 'urutan': urutan}
            )

    def _make_piutang_db(self, jumlah_pokok, jatuh_tempo_delta_days):
        akun, _ = Akun.objects.get_or_create(
            kode_akun='1100', defaults={'nama': 'Piutang Usaha', 'kategori_id': 'aset'}
        )
        p = PiutangHeader.objects.create(
            tanggal=date.today(),
            jatuh_tempo=date.today() - timedelta(days=jatuh_tempo_delta_days),
            jumlah_pokok=Decimal(str(jumlah_pokok)),
            jumlah_terbayar=Decimal('0'),
            jenis_jangka_waktu='short_term',
            coa_piutang_account=akun,
            status='open',
        )
        return p

    def test_short_term_overdue_31_days(self):
        p = self._make_piutang_db(1_000_000, jatuh_tempo_delta_days=31)
        result = compute_penyisihan_for_piutang(p)
        assert result['total_penyisihan'] == Decimal('150000.00')  # 15% of 1_000_000

    def test_current_has_zero_penyisihan(self):
        p = self._make_piutang_db(1_000_000, jatuh_tempo_delta_days=-30)  # future
        result = compute_penyisihan_for_piutang(p)
        assert result['total_penyisihan'] == Decimal('0.00')

    def test_breakdown_has_7_entries(self):
        p = self._make_piutang_db(1_000_000, jatuh_tempo_delta_days=10)
        result = compute_penyisihan_for_piutang(p)
        assert len(result['breakdown']) == 7


# ── Task 7: compute_batch_penyisihan ─────────────────────────────────────────

from apps.piutang.services import compute_batch_penyisihan


class ComputeBatchPenyisihanTest(TestCase):
    def setUp(self):
        defaults = [
            ('current', 'Belum JT', '0.00', 1), ('1_30', '1-30', '5.00', 2),
            ('31_60', '31-60', '15.00', 3), ('61_90', '61-90', '25.00', 4),
            ('91_180', '91-180', '50.00', 5), ('181_365', '181-365', '75.00', 6),
            ('over_365', '>365', '100.00', 7),
        ]
        for key, label, rate, urutan in defaults:
            PenyisihanRateConfig.objects.get_or_create(
                bucket_key=key, defaults={'label': label, 'rate_percent': rate, 'urutan': urutan}
            )

    def test_returns_required_keys(self):
        result = compute_batch_penyisihan(date.today())
        assert 'groups' in result
        assert 'unconfigured' in result
        assert 'piutang_count' in result
        assert 'total_target' in result
        assert 'total_delta' in result

    def test_delta_equals_target_minus_existing(self):
        result = compute_batch_penyisihan(date.today())
        total_existing = sum(g['saldo_existing'] for g in result['groups'])
        assert result['total_delta'] == result['total_target'] - total_existing


# ── Task 1: periode_label field ──────────────────────────────────────────────

from apps.piutang.models import PiutangPenyisihan
from apps.jurnal.models import JurnalHeader


class ComputeBatchPenyisihanSaldoSourceTest(TestCase):
    def setUp(self):
        defaults = [
            ('current', 'Belum JT', '0.00', 1), ('1_30', '1-30', '5.00', 2),
            ('31_60', '31-60', '15.00', 3), ('61_90', '61-90', '25.00', 4),
            ('91_180', '91-180', '50.00', 5), ('181_365', '181-365', '75.00', 6),
            ('over_365', '>365', '100.00', 7),
        ]
        for key, label, rate, urutan in defaults:
            PenyisihanRateConfig.objects.get_or_create(
                bucket_key=key, defaults={'label': label, 'rate_percent': rate, 'urutan': urutan}
            )
        self.coa_all = Akun.objects.create(kategori_id='kewajiban', nama='Cad Piutang', kode_akun='2.1.8')
        self.coa_exp = Akun.objects.create(kategori_id='beban', nama='Beban Penyisihan', kode_akun='6.1.8')
        self.coa_piutang = Akun.objects.create(kategori_id='aset', nama='Piutang Usaha', kode_akun='1.1.8')

    def test_saldo_existing_reads_from_piutang_penyisihan_model(self):
        from apps.piutang.models import PiutangPenyisihan, PiutangHeader
        from apps.jurnal.models import JurnalHeader
        jh = JurnalHeader.objects.create(
            tanggal=date(2026, 5, 31), nomor_transaksi='TRX-TEST-001',
            uraian_transaksi='Batch test', is_penyesuaian=True,
        )
        PiutangPenyisihan.objects.create(
            tanggal=date(2026, 5, 31), jenis='batch', jumlah=Decimal('500000'),
            allowance_account=self.coa_all, expense_account=self.coa_exp,
            jurnal_header=jh, periode_label='2026-05',
        )
        PiutangHeader.objects.create(
            jumlah_pokok=Decimal('1000000'), status='open',
            coa_piutang_account=self.coa_piutang,
            penyisihan_allowance_account=self.coa_all,
            penyisihan_expense_account=self.coa_exp,
            jatuh_tempo=date(2026, 6, 30),
        )
        result = compute_batch_penyisihan(date(2026, 6, 30))
        group = next(g for g in result['groups'] if g['allowance_account'] == self.coa_all)
        self.assertEqual(group['saldo_existing'], Decimal('500000.00'))

    def test_no_duplicate_batch_same_periode(self):
        from apps.piutang.services import create_batch_penyisihan_journal
        from apps.piutang.models import PiutangPenyisihan
        from apps.jurnal.models import JurnalHeader
        jh = JurnalHeader.objects.create(
            tanggal=date(2026, 6, 1), nomor_transaksi='TRX-TEST-002',
            uraian_transaksi='Batch test', is_penyesuaian=True,
        )
        PiutangPenyisihan.objects.create(
            tanggal=date(2026, 6, 1), jenis='batch', jumlah=Decimal('100000'),
            allowance_account=self.coa_all, expense_account=self.coa_exp,
            jurnal_header=jh, periode_label='2026-06',
        )
        batch_data = {
            'groups': [
                {
                    'allowance_account': self.coa_all,
                    'expense_account': self.coa_exp,
                    'delta': Decimal('20000'),
                    'entitas_bisnis': None,
                }
            ],
        }
        with self.assertRaises(ValueError):
            create_batch_penyisihan_journal(
                batch_data=batch_data,
                tanggal=date(2026, 6, 30),
                periode_label='2026-06',
            )


class PiutangPenyisihanPeriodeLabelTest(TestCase):
    def test_periode_label_field_exists(self):
        self.assertTrue(hasattr(PiutangPenyisihan, 'periode_label'))

    def test_periode_label_blank_by_default(self):
        f = make_fixtures()
        coa_all = Akun.objects.create(kategori_id='kewajiban', nama='Cad Piutang', kode_akun='2.1.9')
        coa_exp = Akun.objects.create(kategori_id='beban', nama='Beban Penyisihan', kode_akun='6.1.9')
        jh = JurnalHeader.objects.create(
            tanggal=date(2026, 6, 1), nomor_transaksi='TEST-001',
            uraian_transaksi='Test', is_penyesuaian=True,
        )
        p = PiutangPenyisihan.objects.create(
            tanggal=date(2026, 6, 1), jenis='batch', jumlah=Decimal('100'),
            allowance_account=coa_all, expense_account=coa_exp,
            jurnal_header=jh,
        )
        self.assertEqual(p.periode_label, '')


# ── Task 4 & 5: get_aging_schedule_report / get_aging_schedule_workbook ───────

from apps.piutang.services import get_aging_schedule_report, get_aging_schedule_workbook


class GetAgingScheduleReportTest(TestCase):
    def setUp(self):
        self.f = make_fixtures()

    def test_returns_required_keys(self):
        result = get_aging_schedule_report()
        for key in ('as_of_date', 'short_term', 'long_term', 'combined_bucket_totals', 'grand_total'):
            self.assertIn(key, result)

    def test_section_has_rows_bucket_totals_grand_total(self):
        result = get_aging_schedule_report()
        for section_key in ('short_term', 'long_term'):
            section = result[section_key]
            for k in ('rows', 'bucket_totals', 'grand_total'):
                self.assertIn(k, section)

    def test_short_term_piutang_appears_in_short_term_section(self):
        from datetime import timedelta
        p = create_manual_piutang(
            tanggal=date.today(), entitas_bisnis=None, debitur='PT Test', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'],
            jatuh_tempo=date.today() - timedelta(days=45),
            details=[{'deskripsi': 'X', 'jumlah': Decimal('1000000')}],
        )
        p.status = 'overdue'
        p.save()
        result = get_aging_schedule_report()
        nomors = [r['nomor_piutang'] for r in result['short_term']['rows']]
        self.assertIn(p.nomor_piutang, nomors)
        nomors_lt = [r['nomor_piutang'] for r in result['long_term']['rows']]
        self.assertNotIn(p.nomor_piutang, nomors_lt)

    def test_row_has_required_keys(self):
        from datetime import timedelta
        p = create_manual_piutang(
            tanggal=date.today(), entitas_bisnis=None, debitur='PT Test', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'],
            jatuh_tempo=date.today() - timedelta(days=10),
            details=[{'deskripsi': 'X', 'jumlah': Decimal('500000')}],
        )
        p.status = 'overdue'
        p.save()
        result = get_aging_schedule_report()
        row = result['short_term']['rows'][0]
        for k in ('nomor_piutang', 'debitur', 'tanggal', 'jatuh_tempo_efektif',
                  'jumlah_piutang', 'bucket_key', 'bucket_label', 'hari_lewat',
                  'angsuran_no', 'piutang_pk'):
            self.assertIn(k, row)


class GetAgingScheduleWorkbookTest(TestCase):
    def test_returns_workbook_with_two_sheets(self):
        report = get_aging_schedule_report()
        wb = get_aging_schedule_workbook(report)
        import openpyxl
        self.assertIsInstance(wb, openpyxl.Workbook)
        self.assertIn('Jangka Pendek', wb.sheetnames)
        self.assertIn('Jangka Panjang', wb.sheetnames)

    def test_header_row_has_correct_columns(self):
        report = get_aging_schedule_report()
        wb = get_aging_schedule_workbook(report)
        ws = wb['Jangka Pendek']
        headers = [cell.value for cell in ws[1]]
        self.assertIn('No Piutang', headers)
        self.assertIn('Total', headers)


# ── Task 8: periode_bulan/periode_tahun fields ──────────────────────────────

class PiutangReklasifikasiPeriodesTest(TestCase):
    def test_periode_fields_exist(self):
        from apps.piutang.models import PiutangReklasifikasi
        self.assertTrue(hasattr(PiutangReklasifikasi, 'periode_bulan'))
        self.assertTrue(hasattr(PiutangReklasifikasi, 'periode_tahun'))


# ── Task 9: compute_bagian_lancar fix + create_reklasifikasi_bagian_lancar ─────

class ComputeBagianLancarFixedTest(TestCase):
    def setUp(self):
        self.f = make_fixtures()

    def test_long_term_returns_only_installments_within_12_months(self):
        today = date.today()
        jatuh_tempo = today.replace(year=today.year + 2)
        p = create_manual_piutang(
            tanggal=today, entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'],
            jatuh_tempo=jatuh_tempo,
            jenis_jangka_waktu='long_term',
            details=[{'deskripsi': 'X', 'jumlah': Decimal('24000000')}],
        )
        p.status = 'open'
        p.save()
        bagian = compute_bagian_lancar(p)
        self.assertGreater(bagian, Decimal('0'))
        self.assertLess(bagian, Decimal('24000000'))

    def test_short_term_still_returns_full_amount_within_cutoff(self):
        p = create_manual_piutang(
            tanggal=date.today(), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'],
            jatuh_tempo=date.today() + timedelta(days=30),
            details=[{'deskripsi': 'X', 'jumlah': Decimal('500000')}],
        )
        p.status = 'open'
        p.save()
        self.assertEqual(compute_bagian_lancar(p), Decimal('500000'))


class CreateReklasifikasiBagianLancarTest(TestCase):
    def setUp(self):
        self.f = make_fixtures()
        self.coa_lt = Akun.objects.create(kategori_id='aset', nama='Piutang JK Panjang', kode_akun='1.2.2')
        today = date.today()
        jatuh_tempo = today.replace(year=today.year + 2)
        self.piutang = create_manual_piutang(
            tanggal=today, entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.coa_lt,
            jatuh_tempo=jatuh_tempo,
            jenis_jangka_waktu='long_term',
            details=[{'deskripsi': 'X', 'jumlah': Decimal('24000000')}],
        )
        self.piutang.status = 'open'
        self.piutang.save()

    def test_creates_reklasifikasi_record(self):
        from apps.piutang.services import create_reklasifikasi_bagian_lancar
        rkl = create_reklasifikasi_bagian_lancar(
            piutang=self.piutang,
            dari_akun=self.coa_lt,
            ke_akun=self.f['coa_piutang'],
            tanggal=date.today(),
        )
        from apps.piutang.models import PiutangReklasifikasi
        self.assertTrue(PiutangReklasifikasi.objects.filter(pk=rkl.pk).exists())
        self.assertEqual(rkl.periode_bulan, date.today().month)
        self.assertEqual(rkl.periode_tahun, date.today().year)

    def test_raises_on_duplicate_periode(self):
        from apps.piutang.services import create_reklasifikasi_bagian_lancar
        create_reklasifikasi_bagian_lancar(
            piutang=self.piutang, dari_akun=self.coa_lt,
            ke_akun=self.f['coa_piutang'], tanggal=date.today(),
        )
        with self.assertRaises(ValueError):
            create_reklasifikasi_bagian_lancar(
                piutang=self.piutang, dari_akun=self.coa_lt,
                ke_akun=self.f['coa_piutang'], tanggal=date.today(),
            )


# ── Task 11: update_penyisihan_individual ────────────────────────────────────

class UpdatePenyisihanIndividualTest(TestCase):
    def setUp(self):
        self.f = make_fixtures()
        defaults = [
            ('current', 'Belum JT', '0.00', 1), ('1_30', '1-30', '5.00', 2),
            ('31_60', '31-60', '15.00', 3), ('61_90', '61-90', '25.00', 4),
            ('91_180', '91-180', '50.00', 5), ('181_365', '181-365', '75.00', 6),
            ('over_365', '>365', '100.00', 7),
        ]
        for key, label, rate, urutan in defaults:
            PenyisihanRateConfig.objects.get_or_create(
                bucket_key=key, defaults={'label': label, 'rate_percent': rate, 'urutan': urutan}
            )
        self.coa_all = Akun.objects.create(kategori_id='kewajiban', nama='Cad Piutang', kode_akun='2.1.7')
        self.coa_exp = Akun.objects.create(kategori_id='beban', nama='Beban Penyisihan', kode_akun='6.1.7')
        self.piutang = create_manual_piutang(
            tanggal=date.today(), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'],
            jatuh_tempo=date.today() - timedelta(days=40),
            details=[{'deskripsi': 'X', 'jumlah': Decimal('1000000')}],
        )
        self.piutang.status = 'overdue'
        self.piutang.save()
        from apps.piutang.services import create_penyisihan_journal
        self.entry = create_penyisihan_journal(
            piutang=self.piutang,
            allowance_account=self.coa_all,
            expense_account=self.coa_exp,
            tanggal=date.today(),
        )

    def test_update_creates_new_entry_and_removes_old(self):
        from apps.piutang.services import update_penyisihan_individual
        from apps.piutang.models import PiutangPenyisihan
        old_pk = self.entry.pk
        new_entry = update_penyisihan_individual(
            existing_entry=self.entry,
            allowance_account=self.coa_all,
            expense_account=self.coa_exp,
            tanggal=date.today(),
        )
        self.assertFalse(PiutangPenyisihan.objects.filter(pk=old_pk).exists())
        self.assertTrue(PiutangPenyisihan.objects.filter(pk=new_entry.pk).exists())


# ── Task 12: auto_reverse_penyisihan_on_payment ──────────────────────────────

class AutoReversePenyisihanOnPaymentTest(TestCase):
    def setUp(self):
        self.f = make_fixtures()
        defaults = [
            ('current', 'Belum JT', '0.00', 1), ('1_30', '1-30', '5.00', 2),
            ('31_60', '31-60', '15.00', 3), ('61_90', '61-90', '25.00', 4),
            ('91_180', '91-180', '50.00', 5), ('181_365', '181-365', '75.00', 6),
            ('over_365', '>365', '100.00', 7),
        ]
        for key, label, rate, urutan in defaults:
            PenyisihanRateConfig.objects.get_or_create(
                bucket_key=key, defaults={'label': label, 'rate_percent': rate, 'urutan': urutan}
            )
        self.coa_all = Akun.objects.create(kategori_id='kewajiban', nama='Cad Piutang', kode_akun='2.1.6')
        self.coa_exp = Akun.objects.create(kategori_id='beban', nama='Beban Penyisihan', kode_akun='6.1.6')
        self.piutang = create_manual_piutang(
            tanggal=date.today(), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'],
            jatuh_tempo=date.today() - timedelta(days=40),
            details=[{'deskripsi': 'X', 'jumlah': Decimal('1000000')}],
        )
        self.piutang.status = 'overdue'
        self.piutang.save()
        from apps.piutang.services import create_penyisihan_journal
        create_penyisihan_journal(
            piutang=self.piutang,
            allowance_account=self.coa_all,
            expense_account=self.coa_exp,
            tanggal=date.today(),
        )

    def test_manual_penyisihan_reversed_on_full_payment(self):
        from apps.piutang.models import PiutangPenyisihan
        initial_count = PiutangPenyisihan.objects.filter(piutang_header=self.piutang, jenis='manual').count()
        self.assertEqual(initial_count, 1)
        create_piutang_payment(
            self.piutang,
            {'tanggal_terima': date.today(), 'jumlah_diterima': Decimal('1000000'),
             'payment_account': self.f['coa_kas'], 'metode_penerimaan': 'transfer',
             'nomor_referensi': '', 'catatan': ''},
        )
        self.piutang.refresh_from_db()
        self.assertEqual(self.piutang.status, 'paid')
        remaining = PiutangPenyisihan.objects.filter(piutang_header=self.piutang, jenis='manual').count()
        self.assertEqual(remaining, 0)

    def test_penyisihan_not_reversed_on_partial_payment(self):
        from apps.piutang.models import PiutangPenyisihan
        create_piutang_payment(
            self.piutang,
            {'tanggal_terima': date.today(), 'jumlah_diterima': Decimal('500000'),
             'payment_account': self.f['coa_kas'], 'metode_penerimaan': 'transfer',
             'nomor_referensi': '', 'catatan': ''},
        )
        remaining = PiutangPenyisihan.objects.filter(piutang_header=self.piutang, jenis='manual').count()
        self.assertEqual(remaining, 1)


# ── Task 16: compute_present_value & compute_amortization_schedule_pv ─────────

class ComputePresentValueTest(TestCase):
    def test_pv_equals_face_at_zero_discount(self):
        f = make_fixtures()
        p = create_manual_piutang(
            tanggal=date(2026, 1, 1), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=f['coa_piutang'],
            jatuh_tempo=date(2028, 1, 1),
            jenis_jangka_waktu='long_term',
            details=[{'deskripsi': 'X', 'jumlah': Decimal('12000000')}],
        )
        pv = compute_present_value(p, market_rate=Decimal('0'))
        self.assertEqual(pv, Decimal('12000000'))

    def test_pv_less_than_face_at_positive_rate(self):
        f = make_fixtures()
        p = create_manual_piutang(
            tanggal=date(2026, 1, 1), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=f['coa_piutang'],
            jatuh_tempo=date(2028, 1, 1),
            jenis_jangka_waktu='long_term',
            details=[{'deskripsi': 'X', 'jumlah': Decimal('12000000')}],
        )
        pv = compute_present_value(p, market_rate=Decimal('12'))
        self.assertLess(pv, Decimal('12000000'))
        self.assertGreater(pv, Decimal('0'))


class ComputeAmortizationSchedulePvTest(TestCase):
    def test_returns_list_with_period_and_interest_keys(self):
        f = make_fixtures()
        p = create_manual_piutang(
            tanggal=date(2026, 1, 1), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=f['coa_piutang'],
            jatuh_tempo=date(2028, 1, 1),
            jenis_jangka_waktu='long_term',
            details=[{'deskripsi': 'X', 'jumlah': Decimal('12000000')}],
        )
        p.nilai_wajar_awal = compute_present_value(p, Decimal('12'))
        p.pv_discount_rate = Decimal('12')
        rows = compute_amortization_schedule_pv(p)
        self.assertIsInstance(rows, list)
        if rows:
            row = rows[0]
            for k in ('periode', 'tanggal', 'bunga_efektif', 'carrying_value'):
                self.assertIn(k, row)

    def test_schedule_has_bunga_efektif_gross(self):
        f = make_fixtures()
        p = create_manual_piutang(
            tanggal=date(2026, 1, 1), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=f['coa_piutang'],
            jatuh_tempo=date(2028, 1, 1),
            jenis_jangka_waktu='long_term',
            details=[{'deskripsi': 'X', 'jumlah': Decimal('12000000')}],
        )
        p.nilai_wajar_awal = compute_present_value(p, Decimal('12'))
        p.pv_discount_rate = Decimal('12')
        rows = compute_amortization_schedule_pv(p)
        self.assertTrue(len(rows) > 0)
        for row in rows:
            self.assertIn('bunga_efektif_gross', row)
            self.assertGreaterEqual(row['bunga_efektif_gross'], Decimal('0'))
            self.assertGreaterEqual(row['bunga_efektif_gross'], row['bunga_efektif'])


class PiutangHeaderPostingFieldsTest(TestCase):
    def test_pending_approval_is_valid_status_choice(self):
        from apps.piutang.models import PiutangHeader
        valid_keys = [c[0] for c in PiutangHeader.STATUS_CHOICES]
        self.assertIn('pending_approval', valid_keys)

    def test_is_approval_required_field_exists(self):
        f = make_fixtures()
        p = create_manual_piutang(
            tanggal=date.today(), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=f['coa_piutang'], jatuh_tempo=None,
            details=[{'deskripsi': 'X', 'jumlah': Decimal('1000000'), 'revenue_account': None}],
        )
        self.assertFalse(p.is_approval_required)
        p.is_approval_required = True
        p.save(update_fields=['is_approval_required'])
        p.refresh_from_db()
        self.assertTrue(p.is_approval_required)

    def test_can_post_true_when_draft_no_approval(self):
        f = make_fixtures()
        p = create_manual_piutang(
            tanggal=date.today(), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=f['coa_piutang'], jatuh_tempo=None,
            details=[{'deskripsi': 'X', 'jumlah': Decimal('1000000'), 'revenue_account': None}],
        )
        self.assertTrue(p.can_post)

    def test_can_submit_approval_true_when_draft_with_approval(self):
        f = make_fixtures()
        p = create_manual_piutang(
            tanggal=date.today(), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=f['coa_piutang'], jatuh_tempo=None,
            details=[{'deskripsi': 'X', 'jumlah': Decimal('1000000'), 'revenue_account': None}],
        )
        p.is_approval_required = True
        p.save(update_fields=['is_approval_required'])
        self.assertFalse(p.can_post)
        self.assertTrue(p.can_submit_approval)


class PostPiutangTest(TestCase):
    def setUp(self):
        self.f = make_fixtures()
        self.coa_rev = Akun.objects.create(
            kategori_id='pendapatan', nama='Pendapatan Jasa 2', kode_akun='4.1.2',
        )
        self.piutang = create_manual_piutang(
            tanggal=date.today(), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'], jatuh_tempo=None,
            details=[{'deskripsi': 'X', 'jumlah': Decimal('1000000'), 'revenue_account': self.coa_rev}],
        )

    def test_post_piutang_sets_status_open(self):
        from apps.piutang.services import post_piutang
        post_piutang(self.piutang)
        self.piutang.refresh_from_db()
        self.assertEqual(self.piutang.status, 'open')

    def test_post_piutang_creates_journal(self):
        from apps.piutang.services import post_piutang
        from apps.jurnal.models import JurnalHeader
        count_before = JurnalHeader.objects.count()
        post_piutang(self.piutang)
        self.assertEqual(JurnalHeader.objects.count(), count_before + 1)

    def test_post_piutang_raises_if_not_draft(self):
        from apps.piutang.services import post_piutang
        self.piutang.status = 'open'
        self.piutang.save(update_fields=['status'])
        with self.assertRaises(ValueError):
            post_piutang(self.piutang)

    def test_post_piutang_raises_if_detail_missing_revenue_account(self):
        from apps.piutang.services import post_piutang
        piutang2 = create_manual_piutang(
            tanggal=date.today(), entitas_bisnis=None, debitur='Y', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'], jatuh_tempo=None,
            details=[{'deskripsi': 'Y', 'jumlah': Decimal('500000'), 'revenue_account': None}],
        )
        with self.assertRaises(ValueError):
            post_piutang(piutang2)

    def test_post_piutang_locks_piutang(self):
        from apps.piutang.services import post_piutang
        post_piutang(self.piutang)
        self.piutang.refresh_from_db()
        self.assertTrue(self.piutang.is_locked)


class SubmitForApprovalTest(TestCase):
    def setUp(self):
        self.f = make_fixtures()
        self.piutang = create_manual_piutang(
            tanggal=date.today(), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'], jatuh_tempo=None,
            details=[{'deskripsi': 'X', 'jumlah': Decimal('1000000'), 'revenue_account': None}],
        )

    def test_submit_sets_pending_approval(self):
        from apps.piutang.services import submit_for_approval
        submit_for_approval(self.piutang)
        self.piutang.refresh_from_db()
        self.assertEqual(self.piutang.status, 'pending_approval')

    def test_submit_raises_if_not_draft(self):
        from apps.piutang.services import submit_for_approval
        self.piutang.status = 'pending_approval'
        self.piutang.save(update_fields=['status'])
        with self.assertRaises(ValueError):
            submit_for_approval(self.piutang)


class ApprovePiutangTest(TestCase):
    def setUp(self):
        self.f = make_fixtures()
        self.coa_rev = Akun.objects.create(
            kategori_id='pendapatan', nama='Pendapatan Jasa 3', kode_akun='4.1.3',
        )
        self.piutang = create_manual_piutang(
            tanggal=date.today(), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'], jatuh_tempo=None,
            details=[{'deskripsi': 'X', 'jumlah': Decimal('1000000'), 'revenue_account': self.coa_rev}],
        )
        from apps.piutang.services import submit_for_approval
        self.piutang.is_approval_required = True
        self.piutang.save(update_fields=['is_approval_required'])
        submit_for_approval(self.piutang)

    def test_approve_sets_status_open(self):
        from apps.piutang.services import approve_piutang
        approve_piutang(self.piutang)
        self.piutang.refresh_from_db()
        self.assertEqual(self.piutang.status, 'open')

    def test_approve_records_approved_by_and_at(self):
        from apps.piutang.services import approve_piutang
        from django.contrib.auth import get_user_model
        User = get_user_model()
        approver = User.objects.create_user(email='approver_test@test.com', password='x')
        approve_piutang(self.piutang, user=approver)
        self.piutang.refresh_from_db()
        self.assertEqual(self.piutang.approved_by, approver)
        self.assertIsNotNone(self.piutang.approved_at)

    def test_approve_creates_journal(self):
        from apps.piutang.services import approve_piutang
        from apps.jurnal.models import JurnalHeader
        count_before = JurnalHeader.objects.count()
        approve_piutang(self.piutang)
        self.assertEqual(JurnalHeader.objects.count(), count_before + 1)

    def test_approve_raises_if_not_pending(self):
        from apps.piutang.services import approve_piutang
        self.piutang.status = 'draft'
        self.piutang.save(update_fields=['status'])
        with self.assertRaises(ValueError):
            approve_piutang(self.piutang)

    def test_approve_records_approval_metadata(self):
        from apps.piutang.services import approve_piutang
        from django.contrib.auth import get_user_model
        User = get_user_model()
        approver = User.objects.create_user(email='approver@test.com', password='pass')
        approve_piutang(self.piutang, user=approver)
        self.piutang.refresh_from_db()
        self.assertEqual(self.piutang.approved_by, approver)
        self.assertIsNotNone(self.piutang.approved_at)


class RejectPiutangTest(TestCase):
    def setUp(self):
        self.f = make_fixtures()
        self.piutang = create_manual_piutang(
            tanggal=date.today(), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'], jatuh_tempo=None,
            details=[{'deskripsi': 'X', 'jumlah': Decimal('1000000'), 'revenue_account': None}],
        )
        from apps.piutang.services import submit_for_approval
        self.piutang.is_approval_required = True
        self.piutang.save(update_fields=['is_approval_required'])
        submit_for_approval(self.piutang)

    def test_reject_sets_status_draft(self):
        from apps.piutang.services import reject_piutang
        reject_piutang(self.piutang, alasan='Salah nominal')
        self.piutang.refresh_from_db()
        self.assertEqual(self.piutang.status, 'draft')

    def test_reject_raises_if_not_pending(self):
        from apps.piutang.services import reject_piutang
        self.piutang.status = 'draft'
        self.piutang.save(update_fields=['status'])
        with self.assertRaises(ValueError):
            reject_piutang(self.piutang)


class PostPiutangPSAK71Test(TestCase):
    def setUp(self):
        self.f = make_fixtures()
        self.coa_rev = Akun.objects.create(
            kategori_id='pendapatan', nama='Pendapatan PSAK71', kode_akun='4.1.71',
        )
        self.coa_deferred = Akun.objects.create(
            kategori_id='kewajiban', nama='Pend. Bunga Ditangguhkan', kode_akun='1.3.5',
        )
        self.coa_income = Akun.objects.create(
            kategori_id='pendapatan', nama='Pendapatan Bunga Efektif', kode_akun='4.2.1',
        )

    def _make_pv_piutang(self):
        from apps.piutang.services import post_piutang
        p = create_manual_piutang(
            tanggal=date(2026, 1, 1), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'],
            jatuh_tempo=date(2028, 1, 1),
            jenis_jangka_waktu='long_term',
            pv_discount_rate=Decimal('12'),
            deferred_income_account=self.coa_deferred,
            interest_income_account=self.coa_income,
            details=[{
                'deskripsi': 'X', 'jumlah': Decimal('12000000'),
                'revenue_account': self.coa_rev,
            }],
        )
        post_piutang(p)
        p.refresh_from_db()
        return p

    def test_posting_debits_piutang_at_fair_value_not_face(self):
        from apps.jurnal.models import JurnalDetail
        p = self._make_pv_piutang()
        pv = p.nilai_wajar_awal
        # The initial posting journal should debit piutang at PV, not 12_000_000
        debit_lines = JurnalDetail.objects.filter(
            akun=self.f['coa_piutang'], debit__gt=0
        )
        self.assertEqual(debit_lines.count(), 1)
        self.assertAlmostEqual(float(debit_lines.first().debit), float(pv), places=0)
        self.assertLess(debit_lines.first().debit, Decimal('12000000'))

    def test_posting_does_not_create_deferred_income_credit(self):
        from apps.jurnal.models import JurnalDetail
        p = self._make_pv_piutang()
        deferred_credits = JurnalDetail.objects.filter(
            akun=self.coa_deferred, kredit__gt=0
        )
        self.assertEqual(deferred_credits.count(), 0)

    def test_posting_journal_is_balanced(self):
        from apps.jurnal.models import JurnalDetail, JurnalHeader
        p = self._make_pv_piutang()
        journal = JurnalHeader.objects.filter(
            uraian_transaksi__startswith=f'Pengakuan Piutang {p.nomor_piutang}'
        ).first()
        self.assertIsNotNone(journal)
        total_debit = sum(d.debit for d in journal.details.all())
        total_kredit = sum(d.kredit for d in journal.details.all())
        self.assertAlmostEqual(float(total_debit), float(total_kredit), places=2)


class PvCarryingValuePSAK71Test(TestCase):
    """Under PSAK 71, carrying value = net debit on piutang accounts from all journals."""

    def setUp(self):
        self.f = make_fixtures()
        self.coa_rev = Akun.objects.create(
            kategori_id='pendapatan', nama='Pend PSAK71 CV', kode_akun='4.1.72',
        )
        self.coa_deferred = Akun.objects.create(
            kategori_id='kewajiban', nama='PBD CV', kode_akun='1.3.51',
        )
        self.coa_income = Akun.objects.create(
            kategori_id='pendapatan', nama='PBE CV', kode_akun='4.2.11',
        )

    def _posted_pv_piutang(self):
        from apps.piutang.services import post_piutang
        p = create_manual_piutang(
            tanggal=date(2026, 1, 1), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'],
            jatuh_tempo=date(2028, 1, 1),
            jenis_jangka_waktu='long_term',
            pv_discount_rate=Decimal('12'),
            deferred_income_account=self.coa_deferred,
            interest_income_account=self.coa_income,
            details=[{
                'deskripsi': 'X', 'jumlah': Decimal('12000000'),
                'revenue_account': self.coa_rev,
            }],
        )
        post_piutang(p)
        p.refresh_from_db()
        return p

    def test_carrying_value_equals_pv_immediately_after_posting(self):
        from apps.piutang.services import _pv_carrying_value
        p = self._posted_pv_piutang()
        cv = _pv_carrying_value(p)
        self.assertAlmostEqual(float(cv), float(p.nilai_wajar_awal), places=0)

    def test_carrying_value_reduces_after_payment(self):
        from apps.piutang.services import _pv_carrying_value
        p = self._posted_pv_piutang()
        cv_before = _pv_carrying_value(p)
        create_piutang_payment(
            p,
            {'tanggal_terima': date(2026, 2, 1), 'jumlah_diterima': Decimal('500000'),
             'payment_account': self.f['coa_kas'], 'metode_penerimaan': 'transfer',
             'nomor_referensi': '', 'catatan': ''},
        )
        p.refresh_from_db()
        cv_after = _pv_carrying_value(p)
        self.assertLess(cv_after, cv_before)


class PvAdjustmentJournalPSAK71Test(TestCase):
    def setUp(self):
        self.f = make_fixtures()
        self.coa_rev = Akun.objects.create(
            kategori_id='pendapatan', nama='Pend Adj', kode_akun='4.1.73',
        )
        self.coa_deferred = Akun.objects.create(
            kategori_id='kewajiban', nama='PBD Adj', kode_akun='1.3.52',
        )
        self.coa_income = Akun.objects.create(
            kategori_id='pendapatan', nama='PBE Adj', kode_akun='4.2.12',
        )
        from apps.piutang.services import post_piutang
        self.p = create_manual_piutang(
            tanggal=date(2026, 1, 1), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'],
            jatuh_tempo=date(2028, 1, 1),
            jenis_jangka_waktu='long_term',
            pv_discount_rate=Decimal('12'),
            deferred_income_account=self.coa_deferred,
            interest_income_account=self.coa_income,
            details=[{
                'deskripsi': 'X', 'jumlah': Decimal('12000000'),
                'revenue_account': self.coa_rev,
            }],
        )
        post_piutang(self.p)
        self.p.refresh_from_db()

    def test_amortization_journal_debits_piutang_not_deferred(self):
        from apps.piutang.services import create_pv_adjustment_journal
        from apps.jurnal.models import JurnalDetail
        create_pv_adjustment_journal(
            piutang=self.p,
            interest_income_account=self.coa_income,
            tanggal=date(2026, 2, 1),
            periode_no=1,
        )
        deferred_hits = JurnalDetail.objects.filter(akun=self.coa_deferred)
        self.assertEqual(deferred_hits.count(), 0)
        piutang_debits = JurnalDetail.objects.filter(
            akun=self.f['coa_piutang'], debit__gt=0
        ).exclude(
            jurnal_header__uraian_transaksi__startswith='Pengakuan'
        )
        self.assertGreater(piutang_debits.count(), 0)

    def test_amortization_journal_is_balanced(self):
        from apps.piutang.services import create_pv_adjustment_journal
        journal = create_pv_adjustment_journal(
            piutang=self.p,
            interest_income_account=self.coa_income,
            tanggal=date(2026, 2, 1),
            periode_no=1,
        )
        total_debit = sum(d.debit for d in journal.details.all())
        total_kredit = sum(d.kredit for d in journal.details.all())
        self.assertAlmostEqual(float(total_debit), float(total_kredit), places=2)


class PvAccrualJournalPSAK71Test(TestCase):
    def setUp(self):
        self.f = make_fixtures()
        self.coa_rev = Akun.objects.create(
            kategori_id='pendapatan', nama='Pend Akrual', kode_akun='4.1.74',
        )
        self.coa_deferred = Akun.objects.create(
            kategori_id='kewajiban', nama='PBD Akrual', kode_akun='1.3.53',
        )
        self.coa_income = Akun.objects.create(
            kategori_id='pendapatan', nama='PBE Akrual', kode_akun='4.2.13',
        )
        from apps.piutang.services import post_piutang
        self.p = create_manual_piutang(
            tanggal=date(2026, 1, 1), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'],
            jatuh_tempo=date(2028, 1, 1),
            jenis_jangka_waktu='long_term',
            pv_discount_rate=Decimal('12'),
            deferred_income_account=self.coa_deferred,
            interest_income_account=self.coa_income,
            details=[{
                'deskripsi': 'X', 'jumlah': Decimal('12000000'),
                'revenue_account': self.coa_rev,
            }],
        )
        post_piutang(self.p)
        self.p.refresh_from_db()

    def test_accrual_debits_piutang_not_deferred(self):
        from apps.piutang.services import create_pv_accrual_journal
        from apps.jurnal.models import JurnalDetail
        create_pv_accrual_journal(
            piutang=self.p,
            tanggal=date(2026, 3, 31),
            interest_income_account=self.coa_income,
        )
        deferred_hits = JurnalDetail.objects.filter(akun=self.coa_deferred)
        self.assertEqual(deferred_hits.count(), 0)
        piutang_debits = JurnalDetail.objects.filter(
            akun=self.f['coa_piutang'], debit__gt=0,
            jurnal_header__is_penyesuaian=True,
        )
        self.assertGreater(piutang_debits.count(), 0)

    def test_accrual_journal_is_balanced(self):
        from apps.piutang.services import create_pv_accrual_journal
        j = create_pv_accrual_journal(
            piutang=self.p,
            tanggal=date(2026, 3, 31),
            interest_income_account=self.coa_income,
        )
        total_debit = sum(d.debit for d in j.details.all())
        total_kredit = sum(d.kredit for d in j.details.all())
        self.assertAlmostEqual(float(total_debit), float(total_kredit), places=2)


class PaymentJournalPSAK71Test(TestCase):
    def setUp(self):
        self.f = make_fixtures()
        self.coa_rev = Akun.objects.create(
            kategori_id='pendapatan', nama='Pend Pay', kode_akun='4.1.75',
        )
        self.coa_deferred = Akun.objects.create(
            kategori_id='kewajiban', nama='PBD Pay', kode_akun='1.3.54',
        )
        self.coa_income = Akun.objects.create(
            kategori_id='pendapatan', nama='PBE Pay', kode_akun='4.2.14',
        )
        from apps.piutang.services import post_piutang
        self.p = create_manual_piutang(
            tanggal=date(2026, 1, 1), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'],
            jatuh_tempo=date(2027, 1, 1),
            jenis_jangka_waktu='long_term',
            pv_discount_rate=Decimal('12'),
            deferred_income_account=self.coa_deferred,
            interest_income_account=self.coa_income,
            details=[{
                'deskripsi': 'X', 'jumlah': Decimal('12000000'),
                'revenue_account': self.coa_rev,
            }],
        )
        post_piutang(self.p)
        self.p.refresh_from_db()

    def test_payment_credits_piutang_for_full_cash_flow(self):
        from apps.jurnal.models import JurnalDetail
        create_piutang_payment(
            self.p,
            {'tanggal_terima': date(2026, 2, 1), 'jumlah_diterima': Decimal('600000'),
             'payment_account': self.f['coa_kas'], 'metode_penerimaan': 'transfer',
             'nomor_referensi': '', 'catatan': ''},
        )
        payment_credits = JurnalDetail.objects.filter(
            akun=self.f['coa_piutang'],
            kredit=Decimal('600000'),
            jurnal_header__uraian_transaksi__startswith='Penerimaan Piutang',
        )
        self.assertGreater(payment_credits.count(), 0)

    def test_payment_does_not_credit_deferred_account(self):
        from apps.jurnal.models import JurnalDetail
        create_piutang_payment(
            self.p,
            {'tanggal_terima': date(2026, 2, 1), 'jumlah_diterima': Decimal('600000'),
             'payment_account': self.f['coa_kas'], 'metode_penerimaan': 'transfer',
             'nomor_referensi': '', 'catatan': ''},
        )
        deferred_hits = JurnalDetail.objects.filter(akun=self.coa_deferred)
        self.assertEqual(deferred_hits.count(), 0)

    def test_pre_payment_eir_journal_debits_piutang(self):
        from apps.jurnal.models import JurnalDetail
        create_piutang_payment(
            self.p,
            {'tanggal_terima': date(2026, 2, 1), 'jumlah_diterima': Decimal('600000'),
             'payment_account': self.f['coa_kas'], 'metode_penerimaan': 'transfer',
             'nomor_referensi': '', 'catatan': ''},
        )
        eir_debits = JurnalDetail.objects.filter(
            akun=self.f['coa_piutang'],
            debit__gt=0,
            jurnal_header__uraian_transaksi__startswith='Amortisasi PV Piutang',
        )
        self.assertGreater(eir_debits.count(), 0)


class SettlementCatchupRemovedTest(TestCase):
    """Under PSAK 71, full payment should NOT create a separate catchup journal."""

    def setUp(self):
        self.f = make_fixtures()
        self.coa_rev = Akun.objects.create(
            kategori_id='pendapatan', nama='Pend Catchup', kode_akun='4.1.76',
        )
        self.coa_deferred = Akun.objects.create(
            kategori_id='kewajiban', nama='PBD Catchup', kode_akun='1.3.55',
        )
        self.coa_income = Akun.objects.create(
            kategori_id='pendapatan', nama='PBE Catchup', kode_akun='4.2.15',
        )
        from apps.piutang.services import post_piutang
        self.p = create_manual_piutang(
            tanggal=date(2026, 1, 1), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.f['coa_piutang'],
            jatuh_tempo=date(2027, 1, 1),
            jenis_jangka_waktu='long_term',
            pv_discount_rate=Decimal('12'),
            deferred_income_account=self.coa_deferred,
            interest_income_account=self.coa_income,
            details=[{
                'deskripsi': 'X', 'jumlah': Decimal('12000000'),
                'revenue_account': self.coa_rev,
            }],
        )
        post_piutang(self.p)
        self.p.refresh_from_db()

    def test_full_payment_creates_no_pelunasan_catchup_journal(self):
        from apps.jurnal.models import JurnalHeader
        create_piutang_payment(
            self.p,
            {'tanggal_terima': date(2026, 6, 1),
             'jumlah_diterima': self.p.jumlah_pokok,
             'payment_account': self.f['coa_kas'], 'metode_penerimaan': 'transfer',
             'nomor_referensi': '', 'catatan': ''},
        )
        catchup = JurnalHeader.objects.filter(
            uraian_transaksi__contains='Pelunasan'
        )
        self.assertEqual(catchup.count(), 0)


class ReklasifikasiPSAK71Test(TestCase):
    def setUp(self):
        self.f = make_fixtures()
        self.coa_lt = Akun.objects.create(
            kategori_id='aset', nama='Piutang JK Panjang PSAK71', kode_akun='1.3.3',
        )
        self.coa_bl = Akun.objects.create(
            kategori_id='aset', nama='Piutang Bagian Lancar PSAK71', kode_akun='1.1.9',
        )
        self.coa_rev = Akun.objects.create(
            kategori_id='pendapatan', nama='Pend Rkl', kode_akun='4.1.77',
        )
        self.coa_deferred = Akun.objects.create(
            kategori_id='kewajiban', nama='PBD Rkl', kode_akun='1.3.56',
        )
        self.coa_deferred_bl = Akun.objects.create(
            kategori_id='kewajiban', nama='PBD BL Rkl', kode_akun='1.1.11',
        )
        self.coa_income = Akun.objects.create(
            kategori_id='pendapatan', nama='PBE Rkl', kode_akun='4.2.16',
        )
        from apps.piutang.services import post_piutang
        self.p = create_manual_piutang(
            tanggal=date(2026, 1, 1), entitas_bisnis=None, debitur='X', deskripsi='',
            coa_piutang_account=self.coa_lt,
            jatuh_tempo=date(2028, 1, 1),
            jenis_jangka_waktu='long_term',
            pv_discount_rate=Decimal('12'),
            deferred_income_account=self.coa_deferred,
            interest_income_account=self.coa_income,
            coa_piutang_lancar_account=self.coa_bl,
            deferred_income_lancar_account=self.coa_deferred_bl,
            details=[{
                'deskripsi': 'X', 'jumlah': Decimal('24000000'),
                'revenue_account': self.coa_rev,
            }],
        )
        post_piutang(self.p)
        self.p.refresh_from_db()

    def test_reklasifikasi_journal_has_exactly_two_lines(self):
        from apps.piutang.services import create_reklasifikasi_bagian_lancar
        rkl = create_reklasifikasi_bagian_lancar(
            piutang=self.p,
            dari_akun=self.coa_lt,
            ke_akun=self.coa_bl,
            tanggal=date(2026, 12, 31),
            dari_akun_deferred=self.coa_deferred,
            ke_akun_deferred=self.coa_deferred_bl,
        )
        details = list(rkl.jurnal.details.all())
        self.assertEqual(len(details), 2)

    def test_reklasifikasi_debits_current_account_not_deferred(self):
        from apps.piutang.services import create_reklasifikasi_bagian_lancar
        from apps.jurnal.models import JurnalDetail
        rkl = create_reklasifikasi_bagian_lancar(
            piutang=self.p,
            dari_akun=self.coa_lt,
            ke_akun=self.coa_bl,
            tanggal=date(2026, 12, 31),
            dari_akun_deferred=self.coa_deferred,
            ke_akun_deferred=self.coa_deferred_bl,
        )
        deferred_hits = JurnalDetail.objects.filter(
            jurnal_header=rkl.jurnal,
            akun__in=[self.coa_deferred, self.coa_deferred_bl],
        )
        self.assertEqual(deferred_hits.count(), 0)
        dr_line = rkl.jurnal.details.get(debit__gt=0)
        self.assertEqual(dr_line.akun, self.coa_bl)

    def test_reklasifikasi_amount_is_carrying_not_nominal(self):
        from apps.piutang.services import create_reklasifikasi_bagian_lancar
        rkl = create_reklasifikasi_bagian_lancar(
            piutang=self.p,
            dari_akun=self.coa_lt,
            ke_akun=self.coa_bl,
            tanggal=date(2026, 12, 31),
        )
        self.assertLess(rkl.jumlah, Decimal('24000000'))
        self.assertGreater(rkl.jumlah, Decimal('0'))


# ── Task 2: build_piutang canonical factory ───────────────────────────────────

class BuildPiutangServiceTest(TestCase):
    def _akun(self, kode, nama, kategori_id='aset'):
        from apps.master_data.models import Akun
        return Akun.objects.create(kode_akun=kode, nama=nama, kategori_id=kategori_id)

    def test_build_piutang_creates_full_header_and_details(self):
        from decimal import Decimal
        from datetime import date
        from apps.piutang.services import build_piutang
        from apps.piutang.models import PiutangHeader

        akun_piutang = self._akun('1.1.4', 'Piutang Usaha')
        akun_pend = self._akun('4.1.1', 'Pendapatan Jasa', kategori_id='pendapatan')
        payload = {
            'tanggal': date(2026, 1, 10),
            'debitur': 'PT Maju',
            'deskripsi': 'Piutang uji',
            'coa_piutang_account': akun_piutang,
            'jatuh_tempo': date(2026, 3, 10),
            'jenis_bunga': 'flat',
            'suku_bunga': Decimal('12'),
            'kategori_pengukuran': 'amortised_cost',
        }
        details = [{'deskripsi': 'Baris 1', 'jumlah': Decimal('1000'), 'revenue_account': akun_pend}]
        piutang = build_piutang(payload, source='manual', source_obj=None, details=details, user=None)

        self.assertIsInstance(piutang, PiutangHeader)
        self.assertEqual(piutang.jumlah_pokok, Decimal('1000'))
        self.assertEqual(piutang.debitur, 'PT Maju')
        self.assertEqual(piutang.jenis_bunga, 'flat')
        self.assertEqual(piutang.suku_bunga, Decimal('12'))
        self.assertEqual(piutang.details.count(), 1)
        self.assertEqual(piutang.status, 'draft')  # manual default

    def test_build_piutang_pendapatan_source_status_open(self):
        from decimal import Decimal
        from datetime import date
        from apps.piutang.services import build_piutang
        akun_piutang = self._akun('1.1.5', 'Piutang B')
        payload = {'tanggal': date(2026, 1, 10), 'coa_piutang_account': akun_piutang}
        details = [{'deskripsi': 'x', 'jumlah': Decimal('500'), 'revenue_account': None}]
        piutang = build_piutang(payload, source='pendapatan', source_obj=None, details=details, user=None)
        self.assertEqual(piutang.status, 'open')
        self.assertEqual(piutang.source_type, 'from_pendapatan')


# ── Task 5 refactor: piutang form partial + piutang_form.js ─────────────────

class PiutangFormRendersTest(TestCase):
    def setUp(self):
        from django.contrib.auth import get_user_model
        self.user = get_user_model().objects.create_user(
            email='testpiutang@example.com', password='p', name='Test Piutang'
        )
        self.client.force_login(self.user)

    def test_create_page_renders_wizard(self):
        from django.urls import reverse
        url = reverse('piutang:create')
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'piutang_form.js')
