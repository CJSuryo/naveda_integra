from __future__ import annotations

import calendar
from datetime import date
from decimal import Decimal

from django.db import transaction
from django.db.models import Sum
from django.utils import timezone

from apps.jurnal.models import JurnalDetail, JurnalHeader

from .models import (
    PiutangAuditLog, PiutangAttachment, PiutangDetail, PiutangHeader,
    PiutangPenerimaan, PiutangReklasifikasi, PiutangWriteOff,
    PiutangECLStagingLog, PiutangModifikasi, PiutangPemulihanWriteOff, PiutangFactoring,
)


# ── Schedule Helpers ──────────────────────────────────────────────────────────

_PERIODE_MONTHS_MAP = {'bulanan': 1, 'triwulanan': 3, 'semesteran': 6, 'tahunan': 12}


def _add_months(d: date, months: int) -> date:
    month = d.month - 1 + months
    year = d.year + month // 12
    month = month % 12 + 1
    day = min(d.day, calendar.monthrange(year, month)[1])
    return d.replace(year=year, month=month, day=day)


def compute_angsuran_schedule(piutang) -> list:
    """Returns installment schedule for a PiutangHeader. Empty list if no jatuh_tempo."""
    if not piutang.jatuh_tempo:
        return []
    periode_months = _PERIODE_MONTHS_MAP.get(piutang.periode_angsuran, 1)
    total_months = (
        (piutang.jatuh_tempo.year - piutang.tanggal.year) * 12
        + (piutang.jatuh_tempo.month - piutang.tanggal.month)
    )
    if total_months <= 0:
        return []
    n = max(1, round(total_months / periode_months))
    total = float(piutang.jumlah_pokok)
    jenis = piutang.jenis_bunga
    r = float(piutang.suku_bunga) / 100 / 12 * periode_months if jenis != 'tanpa_bunga' else 0.0

    rows = []
    sisa = total
    if jenis == 'anuitas' and r > 0:
        pmt = total * r / (1 - (1 + r) ** (-n))
        for i in range(n):
            bunga_i = sisa * r
            pokok_i = pmt - bunga_i
            if i == n - 1:
                pokok_i = sisa
            angsuran_i = pokok_i + bunga_i
            sisa -= pokok_i
            rows.append({
                'no': i + 1,
                'tanggal': _add_months(piutang.tanggal, (i + 1) * periode_months),
                'pokok': Decimal(str(int(round(pokok_i, 0)))),
                'bunga': Decimal(str(int(round(bunga_i, 0)))),
                'angsuran': Decimal(str(int(round(angsuran_i, 0)))),
                'sisa_pokok': Decimal(str(int(round(max(0.0, sisa), 0)))),
            })
    else:
        pk_unit = round(total / n, 0)
        bng_unit = round(total * r, 0) if jenis == 'flat' else 0.0
        cumulative_pk = 0.0
        for i in range(n):
            pk_i = round(total - cumulative_pk, 0) if i == n - 1 else pk_unit
            cumulative_pk += pk_i
            sisa -= pk_i
            ang_i = pk_i + bng_unit
            rows.append({
                'no': i + 1,
                'tanggal': _add_months(piutang.tanggal, (i + 1) * periode_months),
                'pokok': Decimal(str(int(pk_i))),
                'bunga': Decimal(str(int(bng_unit))),
                'angsuran': Decimal(str(int(ang_i))),
                'sisa_pokok': Decimal(str(int(round(max(0.0, sisa), 0)))),
            })

    # Payment matching: direct via angsuran_no; unallocated pool fills in order
    direct_paid: dict[int, float] = {}
    unallocated = 0.0
    penerimaan_qs = piutang.penerimaan.all() if piutang.pk else []
    for p in penerimaan_qs:
        if p.angsuran_no:
            direct_paid[p.angsuran_no] = direct_paid.get(p.angsuran_no, 0.0) + float(p.jumlah_diterima)
        else:
            unallocated += float(p.jumlah_diterima)

    today = date.today()
    for row in rows:
        ang = float(row['angsuran'])
        no = row['no']
        paid = direct_paid.get(no, 0.0)
        if paid < ang - 1.0 and unallocated > 0:
            apply = min(unallocated, max(0.0, ang - paid))
            paid += apply
            unallocated = max(0.0, unallocated - apply)
        row['paid'] = Decimal(str(round(paid, 0)))
        row['sisa_bayar'] = Decimal(str(round(max(0.0, ang - paid), 0)))
        if paid >= ang - 1.0:
            row['status'] = 'lunas'
            row['sisa_bayar'] = Decimal('0')
        elif paid > 1.0:
            row['status'] = 'sebagian'
        elif row['tanggal'] < today:
            row['status'] = 'jatuh_tempo'
        else:
            row['status'] = 'akan_datang'
    return rows


def compute_effective_dpd(piutang, schedule: list | None = None) -> int:
    """
    Returns the effective Days Past Due for PSAK 71 ECL staging purposes.

    For long-term installment loans the final jatuh_tempo is the balloon/last
    payment date — not the first missed installment.  PSAK 71 / IFRS 9 measures
    DPD from the earliest contractual payment that has not been settled.

    Pass an already-computed schedule (from compute_angsuran_schedule) to avoid
    recomputing it; if omitted the schedule is computed here.
    """
    if piutang.status in ('paid', 'cancelled', 'written_off'):
        return 0

    # Short-term or no due date: fall back to the simple jatuh_tempo-based DPD
    if piutang.jenis_jangka_waktu != 'long_term' or not piutang.jatuh_tempo:
        return piutang.days_overdue

    today = date.today()
    sched = schedule if schedule is not None else compute_angsuran_schedule(piutang)
    if not sched:
        return piutang.days_overdue

    overdue_dates = [
        r['tanggal'] for r in sched
        if r['tanggal'] < today and r['sisa_bayar'] > 0
    ]
    if not overdue_dates:
        return 0
    return (today - min(overdue_dates)).days


# ── Audit Helper ──────────────────────────────────────────────────────────────

def _log(piutang: PiutangHeader, action: str, user=None, before=None, after=None, notes=''):
    PiutangAuditLog.objects.create(
        piutang_header=piutang,
        nomor_piutang=piutang.nomor_piutang,
        action=action,
        user=user,
        before_json=before or {},
        after_json=after or {},
        notes=notes,
    )


def _snapshot(piutang: PiutangHeader) -> dict:
    return {
        'status': piutang.status,
        'jumlah_pokok': str(piutang.jumlah_pokok),
        'jumlah_terbayar': str(piutang.jumlah_terbayar),
    }


def _next_piutang_journal_number(prefix: str) -> str:
    with transaction.atomic():
        last = (
            JurnalHeader.objects
            .select_for_update()
            .filter(nomor_transaksi__startswith=f'{prefix}-')
            .order_by('-nomor_transaksi')
            .values_list('nomor_transaksi', flat=True)
            .first()
        )
        seq = 1
        if last:
            try:
                seq = int(last.rsplit('-', 1)[1]) + 1
            except (ValueError, IndexError):
                seq = 1
        return f'{prefix}-{seq:04d}'


# ── Create ────────────────────────────────────────────────────────────────────

# Mapping source → (source_type value on PiutangHeader, default status)
_PIUTANG_SOURCE_MAP = {
    'manual': ('manual', 'draft'),
    'pendapatan': ('from_pendapatan', 'open'),
    'sales': ('from_sales', 'open'),
}


def build_piutang(payload: dict, *, source: str, source_obj, details: list, user=None) -> PiutangHeader:
    """Canonical piutang factory used by every module.

    payload: dict of PiutangHeader header fields (debitur, coa_piutang_account, credit terms…).
    details: list of {'deskripsi', 'jumlah', 'revenue_account'(, 'sub_transaction_type')}.
    source: 'manual' | 'pendapatan' | 'sales'.
    source_obj: the originating header (PendapatanHeader / SalesHeader) or None.

    Does NOT post an AR journal — posting is a separate step (manual) or already
    booked by the originating module's confirm (pendapatan).
    """
    if not details:
        raise ValueError('Minimal satu detail piutang diperlukan.')
    total = sum(Decimal(str(d['jumlah'])) for d in details)
    if total <= 0:
        raise ValueError('Total piutang harus lebih besar dari 0.')

    source_type, default_status = _PIUTANG_SOURCE_MAP[source]
    header_kwargs = dict(payload)
    header_kwargs.setdefault('status', default_status)
    header_kwargs['source_type'] = source_type
    if source == 'pendapatan':
        header_kwargs['source_pendapatan'] = source_obj
    elif source == 'sales':
        header_kwargs['source_sales'] = source_obj
    header_kwargs['jumlah_pokok'] = total
    if user is not None:
        header_kwargs['created_by'] = user

    with transaction.atomic():
        piutang = PiutangHeader.objects.create(**header_kwargs)
        PiutangDetail.objects.bulk_create([
            PiutangDetail(
                piutang_header=piutang,
                deskripsi=d.get('deskripsi', ''),
                jumlah=Decimal(str(d['jumlah'])),
                revenue_account=d.get('revenue_account'),
                sub_transaction_type=d.get('sub_transaction_type'),
            )
            for d in details
        ])
        _log(piutang, 'CREATED', user=user, after=_snapshot(piutang))
    return piutang


def create_manual_piutang(
    tanggal,
    entitas_bisnis,
    debitur: str,
    deskripsi: str,
    coa_piutang_account,
    jatuh_tempo,
    details: list,
    jenis_jangka_waktu: str = 'short_term',
    jenis_bunga: str = 'tanpa_bunga',
    suku_bunga: Decimal = Decimal('0'),
    periode_angsuran: str = 'bulanan',
    is_approval_required: bool = False,
    pv_discount_rate=None,
    deferred_income_account=None,
    interest_income_account=None,
    coa_piutang_lancar_account=None,
    deferred_income_lancar_account=None,
    standar_akuntansi: str = '',
    kategori_pengukuran: str = 'amortised_cost',
    business_model: str = '',
    sppi_test_passed=None,
    biaya_transaksi=None,
    biaya_transaksi_account=None,
    agunan_jenis: str = '',
    agunan_nilai=None,
    user=None,
) -> PiutangHeader:
    payload = {
        'tanggal': tanggal,
        'entitas_bisnis': entitas_bisnis,
        'debitur': debitur,
        'deskripsi': deskripsi,
        'coa_piutang_account': coa_piutang_account,
        'jatuh_tempo': jatuh_tempo,
        'jenis_jangka_waktu': jenis_jangka_waktu,
        'jenis_bunga': jenis_bunga,
        'suku_bunga': suku_bunga,
        'periode_angsuran': periode_angsuran,
        'is_approval_required': is_approval_required,
        'pv_discount_rate': pv_discount_rate,
        'deferred_income_account': deferred_income_account,
        'interest_income_account': interest_income_account,
        'coa_piutang_lancar_account': coa_piutang_lancar_account,
        'deferred_income_lancar_account': deferred_income_lancar_account,
        'standar_akuntansi': standar_akuntansi or '',
        'kategori_pengukuran': kategori_pengukuran or 'amortised_cost',
        'business_model': business_model or '',
        'sppi_test_passed': sppi_test_passed,
        'biaya_transaksi': biaya_transaksi or Decimal('0'),
        'biaya_transaksi_account': biaya_transaksi_account,
        'agunan_jenis': agunan_jenis or '',
        'agunan_nilai': agunan_nilai,
    }
    return build_piutang(payload, source='manual', source_obj=None, details=details, user=user)


# ── Stubs for callers that will be implemented in later phases ─────────────────

def create_piutang_from_sales(sales_header, user=None) -> PiutangHeader:
    total = Decimal('0')
    details = []
    for eb_group in sales_header.entitas_groups.select_related('entitas_bisnis').all():
        for item in eb_group.items.select_related('revenue_account').all():
            total += item.total_sales
            details.append({
                'deskripsi': str(item.item),
                'jumlah': item.total_sales,
                'revenue_account': item.revenue_account,
            })

    if total <= 0:
        raise ValueError('Total credit sales harus lebih besar dari 0.')

    coa_piutang = (
        sales_header.entitas_groups.first().payment_account
        if sales_header.entitas_groups.exists()
        else None
    )
    if not coa_piutang:
        raise ValueError('Payment account (akun piutang) diperlukan pada SalesEntitasBisnis.')

    eb = sales_header.entitas_groups.first().entitas_bisnis if sales_header.entitas_groups.exists() else None

    with transaction.atomic():
        piutang = PiutangHeader.objects.create(
            tanggal=sales_header.tanggal,
            entitas_bisnis=eb,
            debitur=str(eb) if eb else '',
            deskripsi=f'Piutang dari Sales {sales_header.transaction_id}',
            source_type='from_sales',
            source_sales=sales_header,
            jumlah_pokok=total,
            status='open',
            coa_piutang_account=coa_piutang,
            created_by=user,
        )
        PiutangDetail.objects.bulk_create([
            PiutangDetail(
                piutang_header=piutang,
                deskripsi=d['deskripsi'],
                jumlah=d['jumlah'],
                revenue_account=d.get('revenue_account'),
            )
            for d in details
        ])
        _log(piutang, 'CREATED', user=user, after=_snapshot(piutang))
    return piutang


def create_piutang_from_pendapatan(pendapatan_header, user=None) -> PiutangHeader:
    from apps.pendapatan.services import pendapatan_to_piutang_payload
    payload, details = pendapatan_to_piutang_payload(pendapatan_header)
    return build_piutang(payload, source='pendapatan', source_obj=pendapatan_header,
                         details=details, user=user)


# ── Payment ───────────────────────────────────────────────────────────────────

def _piutang_journal_ids(piutang) -> set:
    """
    Collect all JurnalHeader PKs that belong to this piutang.

    Uses FK relationships where they exist (penerimaan, reklasifikasi, write-off)
    and falls back to safe uraian patterns only when no FK is available.
    All uraian-based patterns use the full nomor followed by a non-digit character
    (' —' or end-of-string) to prevent PIU-001 from matching PIU-0010.
    """
    from django.db.models import Q
    nom = piutang.nomor_piutang
    ids: set = set()

    # FK-linked: penerimaan (payment) journals
    ids.update(
        piutang.penerimaan.filter(jurnal_header__isnull=False)
        .values_list('jurnal_header_id', flat=True)
    )

    # FK-linked: reklasifikasi journals; collect PKs for reversal lookup too
    rkl_pks = []
    for rkl_pk, j_pk in piutang.reklasifikasi_entries.filter(
        jurnal__isnull=False
    ).values_list('pk', 'jurnal_id'):
        ids.add(j_pk)
        rkl_pks.append(rkl_pk)

    # Reklasifikasi reversal journals: deterministic nomor_transaksi TRX-PIU-RKLR-{pk}
    if rkl_pks:
        ids.update(
            JurnalHeader.objects.filter(
                nomor_transaksi__in=[f'TRX-PIU-RKLR-{pk}' for pk in rkl_pks]
            ).values_list('pk', flat=True)
        )

    # Uraian-based journals (no FK available).
    # Each pattern is anchored so that nom is not a bare prefix of a longer nomor:
    # - exact match (=) for uraian where nom is at the END
    # - startswith with ' —' suffix guard for uraian where nom is followed by extra text
    ids.update(
        JurnalHeader.objects.filter(
            Q(uraian_transaksi=f'Pengakuan Piutang {nom}') |
            Q(uraian_transaksi=f'Write-Off Piutang {nom}') |
            Q(uraian_transaksi=f'Reversal Penerimaan {nom}') |
            Q(uraian_transaksi__startswith=f'Penerimaan Piutang {nom} —') |
            Q(uraian_transaksi__startswith=f'Amortisasi PV Piutang {nom} —') |
            Q(uraian_transaksi__startswith=f'Akrual PV Piutang {nom} —') |
            Q(uraian_transaksi__startswith=f'Balik Akrual PV Piutang {nom} —')
        ).values_list('pk', flat=True)
    )

    return ids


def _net_debit_balance_for_piutang(account, piutang) -> Decimal:
    """
    Net debit balance on `account` across all journals for this piutang.
    Positive = normal asset balance; negative = net credit (normal for contra-assets).

    Collects journal IDs via FK relationships and safe uraian patterns (no bare
    __contains substring match) to prevent one piutang's nomor from accidentally
    matching another piutang whose nomor is a longer string starting with the same prefix.
    """
    if account is None:
        return Decimal('0')
    journal_ids = _piutang_journal_ids(piutang)
    if not journal_ids:
        return Decimal('0')
    result = (
        JurnalDetail.objects
        .filter(akun=account, jurnal_header_id__in=journal_ids)
        .aggregate(debit=Sum('debit'), kredit=Sum('kredit'))
    )
    return (result['debit'] or Decimal('0')) - (result['kredit'] or Decimal('0'))


def create_piutang_payment(piutang: PiutangHeader, data: dict, user=None) -> PiutangPenerimaan:
    jumlah = Decimal(str(data['jumlah_diterima']))
    with transaction.atomic():
        # Lock the piutang row for the duration of this transaction to prevent concurrent payment races
        piutang = PiutangHeader.objects.select_for_update().get(pk=piutang.pk)
        # For interest-bearing piutang, validate against total remaining schedule cash flows
        # (principal + remaining contractual interest), not just sisa_piutang which is principal-only.
        if piutang.jenis_bunga != 'tanpa_bunga' and piutang.jatuh_tempo:
            _schedule = compute_angsuran_schedule(piutang)
            sisa_tagihan = sum(r['sisa_bayar'] for r in _schedule) if _schedule else piutang.sisa_piutang
        else:
            sisa_tagihan = piutang.sisa_piutang
        if jumlah > sisa_tagihan:
            raise ValueError(
                f'Jumlah diterima ({jumlah:,.0f}) melebihi sisa tagihan ({sisa_tagihan:,.0f}).'
            )
        penerimaan = PiutangPenerimaan.objects.create(
            piutang_header=piutang,
            tanggal_terima=data['tanggal_terima'],
            jumlah_diterima=jumlah,
            angsuran_no=data.get('angsuran_no'),
            payment_account=data['payment_account'],
            metode_penerimaan=data.get('metode_penerimaan', 'transfer'),
            nomor_referensi=data.get('nomor_referensi', ''),
            catatan=data.get('catatan', ''),
            created_by=user,
        )
        jurnal = _create_payment_journal(piutang, penerimaan)
        penerimaan.jurnal_header = jurnal
        penerimaan.save(update_fields=['jurnal_header'])

        total_paid = piutang.penerimaan.aggregate(s=Sum('jumlah_diterima'))['s'] or Decimal('0')
        piutang.jumlah_terbayar = total_paid
        if total_paid >= piutang.jumlah_pokok:
            piutang.status = 'paid'
        elif total_paid > 0:
            piutang.status = 'partial'
        piutang.save(update_fields=['jumlah_terbayar', 'status'])

        if piutang.status == 'paid':
            auto_reverse_penyisihan_on_payment(piutang, user=user)

        _log(piutang, 'PAYMENT', user=user, after=_snapshot(piutang))
    return penerimaan


def _create_payment_journal(piutang: PiutangHeader, penerimaan: PiutangPenerimaan) -> JurnalHeader:
    tanggal = penerimaan.tanggal_terima
    jumlah = penerimaan.jumlah_diterima

    # ── PSAK 71: Pre-payment EIR accrual ────────────────────────────────────────
    # Recognise effective interest from last amortisation to payment date.
    # Stage 3 (PSAK): interest on net carrying. Otherwise: gross carrying.
    # Dr. Piutang / Cr. Pendapatan Bunga Efektif (increases carrying amount).
    if (piutang.is_pv_adjusted
            and piutang.interest_income_account_id
            and piutang.pv_discount_rate):
        from_date = _pv_last_amortization_date(piutang)
        bunga_gross = _pv_effective_interest_days(piutang, from_date, tanggal)
        if piutang.stage_ecl == 3 and get_standar_akuntansi(piutang) == 'psak':
            gross_ca = _pv_carrying_value(piutang)
            net_ca = _get_ecl_net_carrying(piutang)
            ratio = (net_ca / gross_ca) if gross_ca > Decimal('0') else Decimal('1')
            bunga_efektif = (bunga_gross * ratio).quantize(Decimal('0.0001'))
        else:
            bunga_efektif = bunga_gross
        if abs(bunga_efektif) >= Decimal('0.005'):
            ar_account_eir = piutang.coa_piutang_account
            amort_nomor = _next_piutang_journal_number('TRX-PIU-PV')
            amort_header = JurnalHeader.objects.create(
                tanggal=tanggal,
                nomor_transaksi=amort_nomor,
                uraian_transaksi=(
                    f'Amortisasi PV Piutang {piutang.nomor_piutang} — '
                    f'{from_date} s.d. {tanggal}'
                ),
                entitas_bisnis=piutang.entitas_bisnis,
                is_penyesuaian=False,
            )
            JurnalDetail.objects.bulk_create([
                JurnalDetail(
                    jurnal_header=amort_header,
                    akun=ar_account_eir,
                    debit=bunga_efektif,
                    kredit=Decimal('0'),
                ),
                JurnalDetail(
                    jurnal_header=amort_header,
                    akun=piutang.interest_income_account,
                    debit=Decimal('0'),
                    kredit=bunga_efektif,
                ),
            ])

    # ── Payment journal ──────────────────────────────────────────────────────────
    ar_account = (
        piutang.coa_piutang_lancar_account
        if piutang.coa_piutang_lancar_account_id
        else piutang.coa_piutang_account
    )
    nomor = _next_piutang_journal_number('TRX-PIU-P')
    header = JurnalHeader.objects.create(
        tanggal=tanggal,
        nomor_transaksi=nomor,
        uraian_transaksi=f'Penerimaan Piutang {piutang.nomor_piutang} — {piutang.entitas_display}',
        entitas_bisnis=piutang.entitas_bisnis,
        is_penyesuaian=False,
    )

    if piutang.is_pv_adjusted and piutang.pv_discount_rate:
        # PSAK 71: full cash flow reduces carrying amount.
        # Contractual interest was already in the carrying amount via EIR; no split needed.
        if piutang.coa_piutang_lancar_account_id:
            bal_lancar = max(
                Decimal('0'),
                _net_debit_balance_for_piutang(piutang.coa_piutang_lancar_account, piutang),
            )
            credit_lancar = min(jumlah, bal_lancar)
            credit_lt = jumlah - credit_lancar
            payment_lines = [
                JurnalDetail(jurnal_header=header, akun=penerimaan.payment_account,
                             debit=jumlah, kredit=Decimal('0')),
            ]
            if credit_lancar > 0:
                payment_lines.append(JurnalDetail(
                    jurnal_header=header, akun=piutang.coa_piutang_lancar_account,
                    debit=Decimal('0'), kredit=credit_lancar,
                ))
            if credit_lt > 0:
                payment_lines.append(JurnalDetail(
                    jurnal_header=header, akun=piutang.coa_piutang_account,
                    debit=Decimal('0'), kredit=credit_lt,
                ))
            JurnalDetail.objects.bulk_create(payment_lines)
        else:
            JurnalDetail.objects.bulk_create([
                JurnalDetail(jurnal_header=header, akun=penerimaan.payment_account,
                             debit=jumlah, kredit=Decimal('0')),
                JurnalDetail(jurnal_header=header, akun=ar_account,
                             debit=Decimal('0'), kredit=jumlah),
            ])
        return header

    # ── Non-PV-adjusted: existing logic ──────────────────────────────────────────
    # For interest-bearing piutang with angsuran_no: split payment bunga dulu, sisanya pokok
    if (piutang.jenis_bunga != 'tanpa_bunga'
            and penerimaan.angsuran_no
            and piutang.interest_income_account_id):
        schedule = compute_angsuran_schedule(piutang)
        installment = next((r for r in schedule if r['no'] == penerimaan.angsuran_no), None)
        if installment:
            bunga_kontrak = installment['bunga']
            if jumlah >= installment['angsuran']:
                pokok_paid = installment['pokok']
                bunga_paid = bunga_kontrak
            elif jumlah >= bunga_kontrak:
                bunga_paid = bunga_kontrak
                pokok_paid = jumlah - bunga_kontrak
            else:
                bunga_paid = jumlah
                pokok_paid = Decimal('0')
            entries = [
                JurnalDetail(jurnal_header=header, akun=penerimaan.payment_account,
                             debit=jumlah, kredit=Decimal('0')),
            ]
            if pokok_paid > 0:
                entries.append(JurnalDetail(
                    jurnal_header=header, akun=ar_account,
                    debit=Decimal('0'), kredit=pokok_paid,
                ))
            if bunga_paid > 0:
                entries.append(JurnalDetail(
                    jurnal_header=header, akun=piutang.interest_income_account,
                    debit=Decimal('0'), kredit=bunga_paid,
                ))
            JurnalDetail.objects.bulk_create(entries)
            return header

    # Default non-PV: Dr. Kas / Cr. Piutang (tanpa_bunga or no angsuran_no)
    if piutang.coa_piutang_lancar_account_id:
        bal_lancar = max(
            Decimal('0'),
            _net_debit_balance_for_piutang(piutang.coa_piutang_lancar_account, piutang),
        )
        credit_lancar = min(jumlah, bal_lancar)
        credit_lt = jumlah - credit_lancar
        payment_lines = [
            JurnalDetail(jurnal_header=header, akun=penerimaan.payment_account,
                         debit=jumlah, kredit=Decimal('0')),
        ]
        if credit_lancar > 0:
            payment_lines.append(JurnalDetail(
                jurnal_header=header, akun=piutang.coa_piutang_lancar_account,
                debit=Decimal('0'), kredit=credit_lancar,
            ))
        if credit_lt > 0:
            payment_lines.append(JurnalDetail(
                jurnal_header=header, akun=piutang.coa_piutang_account,
                debit=Decimal('0'), kredit=credit_lt,
            ))
        JurnalDetail.objects.bulk_create(payment_lines)
    else:
        JurnalDetail.objects.bulk_create([
            JurnalDetail(jurnal_header=header, akun=penerimaan.payment_account,
                         debit=jumlah, kredit=Decimal('0')),
            JurnalDetail(jurnal_header=header, akun=ar_account,
                         debit=Decimal('0'), kredit=jumlah),
        ])
    return header


def auto_reverse_penyisihan_on_payment(piutang: PiutangHeader, user=None) -> None:
    from .models import PiutangPenyisihan
    if piutang.status != 'paid':
        return
    entries = list(
        PiutangPenyisihan.objects.filter(piutang_header=piutang, jenis='manual')
        .select_related('jurnal_header')
    )
    for entry in entries:
        reverse_penyisihan_journal(entry, user=user)


def compute_bagian_lancar(piutang: PiutangHeader) -> Decimal:
    if not piutang.jatuh_tempo:
        return Decimal('0')
    today = timezone.now().date()
    cutoff = today.replace(year=today.year + 1)

    # SAK EMKM: no amortised cost — use nominal carrying (sisa_piutang) only
    if get_standar_akuntansi(piutang) == 'sak_emkm':
        if piutang.jenis_jangka_waktu != 'long_term':
            return piutang.sisa_piutang if piutang.jatuh_tempo <= cutoff else Decimal('0')
        schedule = compute_angsuran_schedule(piutang)
        if not schedule:
            return piutang.sisa_piutang if piutang.jatuh_tempo <= cutoff else Decimal('0')
        return sum(
            (row['sisa_bayar'] for row in schedule if row['status'] != 'lunas' and row['tanggal'] <= cutoff),
            Decimal('0'),
        )

    # PV-adjusted: carrying amount of current portion via EIR discounting (consistent with reklasifikasi)
    if piutang.is_pv_adjusted and piutang.pv_discount_rate:
        if piutang.jenis_jangka_waktu != 'long_term' and piutang.jatuh_tempo > cutoff:
            return Decimal('0')
        return _compute_rkl_detail(piutang, today)

    if piutang.jenis_jangka_waktu != 'long_term':
        return piutang.sisa_piutang if piutang.jatuh_tempo <= cutoff else Decimal('0')
    schedule = compute_angsuran_schedule(piutang)
    if not schedule:
        return piutang.sisa_piutang if piutang.jatuh_tempo <= cutoff else Decimal('0')
    return sum(
        (row['sisa_bayar'] for row in schedule if row['status'] != 'lunas' and row['tanggal'] <= cutoff),
        Decimal('0'),
    )


def _compute_rkl_detail(piutang: PiutangHeader, as_of_date) -> Decimal:
    """
    PSAK 71: carrying amount of current-year installments.

    carrying_current = nominal_current − net_amort_current
    where net_amort_current = Σ(bunga_efektif) [net EIR] for current-window periods.

    Returns Decimal('0') if no current installments exist.
    """
    try:
        cutoff = as_of_date.replace(year=as_of_date.year + 1)
    except ValueError:
        cutoff = as_of_date.replace(year=as_of_date.year + 1, day=28)

    schedule = compute_angsuran_schedule(piutang)
    if schedule:
        all_unpaid = [r for r in schedule if r['status'] != 'lunas']
        current_rows = [
            r for r in all_unpaid
            if as_of_date < r['tanggal'] <= cutoff
        ]
        nominal_current = sum(
            max(
                Decimal('0'),
                row['pokok'] - max(Decimal('0'), row.get('paid', Decimal('0')) - row['bunga']),
            )
            for row in current_rows
        )
    else:
        nominal_current = (
            piutang.sisa_piutang
            if piutang.jatuh_tempo and as_of_date < piutang.jatuh_tempo <= cutoff
            else Decimal('0')
        )

    if nominal_current <= 0:
        return Decimal('0')

    if not piutang.is_pv_adjusted or not piutang.nilai_wajar_awal or not piutang.pv_discount_rate:
        return nominal_current

    # PSAK 71 amortised cost: current portion = Opening Carrying − Closing Carrying,
    # where Closing = carrying after all contractual payments due within 12 months.
    # This equals the carrying reduction expected in the next 12 months under the amort
    # schedule — NOT the PV of those cash flows discounted to today (those differ because
    # the "closing" amount is expressed in future-date terms, not discounted back to now).
    amort = compute_amortization_schedule_pv(piutang)
    if amort:
        opening_carrying = Decimal(str(piutang.nilai_wajar_awal))
        closing_carrying = None
        for row in amort:
            if row['tanggal'] <= as_of_date:
                opening_carrying = row['carrying_value']
            elif row['tanggal'] <= cutoff:
                closing_carrying = row['carrying_value']
        if closing_carrying is None:
            return Decimal('0')
        return max(Decimal('0'), (opening_carrying - closing_carrying).quantize(Decimal('0.0001')))

    # Bullet / no schedule: PV of face at maturity
    i_daily = (1 + float(piutang.pv_discount_rate) / 100) ** (1 / 365) - 1
    if piutang.jatuh_tempo:
        days = (piutang.jatuh_tempo - as_of_date).days
        if days > 0:
            return max(
                Decimal('0'),
                (piutang.sisa_piutang / Decimal(str((1 + i_daily) ** days))).quantize(Decimal('0.0001')),
            )
    return nominal_current


def create_reklasifikasi_bagian_lancar(
    piutang: PiutangHeader,
    dari_akun,
    ke_akun,
    tanggal,
    user=None,
    dari_akun_deferred=None,
    ke_akun_deferred=None,
) -> PiutangReklasifikasi:
    """
    PSAK 71: reklasifikasi carrying amount (not nominal + deferred separately).
    dari_akun_deferred / ke_akun_deferred are accepted for signature compatibility but ignored.
    """
    periode_bulan = tanggal.month
    periode_tahun = tanggal.year
    if PiutangReklasifikasi.objects.filter(
        piutang_header=piutang,
        periode_bulan=periode_bulan,
        periode_tahun=periode_tahun,
    ).exists():
        raise ValueError(
            f'Reklasifikasi bagian lancar untuk periode {periode_tahun}-{periode_bulan:02d} sudah ada.'
        )

    carrying_current = _compute_rkl_detail(piutang, tanggal)
    if carrying_current <= 0:
        raise ValueError('Tidak ada bagian lancar yang dapat direklasifikasi.')

    with transaction.atomic():
        nomor = _next_piutang_journal_number('TRX-PIU-RKL')
        jurnal = JurnalHeader.objects.create(
            tanggal=tanggal,
            nomor_transaksi=nomor,
            uraian_transaksi=f'Reklasifikasi Bagian Lancar {piutang.nomor_piutang}',
            entitas_bisnis=piutang.entitas_bisnis,
            is_penyesuaian=False,
        )
        JurnalDetail.objects.bulk_create([
            JurnalDetail(jurnal_header=jurnal, akun=dari_akun,
                         debit=Decimal('0'), kredit=carrying_current),
            JurnalDetail(jurnal_header=jurnal, akun=ke_akun,
                         debit=carrying_current, kredit=Decimal('0')),
        ])

        rkl = PiutangReklasifikasi.objects.create(
            piutang_header=piutang,
            tanggal=tanggal,
            dari_akun=dari_akun,
            ke_akun=ke_akun,
            jumlah=carrying_current,
            jumlah_deferred=None,
            dari_akun_deferred=None,
            ke_akun_deferred=None,
            keterangan=f'Bagian lancar {periode_tahun}-{periode_bulan:02d}',
            jurnal=jurnal,
            periode_bulan=periode_bulan,
            periode_tahun=periode_tahun,
            created_by=user,
        )
        _log(piutang, 'REKLASIFIKASI', user=user,
             after={'jumlah': str(carrying_current)})
    return rkl


def write_off_piutang(piutang: PiutangHeader, data: dict, user=None) -> PiutangWriteOff:
    with transaction.atomic():
        jumlah = piutang.sisa_piutang
        nomor = _next_piutang_journal_number('TRX-PIU-WO')
        metode = data['metode']
        bad_debt = data['bad_debt_account']
        allowance = data.get('allowance_account')

        dr_akun = allowance if metode == 'cadangan' and allowance else bad_debt

        header = JurnalHeader.objects.create(
            tanggal=data['tanggal'],
            nomor_transaksi=nomor,
            uraian_transaksi=f'Write-Off Piutang {piutang.nomor_piutang}',
            entitas_bisnis=piutang.entitas_bisnis,
            is_penyesuaian=False,
        )
        JurnalDetail.objects.bulk_create([
            JurnalDetail(jurnal_header=header, akun=dr_akun, debit=jumlah, kredit=Decimal('0')),
            JurnalDetail(jurnal_header=header, akun=piutang.coa_piutang_account, debit=Decimal('0'), kredit=jumlah),
        ])

        wo = PiutangWriteOff.objects.create(
            piutang_header=piutang,
            tanggal=data['tanggal'],
            jumlah_dihapus=jumlah,
            metode=metode,
            bad_debt_account=bad_debt,
            allowance_account=allowance,
            alasan=data.get('alasan', ''),
            jurnal=header,
            created_by=user,
        )
        piutang.status = 'written_off'
        piutang.is_locked = True
        piutang.save(update_fields=['status', 'is_locked'])
        _log(piutang, 'WRITE_OFF', user=user, after=_snapshot(piutang))
    return wo


def reverse_piutang_payment(penerimaan: PiutangPenerimaan, user=None) -> JurnalHeader:
    with transaction.atomic():
        piutang = PiutangHeader.objects.select_for_update().get(pk=penerimaan.piutang_header_id)
        tanggal_terima = penerimaan.tanggal_terima
        orig = penerimaan.jurnal_header
        nomor = _next_piutang_journal_number('TRX-PIU-PR')
        rev_header = JurnalHeader.objects.create(
            tanggal=timezone.now().date(),
            nomor_transaksi=nomor,
            uraian_transaksi=f'Reversal Penerimaan {piutang.nomor_piutang}',
            entitas_bisnis=piutang.entitas_bisnis,
            is_penyesuaian=True,
        )
        if orig:
            JurnalDetail.objects.bulk_create([
                JurnalDetail(
                    jurnal_header=rev_header,
                    akun=d.akun,
                    debit=d.kredit,
                    kredit=d.debit,
                )
                for d in orig.details.all()
            ])

        # ── PSAK 71: clean up pre-payment EIR journal when payment is reversed ─
        if piutang.is_pv_adjusted:
            nom = piutang.nomor_piutang

            # 1. Period amortisation (date-based): only remove if no other
            #    payment on the same date still holds ownership of this journal.
            other_same_date = (
                PiutangPenerimaan.objects
                .filter(piutang_header=piutang, tanggal_terima=tanggal_terima)
                .exclude(pk=penerimaan.pk)
                .exists()
            )
            if not other_same_date:
                period_amort = JurnalHeader.objects.filter(
                    uraian_transaksi__startswith=f'Amortisasi PV Piutang {nom} — ',
                    uraian_transaksi__endswith=f' s.d. {tanggal_terima}',
                    tanggal=tanggal_terima,
                ).first()
                if period_amort:
                    period_amort.details.all().delete()
                    period_amort.delete()


        penerimaan.delete()

        total_paid = piutang.penerimaan.aggregate(s=Sum('jumlah_diterima'))['s'] or Decimal('0')
        piutang.jumlah_terbayar = total_paid
        if total_paid <= 0:
            if piutang.status not in ('draft', 'cancelled', 'written_off'):
                piutang.status = 'open'
        elif total_paid < piutang.jumlah_pokok:
            piutang.status = 'partial'
        piutang.save(update_fields=['jumlah_terbayar', 'status'])
        _log(piutang, 'REVERSE_PAYMENT', user=user, after=_snapshot(piutang))
    return rev_header


_AGING_BUCKET_KEYS = ['current', '1_30', '31_60', '61_90', '91_180', '181_365', 'over_365']
_AGING_BUCKET_LABELS = {
    'current':  'Belum Jatuh Tempo',
    '1_30':     'Lewat 1–30 Hari',
    '31_60':    'Lewat 31–60 Hari',
    '61_90':    'Lewat 61–90 Hari',
    '91_180':   'Lewat 91–180 Hari',
    '181_365':  'Lewat 181–365 Hari',
    'over_365': 'Lewat > 365 Hari',
}


def _classify_bucket(tanggal, today) -> str:
    if tanggal is None:
        return 'current'
    delta = (today - tanggal).days
    if delta <= 0:
        return 'current'
    elif delta <= 30:
        return '1_30'
    elif delta <= 60:
        return '31_60'
    elif delta <= 90:
        return '61_90'
    elif delta <= 180:
        return '91_180'
    elif delta <= 365:
        return '181_365'
    else:
        return 'over_365'


def _long_term_bagian_lancar_aging(piutang, as_of_date: date) -> list:
    """
    For long-term receivables, returns a list of (amount, due_date) tuples,
    one per unpaid installment due within 12 months of as_of_date. Each installment is
    bucketed by its OWN due date so overdue and current entries land in separate buckets.
    The non-current portion (>12 months) is excluded.
    """
    cutoff = as_of_date.replace(year=as_of_date.year + 1)
    schedule = compute_angsuran_schedule(piutang)

    if not schedule:
        if piutang.jatuh_tempo and piutang.jatuh_tempo <= cutoff:
            return [(piutang.sisa_piutang, piutang.jatuh_tempo)]
        return []

    return [
        (row['sisa_bayar'], row['tanggal'])
        for row in schedule
        if row['status'] != 'lunas' and row['sisa_bayar'] > 0 and row['tanggal'] <= cutoff
    ]


def get_piutang_aging() -> dict:
    today = timezone.now().date()
    buckets = {k: [] for k in _AGING_BUCKET_KEYS}
    qs = (
        PiutangHeader.objects
        .filter(status__in=('open', 'partial', 'overdue'))
        .select_related('entitas_bisnis')
        .prefetch_related('penerimaan')
    )
    for piutang in qs:
        if piutang.jenis_jangka_waktu == 'long_term' and piutang.jatuh_tempo:
            # long-term: each installment within 12 months enters aging in its OWN bucket
            for amount, due_date in _long_term_bagian_lancar_aging(piutang, today):
                key = _classify_bucket(due_date, today)
                buckets[key].append({
                    'piutang': piutang,
                    'angsuran_no': None,
                    'tanggal_angsuran': due_date,
                    'jumlah': amount,
                    'hari_lewat': max(0, (today - due_date).days),
                })
            continue
        key = _classify_bucket(piutang.jatuh_tempo, today)
        buckets[key].append({
            'piutang': piutang,
            'angsuran_no': None,
            'tanggal_angsuran': piutang.jatuh_tempo,
            'jumlah': piutang.sisa_piutang,
            'hari_lewat': max(0, (today - piutang.jatuh_tempo).days) if piutang.jatuh_tempo else 0,
        })
    return buckets


def get_aging_schedule_report(as_of_date=None) -> dict:
    today = as_of_date or date.today()
    short_term_rows = []
    long_term_rows = []

    qs = (
        PiutangHeader.objects
        .filter(status__in=('open', 'partial', 'overdue'))
        .select_related('entitas_bisnis')
        .prefetch_related('penerimaan')
    )
    for piutang in qs:
        if piutang.jenis_jangka_waktu == 'long_term' and piutang.jatuh_tempo:
            # long-term: each installment within 12 months gets its own row, bucketed by its due date
            for amount, due_date in _long_term_bagian_lancar_aging(piutang, today):
                bucket_key = _classify_bucket(due_date, today)
                long_term_rows.append({
                    'nomor_piutang': piutang.nomor_piutang,
                    'debitur': piutang.entitas_display,
                    'tanggal': piutang.tanggal,
                    'jatuh_tempo_efektif': due_date,
                    'jumlah_piutang': amount,
                    'bucket_key': bucket_key,
                    'bucket_label': _AGING_BUCKET_LABELS[bucket_key],
                    'hari_lewat': max(0, (today - due_date).days),
                    'angsuran_no': None,
                    'piutang_pk': piutang.pk,
                })
            continue
        bucket_key = _classify_bucket(piutang.jatuh_tempo, today)
        short_term_rows.append({
            'nomor_piutang': piutang.nomor_piutang,
            'debitur': piutang.entitas_display,
            'tanggal': piutang.tanggal,
            'jatuh_tempo_efektif': piutang.jatuh_tempo,
            'jumlah_piutang': piutang.sisa_piutang,
            'bucket_key': bucket_key,
            'bucket_label': _AGING_BUCKET_LABELS[bucket_key],
            'hari_lewat': max(0, (today - piutang.jatuh_tempo).days) if piutang.jatuh_tempo else 0,
            'angsuran_no': None,
            'piutang_pk': piutang.pk,
        })

    def _section(rows):
        bucket_totals = {k: Decimal('0') for k in _AGING_BUCKET_KEYS}
        for r in rows:
            bucket_totals[r['bucket_key']] += r['jumlah_piutang']
        return {
            'rows': rows,
            'bucket_totals': bucket_totals,
            'grand_total': sum(bucket_totals.values()),
        }

    st = _section(short_term_rows)
    lt = _section(long_term_rows)
    combined = {k: st['bucket_totals'][k] + lt['bucket_totals'][k] for k in _AGING_BUCKET_KEYS}
    return {
        'as_of_date': today,
        'short_term': st,
        'long_term': lt,
        'combined_bucket_totals': combined,
        'grand_total': sum(combined.values()),
    }


def get_aging_schedule_workbook(report_data: dict):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    bucket_labels_ordered = [_AGING_BUCKET_LABELS[k] for k in _AGING_BUCKET_KEYS]
    header_row = (
        ['No Piutang', 'Debitur', 'Tanggal', 'Jatuh Tempo Efektif', 'No Angsuran', 'Total']
        + bucket_labels_ordered
    )
    bold = Font(bold=True)
    fill = PatternFill(start_color='DDEEFF', end_color='DDEEFF', fill_type='solid')

    for section_key, sheet_name in (('short_term', 'Jangka Pendek'), ('long_term', 'Jangka Panjang')):
        ws = wb.create_sheet(title=sheet_name)
        ws.append(header_row)
        for cell in ws[1]:
            cell.font = bold
            cell.fill = fill

        section = report_data[section_key]
        for row in section['rows']:
            ws.append([
                row['nomor_piutang'],
                row['debitur'],
                row['tanggal'].isoformat() if row['tanggal'] else '',
                row['jatuh_tempo_efektif'].isoformat() if row['jatuh_tempo_efektif'] else '',
                row['angsuran_no'] or '',
                float(row['jumlah_piutang']),
            ] + [
                float(row['jumlah_piutang']) if row['bucket_key'] == k else 0
                for k in _AGING_BUCKET_KEYS
            ])

        totals_row = ['', '', '', '', 'TOTAL', float(section['grand_total'])] + [
            float(section['bucket_totals'][k]) for k in _AGING_BUCKET_KEYS
        ]
        ws.append(totals_row)
        for cell in ws[ws.max_row]:
            cell.font = bold

    return wb


def _get_rate_config() -> dict:
    from .models import PenyisihanRateConfig
    return {r.bucket_key: r.rate_percent for r in PenyisihanRateConfig.objects.all()}


def compute_penyisihan_for_piutang(piutang) -> dict:
    rates = _get_rate_config()
    today = date.today()
    bucket_amounts = {k: Decimal('0') for k in _AGING_BUCKET_KEYS}

    if piutang.jenis_jangka_waktu == 'long_term' and piutang.jatuh_tempo:
        # long-term: only bagian lancar (≤12 months) enters aging; each installment in its own bucket.
        for amount, due_date in _long_term_bagian_lancar_aging(piutang, today):
            key = _classify_bucket(due_date, today)
            bucket_amounts[key] += amount
    else:
        key = _classify_bucket(piutang.jatuh_tempo, today)
        bucket_amounts[key] += piutang.sisa_piutang

    breakdown = []
    total = Decimal('0')
    for key in _AGING_BUCKET_KEYS:
        amt = bucket_amounts[key]
        rate = rates.get(key, Decimal('0'))
        penyisihan = (amt * rate / 100).quantize(Decimal('0.01'))
        total += penyisihan
        breakdown.append({
            'bucket_key': key,
            'label': _AGING_BUCKET_LABELS[key],
            'jumlah_piutang': amt,
            'rate': rate,
            'penyisihan': penyisihan,
        })
    return {'total_penyisihan': total, 'breakdown': breakdown}


def _next_penyisihan_journal_number(prefix='TRX-PIU-PSH') -> str:
    last = (
        JurnalHeader.objects
        .select_for_update()
        .filter(nomor_transaksi__startswith=f'{prefix}-')
        .order_by('-nomor_transaksi')
        .values_list('nomor_transaksi', flat=True)
        .first()
    )
    seq = 1
    if last:
        try:
            seq = int(last.rsplit('-', 1)[1]) + 1
        except (ValueError, IndexError):
            seq = 1
    return f'{prefix}-{seq:04d}'


def create_penyisihan_journal(
    piutang, allowance_account, expense_account, tanggal, catatan='', user=None
):
    from .models import PiutangPenyisihan
    piutang_pk = piutang.pk
    with transaction.atomic():
        piutang = PiutangHeader.objects.select_for_update().get(pk=piutang_pk)
        result = compute_penyisihan_for_piutang(piutang)
        total = result['total_penyisihan']
        if total <= 0:
            raise ValueError('Tidak ada penyisihan yang dapat dihitung untuk piutang ini.')
        nomor = _next_penyisihan_journal_number('TRX-PIU-PSH')
        header = JurnalHeader.objects.create(
            tanggal=tanggal,
            nomor_transaksi=nomor,
            uraian_transaksi=f'Penyisihan Piutang {piutang.nomor_piutang}',
            entitas_bisnis=piutang.entitas_bisnis,
            is_penyesuaian=True,
        )
        JurnalDetail.objects.bulk_create([
            JurnalDetail(jurnal_header=header, akun=expense_account, debit=total, kredit=Decimal('0')),
            JurnalDetail(jurnal_header=header, akun=allowance_account, debit=Decimal('0'), kredit=total),
        ])
        entry = PiutangPenyisihan.objects.create(
            piutang_header=piutang,
            tanggal=tanggal,
            jenis='manual',
            jumlah=total,
            allowance_account=allowance_account,
            expense_account=expense_account,
            jurnal_header=header,
            catatan=catatan,
            created_by=user,
        )
        piutang.is_specifically_impaired = True
        piutang.save(update_fields=['is_specifically_impaired'])
        _log(piutang, 'PENYISIHAN', user=user, after={'jumlah': str(total), 'nomor': nomor})
    return entry


def reverse_penyisihan_journal(entry, user=None) -> None:
    from .models import PiutangPenyisihan
    piutang = entry.piutang_header
    with transaction.atomic():
        if entry.jurnal_header_id:
            orig = entry.jurnal_header
            nomor = _next_penyisihan_journal_number('TRX-PIU-PSHR')
            rev = JurnalHeader.objects.create(
                tanggal=timezone.now().date(),
                nomor_transaksi=nomor,
                uraian_transaksi=f'Reversal Penyisihan {piutang.nomor_piutang if piutang else ""}',
                entitas_bisnis=piutang.entitas_bisnis if piutang else None,
                is_penyesuaian=True,
            )
            JurnalDetail.objects.bulk_create([
                JurnalDetail(jurnal_header=rev, akun=d.akun, debit=d.kredit, kredit=d.debit)
                for d in orig.details.all()
            ])
        entry.delete()
        if piutang:
            remaining = PiutangPenyisihan.objects.filter(
                piutang_header=piutang, jenis='manual'
            ).exists()
            if not remaining:
                piutang.is_specifically_impaired = False
                piutang.save(update_fields=['is_specifically_impaired'])
            _log(piutang, 'PENYISIHAN', user=user, notes='Jurnal penyisihan dibatalkan')


def update_penyisihan_individual(
    existing_entry, allowance_account, expense_account, tanggal, catatan='', user=None,
):
    piutang = existing_entry.piutang_header
    with transaction.atomic():
        reverse_penyisihan_journal(existing_entry, user=user)
        new_entry = create_penyisihan_journal(
            piutang=piutang,
            allowance_account=allowance_account,
            expense_account=expense_account,
            tanggal=tanggal,
            catatan=catatan,
            user=user,
        )
    return new_entry


def compute_batch_penyisihan(tanggal) -> dict:
    """
    Compute batch allowance grouped by (entitas_bisnis, allowance_account, expense_account).
    Each piutang must have penyisihan_allowance_account and penyisihan_expense_account set.
    Piutangs without accounts are returned separately as 'unconfigured'.

    Returns:
        groups:        list of per-group dicts with entitas_bisnis, allowance_account,
                       expense_account, target_saldo, saldo_existing, delta, breakdown,
                       and piutang_list.
        unconfigured:  list of PiutangHeader with missing account config.
        piutang_count: total piutangs processed (configured + unconfigured).
        total_target:  sum of target_saldo across groups.
        total_delta:   sum of delta across groups.
    """
    rates = _get_rate_config()
    today = tanggal

    qs = (
        PiutangHeader.objects
        .filter(status__in=('open', 'partial', 'overdue'), is_specifically_impaired=False)
        .select_related(
            'penyisihan_allowance_account', 'penyisihan_expense_account',
            'entitas_bisnis',
        )
        .prefetch_related('penerimaan')
    )

    # group_key → {entitas_bisnis, allowance_account, expense_account, bucket_amounts, piutang_list}
    groups: dict = {}
    unconfigured: list = []
    piutang_count = 0

    for piutang in qs:
        piutang_count += 1

        if not piutang.penyisihan_allowance_account_id or not piutang.penyisihan_expense_account_id:
            unconfigured.append(piutang)
            continue

        gk = (
            piutang.entitas_bisnis_id,
            piutang.penyisihan_allowance_account_id,
            piutang.penyisihan_expense_account_id,
        )
        if gk not in groups:
            groups[gk] = {
                'entitas_bisnis': piutang.entitas_bisnis,
                'allowance_account': piutang.penyisihan_allowance_account,
                'expense_account': piutang.penyisihan_expense_account,
                'bucket_amounts': {k: Decimal('0') for k in _AGING_BUCKET_KEYS},
                'piutang_list': [],
            }
        g = groups[gk]

        if piutang.jenis_jangka_waktu == 'long_term' and piutang.jatuh_tempo:
            row_total = Decimal('0')
            for amount, due_date in _long_term_bagian_lancar_aging(piutang, today):
                key = _classify_bucket(due_date, today)
                g['bucket_amounts'][key] += amount
                row_total += (amount * rates.get(key, Decimal('0')) / 100).quantize(Decimal('0.01'))
            g['piutang_list'].append({'piutang': piutang, 'penyisihan': row_total})
        else:
            key = _classify_bucket(piutang.jatuh_tempo, today)
            amount = piutang.sisa_piutang
            g['bucket_amounts'][key] += amount
            row_total = (amount * rates.get(key, Decimal('0')) / 100).quantize(Decimal('0.01'))
            g['piutang_list'].append({'piutang': piutang, 'penyisihan': row_total})

    from .models import PiutangPenyisihan as _PP
    result_groups = []
    total_target = Decimal('0')
    total_delta = Decimal('0')

    for gk, g in groups.items():
        target_saldo = Decimal('0')
        breakdown = []
        for key in _AGING_BUCKET_KEYS:
            amt = g['bucket_amounts'][key]
            rate = rates.get(key, Decimal('0'))
            penyisihan = (amt * rate / 100).quantize(Decimal('0.01'))
            target_saldo += penyisihan
            breakdown.append({
                'bucket_key': key,
                'label': _AGING_BUCKET_LABELS[key],
                'jumlah_piutang': amt,
                'rate': rate,
                'penyisihan': penyisihan,
            })

        eb_id = gk[0]
        saldo_existing = Decimal(str(
            _PP.objects
            .filter(
                jenis='batch',
                allowance_account=g['allowance_account'],
                entitas_bisnis_id=eb_id,
                tanggal__lte=tanggal,
            )
            .aggregate(s=Sum('jumlah'))['s'] or Decimal('0')
        )).quantize(Decimal('0.01'))
        delta = (target_saldo - saldo_existing).quantize(Decimal('0.01'))

        total_target += target_saldo
        total_delta += delta
        result_groups.append({
            'entitas_bisnis': g['entitas_bisnis'],
            'allowance_account': g['allowance_account'],
            'expense_account': g['expense_account'],
            'target_saldo': target_saldo,
            'saldo_existing': saldo_existing,
            'delta': delta,
            'breakdown': breakdown,
            'piutang_list': g['piutang_list'],
        })

    return {
        'groups': result_groups,
        'unconfigured': unconfigured,
        'piutang_count': piutang_count,
        'total_target': total_target,
        'total_delta': total_delta,
    }


def create_batch_penyisihan_journal(
    batch_data: dict,
    tanggal, catatan='', periode_label='', user=None,
):
    """
    Create one journal per (entitas_bisnis, allowance_account) group with non-zero delta.
    Each journal is attributed to the entitas_bisnis of the piutang group.
    Returns list of PiutangPenyisihan entries created.
    """
    from .models import PiutangPenyisihan
    active_groups = [g for g in batch_data['groups'] if g['delta'] != 0]
    if not active_groups:
        raise ValueError('Semua delta adalah 0, tidak perlu jurnal.')
    created = []
    with transaction.atomic():
        for g in active_groups:
            delta = g['delta']
            abs_delta = abs(delta)
            allowance_account = g['allowance_account']
            expense_account = g['expense_account']
            eb = g.get('entitas_bisnis')
            eb_id = eb.pk if eb else None
            # Uniqueness per (periode_label, entitas_bisnis, allowance_account)
            if periode_label and PiutangPenyisihan.objects.filter(
                jenis='batch',
                periode_label=periode_label,
                entitas_bisnis_id=eb_id,
                allowance_account=allowance_account,
            ).exists():
                eb_name = eb.nama if eb else 'tanpa EB'
                raise ValueError(
                    f'Jurnal penyisihan batch untuk periode {periode_label}, '
                    f'EB {eb_name}, akun {allowance_account.kode_akun} sudah ada.'
                )
            nomor = _next_penyisihan_journal_number('TRX-PIU-PSH-B')
            eb_label = f' [{eb.nama}]' if eb else ''
            header = JurnalHeader.objects.create(
                tanggal=tanggal,
                nomor_transaksi=nomor,
                uraian_transaksi=(
                    f'Penyisihan Piutang Batch — {tanggal}'
                    f'{eb_label} [{allowance_account.kode_akun}]'
                ),
                is_penyesuaian=True,
                entitas_bisnis=eb,
            )
            if delta > 0:
                JurnalDetail.objects.bulk_create([
                    JurnalDetail(jurnal_header=header, akun=expense_account,
                                 debit=abs_delta, kredit=Decimal('0')),
                    JurnalDetail(jurnal_header=header, akun=allowance_account,
                                 debit=Decimal('0'), kredit=abs_delta),
                ])
            else:
                JurnalDetail.objects.bulk_create([
                    JurnalDetail(jurnal_header=header, akun=allowance_account,
                                 debit=abs_delta, kredit=Decimal('0')),
                    JurnalDetail(jurnal_header=header, akun=expense_account,
                                 debit=Decimal('0'), kredit=abs_delta),
                ])
            entry = PiutangPenyisihan.objects.create(
                piutang_header=None,
                entitas_bisnis=eb,
                tanggal=tanggal,
                jenis='batch',
                jumlah=delta,
                allowance_account=allowance_account,
                expense_account=expense_account,
                jurnal_header=header,
                catatan=catatan,
                periode_label=periode_label,
                created_by=user,
            )
            created.append(entry)
    return created


def reverse_batch_penyisihan_journal(entry, user=None):
    """
    Reverse a single PiutangPenyisihan batch entry and its associated JurnalHeader.
    Deletes the journal and the penyisihan entry inside a transaction.
    """
    from .models import PiutangPenyisihan
    from apps.jurnal.models import JurnalHeader as _JH
    with transaction.atomic():
        header = entry.jurnal_header
        entry.delete()
        if header:
            header.delete()


def get_piutang_dashboard_kpi() -> dict:
    today = timezone.now().date()
    month_start = today.replace(day=1)
    outstanding_qs = PiutangHeader.objects.filter(status__in=('open', 'partial', 'overdue'))
    total_outstanding = outstanding_qs.aggregate(s=Sum('jumlah_pokok'))['s'] or Decimal('0')
    total_terbayar = outstanding_qs.aggregate(s=Sum('jumlah_terbayar'))['s'] or Decimal('0')
    total_outstanding -= total_terbayar

    overdue_qs = outstanding_qs.filter(jatuh_tempo__lt=today)
    total_overdue = sum(p.sisa_piutang for p in overdue_qs)

    collected_this_month = (
        PiutangPenerimaan.objects
        .filter(tanggal_terima__gte=month_start)
        .aggregate(s=Sum('jumlah_diterima'))['s'] or Decimal('0')
    )
    collection_rate = (
        collected_this_month / (collected_this_month + total_outstanding) * 100
        if (collected_this_month + total_outstanding) > 0 else Decimal('0')
    )

    aging_buckets = get_piutang_aging()
    rates = _get_rate_config()
    aging_summary = {}
    total_penyisihan_target = Decimal('0')
    for key in _AGING_BUCKET_KEYS:
        total_amt = sum(entry['jumlah'] for entry in aging_buckets[key])
        rate = rates.get(key, Decimal('0'))
        penyisihan = (Decimal(str(total_amt)) * rate / 100).quantize(Decimal('0.01'))
        total_penyisihan_target += penyisihan
        aging_summary[key] = {
            'label': _AGING_BUCKET_LABELS[key],
            'total_outstanding': Decimal(str(total_amt)),
            'rate': rate,
            'penyisihan': penyisihan,
        }
    total_penyisihan_target = total_penyisihan_target.quantize(Decimal('0.01'))
    piutang_neto = (total_outstanding - total_penyisihan_target).quantize(Decimal('0.01'))

    return {
        'total_outstanding': total_outstanding,
        'total_overdue': total_overdue,
        'collected_this_month': collected_this_month,
        'collection_rate': collection_rate.quantize(Decimal('0.01')),
        'total_penyisihan_target': total_penyisihan_target,
        'piutang_neto': piutang_neto,
        'aging_summary': aging_summary,
    }


def get_piutang_disclosure_report(as_of_date=None) -> dict:
    from .models import PiutangPenyisihan
    today = as_of_date or date.today()
    year_start = today.replace(month=1, day=1)

    outstanding_list = list(
        PiutangHeader.objects
        .filter(status__in=('open', 'partial', 'overdue'))
        .select_related('entitas_bisnis')
    )

    def _carrying(p: PiutangHeader) -> Decimal:
        return _pv_carrying_value(p) if p.is_pv_adjusted else p.sisa_piutang

    total_outstanding = sum(_carrying(p) for p in outstanding_list)

    short_term_total = sum(_carrying(p) for p in outstanding_list if p.jenis_jangka_waktu == 'short_term')
    long_term_total = sum(_carrying(p) for p in outstanding_list if p.jenis_jangka_waktu == 'long_term')

    penyisihan_opening = (
        PiutangPenyisihan.objects.filter(tanggal__lt=year_start)
        .aggregate(s=Sum('jumlah'))['s'] or Decimal('0')
    )
    penyisihan_ytd_pos = (
        PiutangPenyisihan.objects
        .filter(tanggal__gte=year_start, tanggal__lte=today, jumlah__gt=0)
        .aggregate(s=Sum('jumlah'))['s'] or Decimal('0')
    )
    penyisihan_ytd_neg = abs(
        PiutangPenyisihan.objects
        .filter(tanggal__gte=year_start, tanggal__lte=today, jumlah__lt=0)
        .aggregate(s=Sum('jumlah'))['s'] or Decimal('0')
    )
    total_penyisihan = (
        PiutangPenyisihan.objects.filter(tanggal__lte=today)
        .aggregate(s=Sum('jumlah'))['s'] or Decimal('0')
    )

    # Write-off total — adapt if PiutangWriteOff doesn't exist
    try:
        write_off_total = (
            PiutangWriteOff.objects.filter(tanggal__gte=year_start, tanggal__lte=today)
            .aggregate(s=Sum('jumlah_dihapus'))['s'] or Decimal('0')
        )
    except Exception:
        write_off_total = Decimal('0')

    # Impaired piutang — adapt if field doesn't exist
    try:
        impaired = [p for p in outstanding_list if p.is_specifically_impaired]
    except AttributeError:
        impaired = []
    impaired_total = sum(_carrying(p) for p in impaired)

    concentration = sorted(
        [{'debitur': p.entitas_display, 'outstanding': _carrying(p)} for p in outstanding_list],
        key=lambda x: x['outstanding'], reverse=True,
    )[:10]
    for item in concentration:
        item['pct'] = (
            (item['outstanding'] / total_outstanding * 100).quantize(Decimal('0.01'))
            if total_outstanding else Decimal('0')
        )

    rates = _get_rate_config()
    aging_buckets = get_piutang_aging()
    aging_summary = []
    for key in _AGING_BUCKET_KEYS:
        bucket_total = sum(Decimal(str(e['jumlah'])) for e in aging_buckets[key])
        rate = rates.get(key, Decimal('0'))
        aging_summary.append({
            'bucket_key': key,
            'label': _AGING_BUCKET_LABELS[key],
            'total': bucket_total,
            'rate': rate,
            'penyisihan': (bucket_total * rate / 100).quantize(Decimal('0.01')),
        })

    return {
        'as_of_date': today,
        'total_outstanding': total_outstanding,
        'short_term_total': short_term_total,
        'long_term_total': long_term_total,
        'total_penyisihan': total_penyisihan,
        'piutang_neto': total_outstanding - total_penyisihan,
        'allowance_reconciliation': {
            'opening': penyisihan_opening,
            'additions': penyisihan_ytd_pos,
            'reversals': penyisihan_ytd_neg,
            'write_offs': write_off_total,
            'closing': penyisihan_opening + penyisihan_ytd_pos - penyisihan_ytd_neg,
        },
        'impaired_count': len(impaired),
        'impaired_total': impaired_total,
        'concentration': concentration,
        'aging_summary': aging_summary,
    }


# ── PV Carrying-Value Helpers ────────────────────────────────────────────────


def _pv_carrying_value(piutang: PiutangHeader) -> Decimal:
    """
    PSAK 71: carrying amount = net debit balance of piutang accounts across all journals.
    Includes both coa_piutang_account (LT) and coa_piutang_lancar_account (current) so
    reklasifikasi transfers (which shift balance between the two) are transparent.
    """
    if not piutang.is_pv_adjusted or not piutang.nilai_wajar_awal:
        return piutang.sisa_piutang
    account_ids = [piutang.coa_piutang_account_id]
    if piutang.coa_piutang_lancar_account_id:
        account_ids.append(piutang.coa_piutang_lancar_account_id)
    journal_ids = _piutang_journal_ids(piutang)
    if not journal_ids:
        return piutang.nilai_wajar_awal
    result = (
        JurnalDetail.objects
        .filter(akun_id__in=account_ids, jurnal_header_id__in=journal_ids)
        .aggregate(debit=Sum('debit'), kredit=Sum('kredit'))
    )
    net = (result['debit'] or Decimal('0')) - (result['kredit'] or Decimal('0'))
    return max(Decimal('0'), net.quantize(Decimal('0.0001')))


def _pv_last_amortization_date(piutang: PiutangHeader) -> date:
    """
    Date through which effective interest has been recognised.
    Uses the latest Amortization or Accrual journal date (reversals do NOT reset this —
    the reversal un-recognises income but interest still ran until the accrual date).
    Falls back to piutang.tanggal (inception) if no journals exist yet.
    """
    nom = piutang.nomor_piutang
    from django.db.models import Q
    last = (
        JurnalHeader.objects
        .filter(
            Q(uraian_transaksi__startswith=f'Amortisasi PV Piutang {nom} —') |
            Q(uraian_transaksi__startswith=f'Akrual PV Piutang {nom} —')
        )
        .order_by('-tanggal', '-nomor_transaksi')
        .values_list('tanggal', flat=True)
        .first()
    )
    return last if last else piutang.tanggal


def _pv_effective_interest_days(piutang: PiutangHeader, from_date: date, to_date: date) -> Decimal:
    """
    Effective interest for actual days elapsed between from_date and to_date,
    using compound daily rate: i_daily = (1 + annual_rate)^(1/365) − 1.
    Computed on the CURRENT carrying value (pre-payment, so call before updating jumlah_terbayar).
    """
    days = (to_date - from_date).days
    if days <= 0 or not piutang.pv_discount_rate:
        return Decimal('0')
    carrying = _pv_carrying_value(piutang)
    if carrying <= 0:
        return Decimal('0')
    i_daily = (1 + float(piutang.pv_discount_rate) / 100) ** (1 / 365) - 1
    bunga = float(carrying) * i_daily * days
    return Decimal(str(round(bunga, 4)))


# ── Present Value Functions ───────────────────────────────────────────────────

def compute_present_value(piutang: PiutangHeader, market_rate: Decimal) -> Decimal:
    """Compute PV of all future cash flows discounted at market_rate (% per year).

    At 0% the result equals the sum of angsuran (== jumlah_pokok for tanpa_bunga).
    """
    schedule = compute_angsuran_schedule(piutang)
    if not schedule:
        return piutang.jumlah_pokok
    if market_rate == 0:
        total = sum(Decimal(str(row['angsuran'])) for row in schedule)
        return total
    periode_months = _PERIODE_MONTHS_MAP.get(getattr(piutang, 'periode_angsuran', 'bulanan'), 1)
    # Compound rate per period: (1 + annual_rate)^(months/12) - 1  (PSAK 71 effective interest)
    r_per_period = (1 + float(market_rate) / 100) ** (periode_months / 12) - 1
    pv = Decimal('0')
    for row in schedule:
        n = row['no']
        cf = float(row['angsuran'])
        pv += Decimal(str(round(cf / ((1 + r_per_period) ** n), 4)))
    return pv.quantize(Decimal('0.0001'))


def _create_piutang_ar_journal(piutang: PiutangHeader) -> JurnalHeader:
    details = list(piutang.details.select_related('revenue_account').all())
    missing = [d.deskripsi or str(d.pk) for d in details if not d.revenue_account_id]
    if missing:
        raise ValueError(
            f'Akun pendapatan belum diisi untuk detail: {", ".join(missing)}. '
            'Isi akun pendapatan di setiap baris detail sebelum posting.'
        )
    nomor = _next_piutang_journal_number('TRX-PIU-POST')
    header = JurnalHeader.objects.create(
        tanggal=piutang.tanggal,
        nomor_transaksi=nomor,
        uraian_transaksi=f'Pengakuan Piutang {piutang.nomor_piutang}',
        entitas_bisnis=piutang.entitas_bisnis,
        is_penyesuaian=False,
    )

    if piutang.is_pv_adjusted and piutang.nilai_wajar_awal:
        # PSAK 71 amortised cost: Dr. Piutang at fair value / Cr. Revenue at fair value.
        # No deferred income — the discount is implicit in the carrying amount.
        pv = piutang.nilai_wajar_awal
        JurnalDetail.objects.create(
            jurnal_header=header,
            akun=piutang.coa_piutang_account,
            debit=pv,
            kredit=Decimal('0'),
        )
        total_detail = sum(d.jumlah for d in details) or pv
        cumulative = Decimal('0')
        detail_entries = []
        for i, detail in enumerate(details):
            if i == len(details) - 1:
                kredit = pv - cumulative
            else:
                kredit = (detail.jumlah / total_detail * pv).quantize(Decimal('0.0001'))
            cumulative += kredit
            detail_entries.append(JurnalDetail(
                jurnal_header=header,
                akun=detail.revenue_account,
                debit=Decimal('0'),
                kredit=kredit,
            ))
        JurnalDetail.objects.bulk_create(detail_entries)
    else:
        # Standard method: Dr. Piutang (face) / Cr. Revenue (face)
        JurnalDetail.objects.create(
            jurnal_header=header,
            akun=piutang.coa_piutang_account,
            debit=piutang.jumlah_pokok,
            kredit=Decimal('0'),
        )
        JurnalDetail.objects.bulk_create([
            JurnalDetail(
                jurnal_header=header,
                akun=detail.revenue_account,
                debit=Decimal('0'),
                kredit=detail.jumlah,
            )
            for detail in details
        ])

    # ── PSAK 71 / SAK EP: biaya transaksi dikapitalisasi ke nilai tercatat ───
    # Dr. Piutang (biaya_transaksi) / Cr. Akun Offset Biaya Transaksi (Kas/Bank)
    standar = get_standar_akuntansi(piutang)
    if (standar in ('psak', 'sak_ep')
            and piutang.biaya_transaksi
            and piutang.biaya_transaksi > Decimal('0')
            and piutang.biaya_transaksi_account_id):
        JurnalDetail.objects.bulk_create([
            JurnalDetail(
                jurnal_header=header,
                akun=piutang.coa_piutang_account,
                debit=piutang.biaya_transaksi,
                kredit=Decimal('0'),
            ),
            JurnalDetail(
                jurnal_header=header,
                akun=piutang.biaya_transaksi_account,
                debit=Decimal('0'),
                kredit=piutang.biaya_transaksi,
            ),
        ])

    return header


def apply_pv_assessment(piutang: PiutangHeader) -> list[str]:
    """Compute and apply PV adjustment on piutang if eligible.

    Sets nilai_wajar_awal and is_pv_adjusted directly on the instance (caller must save).
    Returns list of field names that were modified (empty list if no change needed).
    Skipped for SAK EMKM or when pv_discount_rate is not set.
    """
    if get_standar_akuntansi(piutang) == 'sak_emkm':
        return []
    if piutang.pv_discount_rate and not piutang.nilai_wajar_awal:
        piutang.nilai_wajar_awal = compute_present_value(piutang, piutang.pv_discount_rate)
        piutang.is_pv_adjusted = True
        return ['nilai_wajar_awal', 'is_pv_adjusted']
    return []


def post_piutang(piutang: PiutangHeader, user=None) -> JurnalHeader:
    with transaction.atomic():
        piutang = PiutangHeader.objects.select_for_update().get(pk=piutang.pk)
        if piutang.status != 'draft':
            raise ValueError(
                f'Hanya piutang berstatus draft yang dapat di-post. Status saat ini: {piutang.get_status_display()}.'
            )
        update_fields = ['status', 'is_locked']
        update_fields += apply_pv_assessment(piutang)
        jurnal = _create_piutang_ar_journal(piutang)
        piutang.status = 'open'
        piutang.is_locked = True
        piutang.save(update_fields=update_fields)
        _log(piutang, 'POSTED', user=user, after={'status': 'open', 'jurnal': jurnal.nomor_transaksi})
    return jurnal


def submit_for_approval(piutang: PiutangHeader, user=None) -> None:
    with transaction.atomic():
        piutang = PiutangHeader.objects.select_for_update().get(pk=piutang.pk)
        if piutang.status != 'draft':
            raise ValueError('Hanya piutang berstatus draft yang dapat disubmit untuk approval.')
        piutang.status = 'pending_approval'
        piutang.save(update_fields=['status'])
        _log(piutang, 'SUBMITTED', user=user, after={'status': 'pending_approval'})


def approve_piutang(piutang: PiutangHeader, user=None) -> JurnalHeader:
    with transaction.atomic():
        piutang = PiutangHeader.objects.select_for_update().get(pk=piutang.pk)
        if piutang.status != 'pending_approval':
            raise ValueError('Hanya piutang berstatus pending_approval yang dapat disetujui.')
        update_fields = ['status', 'is_locked', 'approved_by', 'approved_at']
        update_fields += apply_pv_assessment(piutang)
        jurnal = _create_piutang_ar_journal(piutang)
        piutang.status = 'open'
        piutang.is_locked = True
        piutang.approved_by = user
        piutang.approved_at = timezone.now()
        piutang.save(update_fields=update_fields)
        _log(piutang, 'APPROVED', user=user, after={'status': 'open', 'jurnal': jurnal.nomor_transaksi})
    return jurnal


def reject_piutang(piutang: PiutangHeader, user=None, alasan: str = '') -> None:
    with transaction.atomic():
        piutang = PiutangHeader.objects.select_for_update().get(pk=piutang.pk)
        if piutang.status != 'pending_approval':
            raise ValueError('Hanya piutang berstatus pending_approval yang dapat ditolak.')
        piutang.status = 'draft'
        piutang.save(update_fields=['status'])
        _log(piutang, 'REJECTED', user=user, after={'status': 'draft', 'alasan': alasan})


def compute_amortization_schedule_pv(piutang: PiutangHeader) -> list:
    """Effective-interest amortization table using nilai_wajar_awal and pv_discount_rate."""
    if not piutang.nilai_wajar_awal or not piutang.pv_discount_rate:
        return []
    schedule = compute_angsuran_schedule(piutang)
    if not schedule:
        return []
    periode_months = _PERIODE_MONTHS_MAP.get(getattr(piutang, 'periode_angsuran', 'bulanan'), 1)
    # Compound rate per period: (1 + annual_rate)^(months/12) - 1
    r_per_period = (1 + float(piutang.pv_discount_rate) / 100) ** (periode_months / 12) - 1
    carrying = float(piutang.nilai_wajar_awal)
    rows = []
    for row in schedule:
        bunga_efektif_gross = carrying * r_per_period
        bunga_kontrak = float(row.get('bunga') or 0)
        # Net discount amortization = effective − contractual coupon.
        # Can be negative when contractual coupon > effective interest
        # (e.g. late periods of a flat-rate loan where CA has declined below face).
        # Do NOT clamp — over-clamping causes cumulative over-recognition of discount.
        net_amortization = bunga_efektif_gross - bunga_kontrak
        cf = float(row['angsuran'])
        # Carrying evolves: add full effective, subtract full angsuran cash flow
        carrying = carrying + bunga_efektif_gross - cf
        rows.append({
            'periode': row['no'],
            'tanggal': row['tanggal'],
            'bunga_efektif': Decimal(str(round(net_amortization, 4))),
            'bunga_efektif_gross': Decimal(str(round(bunga_efektif_gross, 4))),
            'cash_flow': row['angsuran'],
            'carrying_value': Decimal(str(round(carrying, 4))),
        })
    return rows


def create_pv_adjustment_journal(
    piutang: PiutangHeader,
    interest_income_account=None,
    tanggal=None,
    catatan='',
    user=None,
    periode_no: int | None = None,
) -> JurnalHeader:
    """PSAK 71: Dr. Piutang (gross EIR) / Cr. Pendapatan Bunga Efektif."""
    income_account = interest_income_account or piutang.interest_income_account
    if not income_account:
        raise ValueError('Akun Pendapatan Bunga Efektif diperlukan untuk amortisasi.')
    if not piutang.is_pv_adjusted or not piutang.nilai_wajar_awal or not piutang.pv_discount_rate:
        raise ValueError('Piutang belum disesuaikan nilai wajar (PV).')
    if tanggal is None:
        tanggal = timezone.now().date()
    amort = compute_amortization_schedule_pv(piutang)
    if not amort:
        raise ValueError('Jadwal amortisasi PV tidak tersedia.')

    if periode_no is None:
        # Count only periodic journals ('— Periode N') to avoid matching
        # pre-payment EIR journals ('— {from_date} s.d. {to_date}').
        recorded = JurnalHeader.objects.filter(
            uraian_transaksi__startswith=f'Amortisasi PV Piutang {piutang.nomor_piutang} — Periode',
        ).count()
        periode_no = recorded + 1

    if periode_no < 1 or periode_no > len(amort):
        raise ValueError(f'Periode {periode_no} tidak valid (total {len(amort)} periode).')

    row = amort[periode_no - 1]
    bunga_gross = row['bunga_efektif_gross']
    if bunga_gross <= Decimal('0.005'):
        raise ValueError('Bunga efektif nol, tidak perlu jurnal.')

    # PSAK 71 Stage 3: interest income on net carrying (gross − ECL allowance)
    if piutang.stage_ecl == 3 and get_standar_akuntansi(piutang) == 'psak':
        gross_ca = _pv_carrying_value(piutang)
        net_ca = _get_ecl_net_carrying(piutang)
        ratio = (net_ca / gross_ca) if gross_ca > Decimal('0') else Decimal('1')
        bunga = (bunga_gross * ratio).quantize(Decimal('0.0001'))
    else:
        bunga = bunga_gross

    if bunga <= Decimal('0.005'):
        raise ValueError('Bunga efektif nol, tidak perlu jurnal.')

    ar_account = piutang.coa_piutang_account
    with transaction.atomic():
        nomor = _next_piutang_journal_number('TRX-PIU-PV')
        stage_label = f' [Stage {piutang.stage_ecl}]' if piutang.stage_ecl == 3 else ''
        header = JurnalHeader.objects.create(
            tanggal=tanggal,
            nomor_transaksi=nomor,
            uraian_transaksi=f'Amortisasi PV Piutang {piutang.nomor_piutang} — Periode {periode_no}{stage_label}',
            entitas_bisnis=piutang.entitas_bisnis,
            is_penyesuaian=True,
        )
        JurnalDetail.objects.bulk_create([
            JurnalDetail(
                jurnal_header=header,
                akun=ar_account,
                debit=bunga,
                kredit=Decimal('0'),
            ),
            JurnalDetail(
                jurnal_header=header,
                akun=income_account,
                debit=Decimal('0'),
                kredit=bunga,
            ),
        ])
        _log(
            piutang, 'EDITED', user=user,
            after={'pv_amortisasi': str(bunga), 'periode': periode_no},
        )
    return header


def create_pv_accrual_journal(
    piutang: PiutangHeader,
    tanggal: date,
    interest_income_account=None,
    catatan: str = '',
    user=None,
) -> JurnalHeader:
    """
    PSAK 71: period-end accrual journal.
    Dr. Piutang / Cr. Pendapatan Bunga Efektif.
    Stage 3 (PSAK): interest accrued on net carrying (gross − ECL allowance).
    Must be paired with a reversal at start of next period.
    """
    if not piutang.is_pv_adjusted or not piutang.pv_discount_rate:
        raise ValueError('Piutang belum disesuaikan nilai wajar (PV).')
    income_account = interest_income_account or piutang.interest_income_account
    if not income_account:
        raise ValueError('Akun Pendapatan Bunga Efektif diperlukan.')

    from_date = _pv_last_amortization_date(piutang)
    bunga_gross = _pv_effective_interest_days(piutang, from_date, tanggal)
    if bunga_gross <= Decimal('0'):
        raise ValueError('Tidak ada selisih bunga efektif yang dapat diakrualkan untuk periode ini.')

    # PSAK 71 Stage 3: accrue on net carrying
    if piutang.stage_ecl == 3 and get_standar_akuntansi(piutang) == 'psak':
        gross_ca = _pv_carrying_value(piutang)
        net_ca = _get_ecl_net_carrying(piutang)
        ratio = (net_ca / gross_ca) if gross_ca > Decimal('0') else Decimal('1')
        bunga = (bunga_gross * ratio).quantize(Decimal('0.0001'))
    else:
        bunga = bunga_gross

    if bunga <= Decimal('0'):
        raise ValueError('Tidak ada selisih bunga efektif yang dapat diakrualkan untuk periode ini.')

    nom = piutang.nomor_piutang
    n_accrual = JurnalHeader.objects.filter(
        uraian_transaksi__startswith=f'Akrual PV Piutang {nom} —'
    ).count()
    n_reversal = JurnalHeader.objects.filter(
        uraian_transaksi__startswith=f'Balik Akrual PV Piutang {nom} —'
    ).count()
    if n_accrual > n_reversal:
        raise ValueError(
            'Masih ada jurnal akrual yang belum dibalik. Balik akrual sebelumnya terlebih dahulu.'
        )

    ar_account = piutang.coa_piutang_account
    with transaction.atomic():
        nomor = _next_piutang_journal_number('TRX-PIU-PV')
        header = JurnalHeader.objects.create(
            tanggal=tanggal,
            nomor_transaksi=nomor,
            uraian_transaksi=f'Akrual PV Piutang {nom} — s.d. {tanggal}',
            entitas_bisnis=piutang.entitas_bisnis,
            is_penyesuaian=True,
        )
        JurnalDetail.objects.bulk_create([
            JurnalDetail(
                jurnal_header=header,
                akun=ar_account,
                debit=bunga,
                kredit=Decimal('0'),
            ),
            JurnalDetail(
                jurnal_header=header,
                akun=income_account,
                debit=Decimal('0'),
                kredit=bunga,
            ),
        ])
        _log(piutang, 'EDITED', user=user,
             after={'pv_akrual': str(bunga), 'tanggal': str(tanggal)})
    return header


def create_pv_accrual_reversal(
    piutang: PiutangHeader,
    tanggal: date,
    catatan: str = '',
    user=None,
) -> JurnalHeader:
    """
    Reverse the last un-reversed period-end accrual journal.
    Swaps debit/credit of the original accrual entry.
    """
    nom = piutang.nomor_piutang
    last_accrual = (
        JurnalHeader.objects
        .filter(uraian_transaksi__startswith=f'Akrual PV Piutang {nom} —')
        .order_by('-tanggal', '-nomor_transaksi')
        .prefetch_related('details')
        .first()
    )
    if not last_accrual:
        raise ValueError('Tidak ada jurnal akrual PV yang dapat dibalik.')

    n_reversal = JurnalHeader.objects.filter(
        uraian_transaksi__startswith=f'Balik Akrual PV Piutang {nom} —'
    ).count()
    n_accrual = JurnalHeader.objects.filter(
        uraian_transaksi__startswith=f'Akrual PV Piutang {nom} —'
    ).count()
    if n_reversal >= n_accrual:
        raise ValueError('Semua jurnal akrual PV sudah dibalik.')

    with transaction.atomic():
        nomor = _next_piutang_journal_number('TRX-PIU-PV')
        reversal = JurnalHeader.objects.create(
            tanggal=tanggal,
            nomor_transaksi=nomor,
            uraian_transaksi=f'Balik Akrual PV Piutang {nom} — {last_accrual.nomor_transaksi}',
            entitas_bisnis=piutang.entitas_bisnis,
            is_penyesuaian=True,
        )
        JurnalDetail.objects.bulk_create([
            JurnalDetail(
                jurnal_header=reversal,
                akun=d.akun,
                debit=d.kredit,    # swap
                kredit=d.debit,
            )
            for d in last_accrual.details.all()
        ])
        _log(piutang, 'EDITED', user=user,
             after={'pv_balik_akrual': last_accrual.nomor_transaksi, 'tanggal': str(tanggal)})
    return reversal


# ════════════════════════════════════════════════════════════════════════════════
#  MODE STANDAR AKUNTANSI
# ════════════════════════════════════════════════════════════════════════════════

def get_standar_akuntansi(piutang: PiutangHeader) -> str:
    """
    Resolves effective standar akuntansi for a piutang:
      1. piutang.standar_akuntansi (override)
      2. piutang.entitas_bisnis.standar_akuntansi
      3. Default 'psak'
    Returns one of: 'psak', 'sak_ep', 'sak_emkm'.
    """
    if piutang.standar_akuntansi:
        return piutang.standar_akuntansi
    eb = piutang.entitas_bisnis
    if eb is not None:
        sa = getattr(eb, 'standar_akuntansi', None)
        if sa:
            return sa
    return 'psak'


# ════════════════════════════════════════════════════════════════════════════════
#  ECL STAGING — GENERAL APPROACH (PSAK 71)
# ════════════════════════════════════════════════════════════════════════════════

_STAGE_TRIGGER_SICR = 30    # hari lewat jatuh tempo → Stage 2 (SICR)
_STAGE_TRIGGER_IMPAIRED = 90  # hari lewat jatuh tempo → Stage 3


def assess_ecl_stage(piutang: PiutangHeader, as_of_date=None) -> int:
    """
    Auto-assess recommended ECL stage based on days past due.
    Returns 1, 2, or 3.
    Only meaningful for PSAK General Approach.

    Uses compute_effective_dpd so that installment loans measure DPD from the
    earliest missed installment, not the final balloon/maturity date.
    """
    if piutang.status in ('paid', 'cancelled', 'written_off'):
        return 1
    days = compute_effective_dpd(piutang)
    if days >= _STAGE_TRIGGER_IMPAIRED:
        return 3
    if days >= _STAGE_TRIGGER_SICR:
        return 2
    return 1


def update_ecl_stage(
    piutang: PiutangHeader,
    new_stage: int,
    alasan: str = '',
    is_auto: bool = True,
    user=None,
) -> PiutangECLStagingLog | None:
    """
    Update ECL stage on a piutang and write a staging log entry.
    Returns None if stage is already equal to new_stage.
    """
    if new_stage not in (1, 2, 3):
        raise ValueError('Stage ECL harus 1, 2, atau 3.')
    today = timezone.now().date()
    with transaction.atomic():
        piutang = PiutangHeader.objects.select_for_update().get(pk=piutang.pk)
        if piutang.stage_ecl == new_stage:
            return None
        dpd = compute_effective_dpd(piutang)
        log = PiutangECLStagingLog.objects.create(
            piutang_header=piutang,
            stage_dari=piutang.stage_ecl,
            stage_ke=new_stage,
            tanggal=today,
            days_past_due=dpd,
            alasan=alasan or f'Otomatis: {dpd} hari lewat jatuh tempo angsuran',
            is_auto=is_auto,
            created_by=user,
        )
        piutang.stage_ecl = new_stage
        piutang.stage_ecl_tanggal = today
        piutang.save(update_fields=['stage_ecl', 'stage_ecl_tanggal'])
        _log(piutang, 'ECL_STAGING', user=user, after={
            'stage_dari': log.stage_dari, 'stage_ke': new_stage, 'alasan': log.alasan,
        })
    return log


def batch_assess_ecl_stages(
    entitas_bisnis=None,
    as_of_date=None,
    dry_run: bool = False,
    user=None,
) -> list:
    """
    Bulk auto-assess ECL stages for all active PSAK piutang.
    Returns list of dicts describing changes made (or that would be made if dry_run=True).
    """
    today = as_of_date or timezone.now().date()
    qs = PiutangHeader.objects.filter(status__in=('open', 'partial', 'overdue'))
    if entitas_bisnis is not None:
        qs = qs.filter(entitas_bisnis=entitas_bisnis)

    results = []
    for piutang in qs.select_related('entitas_bisnis').prefetch_related('penerimaan'):
        if get_standar_akuntansi(piutang) != 'psak':
            continue
        dpd = compute_effective_dpd(piutang)
        recommended = assess_ecl_stage(piutang, today)
        if piutang.stage_ecl != recommended:
            results.append({
                'piutang': piutang,
                'stage_dari': piutang.stage_ecl,
                'stage_ke': recommended,
                'days_past_due': dpd,
            })
            if not dry_run:
                update_ecl_stage(
                    piutang, recommended,
                    alasan=f'Batch assess: {dpd} hari lewat jatuh tempo angsuran',
                    is_auto=True, user=user,
                )
    return results


# ════════════════════════════════════════════════════════════════════════════════
#  ECL GENERAL APPROACH — PD × LGD × EAD  (PSAK 71)
# ════════════════════════════════════════════════════════════════════════════════

def _get_ecl_net_carrying(piutang: PiutangHeader) -> Decimal:
    """Net carrying = gross carrying − accumulated ECL allowance. Used for Stage 3 EIR."""
    from .models import PiutangPenyisihan
    gross = _pv_carrying_value(piutang) if piutang.is_pv_adjusted else piutang.sisa_piutang
    total_psh = Decimal(str(
        PiutangPenyisihan.objects.filter(piutang_header=piutang)
        .aggregate(s=Sum('jumlah'))['s'] or Decimal('0')
    ))
    return max(Decimal('0'), gross - total_psh)


def compute_ecl_general_approach(
    piutang: PiutangHeader,
    pd_rate: Decimal,
    lgd_rate: Decimal,
    forward_looking_adj: Decimal = Decimal('1.0'),
    as_of_date=None,
) -> dict:
    """
    PSAK 71 General Approach ECL = PD × LGD × EAD × forward_looking_adj.

    Stage 1: 12-month PD.
    Stage 2/3: lifetime PD = 1 − (1 − pd_annual)^remaining_years.

    Returns dict: ecl_amount, ead, stage, ecl_horizon, pd_effective (%), ...
    """
    today = as_of_date or timezone.now().date()
    stage = piutang.stage_ecl or assess_ecl_stage(piutang, today)

    # EAD = carrying amount (Stage 3: net of allowance)
    if stage == 3:
        ead = _get_ecl_net_carrying(piutang)
    elif piutang.is_pv_adjusted:
        ead = _pv_carrying_value(piutang)
    else:
        ead = piutang.sisa_piutang

    if stage == 1:
        ecl_horizon = '12-bulan'
        pd_effective = pd_rate / 100
    else:
        ecl_horizon = 'seumur-hidup'
        if piutang.jatuh_tempo and piutang.jatuh_tempo > today:
            remaining_years = Decimal(str((piutang.jatuh_tempo - today).days / 365.25))
        else:
            remaining_years = Decimal('1')
        pd_fl = float(pd_rate) / 100
        pd_effective = Decimal(str(round(1 - (1 - pd_fl) ** float(remaining_years), 8)))

    ecl_amount = (ead * pd_effective * (lgd_rate / 100) * forward_looking_adj).quantize(Decimal('0.01'))

    return {
        'stage': stage,
        'ecl_horizon': ecl_horizon,
        'ead': ead,
        'pd_rate_input': pd_rate,
        'lgd_rate': lgd_rate,
        'pd_effective_pct': (pd_effective * 100).quantize(Decimal('0.0001')),
        'forward_looking_adj': forward_looking_adj,
        'ecl_amount': ecl_amount,
    }


def create_penyisihan_ecl_general(
    piutang: PiutangHeader,
    pd_rate: Decimal,
    lgd_rate: Decimal,
    allowance_account,
    expense_account,
    tanggal,
    forward_looking_adj: Decimal = Decimal('1.0'),
    catatan: str = '',
    user=None,
):
    """
    Create ECL penyisihan journal using General Approach (PSAK 71 only).
    Uses PD × LGD × EAD instead of flat aging-bucket rate.
    """
    from .models import PiutangPenyisihan
    standar = get_standar_akuntansi(piutang)
    if standar != 'psak':
        raise ValueError(
            f'General Approach ECL hanya tersedia untuk mode PSAK. '
            f'Piutang ini menggunakan {standar.upper()}. '
            f'Gunakan create_penyisihan_journal() untuk Simplified Approach.'
        )
    result = compute_ecl_general_approach(piutang, pd_rate, lgd_rate, forward_looking_adj)
    total = result['ecl_amount']
    if total <= Decimal('0'):
        raise ValueError('ECL General Approach bernilai 0. Periksa parameter PD dan LGD.')

    keterangan = (
        f'ECL General Approach {piutang.nomor_piutang} — '
        f'Stage {result["stage"]}, PD {pd_rate}%, LGD {lgd_rate}%, '
        f'FwdAdj {forward_looking_adj}'
    )
    with transaction.atomic():
        piutang = PiutangHeader.objects.select_for_update().get(pk=piutang.pk)
        nomor = _next_penyisihan_journal_number('TRX-PIU-ECL')
        header = JurnalHeader.objects.create(
            tanggal=tanggal,
            nomor_transaksi=nomor,
            uraian_transaksi=keterangan,
            entitas_bisnis=piutang.entitas_bisnis,
            is_penyesuaian=True,
        )
        JurnalDetail.objects.bulk_create([
            JurnalDetail(jurnal_header=header, akun=expense_account, debit=total, kredit=Decimal('0')),
            JurnalDetail(jurnal_header=header, akun=allowance_account, debit=Decimal('0'), kredit=total),
        ])
        entry = PiutangPenyisihan.objects.create(
            piutang_header=piutang,
            tanggal=tanggal,
            jenis='manual',
            jumlah=total,
            allowance_account=allowance_account,
            expense_account=expense_account,
            jurnal_header=header,
            catatan=catatan or keterangan,
            created_by=user,
        )
        piutang.is_specifically_impaired = True
        piutang.save(update_fields=['is_specifically_impaired'])
        _log(piutang, 'PENYISIHAN', user=user, after={
            'metode': 'general_approach', 'ecl': str(total),
            'stage': result['stage'], 'nomor': nomor,
        })
    return entry


# ════════════════════════════════════════════════════════════════════════════════
#  MODIFIKASI PIUTANG  (PSAK 71 / SAK EP)
# ════════════════════════════════════════════════════════════════════════════════

def compute_modification_pv(
    piutang: PiutangHeader,
    new_cashflows: list,
) -> Decimal:
    """
    PV of new contractual cash flows discounted at the original EIR.
    new_cashflows: [{'tanggal': date, 'jumlah': Decimal}, ...]
    """
    if not piutang.pv_discount_rate:
        raise ValueError('EIR asli (pv_discount_rate) diperlukan untuk menghitung PV modifikasi.')
    today = timezone.now().date()
    i_daily = (1 + float(piutang.pv_discount_rate) / 100) ** (1 / 365.25) - 1
    pv = Decimal('0')
    for cf in new_cashflows:
        days = (cf['tanggal'] - today).days
        if days <= 0:
            pv += Decimal(str(cf['jumlah']))
        else:
            pv += Decimal(str(cf['jumlah'])) / Decimal(str((1 + i_daily) ** days))
    return pv.quantize(Decimal('0.0001'))


def process_piutang_modification(
    piutang: PiutangHeader,
    new_cashflows: list,
    gain_loss_account,
    tanggal,
    eir_baru: Decimal = None,
    deskripsi: str = '',
    user=None,
) -> PiutangModifikasi:
    """
    PSAK 71 / SAK EP: record a loan modification.
    Calculates modification gain/loss = PV(new CF at original EIR) − carrying amount.
    Journals the difference to gain_loss_account.

    new_cashflows: [{'tanggal': date, 'jumlah': Decimal}, ...]
    eir_baru: if provided, updates pv_discount_rate for future amortization.
    """
    standar = get_standar_akuntansi(piutang)
    if standar == 'sak_emkm':
        raise ValueError('Modification accounting tidak diterapkan untuk SAK EMKM.')
    if piutang.status in ('paid', 'cancelled', 'written_off'):
        raise ValueError('Piutang sudah tidak aktif. Tidak dapat dimodifikasi.')

    carrying = _pv_carrying_value(piutang) if piutang.is_pv_adjusted else piutang.sisa_piutang
    pv_baru = compute_modification_pv(piutang, new_cashflows)
    gl = (pv_baru - carrying).quantize(Decimal('0.0001'))

    with transaction.atomic():
        piutang = PiutangHeader.objects.select_for_update().get(pk=piutang.pk)
        jurnal = None
        if abs(gl) >= Decimal('0.01'):
            nomor = _next_piutang_journal_number('TRX-PIU-MOD')
            jurnal = JurnalHeader.objects.create(
                tanggal=tanggal,
                nomor_transaksi=nomor,
                uraian_transaksi=f'Modifikasi Piutang {piutang.nomor_piutang} — GL {gl:+.0f}',
                entitas_bisnis=piutang.entitas_bisnis,
                is_penyesuaian=True,
            )
            ar_account = piutang.coa_piutang_account
            if gl > Decimal('0'):
                JurnalDetail.objects.bulk_create([
                    JurnalDetail(jurnal_header=jurnal, akun=ar_account, debit=gl, kredit=Decimal('0')),
                    JurnalDetail(jurnal_header=jurnal, akun=gain_loss_account, debit=Decimal('0'), kredit=gl),
                ])
            else:
                abs_gl = abs(gl)
                JurnalDetail.objects.bulk_create([
                    JurnalDetail(jurnal_header=jurnal, akun=gain_loss_account, debit=abs_gl, kredit=Decimal('0')),
                    JurnalDetail(jurnal_header=jurnal, akun=ar_account, debit=Decimal('0'), kredit=abs_gl),
                ])

        update_fields = ['nilai_wajar_awal', 'is_pv_adjusted']
        piutang.nilai_wajar_awal = pv_baru
        piutang.is_pv_adjusted = True
        if eir_baru is not None:
            piutang.pv_discount_rate = eir_baru
            update_fields.append('pv_discount_rate')
        piutang.save(update_fields=update_fields)

        mod = PiutangModifikasi.objects.create(
            piutang_header=piutang,
            tanggal=tanggal,
            carrying_amount_lama=carrying,
            pv_syarat_baru=pv_baru,
            modification_gain_loss=gl,
            eir_baru=eir_baru,
            deskripsi_perubahan=deskripsi,
            gain_loss_account=gain_loss_account,
            jurnal=jurnal,
            created_by=user,
        )
        _log(piutang, 'MODIFIKASI', user=user, after={
            'carrying_lama': str(carrying), 'pv_baru': str(pv_baru), 'gl': str(gl),
        })
    return mod


# ════════════════════════════════════════════════════════════════════════════════
#  PEMULIHAN WRITE-OFF  (semua standar)
# ════════════════════════════════════════════════════════════════════════════════

def recover_written_off_piutang(
    piutang: PiutangHeader,
    jumlah_dipulihkan: Decimal,
    kas_account,
    recovery_income_account,
    tanggal,
    catatan: str = '',
    user=None,
) -> PiutangPemulihanWriteOff:
    """
    Catat pemulihan piutang yang sebelumnya sudah di-write-off.
    Jurnal: Dr Kas/Bank / Cr Pendapatan Pemulihan Piutang.
    Berlaku untuk semua standar (PSAK, SAK EP, SAK EMKM).
    """
    if piutang.status != 'written_off':
        raise ValueError('Hanya piutang berstatus written_off yang dapat dipulihkan.')
    if jumlah_dipulihkan <= Decimal('0'):
        raise ValueError('Jumlah pemulihan harus lebih dari nol.')
    write_off = getattr(piutang, 'write_off', None)
    if write_off and jumlah_dipulihkan > write_off.jumlah_dihapus:
        raise ValueError(
            f'Jumlah pemulihan ({jumlah_dipulihkan:,.0f}) melebihi '
            f'jumlah yang pernah dihapusbukukan ({write_off.jumlah_dihapus:,.0f}).'
        )
    with transaction.atomic():
        piutang = PiutangHeader.objects.select_for_update().get(pk=piutang.pk)
        nomor = _next_piutang_journal_number('TRX-PIU-REC')
        header = JurnalHeader.objects.create(
            tanggal=tanggal,
            nomor_transaksi=nomor,
            uraian_transaksi=f'Pemulihan Write-Off Piutang {piutang.nomor_piutang}',
            entitas_bisnis=piutang.entitas_bisnis,
            is_penyesuaian=False,
        )
        JurnalDetail.objects.bulk_create([
            JurnalDetail(jurnal_header=header, akun=kas_account,
                         debit=jumlah_dipulihkan, kredit=Decimal('0')),
            JurnalDetail(jurnal_header=header, akun=recovery_income_account,
                         debit=Decimal('0'), kredit=jumlah_dipulihkan),
        ])
        entry = PiutangPemulihanWriteOff.objects.create(
            piutang_header=piutang,
            tanggal=tanggal,
            jumlah_dipulihkan=jumlah_dipulihkan,
            kas_account=kas_account,
            recovery_income_account=recovery_income_account,
            catatan=catatan,
            jurnal=header,
            created_by=user,
        )
        _log(piutang, 'RECOVERY', user=user, after={
            'jumlah': str(jumlah_dipulihkan), 'nomor': nomor,
        })
    return entry


# ════════════════════════════════════════════════════════════════════════════════
#  FACTORING / DERECOGNITION  (PSAK 71 / SAK EP)
# ════════════════════════════════════════════════════════════════════════════════

def create_factoring_derecognition(
    piutang: PiutangHeader,
    nilai_transfer: Decimal,
    hasil_analisis: str,
    pihak_penerima: str,
    kas_account,
    gain_loss_account,
    tanggal,
    continuing_involvement_amount: Decimal = None,
    analisis_detail: str = '',
    user=None,
) -> PiutangFactoring:
    """
    PSAK 71 / SAK EP: catat transfer/anjak piutang.

    hasil_analisis:
      'derecognized'     → semua risiko & manfaat ditransfer; piutang diderecognize, jurnal dibuat.
      'continuing'       → sebagian risiko ditahan; continuing involvement dicatat, piutang tetap.
      'not_derecognized' → risiko ditahan penuh; hanya dokumentasi, tidak ada derecognition.

    Jika 'derecognized': Dr Kas / Cr Piutang / Dr|Cr Gain/Loss Derecognition.
    """
    standar = get_standar_akuntansi(piutang)
    if standar == 'sak_emkm':
        raise ValueError(
            'Analisis derecognition factoring tidak diperlukan untuk SAK EMKM. '
            'Cukup catat penerimaan kas secara langsung.'
        )
    if hasil_analisis not in ('derecognized', 'continuing', 'not_derecognized'):
        raise ValueError(
            "hasil_analisis harus salah satu dari: 'derecognized', 'continuing', 'not_derecognized'."
        )
    if piutang.status in ('paid', 'cancelled'):
        raise ValueError('Piutang sudah lunas atau dibatalkan.')

    carrying = _pv_carrying_value(piutang) if piutang.is_pv_adjusted else piutang.sisa_piutang
    gl = (nilai_transfer - carrying).quantize(Decimal('0.0001'))

    jurnal = None
    with transaction.atomic():
        piutang = PiutangHeader.objects.select_for_update().get(pk=piutang.pk)

        if hasil_analisis == 'derecognized':
            nomor = _next_piutang_journal_number('TRX-PIU-FAC')
            jurnal = JurnalHeader.objects.create(
                tanggal=tanggal,
                nomor_transaksi=nomor,
                uraian_transaksi=f'Derecognition Factoring {piutang.nomor_piutang}',
                entitas_bisnis=piutang.entitas_bisnis,
                is_penyesuaian=False,
            )
            ar_account = piutang.coa_piutang_account
            lines = [
                JurnalDetail(jurnal_header=jurnal, akun=kas_account,
                             debit=nilai_transfer, kredit=Decimal('0')),
                JurnalDetail(jurnal_header=jurnal, akun=ar_account,
                             debit=Decimal('0'), kredit=carrying),
            ]
            if gl > Decimal('0.005'):
                lines.append(JurnalDetail(
                    jurnal_header=jurnal, akun=gain_loss_account,
                    debit=Decimal('0'), kredit=gl,
                ))
            elif gl < Decimal('-0.005'):
                lines.append(JurnalDetail(
                    jurnal_header=jurnal, akun=gain_loss_account,
                    debit=abs(gl), kredit=Decimal('0'),
                ))
            JurnalDetail.objects.bulk_create(lines)
            piutang.status = 'cancelled'
            piutang.save(update_fields=['status'])

        entry = PiutangFactoring.objects.create(
            piutang_header=piutang,
            tanggal=tanggal,
            nilai_transfer=nilai_transfer,
            hasil_analisis=hasil_analisis,
            continuing_involvement_amount=continuing_involvement_amount,
            gain_loss_derecognition=gl if hasil_analisis == 'derecognized' else None,
            pihak_penerima=pihak_penerima,
            analisis_detail=analisis_detail,
            jurnal=jurnal,
            created_by=user,
        )
        _log(piutang, 'FACTORING', user=user, after={
            'hasil_analisis': hasil_analisis,
            'nilai_transfer': str(nilai_transfer),
            'gl': str(gl),
        })
    return entry


# ════════════════════════════════════════════════════════════════════════════════
#  ECL ROLL-FORWARD TABLE  (Disclosure PSAK 71)
# ════════════════════════════════════════════════════════════════════════════════

def get_ecl_rollforward_table(
    entitas_bisnis=None,
    periode_awal: date = None,
    periode_akhir: date = None,
) -> dict:
    """
    PSAK 71 Disclosure: ECL roll-forward table.

    Returns:
      gross_carrying_by_stage  — carrying amount outstanding per stage per periode_akhir
      ecl_balance_by_stage     — accumulated penyisihan per stage per periode_akhir
      ecl_movement             — additions, reversals, net during the period
      stage_movements_count    — jumlah perubahan stage selama periode
      staging_logs             — list detail perubahan stage
    """
    from .models import PiutangPenyisihan
    today = date.today()
    if periode_akhir is None:
        periode_akhir = today
    if periode_awal is None:
        periode_awal = periode_akhir.replace(day=1)

    qs = PiutangHeader.objects.filter(
        status__in=('open', 'partial', 'overdue', 'written_off'),
    ).select_related('entitas_bisnis').prefetch_related('penyisihan_entries')
    if entitas_bisnis is not None:
        qs = qs.filter(entitas_bisnis=entitas_bisnis)

    stage_carrying: dict[int | None, Decimal] = {1: Decimal('0'), 2: Decimal('0'), 3: Decimal('0'), None: Decimal('0')}
    stage_ecl: dict[int | None, Decimal] = {1: Decimal('0'), 2: Decimal('0'), 3: Decimal('0'), None: Decimal('0')}

    for piutang in qs:
        carrying = _pv_carrying_value(piutang) if piutang.is_pv_adjusted else piutang.sisa_piutang
        s = piutang.stage_ecl
        stage_carrying[s] = stage_carrying.get(s, Decimal('0')) + carrying
        psh = Decimal(str(
            PiutangPenyisihan.objects.filter(
                piutang_header=piutang, tanggal__lte=periode_akhir,
            ).aggregate(v=Sum('jumlah'))['v'] or Decimal('0')
        ))
        stage_ecl[s] = stage_ecl.get(s, Decimal('0')) + psh

    psh_qs = PiutangPenyisihan.objects.filter(
        tanggal__gte=periode_awal, tanggal__lte=periode_akhir,
    )
    if entitas_bisnis is not None:
        psh_qs = psh_qs.filter(entitas_bisnis=entitas_bisnis)
    ecl_additions = Decimal(str(
        psh_qs.filter(jumlah__gt=0).aggregate(v=Sum('jumlah'))['v'] or Decimal('0')
    ))
    ecl_reversals = abs(Decimal(str(
        psh_qs.filter(jumlah__lt=0).aggregate(v=Sum('jumlah'))['v'] or Decimal('0')
    )))

    staging_qs = PiutangECLStagingLog.objects.filter(
        tanggal__gte=periode_awal, tanggal__lte=periode_akhir,
    )
    if entitas_bisnis is not None:
        staging_qs = staging_qs.filter(piutang_header__entitas_bisnis=entitas_bisnis)

    return {
        'periode_awal': periode_awal,
        'periode_akhir': periode_akhir,
        'gross_carrying_by_stage': {
            'stage_1': stage_carrying[1],
            'stage_2': stage_carrying[2],
            'stage_3': stage_carrying[3],
            'unstaged': stage_carrying[None],
            'total': sum(stage_carrying.values()),
        },
        'ecl_balance_by_stage': {
            'stage_1': stage_ecl[1],
            'stage_2': stage_ecl[2],
            'stage_3': stage_ecl[3],
            'unstaged': stage_ecl[None],
            'total': sum(stage_ecl.values()),
        },
        'ecl_movement': {
            'additions': ecl_additions,
            'reversals': ecl_reversals,
            'net': ecl_additions - ecl_reversals,
        },
        'stage_movements_count': staging_qs.count(),
        'staging_logs': list(staging_qs.values(
            'piutang_header__nomor_piutang', 'stage_dari', 'stage_ke',
            'tanggal', 'alasan', 'is_auto',
        )),
    }
