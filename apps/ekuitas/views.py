"""Ekuitas views — Modal Disetor CRUD + journal integration."""
import json
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

from apps.entitas_bisnis.models import EntitasBisnis
from apps.purchase.views import _get_eb_tree, _resolve_eb_lv1_ids

from .models import ModalDisetor, Pemilik
from .services import (
    create_modal_disetor_batch,
    delete_modal_disetor_group,
    get_group_siblings,
)


@login_required
def ekuitas_list(request: HttpRequest) -> HttpResponse:
    qs = list(ModalDisetor.objects.select_related('entitas_bisnis', 'pemilik', 'jurnal_header').all())
    eb_filter_list = [v for v in request.GET.getlist('entitas_bisnis') if v]
    if eb_filter_list:
        lv1_ids = _resolve_eb_lv1_ids(eb_filter_list)
        qs = [r for r in qs if r.entitas_bisnis_id in lv1_ids]
    if eb_filter_list:
        total_all = sum(r.jumlah_modal for r in qs) or Decimal('1')
        for r in qs:
            r.persentase = (r.jumlah_modal / total_all * 100).quantize(Decimal('0.01'))
    else:
        for r in qs:
            r.persentase = None
    return render(request, 'ekuitas/ekuitas_list.html', {
        'records': qs,
        'eb_tree': _get_eb_tree(),
        'eb_filter_list': eb_filter_list,
    })


@login_required
def ekuitas_history(request: HttpRequest) -> HttpResponse:
    eb_filter = request.GET.get('entitas_bisnis', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    records: list = []
    total_modal = Decimal('0')
    selected_eb = None
    if eb_filter:
        qs = ModalDisetor.objects.select_related('entitas_bisnis', 'pemilik', 'jurnal_header').filter(
            entitas_bisnis_id=eb_filter,
        )
        if tanggal_sampai:
            qs = qs.filter(tanggal_setor__lte=tanggal_sampai)
        records = list(qs.order_by('pemilik__nama', 'tanggal_setor'))
        total_modal = sum(r.jumlah_modal for r in records) or Decimal('1')
        for r in records:
            r.persentase = (r.jumlah_modal / total_modal * 100).quantize(Decimal('0.01'))
        try:
            selected_eb = EntitasBisnis.objects.get(pk=eb_filter)
        except EntitasBisnis.DoesNotExist:
            pass
    return render(request, 'ekuitas/ekuitas_history.html', {
        'records': records,
        'total_modal': total_modal,
        'entitas_list': EntitasBisnis.objects.filter(status_aktif=True).order_by('nama'),
        'eb_filter': eb_filter,
        'tanggal_sampai': tanggal_sampai,
        'selected_eb': selected_eb,
    })


@login_required
def ekuitas_detail(request: HttpRequest, pk: int) -> HttpResponse:
    record = get_object_or_404(
        ModalDisetor.objects.select_related('entitas_bisnis', 'pemilik', 'jurnal_header')
                            .prefetch_related('debit_lines__akun'),
        pk=pk,
    )
    total_all = ModalDisetor.objects.filter(
        entitas_bisnis=record.entitas_bisnis,
    ).aggregate(t=Sum('jumlah_modal'))['t'] or Decimal('1')
    persentase = (record.jumlah_modal / total_all * 100).quantize(Decimal('0.01'))
    siblings = get_group_siblings(record)
    return render(request, 'ekuitas/ekuitas_detail.html', {
        'record': record,
        'persentase': persentase,
        'siblings': siblings,
    })


@login_required
def ekuitas_create(request: HttpRequest) -> HttpResponse:
    from apps.purchase.models import FIFOBatch, ItemMasterPurchase
    from apps.inventory.models import InventoryRecord
    from apps.aset_tetap.models import AsetTetapRecord
    from apps.aset_lainnya.models import AsetLainnyaRecord

    AKUN_DETAIL_PREFIXES: dict[str, str] = {
        '1.1.7': 'persediaan',
        '1.1.8': 'persediaan',
        '1.1.9': 'persediaan',
        '1.1.10': 'persediaan',
        '1.2': 'aset_tetap',
        '1.2.3': 'aset_tetap',
        '1.3': 'aset_lainnya',
        '3.1.1': 'modal_disetor',
    }

    eb_list = EntitasBisnis.objects.filter(status_aktif=True).order_by('nama')
    pemilik_list = Pemilik.objects.order_by('nama')
    akun_autocomplete_url = reverse('jurnal:akun_autocomplete')

    if request.method == 'POST':
        eb_id = request.POST.get('entitas_bisnis', '').strip()
        tanggal = request.POST.get('tanggal', '').strip()
        owners_json = request.POST.get('owners_json', '[]')
        debit_json = request.POST.get('debit_lines_json', '[]')

        errors: dict = {}
        if not eb_id:
            errors['entitas_bisnis'] = 'Entitas Bisnis wajib dipilih.'
        if not tanggal:
            errors['tanggal'] = 'Tanggal wajib diisi.'

        try:
            owners = json.loads(owners_json)
        except (ValueError, TypeError):
            owners = []
        if not owners:
            errors['owners'] = 'Minimal 1 pemilik wajib diisi.'

        try:
            debit_lines = json.loads(debit_json)
        except (ValueError, TypeError):
            debit_lines = []
        if not debit_lines:
            errors['debit_lines'] = 'Minimal 1 baris debit akun wajib diisi.'

        if not errors:
            try:
                records = create_modal_disetor_batch(
                    entitas_bisnis_id=int(eb_id),
                    tanggal=tanggal,
                    owners=owners,
                    debit_lines=debit_lines,
                )
                # Process detail rows on debit lines (persediaan, aset_tetap, aset_lainnya)
                eb_id_int = int(eb_id)
                all_item_ids = {
                    int(d['item_id'])
                    for dl in debit_lines
                    for d in dl.get('detail_rows', [])
                    if dl.get('detail_type') in ('persediaan', 'aset_tetap', 'aset_lainnya')
                    and str(d.get('item_id', '')).isdigit()
                }
                if all_item_ids:
                    items_map = {
                        item.pk: item
                        for item in ItemMasterPurchase.objects.filter(pk__in=all_item_ids)
                    }
                    for dl in debit_lines:
                        detail_type = dl.get('detail_type')
                        if not detail_type or detail_type not in ('persediaan', 'aset_tetap', 'aset_lainnya'):
                            continue
                        for d in dl.get('detail_rows', []):
                            try:
                                item_pk = int(str(d.get('item_id', '')))
                                qty = Decimal(str(d.get('qty') or 0))
                                unit_price = Decimal(str(d.get('unit_price') or 0))
                            except (ValueError, TypeError):
                                continue
                            if qty <= 0 or unit_price < 0:
                                continue
                            item = items_map.get(item_pk)
                            if not item:
                                continue
                            if detail_type == 'persediaan':
                                InventoryRecord.objects.create(
                                    item=item, purchase_item=None,
                                    entitas_bisnis_id=eb_id_int,
                                    quantity=qty, unit_price=unit_price, tanggal=tanggal,
                                )
                                FIFOBatch.objects.create(
                                    purchase_item=None, item=item, tanggal=tanggal,
                                    quantity_in=qty, unit_price=unit_price, remaining_qty=qty,
                                )
                            elif detail_type == 'aset_tetap':
                                AsetTetapRecord.objects.create(
                                    item=item, purchase_item=None,
                                    entitas_bisnis_id=eb_id_int,
                                    quantity=qty, harga_perolehan=unit_price,
                                    tanggal_perolehan=tanggal,
                                    masa_manfaat=item.masa_manfaat,
                                    metode_penyusutan=item.metode_penyusutan,
                                )
                            elif detail_type == 'aset_lainnya':
                                AsetLainnyaRecord.objects.create(
                                    item=item, purchase_item=None,
                                    entitas_bisnis_id=eb_id_int,
                                    quantity=qty, harga_perolehan=unit_price,
                                    tanggal_perolehan=tanggal,
                                    masa_manfaat=item.masa_manfaat,
                                    metode_amortisasi=item.metode_amortisasi,
                                )
                messages.success(request, f'{len(records)} data modal disetor berhasil disimpan dan jurnal dibuat.')
                return redirect('ekuitas:list')
            except ValueError as e:
                errors['form'] = str(e)

        return render(request, 'ekuitas/ekuitas_create.html', {
            'eb_list': eb_list,
            'pemilik_list': pemilik_list,
            'errors': errors,
            'posted_eb': eb_id,
            'posted_tanggal': tanggal,
            'akun_autocomplete_url': akun_autocomplete_url,
            'akun_detail_prefixes_json': json.dumps(AKUN_DETAIL_PREFIXES),
        })

    return render(request, 'ekuitas/ekuitas_create.html', {
        'eb_list': eb_list,
        'pemilik_list': pemilik_list,
        'errors': {},
        'posted_tanggal': timezone.now().date().isoformat(),
        'akun_autocomplete_url': akun_autocomplete_url,
        'akun_detail_prefixes_json': json.dumps(AKUN_DETAIL_PREFIXES),
    })


@login_required
def ekuitas_delete(request: HttpRequest, pk: int) -> HttpResponse:
    record = get_object_or_404(
        ModalDisetor.objects.select_related('pemilik', 'entitas_bisnis', 'jurnal_header'),
        pk=pk,
    )
    siblings = get_group_siblings(record)

    if request.method == 'POST':
        jurnal_nomor = record.jurnal_header.nomor_transaksi if record.jurnal_header else None
        try:
            delete_modal_disetor_group(record, request=request)
            if jurnal_nomor:
                messages.success(request, f'Jurnal {jurnal_nomor} dan {len(siblings)} data modal disetor berhasil dihapus.')
            else:
                messages.success(request, 'Modal disetor berhasil dihapus.')
        except Exception as e:
            messages.error(request, f'Gagal menghapus: {e}')
        return redirect('ekuitas:list')

    return render(request, 'ekuitas/ekuitas_delete.html', {
        'record': record,
        'siblings': siblings,
    })


@login_required
def api_pemilik_search(request: HttpRequest) -> JsonResponse:
    term = request.GET.get('term', '').strip()
    qs = Pemilik.objects.all()
    if term:
        qs = qs.filter(nama__icontains=term)
    qs = qs.order_by('nama')[:30]
    return JsonResponse([{'id': p.pk, 'text': p.nama} for p in qs], safe=False)


@login_required
@require_POST
def api_pemilik_create(request: HttpRequest) -> JsonResponse:
    nama = request.POST.get('nama', '').strip()
    if not nama:
        return JsonResponse({'error': 'Nama pemilik wajib diisi.'}, status=400)
    if Pemilik.objects.filter(nama__iexact=nama).exists():
        p = Pemilik.objects.get(nama__iexact=nama)
        return JsonResponse({'id': p.pk, 'text': p.nama})
    p = Pemilik.objects.create(nama=nama)
    return JsonResponse({'id': p.pk, 'text': p.nama})


# ── Export views ─────────────────────────────────────────────────────────────

def _ekuitas_export_qs(request):
    """Return filtered ModalDisetor queryset based on GET params."""
    qs = ModalDisetor.objects.select_related(
        'entitas_bisnis', 'pemilik', 'jurnal_header',
    ).order_by('-tanggal_setor', '-created_at')
    eb_filter = request.GET.get('entitas_bisnis', '')
    tanggal_dari = request.GET.get('tanggal_dari', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    if eb_filter:
        qs = qs.filter(entitas_bisnis_id=eb_filter)
    if tanggal_dari:
        qs = qs.filter(tanggal_setor__gte=tanggal_dari)
    if tanggal_sampai:
        qs = qs.filter(tanggal_setor__lte=tanggal_sampai)
    return qs


@login_required
def ekuitas_export(request: HttpRequest) -> HttpResponse:
    """Export modal disetor list as XLSX with same filters as list page."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    records = list(_ekuitas_export_qs(request))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Modal Disetor'

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin = Side(style='thin')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right_align = Alignment(horizontal='right')

    headers = [
        'Pemilik', 'Entitas Bisnis', 'Jumlah Modal (Rp)',
        'Tanggal Setor', 'No. Jurnal', 'Keterangan',
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')

    for row_num, r in enumerate(records, 2):
        vals = [
            r.pemilik.nama,
            r.entitas_bisnis.nama,
            float(r.jumlah_modal or 0),
            str(r.tanggal_setor),
            r.jurnal_header.nomor_transaksi if r.jurnal_header else '',
            r.keterangan or '',
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.border = thin_border
            if col == 3:
                c.alignment = right_align
                c.number_format = '#,##0'

    col_widths = [28, 28, 22, 14, 20, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="ekuitas.xlsx"'
    wb.save(response)
    return response


@login_required
def ekuitas_export_pdf(request: HttpRequest) -> HttpResponse:
    """Render print-friendly modal disetor list for browser PDF printing."""
    import datetime
    # Fetch ascending (oldest first) for cumulative computation
    records_qs = _ekuitas_export_qs(request).order_by('entitas_bisnis_id', 'tanggal_setor', 'pk')
    records = list(records_qs)

    # Compute per-EB cumulative totals and ownership %
    eb_cumulative: dict[int, Decimal] = {}
    eb_pemilik_cumulative: dict[tuple, Decimal] = {}
    eb_pemilik_prev_pct: dict[tuple, Decimal] = {}

    for r in records:
        eb_id = r.entitas_bisnis_id
        pemilik_id = r.pemilik_id

        eb_cumulative[eb_id] = eb_cumulative.get(eb_id, Decimal('0')) + r.jumlah_modal

        key = (eb_id, pemilik_id)
        eb_pemilik_cumulative[key] = eb_pemilik_cumulative.get(key, Decimal('0')) + r.jumlah_modal

        total_all = eb_cumulative[eb_id]
        pemilik_total = eb_pemilik_cumulative[key]
        pct = (pemilik_total / total_all * 100) if total_all > 0 else Decimal('0')

        prev_pct = eb_pemilik_prev_pct.get(key, Decimal('0'))
        delta_pct = pct - prev_pct
        eb_pemilik_prev_pct[key] = pct

        r._pct = round(pct, 2)
        r._delta_pct = round(delta_pct, 2)

    total_modal = sum(r.jumlah_modal for r in records)
    return render(request, 'ekuitas/ekuitas_export_pdf.html', {
        'records': records,
        'tanggal_dari': request.GET.get('tanggal_dari', ''),
        'tanggal_sampai': request.GET.get('tanggal_sampai', ''),
        'generated_at': datetime.datetime.now(),
        'total_modal': total_modal,
        'total_records': len(records),
    })
