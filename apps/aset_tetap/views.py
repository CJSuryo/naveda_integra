"""Aset Tetap views — CRUD for AsetTetapRecord."""
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django_ratelimit.decorators import ratelimit

from naveda_integra.ratelimit_utils import rate_from
from django.db import models
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render

from apps.entitas_bisnis.models import EntitasBisnis
from apps.master_data.models import Akun
from apps.purchase.models import ItemMasterPurchase, KategoriItem
from apps.purchase.views import _get_eb_tree, _resolve_eb_lv1_ids

from .forms import (
    AsetTetapRecordForm, AssetDisposalForm,
    AssetMaintenanceForm, AssetTransferForm, AssetRevaluationForm,
)
from .models import (
    AsetTetapRecord, AssetDisposal, AssetMaintenance, AssetTransfer, AssetRevaluation, LokasiAset,
)
from . import services, reports
from .services import (
    calculate_depreciation, process_depreciation,
    process_asset_disposal, reverse_asset_disposal,
)
from .documents import list_bukti_aset

DEFAULT_DAYS = 30  # monthly processing default


@login_required
def aset_tetap_list(request: HttpRequest) -> HttpResponse:
    """List all fixed asset records with optional filters."""
    qs = AsetTetapRecord.objects.select_related(
        'item', 'item__kategori', 'item__coa_account', 'entitas_bisnis',
    ).all()

    tanggal_dari = request.GET.get('tanggal_dari', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    item_filter = request.GET.get('item', '')
    eb_filter_list = [v for v in request.GET.getlist('entitas_bisnis') if v]
    kondisi_filter = request.GET.get('kondisi', '')
    kategori_filter = request.GET.get('kategori', '')
    akun_filter = request.GET.get('coa_account', '')

    if tanggal_dari:
        qs = qs.filter(tanggal_perolehan__gte=tanggal_dari)
    if tanggal_sampai:
        qs = qs.filter(tanggal_perolehan__lte=tanggal_sampai)
    if item_filter:
        qs = qs.filter(item_id=item_filter)
    if eb_filter_list:
        qs = qs.filter(entitas_bisnis_id__in=_resolve_eb_lv1_ids(eb_filter_list, request.user))
    if kondisi_filter:
        qs = qs.filter(kondisi=kondisi_filter)
    if kategori_filter:
        qs = qs.filter(item__kategori_id=kategori_filter)
    if akun_filter:
        qs = qs.filter(item__coa_account_id=akun_filter)

    # Build dropdown lists scoped to ATP items
    kategori_list = KategoriItem.objects.filter(
        items__tipe_item='ATP'
    ).distinct().order_by('nama')
    akun_atp_list = Akun.objects.filter(
        item_masters__tipe_item='ATP'
    ).distinct().order_by('kode_akun')

    return render(request, 'aset_tetap/aset_tetap_list.html', {
        'records': qs,
        'items': ItemMasterPurchase.objects.filter(tipe_item='ATP').order_by('item_id'),
        'eb_tree': _get_eb_tree(request.user),
        'entitas_list': EntitasBisnis.objects.filter(status_aktif=True).order_by('nama'),
        'kondisi_choices': AsetTetapRecord.KONDISI_CHOICES,
        'metode_choices': AsetTetapRecord.METODE_PENYUSUTAN_CHOICES,
        'kategori_list': kategori_list,
        'akun_atp_list': akun_atp_list,
        'tanggal_dari': tanggal_dari,
        'tanggal_sampai': tanggal_sampai,
        'item_filter': item_filter,
        'eb_filter_list': eb_filter_list,
        'kondisi_filter': kondisi_filter,
        'kategori_filter': kategori_filter,
        'akun_filter': akun_filter,
    })


@login_required
def aset_tetap_detail(request: HttpRequest, pk: int) -> HttpResponse:
    """Show fixed asset record detail with depreciation preview."""
    from decimal import Decimal
    record = get_object_or_404(
        AsetTetapRecord.objects.select_related('item', 'entitas_bisnis', 'purchase_item'),
        pk=pk,
    )

    # Calculate depreciation preview
    metode = record.metode_penyusutan or (record.item.metode_penyusutan if record.item else '')
    masa = record.masa_manfaat or (record.item.masa_manfaat if record.item else 0) or 0
    needs_activity_input = metode in ('service_hours', 'units_of_production')

    # For time-based methods, calculate current year number
    if masa > 0 and record.tanggal_perolehan:
        from datetime import date
        years_elapsed = (date.today() - record.tanggal_perolehan).days / Decimal('365.25')
        tahun_ke = int(years_elapsed) + 1
    else:
        tahun_ke = 1

    # Preview amount (for time-based methods), default 30 days
    preview_amount = Decimal('0')
    if not needs_activity_input:
        preview_amount = calculate_depreciation(record, tahun_ke=tahun_ke, days=DEFAULT_DAYS)

    # Depreciation journal history
    from apps.jurnal.models import JurnalHeader, JurnalDetail
    from django.db.models import Sum as _Sum
    from datetime import date as date_cls, timedelta
    dep_journals = list(
        JurnalHeader.objects.filter(
            uraian_transaksi__startswith=f'Penyusutan {record.aset_number}',
        ).order_by('-tanggal')
    )

    # Build a map of journal pk -> depreciation amount (debit on akun 5.1.19.xx)
    dep_amounts = {}
    if dep_journals:
        journal_pks = [j.pk for j in dep_journals]
        amounts_qs = (
            JurnalDetail.objects
            .filter(jurnal_header_id__in=journal_pks, akun__kode_akun__startswith='5.1.19')
            .values('jurnal_header_id')
            .annotate(total=_Sum('debit'))
        )
        dep_amounts = {row['jurnal_header_id']: row['total'] for row in amounts_qs}

    last_dep = dep_journals[0] if dep_journals else None
    last_dep_date = last_dep.tanggal if last_dep else None

    # Build combined list for template: [{journal, amount}]
    dep_journal_data = [
        {'journal': j, 'amount': dep_amounts.get(j.pk, Decimal('0'))}
        for j in dep_journals
    ]
    if last_dep_date:
        suggested_start_date = last_dep_date + timedelta(days=1)
    else:
        suggested_start_date = record.tanggal_perolehan

    today = date_cls.today()
    suggested_days = (today - suggested_start_date).days + 1 if today >= suggested_start_date else 0

    from datetime import date as _dc
    disposal_form = AssetDisposalForm(aset=record, initial={'tanggal': _dc.today(), 'quantity': record.quantity})
    disposals = record.disposals.select_related('akun_kas', 'akun_laba_rugi', 'jurnal_header').all()
    dokumen_list = list_bukti_aset(record)

    return render(request, 'aset_tetap/aset_tetap_detail.html', {
        'record': record,
        'preview_amount': preview_amount,
        'tahun_ke': tahun_ke,
        'needs_activity_input': needs_activity_input,
        'metode': metode,
        'can_depreciate': record.nilai_buku > record.nilai_residu,
        'dep_journal_data': dep_journal_data,
        'latest_dep_journal_pk': dep_journals[0].pk if dep_journals else None,
        'default_days': DEFAULT_DAYS,
        'last_dep_date': last_dep_date,
        'suggested_start_date': suggested_start_date,
        'suggested_days': suggested_days,
        'today': today,
        'disposal_form': disposal_form,
        'disposals': disposals,
        'can_dispose': record.status == 'aktif' and record.quantity > 0,
        'dokumen_list': dokumen_list,
    })


@login_required
def aset_tetap_create(request: HttpRequest) -> HttpResponse:
    """Create a new fixed asset record."""
    if request.method == 'POST':
        form = AsetTetapRecordForm(request.POST)
        if form.is_valid():
            record = form.save()
            messages.success(request, f'Aset tetap {record.aset_number} berhasil dibuat.')
            return redirect('aset_tetap:detail', pk=record.pk)
    else:
        form = AsetTetapRecordForm()
    return render(request, 'aset_tetap/aset_tetap_form.html', {'form': form, 'title': 'Tambah Aset Tetap'})


@login_required
def aset_tetap_update(request: HttpRequest, pk: int) -> HttpResponse:
    """Edit an existing fixed asset record."""
    record = get_object_or_404(AsetTetapRecord, pk=pk)
    if request.method == 'POST':
        form = AsetTetapRecordForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            messages.success(request, f'Aset tetap {record.aset_number} berhasil diperbarui.')
            return redirect('aset_tetap:detail', pk=record.pk)
    else:
        form = AsetTetapRecordForm(instance=record)
    return render(request, 'aset_tetap/aset_tetap_form.html', {
        'form': form,
        'record': record,
        'title': f'Edit {record.aset_number}',
    })


@login_required
def aset_tetap_delete(request: HttpRequest, pk: int) -> HttpResponse:
    """Delete a fixed asset record, removing all associated depreciation journals."""
    from apps.jurnal.models import JurnalHeader
    from apps.jurnal.utils import log_jurnal_terhapus

    record = get_object_or_404(AsetTetapRecord.objects.select_related('item', 'entitas_bisnis'), pk=pk)

    # Find all depreciation journals for this asset
    dep_journals = list(
        JurnalHeader.objects.filter(
            uraian_transaksi__startswith=f'Penyusutan {record.aset_number}'
        ).order_by('tanggal')
    )

    if request.method == 'POST':
        from django.db import transaction as db_transaction
        number = record.aset_number
        with db_transaction.atomic():
            for journal in dep_journals:
                log_jurnal_terhapus(journal, 'aset_tetap', request)
                journal.details.all().delete()
                journal.delete()
            record.delete()
        messages.success(request, f'Aset tetap {number} dan {len(dep_journals)} jurnal penyusutan berhasil dihapus.')
        return redirect('aset_tetap:list')

    return render(request, 'aset_tetap/aset_tetap_delete.html', {
        'record': record,
        'dep_journals': dep_journals,
    })


@login_required
def aset_tetap_process_depreciation(request: HttpRequest, pk: int) -> HttpResponse:
    """Process depreciation for a fixed asset record, creating a journal entry."""
    from decimal import Decimal, InvalidOperation

    record = get_object_or_404(
        AsetTetapRecord.objects.select_related('item', 'entitas_bisnis'),
        pk=pk,
    )

    if request.method != 'POST':
        return redirect('aset_tetap:detail', pk=pk)

    metode = record.metode_penyusutan or (record.item.metode_penyusutan if record.item else '')
    tanggal_str = request.POST.get('tanggal', '')

    try:
        from datetime import date as date_cls
        tanggal = date_cls.fromisoformat(tanggal_str) if tanggal_str else None
    except ValueError:
        tanggal = None

    # For activity-based methods, get input values
    if metode == 'service_hours':
        try:
            jam_aktual = Decimal(request.POST.get('jam_aktual', '0'))
        except InvalidOperation:
            messages.error(request, 'Jam kerja aktual harus berupa angka.')
            return redirect('aset_tetap:detail', pk=pk)
        depreciation = calculate_depreciation(record, jam_aktual=jam_aktual)
    elif metode == 'units_of_production':
        try:
            unit_aktual = Decimal(request.POST.get('unit_aktual', '0'))
        except InvalidOperation:
            messages.error(request, 'Unit produksi aktual harus berupa angka.')
            return redirect('aset_tetap:detail', pk=pk)
        depreciation = calculate_depreciation(record, unit_aktual=unit_aktual)
    else:
        # Auto-calculate days from the appropriate start date to tanggal_jurnal
        from apps.jurnal.models import JurnalHeader as _JH
        from datetime import timedelta as _td
        last_dep = _JH.objects.filter(
            uraian_transaksi__startswith=f'Penyusutan {record.aset_number}',
        ).order_by('-tanggal').first()
        if last_dep:
            start_date = last_dep.tanggal + _td(days=1)
        else:
            start_date = record.tanggal_perolehan

        if tanggal is None:
            messages.error(request, 'Tanggal jurnal wajib diisi.')
            return redirect('aset_tetap:detail', pk=pk)
        if tanggal < start_date:
            messages.error(
                request,
                f'Tanggal jurnal ({tanggal}) tidak boleh sebelum tanggal mulai perhitungan ({start_date}).',
            )
            return redirect('aset_tetap:detail', pk=pk)

        hari = (tanggal - start_date).days + 1
        masa = record.masa_manfaat or (record.item.masa_manfaat if record.item else 0) or 0
        if masa > 0 and record.tanggal_perolehan:
            from datetime import date as d
            years_elapsed = (tanggal - record.tanggal_perolehan).days / Decimal('365.25')
            tahun_ke = min(int(years_elapsed) + 1, masa)
        else:
            tahun_ke = 1
        depreciation = calculate_depreciation(record, tahun_ke=tahun_ke, days=hari)

    try:
        header = process_depreciation(record, depreciation, tanggal)
        messages.success(
            request,
            f'Penyusutan {record.aset_number} sebesar Rp {depreciation:,.0f} berhasil diproses. '
            f'Jurnal: {header.nomor_transaksi}',
        )
    except ValueError as e:
        messages.error(request, str(e))

    return redirect('aset_tetap:detail', pk=pk)


@login_required
def aset_tetap_bulk_depreciation(request: HttpRequest) -> HttpResponse:
    """Process depreciation for selected fixed assets in one batch."""
    from datetime import date as date_cls, timedelta as _td

    if request.method != 'POST':
        return redirect('aset_tetap:list')

    tanggal_str = request.POST.get('tanggal', '')
    try:
        tanggal = date_cls.fromisoformat(tanggal_str) if tanggal_str else date_cls.today()
    except ValueError:
        tanggal = date_cls.today()

    # Entitas bisnis filter (required in UI but handled gracefully here)
    eb_id = request.POST.get('entitas_bisnis', '')

    # Selected asset IDs from checkboxes
    selected_ids = request.POST.getlist('record_ids')

    records = AsetTetapRecord.objects.select_related('item', 'entitas_bisnis')
    if selected_ids:
        records = records.filter(pk__in=selected_ids)
    else:
        records = records.all()
    if eb_id:
        records = records.filter(entitas_bisnis_id=eb_id)

    success_count = 0
    skip_count = 0
    error_msgs = []

    for record in records:
        if record.status == 'dilepas':
            skip_count += 1
            continue

        if record.nilai_buku <= record.nilai_residu:
            skip_count += 1
            continue

        metode = record.metode_penyusutan or (record.item.metode_penyusutan if record.item else '')
        if metode in ('service_hours', 'units_of_production'):
            skip_count += 1
            continue

        # Calculate days from last depreciation or acquisition date
        from apps.jurnal.models import JurnalHeader
        last_dep = (
            JurnalHeader.objects
            .filter(
                nomor_transaksi__startswith='TRX-DEP-',
                uraian_transaksi__contains=record.aset_number,
            )
            .order_by('-tanggal')
            .first()
        )
        start_date = (last_dep.tanggal + _td(days=1)) if last_dep else record.tanggal_perolehan
        if tanggal < start_date:
            skip_count += 1
            continue

        hari = (tanggal - start_date).days + 1

        masa = record.masa_manfaat or (record.item.masa_manfaat if record.item else 0) or 0
        tahun_ke = 1
        if masa > 0 and record.tanggal_perolehan:
            years_elapsed = (tanggal - record.tanggal_perolehan).days / Decimal('365.25')
            tahun_ke = min(int(years_elapsed) + 1, masa)

        depreciation = calculate_depreciation(record, tahun_ke=tahun_ke, days=hari)
        if depreciation <= 0:
            skip_count += 1
            continue

        try:
            process_depreciation(record, depreciation, tanggal)
            success_count += 1
        except ValueError as e:
            error_msgs.append(f'{record.aset_number}: {e}')

    if success_count:
        messages.success(
            request,
            f'Bulk penyusutan selesai: {success_count} aset diproses, '
            f'{skip_count} dilewati.',
        )
    else:
        messages.warning(request, f'Tidak ada aset yang diproses. {skip_count} dilewati.')

    for msg in error_msgs[:5]:
        messages.error(request, msg)

    return redirect('aset_tetap:list')


@login_required
def aset_tetap_bulk_preview(request: HttpRequest) -> JsonResponse:
    """AJAX preview: calculate depreciation amounts for all eligible assets."""
    from datetime import date as date_cls, timedelta as _td

    tanggal_str = request.GET.get('tanggal', '')
    try:
        tanggal = date_cls.fromisoformat(tanggal_str) if tanggal_str else None
    except ValueError:
        tanggal = None

    if not tanggal:
        return JsonResponse({'error': 'Tanggal wajib diisi.'}, status=400)

    eb_id = request.GET.get('entitas_bisnis', '')
    metode_filter = request.GET.getlist('metode')  # optional list of metode codes

    from apps.jurnal.models import JurnalHeader

    records = AsetTetapRecord.objects.select_related('item', 'entitas_bisnis').all()
    if eb_id:
        records = records.filter(entitas_bisnis_id=eb_id)
    if metode_filter:
        # Filter by records whose effective metode is in the list
        records = records.filter(
            models.Q(metode_penyusutan__in=metode_filter) |
            models.Q(metode_penyusutan='', item__metode_penyusutan__in=metode_filter) |
            models.Q(metode_penyusutan__isnull=True, item__metode_penyusutan__in=metode_filter)
        )
    preview = []

    for record in records:
        if record.nilai_buku <= record.nilai_residu:
            continue
        metode = record.metode_penyusutan or (record.item.metode_penyusutan if record.item else '')
        if metode in ('service_hours', 'units_of_production'):
            continue

        last_dep = (
            JurnalHeader.objects
            .filter(
                nomor_transaksi__startswith='TRX-DEP-',
                uraian_transaksi__contains=record.aset_number,
            )
            .order_by('-tanggal')
            .first()
        )
        start_date = (last_dep.tanggal + _td(days=1)) if last_dep else record.tanggal_perolehan
        if tanggal < start_date:
            continue

        hari = (tanggal - start_date).days + 1

        masa = record.masa_manfaat or (record.item.masa_manfaat if record.item else 0) or 0
        tahun_ke = 1
        if masa > 0 and record.tanggal_perolehan:
            years_elapsed = (tanggal - record.tanggal_perolehan).days / Decimal('365.25')
            tahun_ke = min(int(years_elapsed) + 1, masa)

        depreciation = calculate_depreciation(record, tahun_ke=tahun_ke, days=hari)
        if depreciation <= 0:
            continue

        new_book_value = record.nilai_buku - depreciation
        preview.append({
            'id': record.pk,
            'aset_number': record.aset_number,
            'item_nama': record.item.nama if record.item else '-',
            'metode': record.get_metode_penyusutan_display(),
            'nilai_buku': float(record.nilai_buku),
            'hari': hari,
            'depreciation': float(depreciation),
            'new_book_value': float(new_book_value),
            'entitas_bisnis': record.entitas_bisnis.nama if record.entitas_bisnis else '-',
        })

    return JsonResponse({'records': preview})


@login_required
def delete_depreciation_journal(request: HttpRequest, pk: int, jurnal_pk: int) -> HttpResponse:
    """Delete a single depreciation journal and reverse its akumulasi_penyusutan."""
    from apps.jurnal.models import JurnalHeader, JurnalDetail
    from apps.jurnal.utils import log_jurnal_terhapus
    from django.db import transaction as db_transaction
    from django.db.models import Sum as _Sum

    record = get_object_or_404(
        AsetTetapRecord.objects.select_related('item', 'entitas_bisnis'),
        pk=pk,
    )
    journal = get_object_or_404(
        JurnalHeader,
        pk=jurnal_pk,
        uraian_transaksi__startswith=f'Penyusutan {record.aset_number}',
    )

    # Validate that this is the latest depreciation journal for this asset
    latest_journal = (
        JurnalHeader.objects
        .filter(uraian_transaksi__startswith=f'Penyusutan {record.aset_number}')
        .order_by('-tanggal', '-pk')
        .first()
    )
    if not latest_journal or latest_journal.pk != journal.pk:
        messages.error(request, 'Hanya jurnal penyusutan terbaru yang dapat dihapus.')
        return redirect('aset_tetap:detail', pk=pk)

    # Get the depreciation amount from the beban penyusutan (debit on 5.1.19.xx)
    agg = (
        JurnalDetail.objects
        .filter(jurnal_header=journal, akun__kode_akun__startswith='5.1.19')
        .aggregate(total=_Sum('debit'))
    )
    dep_amount = agg['total'] or Decimal('0')

    if request.method == 'POST':
        with db_transaction.atomic():
            log_jurnal_terhapus(journal, 'aset_tetap', request)
            journal.details.all().delete()
            journal.delete()
            # Reverse accumulated depreciation
            record.akumulasi_penyusutan = max(
                Decimal('0'),
                record.akumulasi_penyusutan - dep_amount,
            )
            record.save(update_fields=['akumulasi_penyusutan'])
        messages.success(
            request,
            f'Jurnal {journal.nomor_transaksi} dihapus. '
            f'Akumulasi penyusutan dikurangi Rp {dep_amount:,.0f}.',
        )
        return redirect('aset_tetap:detail', pk=pk)

    return render(request, 'aset_tetap/delete_dep_journal_confirm.html', {
        'record': record,
        'journal': journal,
        'dep_amount': dep_amount,
    })


@login_required
def aset_tetap_dispose(request: HttpRequest, pk: int) -> HttpResponse:
    """Proses pelepasan aset dari halaman detail."""
    record = get_object_or_404(
        AsetTetapRecord.objects.select_related('item', 'entitas_bisnis', 'purchase_item'),
        pk=pk,
    )
    if request.method != 'POST':
        return redirect('aset_tetap:detail', pk=pk)

    form = AssetDisposalForm(request.POST, aset=record)
    if not form.is_valid():
        for field, errs in form.errors.items():
            for e in errs:
                messages.error(request, f'{field}: {e}')
        return redirect('aset_tetap:detail', pk=pk)

    disposal = form.save(commit=False)
    disposal.aset = record
    try:
        header = process_asset_disposal(disposal)
        messages.success(
            request,
            f'Pelepasan {record.aset_number} berhasil diproses. Jurnal: {header.nomor_transaksi}.',
        )
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('aset_tetap:detail', pk=pk)


@login_required
def aset_tetap_disposal_delete(request: HttpRequest, pk: int, disposal_pk: int) -> HttpResponse:
    """Batalkan pelepasan aset dan pulihkan aset."""
    record = get_object_or_404(AsetTetapRecord, pk=pk)
    disposal = get_object_or_404(AssetDisposal, pk=disposal_pk, aset=record)
    if request.method == 'POST':
        disposal_number = disposal.disposal_number
        reverse_asset_disposal(disposal, request)
        messages.success(
            request,
            f'Pelepasan {disposal_number} dibatalkan. Aset dipulihkan.',
        )
        return redirect('aset_tetap:detail', pk=pk)
    return render(request, 'aset_tetap/disposal_delete_confirm.html', {
        'record': record,
        'disposal': disposal,
    })


# ── Export views ─────────────────────────────────────────────────────────────

def _aset_tetap_export_qs(request):
    """Return filtered AsetTetapRecord queryset based on GET params."""
    qs = AsetTetapRecord.objects.select_related(
        'item', 'item__kategori', 'entitas_bisnis',
    ).order_by('item__kategori__nama', '-tanggal_perolehan', '-created_at')
    tanggal_dari = request.GET.get('tanggal_dari', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    item_filter = request.GET.get('item', '')
    eb_filter = request.GET.get('entitas_bisnis', '')
    kondisi_filter = request.GET.get('kondisi', '')
    kategori_filter = request.GET.get('kategori', '')
    if tanggal_dari:
        qs = qs.filter(tanggal_perolehan__gte=tanggal_dari)
    if tanggal_sampai:
        qs = qs.filter(tanggal_perolehan__lte=tanggal_sampai)
    if item_filter:
        qs = qs.filter(item_id=item_filter)
    if eb_filter:
        qs = qs.filter(entitas_bisnis_id=eb_filter)
    if kondisi_filter:
        qs = qs.filter(kondisi=kondisi_filter)
    if kategori_filter:
        qs = qs.filter(item__kategori_id=kategori_filter)
    return qs


@login_required
@ratelimit(key='user', rate=rate_from('export'), method='GET', block=True)
def aset_tetap_export(request: HttpRequest) -> HttpResponse:
    """Export aset tetap list as XLSX with same filters as list page."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    records = list(_aset_tetap_export_qs(request))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Aset Tetap'

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin = Side(style='thin')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right_align = Alignment(horizontal='right')

    headers = [
        'No. Aset', 'Item', 'Kategori', 'Entitas Bisnis', 'Tanggal Perolehan',
        'Qty', 'Harga Perolehan (Rp)', 'Total Nilai (Rp)',
        'Akum. Penyusutan (Rp)', 'Nilai Buku (Rp)', 'Kondisi', 'Metode Penyusutan',
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')

    for row_num, r in enumerate(records, 2):
        vals = [
            r.aset_number,
            f'{r.item.item_id} - {r.item.nama}',
            r.item.kategori.nama if r.item.kategori else '',
            r.entitas_bisnis.nama,
            str(r.tanggal_perolehan),
            float(r.quantity or 1),
            float(r.harga_perolehan or 0),
            float(r.total_value or 0),
            float(r.akumulasi_penyusutan or 0),
            float(r.nilai_buku),
            r.get_kondisi_display() if hasattr(r, 'get_kondisi_display') else r.kondisi,
            r.metode_penyusutan or '',
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.border = thin_border
            if col in (6, 7, 8, 9, 10):
                c.alignment = right_align
                c.number_format = '#,##0'

    col_widths = [18, 32, 18, 26, 16, 8, 20, 20, 22, 18, 12, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="aset_tetap.xlsx"'
    wb.save(response)
    return response


@login_required
@ratelimit(key='user', rate=rate_from('export'), method='GET', block=True)
def aset_tetap_export_pdf(request: HttpRequest) -> HttpResponse:
    """Render print-friendly aset tetap list for browser PDF printing."""
    records = list(_aset_tetap_export_qs(request))
    total_nilai = sum(r.total_value for r in records)
    total_akum = sum(r.akumulasi_penyusutan for r in records)
    total_buku = sum(r.nilai_buku for r in records)
    return render(request, 'aset_tetap/aset_tetap_export_pdf.html', {
        'records': records,
        'tanggal_dari': request.GET.get('tanggal_dari', ''),
        'tanggal_sampai': request.GET.get('tanggal_sampai', ''),
        'generated_at': __import__('datetime').datetime.now(),
        'total_nilai': total_nilai,
        'total_akum': total_akum,
        'total_buku': total_buku,
        'total_records': len(records),
    })


# ── Maintenance views ────────────────────────────────────────────────────────

@login_required
def maintenance_list(request: HttpRequest) -> HttpResponse:
    items = AssetMaintenance.objects.select_related('aset').all()
    return render(request, 'aset_tetap/maintenance_list.html', {'items': items})


@login_required
def maintenance_create(request: HttpRequest) -> HttpResponse:
    if request.method == 'POST':
        form = AssetMaintenanceForm(request.POST)
        if form.is_valid():
            mtn = form.save()
            try:
                services.process_asset_maintenance(mtn)
                messages.success(request, f'Maintenance {mtn.maintenance_number} tersimpan.')
                return redirect('aset_tetap:maintenance_list')
            except ValueError as e:
                mtn.delete()
                messages.error(request, str(e))
    else:
        form = AssetMaintenanceForm()
    return render(request, 'aset_tetap/maintenance_form.html', {'form': form})


@login_required
def maintenance_delete(request: HttpRequest, pk: int) -> HttpResponse:
    mtn = get_object_or_404(AssetMaintenance, pk=pk)
    services.reverse_asset_maintenance(mtn, request)
    messages.success(request, 'Maintenance dibatalkan.')
    return redirect('aset_tetap:maintenance_list')


# ── Transfer views ───────────────────────────────────────────────────────────

@login_required
def transfer_list(request: HttpRequest) -> HttpResponse:
    items = AssetTransfer.objects.select_related('aset').all()
    return render(request, 'aset_tetap/transfer_list.html', {'items': items})


@login_required
def transfer_create(request: HttpRequest) -> HttpResponse:
    if request.method == 'POST':
        form = AssetTransferForm(request.POST)
        if form.is_valid():
            trf = form.save()
            try:
                services.process_asset_transfer(trf)
                messages.success(request, f'Transfer {trf.transfer_number} tersimpan.')
                return redirect('aset_tetap:transfer_list')
            except ValueError as e:
                trf.delete()
                messages.error(request, str(e))
    else:
        form = AssetTransferForm()
    return render(request, 'aset_tetap/transfer_form.html', {'form': form})


@login_required
def transfer_delete(request: HttpRequest, pk: int) -> HttpResponse:
    trf = get_object_or_404(AssetTransfer, pk=pk)
    services.reverse_asset_transfer(trf, request)
    messages.success(request, 'Transfer dibatalkan.')
    return redirect('aset_tetap:transfer_list')


# ── Revaluation views ────────────────────────────────────────────────────────

@login_required
def revaluation_list(request: HttpRequest) -> HttpResponse:
    items = AssetRevaluation.objects.select_related('aset').all()
    return render(request, 'aset_tetap/revaluation_list.html', {'items': items})


@login_required
def revaluation_create(request: HttpRequest) -> HttpResponse:
    warning = ''
    if request.method == 'POST':
        form = AssetRevaluationForm(request.POST)
        if form.is_valid():
            rev = form.save()
            try:
                services.process_asset_revaluation(rev)
                messages.success(request, f'Revaluasi {rev.revaluation_number} tersimpan.')
                return redirect('aset_tetap:revaluation_list')
            except ValueError as e:
                rev.delete()
                messages.error(request, str(e))
    else:
        form = AssetRevaluationForm()
    return render(request, 'aset_tetap/revaluation_form.html', {'form': form, 'warning': warning})


@login_required
def revaluation_delete(request: HttpRequest, pk: int) -> HttpResponse:
    rev = get_object_or_404(AssetRevaluation, pk=pk)
    services.reverse_asset_revaluation(rev, request)
    messages.success(request, 'Revaluasi dibatalkan.')
    return redirect('aset_tetap:revaluation_list')


# ── Reports ──────────────────────────────────────────────────────────────────

@login_required
def depreciation_schedule(request: HttpRequest, pk: int) -> HttpResponse:
    aset = get_object_or_404(AsetTetapRecord, pk=pk)
    periods = int(request.GET.get('periods', 12))
    rows = reports.depreciation_schedule(aset, periods=periods)
    return render(request, 'aset_tetap/depreciation_schedule.html', {'aset': aset, 'rows': rows})


@login_required
def laporan_penyusutan(request: HttpRequest) -> HttpResponse:
    from apps.purchase.models import KategoriItem
    from apps.entitas_bisnis.models import EntitasBisnisLv3

    kategori_id = request.GET.get('kategori') or None
    lokasi_id = request.GET.get('lokasi') or None
    dept_id = request.GET.get('departemen') or None
    rows = reports.laporan_penyusutan(
        kategori=KategoriItem.objects.filter(pk=kategori_id).first() if kategori_id else None,
        lokasi=LokasiAset.objects.filter(pk=lokasi_id).first() if lokasi_id else None,
        departemen=EntitasBisnisLv3.objects.filter(pk=dept_id).first() if dept_id else None,
    )
    return render(request, 'aset_tetap/laporan_penyusutan.html', {'rows': rows})
