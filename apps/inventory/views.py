"""Inventory views — CRUD for InventoryRecord."""
import json
import math
from collections import defaultdict
from datetime import timedelta
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from .forms import InventoryRecordForm
from .models import InventoryRecord

BULK_TO_SATUAN_MAP = {'RMB': 'RM', 'FGB': 'FG', 'ITMB': 'ITM'}


@login_required
def inventory_list(request: HttpRequest) -> HttpResponse:
    """List all inventory records with optional filters, split by Satuan and Bulk tabs."""
    qs = InventoryRecord.objects.select_related('item', 'entitas_bisnis').all()

    tanggal_dari = request.GET.get('tanggal_dari', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    item_filter = request.GET.get('item', '')
    eb_filter = request.GET.get('entitas_bisnis', '')
    tab = request.GET.get('tab', 'satuan')

    if tanggal_dari:
        qs = qs.filter(tanggal__gte=tanggal_dari)
    if tanggal_sampai:
        qs = qs.filter(tanggal__lte=tanggal_sampai)
    if item_filter:
        qs = qs.filter(item_id=item_filter)
    if eb_filter:
        qs = qs.filter(entitas_bisnis_id=eb_filter)

    BULK_TYPES = ('RMB', 'FGB', 'ITMB')
    SATUAN_TYPES = ('RM', 'FG', 'ITM')

    records_satuan = qs.filter(item__tipe_item__in=SATUAN_TYPES)
    records_bulk = qs.filter(item__tipe_item__in=BULK_TYPES)

    from apps.purchase.models import ItemMasterPurchase
    from apps.entitas_bisnis.models import EntitasBisnis
    from apps.dashboard.models import DashboardInventoryTag

    tagged_qs = DashboardInventoryTag.objects
    if eb_filter:
        tagged_qs = tagged_qs.filter(entitas_bisnis_id=eb_filter)
    tagged_ids = set(tagged_qs.values_list('item_id', flat=True))

    return render(request, 'inventory/inventory_list.html', {
        'records_satuan': records_satuan,
        'records_bulk': records_bulk,
        'items': ItemMasterPurchase.objects.filter(
            tipe_item__in=list(SATUAN_TYPES) + list(BULK_TYPES),
        ).order_by('item_id'),
        'entitas_list': EntitasBisnis.objects.filter(status_aktif=True).order_by('nama'),
        'tanggal_dari': tanggal_dari,
        'tanggal_sampai': tanggal_sampai,
        'item_filter': item_filter,
        'eb_filter': eb_filter,
        'active_tab': tab,
        'tagged_item_ids': tagged_ids,
    })


@login_required
def inventory_detail(request: HttpRequest, pk: int) -> HttpResponse:
    """Show inventory record detail with mutation history."""
    record = get_object_or_404(
        InventoryRecord.objects.select_related(
            'item', 'entitas_bisnis',
            'purchase_item__purchase_eb__purchase_header',
        ),
        pk=pk,
    )

    # Build mutation history for this inventory record
    # Purchase or saldo-awal inflow (the transaction that created this record)
    is_bulk = record.item.tipe_item in BULK_TO_SATUAN_MAP
    mutations = []
    if record.purchase_item:
        mutations.append({
            'tanggal': record.tanggal,
            'tipe': 'Masuk (Purchase)',
            'referensi': record.purchase_item.purchase_eb.purchase_header.transaction_id,
            'url': None,
            'quantity': record.purchase_item.total_value if is_bulk else record.purchase_item.quantity,
            'keterangan': f'Pembelian via {record.purchase_item.purchase_eb.purchase_header.transaction_id}',
        })
    else:
        # Saldo awal — compute original value/quantity from remaining + all consumed allocations
        from decimal import Decimal
        from django.db.models import Sum
        from apps.sales.models import SalesItemFIFOAllocation as _SIFA
        if is_bulk:
            consumed_total = (
                _SIFA.objects
                .filter(inventory_record=record)
                .aggregate(total=Sum('cogs_amount'))['total'] or Decimal('0')
            )
            original_value = record.total_value + consumed_total
            mutations.append({
                'tanggal': record.tanggal,
                'tipe': 'Masuk (Saldo Awal)',
                'referensi': f'Saldo Awal {record.tanggal}',
                'url': None,
                'quantity': original_value,
                'keterangan': 'Saldo awal persediaan',
            })
        else:
            consumed_total = (
                _SIFA.objects
                .filter(inventory_record=record)
                .aggregate(total=Sum('quantity_consumed'))['total'] or Decimal('0')
            )
            original_qty = record.quantity + consumed_total
            mutations.append({
                'tanggal': record.tanggal,
                'tipe': 'Masuk (Saldo Awal)',
                'referensi': f'Saldo Awal {record.tanggal}',
                'url': None,
                'quantity': original_qty,
                'keterangan': 'Saldo awal persediaan',
            })

    # Sales outflows — via SalesItemFIFOAllocation for accurate per-batch quantities
    from apps.sales.models import SalesItemFIFOAllocation
    alloc_qs = (
        SalesItemFIFOAllocation.objects
        .filter(inventory_record=record)
        .select_related(
            'sales_item__sales_eb__sales_header',
            'sales_item__sales_eb__entitas_bisnis',
        )
        .order_by('sales_item__sales_eb__sales_header__tanggal')
    )
    for alloc in alloc_qs:
        sh = alloc.sales_item.sales_eb.sales_header
        mutations.append({
            'tanggal': sh.tanggal,
            'tipe': 'Keluar (Sales)',
            'referensi': sh.transaction_id,
            'url': f'/sales/{sh.pk}/',
            'quantity': alloc.cogs_amount if is_bulk else alloc.quantity_consumed,
            'keterangan': f'Penjualan via {sh.transaction_id} ({alloc.sales_item.sales_eb.entitas_bisnis.nama})',
        })

    context = {
        'record': record,
        'mutations': mutations,
    }

    # For bulk items, provide satuan items for conversion modal
    if record.item.tipe_item in BULK_TO_SATUAN_MAP:
        from apps.purchase.models import ItemMasterPurchase
        satuan_tipe = BULK_TO_SATUAN_MAP[record.item.tipe_item]
        context['satuan_items'] = ItemMasterPurchase.objects.filter(
            tipe_item=satuan_tipe,
        ).order_by('item_id')
        context['is_bulk'] = True

    return render(request, 'inventory/inventory_detail.html', context)


@login_required
def inventory_create(request: HttpRequest) -> HttpResponse:
    """Manually create a new InventoryRecord (e.g. saldo awal)."""
    if request.method == 'POST':
        form = InventoryRecordForm(request.POST)
        if form.is_valid():
            record = form.save()
            messages.success(request, f'Inventory record {record.inventory_number} berhasil dibuat.')
            return redirect('inventory:detail', pk=record.pk)
    else:
        form = InventoryRecordForm()
    return render(request, 'inventory/inventory_form.html', {'form': form, 'title': 'Tambah Inventory Record'})


@login_required
def inventory_update(request: HttpRequest, pk: int) -> HttpResponse:
    """Edit an existing InventoryRecord."""
    record = get_object_or_404(InventoryRecord, pk=pk)
    if request.method == 'POST':
        form = InventoryRecordForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            messages.success(request, f'Inventory record {record.inventory_number} berhasil diperbarui.')
            return redirect('inventory:detail', pk=record.pk)
    else:
        form = InventoryRecordForm(instance=record)
    return render(request, 'inventory/inventory_form.html', {
        'form': form,
        'record': record,
        'title': f'Edit {record.inventory_number}',
    })


@login_required
def inventory_delete(request: HttpRequest, pk: int) -> HttpResponse:
    """Delete an InventoryRecord with confirmation page."""
    record = get_object_or_404(
        InventoryRecord.objects.select_related('item', 'entitas_bisnis', 'purchase_item'),
        pk=pk,
    )
    if request.method == 'POST':
        number = record.inventory_number
        record.delete()
        messages.success(request, f'Inventory record {number} berhasil dihapus.')
        return redirect('inventory:list')
    return render(request, 'inventory/inventory_delete.html', {'record': record})


@login_required
def convert_bulk_to_satuan(request: HttpRequest, pk: int) -> HttpResponse:
    """Convert a bulk inventory record to satuan inventory records.

    Deducts value from the bulk record and creates a new satuan InventoryRecord
    with the specified quantity and unit price. Also creates corresponding
    satuan FIFO batch.
    """
    from decimal import Decimal
    from apps.purchase.models import ItemMasterPurchase, FIFOBatch

    record = get_object_or_404(
        InventoryRecord.objects.select_related('item', 'entitas_bisnis'),
        pk=pk,
    )

    if record.item.tipe_item not in BULK_TO_SATUAN_MAP:
        messages.error(request, 'Record ini bukan tipe bulk.')
        return redirect('inventory:detail', pk=pk)

    if request.method != 'POST':
        return redirect('inventory:detail', pk=pk)

    satuan_tipe = BULK_TO_SATUAN_MAP[record.item.tipe_item]

    # Get or validate the target satuan item
    satuan_item_id = request.POST.get('satuan_item_id', '')
    quantity = request.POST.get('quantity', '')
    unit_price = request.POST.get('unit_price', '')

    # Validation
    errors = []
    if not satuan_item_id:
        errors.append('Item satuan wajib dipilih.')
    if not quantity:
        errors.append('Jumlah wajib diisi.')
    if not unit_price:
        errors.append('Harga satuan wajib diisi.')

    if errors:
        messages.error(request, ' '.join(errors))
        return redirect('inventory:detail', pk=pk)

    try:
        qty = Decimal(quantity)
        price = Decimal(unit_price)
    except Exception:
        messages.error(request, 'Jumlah dan harga harus berupa angka.')
        return redirect('inventory:detail', pk=pk)

    if qty <= 0 or price <= 0:
        messages.error(request, 'Jumlah dan harga harus lebih dari 0.')
        return redirect('inventory:detail', pk=pk)

    total_deduction = qty * price

    if total_deduction > record.total_value:
        messages.error(
            request,
            f'Nilai konversi ({total_deduction:,.0f}) melebihi sisa nilai bulk ({record.total_value:,.0f}).',
        )
        return redirect('inventory:detail', pk=pk)

    # Validate satuan item exists and has correct type
    try:
        satuan_item = ItemMasterPurchase.objects.get(pk=satuan_item_id, tipe_item=satuan_tipe)
    except ItemMasterPurchase.DoesNotExist:
        messages.error(request, f'Item satuan tipe {satuan_tipe} tidak ditemukan.')
        return redirect('inventory:detail', pk=pk)

    from django.db import transaction as db_transaction
    with db_transaction.atomic():
        # Deduct from bulk record
        record.total_value -= total_deduction
        record.unit_price = record.total_value  # bulk: unit_price tracks total value
        record.save()

        # Also deduct from bulk FIFO batch
        bulk_batches = (
            FIFOBatch.objects
            .filter(item=record.item, remaining_qty__gt=0)
            .order_by('tanggal', 'created_at')
            .select_for_update()
        )
        remaining_deduct = total_deduction
        for batch in bulk_batches:
            if remaining_deduct <= 0:
                break
            batch_value = batch.remaining_qty * batch.unit_price
            deduct = min(batch_value, remaining_deduct)
            if deduct > 0 and batch.unit_price:
                batch.remaining_qty = (batch_value - deduct) / batch.unit_price
                batch.save()
                remaining_deduct -= deduct

        # Create satuan InventoryRecord
        new_record = InventoryRecord(
            item=satuan_item,
            entitas_bisnis=record.entitas_bisnis,
            quantity=qty,
            unit_price=price,
            tanggal=record.tanggal,
            metode_alokasi=record.metode_alokasi,
        )
        new_record.save()

        # Create satuan FIFO batch
        FIFOBatch.objects.create(
            item=satuan_item,
            entitas_bisnis=record.entitas_bisnis,
            tanggal=record.tanggal,
            quantity_in=qty,
            remaining_qty=qty,
            unit_price=price,
        )

    messages.success(
        request,
        f'Berhasil konversi {qty:,.0f} unit {satuan_item.nama} '
        f'(Rp {total_deduction:,.0f}) dari {record.inventory_number}.',
    )
    return redirect('inventory:detail', pk=pk)


# ── Laporan Persediaan ───────────────────────────────────────────────────────

INVENTORY_TIPE_ITEMS = ('RM', 'FG', 'ITM', 'RMB', 'FGB', 'ITMB')
BULK_TYPES = ('RMB', 'FGB', 'ITMB')

_TIPE_LABEL = {
    'RM': 'Raw Material', 'FG': 'Finished Good', 'ITM': 'Item Lainnya',
    'RMB': 'Raw Material (Bulk)', 'FGB': 'Finished Good (Bulk)', 'ITMB': 'Item Lainnya (Bulk)',
}


def _format_rp(val: Decimal | int | float) -> str:
    """Format Rupiah as shortened string (jt/rb)."""
    v = float(val)
    if abs(v) >= 1_000_000:
        return f'{v / 1_000_000:,.2f}jt'
    if abs(v) >= 1_000:
        return f'{v / 1_000:,.0f}rb'
    return f'{v:,.0f}'


@login_required
def laporan_persediaan(request: HttpRequest) -> HttpResponse:
    """Comprehensive inventory report page."""
    from apps.entitas_bisnis.models import EntitasBisnis
    from apps.purchase.models import FIFOBatch, ItemMasterPurchase
    from apps.sales.models import SalesItemFIFOAllocation

    today = timezone.now().date()

    # ── Filters ──────────────────────────────────────────────────────────
    eb_filter = request.GET.get('entitas_bisnis', '')
    tipe_filter = request.GET.get('tipe', '')
    tanggal_dari = request.GET.get('tanggal_dari', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')

    eb_list = EntitasBisnis.objects.filter(status_aktif=True).order_by('nama')

    # Base queryset for inventory records
    ir_qs = (
        InventoryRecord.objects
        .filter(item__tipe_item__in=INVENTORY_TIPE_ITEMS)
        .select_related('item', 'entitas_bisnis', 'purchase_item__purchase_eb__purchase_header')
    )
    if eb_filter:
        ir_qs = ir_qs.filter(entitas_bisnis_id=eb_filter)
    if tipe_filter:
        ir_qs = ir_qs.filter(item__tipe_item=tipe_filter)
    if tanggal_dari:
        ir_qs = ir_qs.filter(tanggal__gte=tanggal_dari)
    if tanggal_sampai:
        ir_qs = ir_qs.filter(tanggal__lte=tanggal_sampai)

    records = list(ir_qs)

    # ── Compute as_of_date for aging/slow-dead/expiry (default: today) ──
    from datetime import date as _isodate
    try:
        as_of_date = _isodate.fromisoformat(tanggal_sampai) if tanggal_sampai else today
    except ValueError:
        as_of_date = today

    # FIFOBatch base queryset (active)
    fb_qs = FIFOBatch.objects.filter(
        remaining_qty__gt=0, item__tipe_item__in=INVENTORY_TIPE_ITEMS,
    ).select_related('item', 'purchase_item__purchase_eb__purchase_header')

    # Sales allocations for outflows
    alloc_qs = SalesItemFIFOAllocation.objects.filter(
        inventory_record__item__tipe_item__in=INVENTORY_TIPE_ITEMS,
    ).select_related(
        'sales_item__sales_eb__sales_header',
        'sales_item__sales_eb__entitas_bisnis',
        'sales_item__item',
        'inventory_record',
    )
    if eb_filter:
        alloc_qs = alloc_qs.filter(
            sales_item__sales_eb__entitas_bisnis_id=eb_filter,
        )

    # Pre-build allocation lookup by inventory record (used throughout)
    alloc_by_ir: dict = defaultdict(list)
    for _a in alloc_qs:
        alloc_by_ir[_a.inventory_record_id].append(_a)

    # ── ① Metric Cards ──────────────────────────────────────────────────
    total_nilai = sum(r.total_value for r in records)
    # total_masuk = ORIGINAL purchase value (not current remaining; items consumed show 0 otherwise)
    total_masuk = sum(_original_masuk_value(r, alloc_by_ir) for r in records)
    total_keluar = sum(a.cogs_amount for a in alloc_qs)
    active_items = len({r.item_id for r in records if r.total_value > 0})

    # Expiry alerts (evaluated as of as_of_date)
    expiry_soon = []  # items expiring within 30 days
    expiry_past = []  # already expired
    for r in records:
        if r.tanggal_kadaluarsa:
            days_left = (r.tanggal_kadaluarsa - as_of_date).days
            if days_left < 0:
                expiry_past.append((r, days_left))
            elif days_left <= 7:
                expiry_soon.append((r, days_left, 'danger'))
            elif days_left <= 30:
                expiry_soon.append((r, days_left, 'warning'))

    # Reorder Point alerts (evaluated as of as_of_date)
    rop_alerts = []
    for r in records:
        if r.lead_time_days and r.total_value > 0:
            daily_usage = _calc_daily_usage(r, alloc_qs)
            rop = Decimal(r.lead_time_days) * daily_usage
            if rop > 0:
                is_bulk = r.item.tipe_item in BULK_TYPES
                current = r.total_value if is_bulk else r.quantity
                if current < rop:
                    rop_alerts.append((r, current, rop))

    items_need_action = len(expiry_past) + len([e for e in expiry_soon if e[2] == 'danger']) + len(rop_alerts)

    # Slow/Dead stock (evaluated as of as_of_date)
    slow_dead = _calc_slow_dead(records, alloc_qs, as_of_date)
    dead_count = sum(1 for s in slow_dead if s['status'] == 'dead')
    slow_count = sum(1 for s in slow_dead if s['status'] == 'slow')

    metrics = {
        'total_nilai': total_nilai,
        'total_masuk': total_masuk,
        'total_keluar': total_keluar,
        'active_items': active_items,
        'items_need_action': items_need_action,
        'slow_dead_count': dead_count + slow_count,
        'masuk_count': len(records),
        'keluar_count': alloc_qs.count(),
    }

    # ── ② Chart data ────────────────────────────────────────────────────
    # Proportion by item type
    tipe_values = defaultdict(Decimal)
    for r in records:
        base_tipe = r.item.tipe_item.replace('B', '')  # RMB→RM, FGB→FG
        if base_tipe not in ('RM', 'FG', 'ITM'):
            base_tipe = r.item.tipe_item[:2] if len(r.item.tipe_item) >= 2 else r.item.tipe_item
        tipe_values[base_tipe] += r.total_value
    pie_labels = []
    pie_data = []
    for t in ('RM', 'FG', 'ITM'):
        if tipe_values.get(t, 0) > 0:
            pie_labels.append(_TIPE_LABEL.get(t, t))
            pie_data.append(float(tipe_values[t]))

    # Monthly trend (last 6 months)
    months_labels = []
    trend_data = []
    for i in range(5, -1, -1):
        dt = today.replace(day=1) - timedelta(days=i * 30)
        month_start = dt.replace(day=1)
        month_label = month_start.strftime('%b')
        months_labels.append(month_label)
        # Records that existed by end of that month
        if i == 0:
            month_end = today
        else:
            next_month = month_start.replace(day=28) + timedelta(days=4)
            month_end = next_month.replace(day=1) - timedelta(days=1)
        val = sum(
            r.total_value for r in records
            if r.tanggal <= month_end
        )
        trend_data.append(float(val))

    avg_trend = sum(trend_data) / len(trend_data) if trend_data else 0

    # Monthly mutations (masuk vs keluar)
    masuk_monthly = defaultdict(float)
    keluar_monthly = defaultdict(float)
    for r in records:
        key = r.tanggal.strftime('%b')
        if key in months_labels:
            masuk_monthly[key] += float(_original_masuk_value(r, alloc_by_ir))
    for a in alloc_qs:
        key = a.sales_item.sales_eb.sales_header.tanggal.strftime('%b')
        if key in months_labels:
            keluar_monthly[key] += float(a.cogs_amount)
    bar_masuk = [masuk_monthly.get(m, 0) for m in months_labels]
    bar_keluar = [keluar_monthly.get(m, 0) for m in months_labels]

    # Top items by value
    item_values = defaultdict(lambda: {'nama': '', 'value': Decimal(0)})
    for r in records:
        item_values[r.item_id]['nama'] = r.item.nama
        item_values[r.item_id]['value'] += r.total_value
    top_items = sorted(item_values.values(), key=lambda x: x['value'], reverse=True)[:5]
    top_labels = [t['nama'][:20] for t in top_items]
    top_data = [float(t['value']) for t in top_items]

    # Waterfall — saldo berjalan
    waterfall = _calc_waterfall(records, alloc_qs)

    chart_data = {
        'pie_labels': pie_labels,
        'pie_data': pie_data,
        'months_labels': months_labels,
        'trend_data': trend_data,
        'avg_trend': avg_trend,
        'bar_masuk': bar_masuk,
        'bar_keluar': bar_keluar,
        'top_labels': top_labels,
        'top_data': top_data,
    }

    # ── ④ Per-entity breakdown ───────────────────────────────────────────
    eb_breakdown = []
    eb_groups = defaultdict(list)
    for r in records:
        eb_groups[r.entitas_bisnis_id].append(r)
    for eb_id, eb_records in eb_groups.items():
        eb = eb_records[0].entitas_bisnis
        eb_nilai = sum(r.total_value for r in eb_records)
        eb_items = len({r.item_id for r in eb_records})
        eb_masuk = eb_nilai
        eb_dead = sum(1 for s in slow_dead if s['record'].entitas_bisnis_id == eb_id and s['status'] == 'dead')
        eb_exp = sum(
            1 for r in eb_records
            if r.tanggal_kadaluarsa and (r.tanggal_kadaluarsa - today).days < 0
        )
        eb_rop = sum(1 for ra in rop_alerts if ra[0].entitas_bisnis_id == eb_id)
        eb_breakdown.append({
            'nama': eb.nama,
            'nilai': eb_nilai,
            'item_count': eb_items,
            'masuk': eb_masuk,
            'dead_count': eb_dead,
            'exp_count': eb_exp,
            'rop_count': eb_rop,
        })

    # ── ⑦ Aging analysis ────────────────────────────────────────────────
    aging_buckets = _calc_aging(records, as_of_date)

    # ── ⑧ Cost accounting ───────────────────────────────────────────────
    fifo_batches = list(fb_qs)
    total_fifo_value = sum(b.batch_value for b in fifo_batches)
    active_layers = len(fifo_batches)
    eoq_ready = sum(
        1 for r in records
        if r.ordering_cost and r.holding_cost_pct and r.total_value > 0
    )
    total_items_with_value = len([r for r in records if r.total_value > 0])
    dead_value = sum(
        Decimal(str(s['record'].total_value))
        for s in slow_dead if s['status'] == 'dead'
    )

    cost_data = {
        'total_fifo_value': total_fifo_value,
        'active_layers': active_layers,
        'eoq_ready': eoq_ready,
        'total_items': total_items_with_value,
        'dead_value': dead_value,
        'total_keluar': total_keluar,
    }

    # ── ⑨ Posisi Persediaan ─────────────────────────────────────────────
    positions = _calc_positions(records, alloc_qs, as_of_date)

    # ── ⑩ Mutasi Persediaan ─────────────────────────────────────────────
    mutasi_list = _build_mutasi(records, alloc_qs, alloc_by_ir)

    # ── ⑪ FIFO Layers ───────────────────────────────────────────────────
    fifo_layers = []
    for i, batch in enumerate(fifo_batches, 1):
        inv_record = None
        for r in records:
            if r.item_id == batch.item_id:
                inv_record = r
                break
        status_label = 'Aktif'
        status_class = 'success'
        # Check if dead stock
        for s in slow_dead:
            if s['record'].item_id == batch.item_id and s['status'] == 'dead':
                status_label = 'Aktif · dead stock'
                status_class = 'danger'
                break
        # Check if expired
        if inv_record and inv_record.tanggal_kadaluarsa:
            if inv_record.tanggal_kadaluarsa < today:
                status_label = 'Aktif · kadaluarsa'
                status_class = 'danger'
            elif (inv_record.tanggal_kadaluarsa - today).days <= 7:
                status_label = 'Aktif · segera kadaluarsa'
                status_class = 'danger'
        # Check if below ROP
        for ra in rop_alerts:
            if ra[0].item_id == batch.item_id:
                if status_label == 'Aktif':
                    status_label = 'Aktif · ROP alert'
                    status_class = 'warning'
                break

        ref = ''
        if batch.purchase_item:
            ref = batch.purchase_item.purchase_eb.purchase_header.transaction_id

        fifo_layers.append({
            'num': i,
            'inventory_number': inv_record.inventory_number if inv_record else f'{batch.item.item_id}',
            'ref': ref,
            'tanggal': batch.tanggal,
            'value': batch.batch_value,
            'status_label': status_label,
            'status_class': status_class,
        })

    # ── ⑱ Inventory Turnover ────────────────────────────────────────────
    turnover_data = _calc_inventory_turnover(records, alloc_qs, alloc_by_ir)

    context = {
        'metrics': metrics,
        'chart_data_json': json.dumps(chart_data, default=str),
        'waterfall': waterfall,
        'eb_breakdown': eb_breakdown,
        'expiry_past': expiry_past,
        'expiry_soon': expiry_soon,
        'rop_alerts': rop_alerts,
        'slow_dead': slow_dead,
        'aging_buckets': aging_buckets,
        'aging_json': json.dumps(_aging_chart_data(aging_buckets), default=str),
        'cost_data': cost_data,
        'positions': positions,
        'mutasi_list': mutasi_list,
        'fifo_layers': fifo_layers,
        'turnover_data': turnover_data,
        'today': today,
        # Filters
        'eb_list': eb_list,
        'eb_filter': eb_filter,
        'tipe_filter': tipe_filter,
        'tanggal_dari': tanggal_dari,
        'tanggal_sampai': tanggal_sampai,
        'tipe_choices': [
            ('RM', 'Raw Material'),
            ('FG', 'Finished Good'),
            ('ITM', 'Item Lainnya'),
            ('RMB', 'Raw Material (Bulk)'),
            ('FGB', 'Finished Good (Bulk)'),
            ('ITMB', 'Item Lainnya (Bulk)'),
        ],
    }
    return render(request, 'inventory/laporan_persediaan.html', context)


# ── Helper functions ─────────────────────────────────────────────────────────

def _calc_daily_usage(record, alloc_qs) -> Decimal:
    """Estimate daily usage from sales allocations for this item."""
    item_allocs = [a for a in alloc_qs if a.inventory_record_id == record.pk]
    if not item_allocs:
        return Decimal(0)
    is_bulk = record.item.tipe_item in BULK_TYPES
    total_consumed = sum(a.cogs_amount if is_bulk else a.quantity_consumed for a in item_allocs)
    dates = [a.sales_item.sales_eb.sales_header.tanggal for a in item_allocs]
    if not dates:
        return Decimal(0)
    span = (max(dates) - min(dates)).days or 1
    return total_consumed / Decimal(span)


def _calc_slow_dead(records, alloc_qs, as_of_date):
    """Calculate slow-moving and dead stock items as of as_of_date."""
    result = []
    # Group allocations by inventory record
    alloc_by_ir = defaultdict(list)
    for a in alloc_qs:
        alloc_by_ir[a.inventory_record_id].append(a)

    for r in records:
        if r.total_value <= 0:
            continue
        ir_allocs = alloc_by_ir.get(r.pk, [])
        # Only consider allocations on or before as_of_date
        relevant_allocs = [
            a for a in ir_allocs
            if a.sales_item.sales_eb.sales_header.tanggal <= as_of_date
        ]
        if relevant_allocs:
            last_sale_date = max(
                a.sales_item.sales_eb.sales_header.tanggal for a in relevant_allocs
            )
            days_idle = (as_of_date - last_sale_date).days
        else:
            days_idle = (as_of_date - r.tanggal).days

        threshold = r.item.threshold_days_outstanding or 30
        if days_idle > threshold:
            status = 'dead'
            recommendation = 'Evaluasi kebutuhan · pertimbangkan retur ke supplier'
        elif days_idle > (threshold // 2):
            status = 'slow'
            recommendation = 'Pantau pergerakan stok'
        elif days_idle <= 1:
            status = 'new'
            recommendation = f'Pantau {threshold} hari ke depan'
        else:
            continue

        last_out_label = last_sale_date.strftime('%d %b %Y') if ir_allocs else 'Belum pernah'

        result.append({
            'record': r,
            'value': r.total_value,
            'last_out': last_out_label,
            'days_idle': days_idle,
            'status': status,
            'recommendation': recommendation,
        })
    return result


def _calc_waterfall(records, alloc_qs):
    """Build waterfall data: saldo awal, each inflow, each outflow, saldo akhir."""
    # Group by purchase header reference
    inflows = defaultdict(Decimal)
    for r in records:
        if r.purchase_item:
            ref = r.purchase_item.purchase_eb.purchase_header.transaction_id
        else:
            ref = f'SA-{r.tanggal}'
        inflows[ref] += r.total_value

    outflows = defaultdict(Decimal)
    for a in alloc_qs:
        ref = a.sales_item.sales_eb.sales_header.transaction_id
        outflows[ref] += a.cogs_amount

    total_in = sum(inflows.values())
    total_out = sum(outflows.values())
    saldo_akhir = total_in - total_out

    items = [{'label': 'Saldo awal', 'value': Decimal(0), 'type': 'neutral'}]

    # Top 3 inflows by value
    sorted_in = sorted(inflows.items(), key=lambda x: x[1], reverse=True)[:3]
    rest_in = sum(v for _, v in sorted(inflows.items(), key=lambda x: x[1], reverse=True)[3:])
    for ref, val in sorted_in:
        items.append({'label': f'+ {ref}', 'value': val, 'type': 'in'})
    if rest_in > 0:
        items.append({'label': f'+ lainnya ({len(inflows) - 3})', 'value': rest_in, 'type': 'in'})

    # Top 3 outflows
    sorted_out = sorted(outflows.items(), key=lambda x: x[1], reverse=True)[:3]
    rest_out = sum(v for _, v in sorted(outflows.items(), key=lambda x: x[1], reverse=True)[3:])
    for ref, val in sorted_out:
        items.append({'label': f'− {ref}', 'value': val, 'type': 'out'})
    if rest_out > 0:
        items.append({'label': f'− lainnya ({len(outflows) - 3})', 'value': rest_out, 'type': 'out'})

    items.append({'label': 'Saldo akhir', 'value': saldo_akhir, 'type': 'total'})

    # Calculate bidirectional bar positions (center = 50%)
    all_abs = [abs(float(i['value'])) for i in items if i['value'] != 0]
    max_val = float(max(all_abs, default=1)) or 1
    for item in items:
        v = float(item['value'])
        half_pct = int(abs(v) / max_val * 48)  # max 48% each side, leaves 2% gap
        if item['type'] == 'in':
            item['bar_left'] = 50
            item['bar_width'] = half_pct
        elif item['type'] == 'out':
            item['bar_left'] = 50 - half_pct
            item['bar_width'] = half_pct
        elif item['type'] == 'total':
            if v >= 0:
                item['bar_left'] = 50
                item['bar_width'] = half_pct
            else:
                item['bar_left'] = 50 - half_pct
                item['bar_width'] = half_pct
        else:  # neutral
            item['bar_left'] = 50
            item['bar_width'] = 0
        item['display'] = _format_rp(v)

    return items


def _original_masuk_value(record, alloc_by_ir: dict) -> Decimal:
    """Compute original intake value of a record (not current remaining)."""
    ir_allocs = alloc_by_ir.get(record.pk, [])
    consumed_cogs = sum(a.cogs_amount for a in ir_allocs)
    return record.total_value + consumed_cogs


def _calc_aging(records, as_of_date):
    """Calculate aging buckets as of as_of_date: 0-30, 31-60, 61-90, >90 days."""
    buckets = [
        {'label': '0–30 hari', 'value': Decimal(0), 'count': 0, 'color': '#10b981'},
        {'label': '31–60 hari', 'value': Decimal(0), 'count': 0, 'color': '#0054a6'},
        {'label': '61–90 hari', 'value': Decimal(0), 'count': 0, 'color': '#f59e0b'},
        {'label': '> 90 hari', 'value': Decimal(0), 'count': 0, 'color': '#ef4444'},
    ]
    for r in records:
        val = r.total_value
        if val <= 0:
            continue
        age = (as_of_date - r.tanggal).days
        if age <= 30:
            idx = 0
        elif age <= 60:
            idx = 1
        elif age <= 90:
            idx = 2
        else:
            idx = 3
        buckets[idx]['value'] += val
        buckets[idx]['count'] += 1

    total = sum(b['value'] for b in buckets) or Decimal(1)
    for b in buckets:
        b['pct'] = int(float(b['value']) / float(total) * 100)
        b['display'] = _format_rp(b['value'])
    return buckets


def _aging_chart_data(buckets):
    return {
        'labels': [b['label'] for b in buckets],
        'data': [float(b['value']) for b in buckets],
        'colors': [b['color'] for b in buckets],
    }


def _calc_positions(records, alloc_qs, as_of_date):
    """Build position table rows."""
    positions = []
    for r in records:
        if r.total_value <= 0:
            continue
        is_bulk = r.item.tipe_item in BULK_TYPES
        # Calculate EOQ & Reorder Point
        daily_usage = _calc_daily_usage(r, alloc_qs)
        eoq = None
        rop = None
        if r.ordering_cost and r.holding_cost_pct and daily_usage > 0:
            annual_demand = daily_usage * 365
            h = float(r.unit_price) * float(r.holding_cost_pct)
            if h > 0:
                eoq = math.sqrt(2 * float(annual_demand) * float(r.ordering_cost) / h)
                eoq = round(eoq)
        if r.lead_time_days and daily_usage > 0:
            rop = float(r.lead_time_days) * float(daily_usage)
            rop = round(rop)

        current_qty = r.total_value if is_bulk else r.quantity

        # Expiry info (relative to as_of_date)
        exp_info = None
        if r.tanggal_kadaluarsa:
            days_left = (r.tanggal_kadaluarsa - as_of_date).days
            if days_left < 0:
                exp_info = {'date': r.tanggal_kadaluarsa, 'label': 'Sudah kadaluarsa', 'class': 'danger'}
            elif days_left <= 7:
                exp_info = {'date': r.tanggal_kadaluarsa, 'label': f'{days_left} hari lagi', 'class': 'danger'}
            elif days_left <= 30:
                exp_info = {'date': r.tanggal_kadaluarsa, 'label': f'{days_left} hari lagi', 'class': 'warning'}
            else:
                exp_info = {'date': r.tanggal_kadaluarsa, 'label': '', 'class': ''}

        # Below Reorder Point?
        below_rop = rop is not None and float(current_qty) < rop

        # Dead stock days
        dead_days = None
        for s in _quick_idle_days(r, alloc_qs, as_of_date):
            dead_days = s
            break

        positions.append({
            'record': r,
            'tipe_label': _TIPE_LABEL.get(r.item.tipe_item, r.item.tipe_item),
            'is_bulk': is_bulk,
            'value': r.total_value,
            'lead_time': r.lead_time_days,
            'ordering_cost': r.ordering_cost,
            'holding_pct': r.holding_cost_pct,
            'moq': r.moq,
            'eoq': eoq,
            'rop': rop,
            'below_rop': below_rop,
            'exp_info': exp_info,
            'metode': r.get_metode_alokasi_display() or 'FIFO',
            'dead_days': dead_days,
        })
    return positions


def _quick_idle_days(record, alloc_qs, as_of_date):
    """Yield idle days for a single record (generator for lazy eval)."""
    ir_allocs = [a for a in alloc_qs if a.inventory_record_id == record.pk]
    relevant = [
        a for a in ir_allocs
        if a.sales_item.sales_eb.sales_header.tanggal <= as_of_date
    ]
    if relevant:
        last_date = max(a.sales_item.sales_eb.sales_header.tanggal for a in relevant)
        yield (as_of_date - last_date).days
    else:
        yield (as_of_date - record.tanggal).days


def _build_mutasi(records, alloc_qs, alloc_by_ir: dict):
    """Build combined mutation list (masuk + keluar).
    Masuk rows show ORIGINAL purchase value (not current remaining).
    """
    mutasi = []
    for r in records:
        ref = ''
        if r.purchase_item:
            ref = r.purchase_item.purchase_eb.purchase_header.transaction_id
        is_bulk = r.item.tipe_item in BULK_TYPES
        original_value = _original_masuk_value(r, alloc_by_ir)
        mutasi.append({
            'tanggal': r.tanggal,
            'inventory_number': r.inventory_number,
            'item_name': f'{r.item.item_id} - {r.item.nama}',
            'tipe': 'in',
            'tipe_label': 'Masuk',
            'ref': ref,
            'qty': None if is_bulk else r.quantity,
            'value': original_value,
            'keterangan': f'Pembelian via {ref}' if ref else 'Saldo awal',
        })
    for a in alloc_qs:
        sh = a.sales_item.sales_eb.sales_header
        ir = a.inventory_record
        is_bulk = a.sales_item.item.tipe_item in BULK_TYPES
        mutasi.append({
            'tanggal': sh.tanggal,
            'inventory_number': ir.inventory_number,
            'item_name': f'{a.sales_item.item.item_id} - {a.sales_item.item.nama}',
            'tipe': 'out',
            'tipe_label': 'Keluar',
            'ref': sh.transaction_id,
            'qty': None if is_bulk else a.quantity_consumed,
            'value': a.cogs_amount,
            'keterangan': f'Penjualan via {sh.transaction_id}',
        })
    mutasi.sort(key=lambda x: x['tanggal'], reverse=True)
    return mutasi


def _calc_inventory_turnover(records, alloc_qs, alloc_by_ir: dict) -> dict:
    """Calculate inventory turnover metrics per item and per kategori.

    Turnover ratio = COGS (sold value) / original inventory value.
    Top-10 returned for: by sales value, by mutation count, per kategori.
    """
    item_data: dict = {}
    for r in records:
        iid = r.item_id
        if iid not in item_data:
            kat = r.item.kategori.nama if (hasattr(r.item, 'kategori') and r.item.kategori) else '—'
            item_data[iid] = {
                'item_id_str': r.item.item_id,
                'nama': r.item.nama,
                'tipe_label': _TIPE_LABEL.get(r.item.tipe_item, r.item.tipe_item),
                'kategori': kat,
                'orig_value': Decimal(0),
                'cogs': Decimal(0),
                'sell_count': 0,
                'mutation_count': 0,
            }
        item_data[iid]['orig_value'] += _original_masuk_value(r, alloc_by_ir)

    for a in alloc_qs:
        iid = a.sales_item.item_id
        if iid in item_data:
            item_data[iid]['cogs'] += a.cogs_amount
            item_data[iid]['sell_count'] += 1
        item_data[iid]['mutation_count'] = item_data.get(iid, {}).get('mutation_count', 0) + 1

    for d in item_data.values():
        if d['orig_value'] > 0:
            d['turnover_ratio'] = round(float(d['cogs']) / float(d['orig_value']), 2)
        else:
            d['turnover_ratio'] = 0.0

    items_list = list(item_data.values())
    top_by_sales = sorted(
        [d for d in items_list if d['cogs'] > 0],
        key=lambda x: x['cogs'], reverse=True,
    )[:10]
    top_by_count = sorted(
        [d for d in items_list if d['sell_count'] > 0],
        key=lambda x: x['sell_count'], reverse=True,
    )[:10]

    # Per kategori
    kat_data: dict = {}
    for d in items_list:
        k = d['kategori']
        if k not in kat_data:
            kat_data[k] = {'kategori': k, 'cogs': Decimal(0), 'orig_value': Decimal(0), 'item_count': 0, 'sell_count': 0}
        kat_data[k]['cogs'] += d['cogs']
        kat_data[k]['orig_value'] += d['orig_value']
        kat_data[k]['item_count'] += 1
        kat_data[k]['sell_count'] += d['sell_count']
    for kd in kat_data.values():
        if kd['orig_value'] > 0:
            kd['turnover_ratio'] = round(float(kd['cogs']) / float(kd['orig_value']), 2)
        else:
            kd['turnover_ratio'] = 0.0
    top_by_kategori = sorted(kat_data.values(), key=lambda x: x['cogs'], reverse=True)[:10]

    return {
        'top_by_sales': top_by_sales,
        'top_by_count': top_by_count,
        'top_by_kategori': top_by_kategori,
    }


# ── Export views ─────────────────────────────────────────────────────────────

def _inventory_export_qs(request):
    """Return filtered InventoryRecord queryset based on GET params."""
    qs = InventoryRecord.objects.select_related('item', 'entitas_bisnis').order_by('-tanggal', '-created_at')
    tanggal_dari = request.GET.get('tanggal_dari', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    item_filter = request.GET.get('item', '')
    eb_filter = request.GET.get('entitas_bisnis', '')
    if tanggal_dari:
        qs = qs.filter(tanggal__gte=tanggal_dari)
    if tanggal_sampai:
        qs = qs.filter(tanggal__lte=tanggal_sampai)
    if item_filter:
        qs = qs.filter(item_id=item_filter)
    if eb_filter:
        qs = qs.filter(entitas_bisnis_id=eb_filter)
    return qs


@login_required
def inventory_export(request: HttpRequest) -> HttpResponse:
    """Export inventory list as XLSX with same filters as list page."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    records = list(_inventory_export_qs(request))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventory'

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin = Side(style='thin')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right_align = Alignment(horizontal='right')

    headers = ['No. Inventory', 'Item ID', 'Item Nama', 'Tipe', 'Entitas Bisnis',
               'Tanggal', 'Qty', 'Harga/Unit (Rp)', 'Total Nilai (Rp)', 'Metode']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')

    for row_num, r in enumerate(records, 2):
        vals = [
            r.inventory_number,
            r.item.item_id,
            r.item.nama,
            r.item.tipe_item,
            r.entitas_bisnis.nama,
            str(r.tanggal),
            float(r.quantity or 0),
            float(r.unit_price or 0),
            float(r.total_value or 0),
            r.get_metode_alokasi_display() or 'FIFO',
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.border = thin_border
            if col in (7, 8, 9):
                c.alignment = right_align
                c.number_format = '#,##0'

    col_widths = [20, 14, 30, 16, 28, 13, 10, 18, 18, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="inventory.xlsx"'
    wb.save(response)
    return response


@login_required
def inventory_export_pdf(request: HttpRequest) -> HttpResponse:
    """Render print-friendly inventory list for browser PDF printing."""
    records = list(_inventory_export_qs(request))
    total_nilai = sum(r.total_value for r in records)
    return render(request, 'inventory/inventory_export_pdf.html', {
        'records': records,
        'tanggal_dari': request.GET.get('tanggal_dari', ''),
        'tanggal_sampai': request.GET.get('tanggal_sampai', ''),
        'generated_at': timezone.now(),
        'total_nilai': total_nilai,
        'total_records': len(records),
    })
