"""Aset Lainnya views — CRUD for AsetLainnyaRecord."""
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django_ratelimit.decorators import ratelimit

from naveda_integra.ratelimit_utils import rate_from
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from apps.entitas_bisnis.models import EntitasBisnis
from apps.purchase.models import ItemMasterPurchase
from apps.purchase.views import _get_eb_tree, _resolve_eb_lv1_ids

from .forms import AsetLainnyaRecordForm
from .models import AsetLainnyaRecord
from .services import calculate_amortization, process_amortization

DEFAULT_DAYS = 30  # monthly processing default


@login_required
def aset_lainnya_list(request: HttpRequest) -> HttpResponse:
    """List all other asset records with optional filters."""
    qs = AsetLainnyaRecord.objects.select_related('item', 'entitas_bisnis').all()

    tanggal_dari = request.GET.get('tanggal_dari', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    item_filter = request.GET.get('item', '')
    eb_filter_list = [v for v in request.GET.getlist('entitas_bisnis') if v]

    if tanggal_dari:
        qs = qs.filter(tanggal_perolehan__gte=tanggal_dari)
    if tanggal_sampai:
        qs = qs.filter(tanggal_perolehan__lte=tanggal_sampai)
    if item_filter:
        qs = qs.filter(item_id=item_filter)
    if eb_filter_list:
        qs = qs.filter(entitas_bisnis_id__in=_resolve_eb_lv1_ids(eb_filter_list, request.user))

    return render(request, 'aset_lainnya/aset_lainnya_list.html', {
        'records': qs,
        'items': ItemMasterPurchase.objects.filter(tipe_item='ALL').order_by('item_id'),
        'eb_tree': _get_eb_tree(request.user),
        'tanggal_dari': tanggal_dari,
        'tanggal_sampai': tanggal_sampai,
        'item_filter': item_filter,
        'eb_filter_list': eb_filter_list,
    })


@login_required
def aset_lainnya_detail(request: HttpRequest, pk: int) -> HttpResponse:
    """Show other asset record detail with amortization preview."""
    from decimal import Decimal
    record = get_object_or_404(
        AsetLainnyaRecord.objects.select_related('item', 'entitas_bisnis', 'purchase_item'),
        pk=pk,
    )

    metode = record.metode_amortisasi or (record.item.metode_amortisasi if record.item else '')
    needs_activity_input = metode in ('units_of_production', 'revenue_based')

    preview_amount = Decimal('0')
    if not needs_activity_input:
        preview_amount = calculate_amortization(record, days=DEFAULT_DAYS)

    from apps.jurnal.models import JurnalHeader
    amort_journals = JurnalHeader.objects.filter(
        uraian_transaksi__startswith=f'Amortisasi {record.aset_number}',
    ).order_by('-tanggal', '-pk')

    return render(request, 'aset_lainnya/aset_lainnya_detail.html', {
        'record': record,
        'preview_amount': preview_amount,
        'needs_activity_input': needs_activity_input,
        'metode': metode,
        'can_amortize': record.nilai_buku > record.nilai_residu,
        'amort_journals': amort_journals,
        'latest_amort_journal_pk': amort_journals[0].pk if amort_journals else None,
    })


@login_required
def delete_amortization_journal(request: HttpRequest, pk: int, jurnal_pk: int) -> HttpResponse:
    """Delete a single amortization journal for an asset, reversing its accumulated amortization.

    Only the latest (most recent) amortization journal may be deleted.
    """
    from decimal import Decimal
    from apps.jurnal.models import JurnalHeader, JurnalDetail
    from apps.jurnal.utils import log_jurnal_terhapus
    from django.db import transaction as db_transaction
    from django.db.models import Sum as _Sum

    record = get_object_or_404(
        AsetLainnyaRecord.objects.select_related('item', 'entitas_bisnis'),
        pk=pk,
    )
    journal = get_object_or_404(
        JurnalHeader,
        pk=jurnal_pk,
        uraian_transaksi__startswith=f'Amortisasi {record.aset_number}',
    )

    # Validate that this is the latest amortization journal for this asset
    latest_journal = (
        JurnalHeader.objects
        .filter(uraian_transaksi__startswith=f'Amortisasi {record.aset_number}')
        .order_by('-tanggal', '-pk')
        .first()
    )
    if not latest_journal or latest_journal.pk != journal.pk:
        messages.error(request, 'Hanya jurnal amortisasi terbaru yang dapat dihapus.')
        return redirect('aset_lainnya:detail', pk=pk)

    # Determine amortization amount from beban amortisasi debit entries
    agg = (
        JurnalDetail.objects
        .filter(jurnal_header=journal, akun__kode_akun__startswith='5.1.31')
        .aggregate(total=_Sum('debit'))
    )
    amort_amount = agg['total'] or Decimal('0')

    if request.method == 'POST':
        with db_transaction.atomic():
            log_jurnal_terhapus(journal, 'aset_lainnya', request)
            journal.details.all().delete()
            journal.delete()
            record.akumulasi_amortisasi = max(
                Decimal('0'),
                record.akumulasi_amortisasi - amort_amount,
            )
            record.save(update_fields=['akumulasi_amortisasi'])
        messages.success(
            request,
            f'Jurnal {journal.nomor_transaksi} dihapus. '
            f'Akumulasi amortisasi dikurangi Rp {amort_amount:,.0f}.',
        )
        return redirect('aset_lainnya:detail', pk=pk)

    return render(request, 'aset_lainnya/delete_amort_journal_confirm.html', {
        'record': record,
        'journal': journal,
        'amort_amount': amort_amount,
    })


@login_required
def aset_lainnya_create(request: HttpRequest) -> HttpResponse:
    """Create a new other asset record."""
    if request.method == 'POST':
        form = AsetLainnyaRecordForm(request.POST)
        if form.is_valid():
            record = form.save()
            messages.success(request, f'Aset lainnya {record.aset_number} berhasil dibuat.')
            return redirect('aset_lainnya:detail', pk=record.pk)
    else:
        form = AsetLainnyaRecordForm()
    return render(request, 'aset_lainnya/aset_lainnya_form.html', {'form': form, 'title': 'Tambah Aset Lainnya'})


@login_required
def aset_lainnya_update(request: HttpRequest, pk: int) -> HttpResponse:
    """Edit an existing other asset record."""
    record = get_object_or_404(AsetLainnyaRecord, pk=pk)
    if request.method == 'POST':
        form = AsetLainnyaRecordForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            messages.success(request, f'Aset lainnya {record.aset_number} berhasil diperbarui.')
            return redirect('aset_lainnya:detail', pk=record.pk)
    else:
        form = AsetLainnyaRecordForm(instance=record)
    return render(request, 'aset_lainnya/aset_lainnya_form.html', {
        'form': form,
        'record': record,
        'title': f'Edit {record.aset_number}',
    })


@login_required
def aset_lainnya_delete(request: HttpRequest, pk: int) -> HttpResponse:
    """Delete an other asset record, removing all associated amortization journals."""
    from apps.jurnal.models import JurnalHeader
    from apps.jurnal.utils import log_jurnal_terhapus

    record = get_object_or_404(AsetLainnyaRecord.objects.select_related('item', 'entitas_bisnis'), pk=pk)

    # Find all amortization journals for this asset
    amor_journals = list(
        JurnalHeader.objects.filter(
            uraian_transaksi__startswith=f'Amortisasi {record.aset_number}'
        ).order_by('tanggal')
    )

    if request.method == 'POST':
        from django.db import transaction as db_transaction
        number = record.aset_number
        with db_transaction.atomic():
            for journal in amor_journals:
                log_jurnal_terhapus(journal, 'aset_lainnya', request)
                journal.details.all().delete()
                journal.delete()
            record.delete()
        messages.success(request, f'Aset lainnya {number} dan {len(amor_journals)} jurnal amortisasi berhasil dihapus.')
        return redirect('aset_lainnya:list')

    return render(request, 'aset_lainnya/aset_lainnya_delete.html', {
        'record': record,
        'amor_journals': amor_journals,
    })


@login_required
def aset_lainnya_process_amortization(request: HttpRequest, pk: int) -> HttpResponse:
    """Process amortization for an other asset record, creating a journal entry."""
    from decimal import Decimal, InvalidOperation

    record = get_object_or_404(
        AsetLainnyaRecord.objects.select_related('item', 'entitas_bisnis'),
        pk=pk,
    )

    if request.method != 'POST':
        return redirect('aset_lainnya:detail', pk=pk)

    metode = record.metode_amortisasi or (record.item.metode_amortisasi if record.item else '')
    tanggal_str = request.POST.get('tanggal', '')

    try:
        from datetime import date as date_cls
        tanggal = date_cls.fromisoformat(tanggal_str) if tanggal_str else None
    except ValueError:
        tanggal = None

    if metode == 'units_of_production':
        try:
            unit_aktual = Decimal(request.POST.get('unit_aktual', '0'))
        except InvalidOperation:
            messages.error(request, 'Unit produksi aktual harus berupa angka.')
            return redirect('aset_lainnya:detail', pk=pk)
        amortization = calculate_amortization(record, unit_aktual=unit_aktual)
    elif metode == 'revenue_based':
        try:
            pendapatan = Decimal(request.POST.get('pendapatan_aktual', '0'))
        except InvalidOperation:
            messages.error(request, 'Pendapatan aktual harus berupa angka.')
            return redirect('aset_lainnya:detail', pk=pk)
        amortization = calculate_amortization(record, pendapatan_aktual=pendapatan)
    else:
        manual_amount = request.POST.get('amount', '')
        if manual_amount:
            try:
                amortization = Decimal(manual_amount)
            except InvalidOperation:
                messages.error(request, 'Jumlah amortisasi harus berupa angka.')
                return redirect('aset_lainnya:detail', pk=pk)
        else:
            try:
                hari = max(1, int(request.POST.get('hari', str(DEFAULT_DAYS))))
            except (ValueError, TypeError):
                hari = DEFAULT_DAYS
            amortization = calculate_amortization(record, days=hari)

    try:
        header = process_amortization(record, amortization, tanggal)
        messages.success(
            request,
            f'Amortisasi {record.aset_number} sebesar Rp {amortization:,.0f} berhasil diproses. '
            f'Jurnal: {header.nomor_transaksi}',
        )
    except ValueError as e:
        messages.error(request, str(e))

    return redirect('aset_lainnya:detail', pk=pk)


@login_required
def aset_lainnya_bulk_amortization(request: HttpRequest) -> HttpResponse:
    """Process amortization for all eligible other assets in one batch."""
    from decimal import Decimal
    from datetime import date as date_cls

    if request.method != 'POST':
        return redirect('aset_lainnya:list')

    tanggal_str = request.POST.get('tanggal', '')
    try:
        tanggal = date_cls.fromisoformat(tanggal_str) if tanggal_str else date_cls.today()
    except ValueError:
        tanggal = date_cls.today()

    try:
        hari = max(1, int(request.POST.get('hari', str(DEFAULT_DAYS))))
    except (ValueError, TypeError):
        hari = DEFAULT_DAYS

    records = AsetLainnyaRecord.objects.select_related('item', 'entitas_bisnis').all()
    success_count = 0
    skip_count = 0
    error_msgs = []

    for record in records:
        if record.nilai_buku <= record.nilai_residu:
            skip_count += 1
            continue

        metode = record.metode_amortisasi or (record.item.metode_amortisasi if record.item else '')
        # Skip activity-based methods — require manual input per record
        if metode in ('units_of_production', 'revenue_based'):
            skip_count += 1
            continue

        amortization = calculate_amortization(record, days=hari)
        if amortization <= 0:
            skip_count += 1
            continue

        try:
            process_amortization(record, amortization, tanggal)
            success_count += 1
        except ValueError as e:
            error_msgs.append(f'{record.aset_number}: {e}')

    if success_count:
        messages.success(
            request,
            f'Bulk amortisasi selesai: {success_count} aset diproses ({hari} hari), '
            f'{skip_count} dilewati.',
        )
    else:
        messages.warning(request, f'Tidak ada aset yang diproses. {skip_count} dilewati.')

    for msg in error_msgs[:5]:
        messages.error(request, msg)

    return redirect('aset_lainnya:list')


# ── Export views ─────────────────────────────────────────────────────────────

def _aset_lainnya_export_qs(request):
    """Return filtered AsetLainnyaRecord queryset based on GET params."""
    qs = AsetLainnyaRecord.objects.select_related(
        'item', 'item__kategori', 'entitas_bisnis',
    ).order_by('item__kategori__nama', '-tanggal_perolehan', '-created_at')
    tanggal_dari = request.GET.get('tanggal_dari', '')
    tanggal_sampai = request.GET.get('tanggal_sampai', '')
    item_filter = request.GET.get('item', '')
    eb_filter = request.GET.get('entitas_bisnis', '')
    if tanggal_dari:
        qs = qs.filter(tanggal_perolehan__gte=tanggal_dari)
    if tanggal_sampai:
        qs = qs.filter(tanggal_perolehan__lte=tanggal_sampai)
    if item_filter:
        qs = qs.filter(item_id=item_filter)
    if eb_filter:
        qs = qs.filter(entitas_bisnis_id=eb_filter)
    return qs


@login_required
@ratelimit(key='user', rate=rate_from('export'), method='GET', block=True)
def aset_lainnya_export(request: HttpRequest) -> HttpResponse:
    """Export aset lainnya list as XLSX with same filters as list page."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    records = list(_aset_lainnya_export_qs(request))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Aset Lainnya'

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin = Side(style='thin')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right_align = Alignment(horizontal='right')

    headers = [
        'No. Aset', 'Item', 'Entitas Bisnis', 'Tanggal Perolehan',
        'Qty', 'Harga Perolehan (Rp)', 'Total Nilai (Rp)',
        'Akum. Amortisasi (Rp)', 'Nilai Buku (Rp)', 'Metode Amortisasi',
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')

    for row_num, r in enumerate(records, 2):
        vals = [
            r.aset_number,
            f'{r.item.item_id} - {r.item.nama}',
            r.entitas_bisnis.nama,
            str(r.tanggal_perolehan),
            float(r.quantity or 1),
            float(r.harga_perolehan or 0),
            float(r.total_value or 0),
            float(r.akumulasi_amortisasi or 0),
            float(r.nilai_buku),
            r.metode_amortisasi or '',
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.border = thin_border
            if col in (5, 6, 7, 8, 9):
                c.alignment = right_align
                c.number_format = '#,##0'

    col_widths = [18, 34, 26, 16, 8, 20, 20, 22, 18, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="aset_lainnya.xlsx"'
    wb.save(response)
    return response


@login_required
@ratelimit(key='user', rate=rate_from('export'), method='GET', block=True)
def aset_lainnya_export_pdf(request: HttpRequest) -> HttpResponse:
    """Render print-friendly aset lainnya list for browser PDF printing."""
    import datetime
    records = list(_aset_lainnya_export_qs(request))
    total_nilai = sum(r.total_value for r in records)
    total_akum = sum(r.akumulasi_amortisasi for r in records)
    total_buku = sum(r.nilai_buku for r in records)
    return render(request, 'aset_lainnya/aset_lainnya_export_pdf.html', {
        'records': records,
        'tanggal_dari': request.GET.get('tanggal_dari', ''),
        'tanggal_sampai': request.GET.get('tanggal_sampai', ''),
        'generated_at': datetime.datetime.now(),
        'total_nilai': total_nilai,
        'total_akum': total_akum,
        'total_buku': total_buku,
        'total_records': len(records),
    })
